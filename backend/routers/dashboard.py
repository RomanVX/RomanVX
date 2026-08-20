import asyncio
import time as _time
from datetime import datetime, timedelta
from typing import Annotated, Optional

from fastapi import APIRouter, Query, HTTPException
import analytics
import cache
import catalog as _catalog
import cost_store
import ozon_client
import sales_history
import wb_client
import ym_client

# фоновые задачи вкладок (держим ссылки, иначе GC их убьёт)
_dbg_tasks: set = set()


def _dspawn(coro) -> None:
    import heavy
    t = asyncio.get_event_loop().create_task(heavy.guard(coro))
    _dbg_tasks.add(t)
    t.add_done_callback(_dbg_tasks.discard)

router = APIRouter(prefix="/api/dashboard", tags=["dashboard"])


@router.get("/sales_history")
async def sales_history_endpoint(
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
):
    """Накопленная история продаж (день × площадка) из БД — за пределами окон API."""
    return sales_history.get_summary(date_from=date_from, date_to=date_to)

_DATE_FROM = Query(description="YYYY-MM-DD. Overrides ?days.")
_DATE_TO   = Query(description="YYYY-MM-DD. Defaults to today.")
_DAYS      = Query(ge=1, le=365)
_BRAND     = Query()
_CATEGORY  = Query()


def _range(date_from, date_to, days) -> tuple[datetime, datetime]:
    dt_to = datetime.fromisoformat(date_to) if date_to else datetime.utcnow()
    dt_from = datetime.fromisoformat(date_from) if date_from else dt_to - timedelta(days=days)
    return dt_from, dt_to


async def _fetch(dt_from, dt_to, brand, category):
    try:
        sales, orders, stocks = await cache.get_raw_data(dt_from, dt_to)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"WB API error: {exc}")
    sales  = analytics.filter_records(sales,  brand, category)
    orders = analytics.filter_records(orders, brand, category)
    stocks = analytics.filter_records(stocks, brand, category)
    return sales, orders, stocks


def _days(dt_from, dt_to) -> int:
    return max(1, (dt_to - dt_from).days)


@router.get("/filters")
async def get_filters(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, _, stocks = await _fetch(dt_from, dt_to, None, None)
    return analytics.available_filters(stocks, sales)


@router.get("/finance")
async def get_finance(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
    brand:     Annotated[Optional[str], _BRAND]     = None,
    category:  Annotated[Optional[str], _CATEGORY]  = None,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    length = dt_to - dt_from
    sales, orders, _ = await _fetch(dt_from, dt_to, brand, category)
    p_sales, p_orders, _ = await _fetch(dt_from - length, dt_from, brand, category)
    cur  = analytics.finance_aggregate(sales, orders)
    prev = analytics.finance_aggregate(p_sales, p_orders)
    return {
        "cards":     analytics.finance_cards(cur, prev),
        "structure": analytics.revenue_structure(cur),
        "top_skus":  analytics.top_skus(sales, _days(dt_from, dt_to), 5),
        "aggregate": cur,
    }


@router.get("/sales-dynamics")
async def get_sales_dynamics(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
    brand:     Annotated[Optional[str], _BRAND]     = None,
    category:  Annotated[Optional[str], _CATEGORY]  = None,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, orders, _ = await _fetch(dt_from, dt_to, brand, category)
    return analytics.sales_dynamics(sales, orders)


@router.get("/products")
async def get_products(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
    brand:     Annotated[Optional[str], _BRAND]     = None,
    category:  Annotated[Optional[str], _CATEGORY]  = None,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, orders, _ = await _fetch(dt_from, dt_to, brand, category)
    return analytics.products_table(sales, orders, _days(dt_from, dt_to))


@router.get("/warehouses")
async def get_warehouses(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
    brand:     Annotated[Optional[str], _BRAND]     = None,
    category:  Annotated[Optional[str], _CATEGORY]  = None,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, orders, stocks = await _fetch(dt_from, dt_to, brand, category)
    return analytics.warehouses(sales, orders, stocks, _days(dt_from, dt_to))


import logging as _logging
_slog = _logging.getLogger("stocks_table")
_stocks_cache: dict = {}
_stocks_cache_ts: float = 0.0
_STOCKS_TTL = 1800  # 30 минут
_stocks_lock = asyncio.Lock()


@router.get("/stocks_table")
async def get_stocks_table():
    """Multi-marketplace stock status: WB + Ozon + YM per SKU."""
    global _stocks_cache, _stocks_cache_ts
    if _stocks_cache and _time.monotonic() - _stocks_cache_ts < _STOCKS_TTL:
        return _stocks_cache

    # холодный старт: отдаём последний снапшот из БД (свежесть 60с — реальная
    # пересборка случится следующим запросом или фоновым прогревом)
    if _stocks_cache_ts == 0.0:
        import snapshot as _snapmod
        snap = await asyncio.to_thread(_snapmod.load, "stocks_table", None)
        if snap:
            _stocks_cache = snap
            _stocks_cache_ts = _time.monotonic() - _STOCKS_TTL + 60
            return snap

    async with _stocks_lock:
        if _stocks_cache and _time.monotonic() - _stocks_cache_ts < _STOCKS_TTL:
            return _stocks_cache

        dt_to   = datetime.utcnow()
        dt_from = dt_to - timedelta(days=28)

        (wb_sales, _, wb_stocks), oz_stocks, oz_sales, ym_stocks, ym_sales = await asyncio.gather(
            _fetch(dt_from, dt_to, None, None),
            ozon_client.get_stocks(),
            ozon_client.get_sales_28d(),
            ym_client.get_stocks(),
            ym_client.get_sales_28d(),
        )

        _slog.info("stocks_table: wb_sales=%d wb_stocks=%d oz_stocks=%d oz_sales=%d ym_stocks=%d ym_sales=%d",
                   len(wb_sales), len(wb_stocks), len(oz_stocks), len(oz_sales), len(ym_stocks), len(ym_sales))

        names = cost_store.get_names()
        result = analytics.stocks_table_multi(
            wb_sales=wb_sales, wb_stocks=wb_stocks,
            oz_stocks=oz_stocks, oz_sales=oz_sales,
            ym_stocks=ym_stocks, ym_sales=ym_sales,
            names=names, days=28,
        )
        _stocks_cache = result
        _stocks_cache_ts = _time.monotonic()
        import snapshot as _snapmod
        await asyncio.to_thread(_snapmod.save, "stocks_table", result)
        return result

# Группировка остатков как в UI: спец-подгруппы по SKU, затем бренд, иначе «Прочее»
_STOCK_SUBGROUPS = [
    ("Фисты",            {"BMN-0013", "BMN-0028", "BMN-0035", "BMN-0036", "ST-07"}),
    ("Спреи для минета", {"BMN-0115", "BMN-0116", "BMN-0110"}),
]
_STOCK_BRANDS = ["Джага", "Satisfucktion", "Aloe"]
_STOCK_GROUP_ORDER = ["Фисты", "Aloe", "Спреи для минета", "Satisfucktion", "Джага", "Прочее"]


def _stock_group(row: dict) -> str:
    sku = row.get("supplierArticle", "")
    for gname, skus in _STOCK_SUBGROUPS:
        if sku in skus:
            return gname
    brand = row.get("brand", "")
    return brand if brand in _STOCK_BRANDS else "Прочее"


@router.get("/stocks_export", include_in_schema=False)
async def export_stocks_excel():
    """Выгрузка остатков в Excel: группы как в UI, все площадки, цветные «дней до OOS»."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response

    rows = await get_stocks_table()

    # группируем и сортируем как в UI (внутри группы — по средней скорости продаж)
    groups: dict[str, list] = {}
    for row in rows:
        groups.setdefault(_stock_group(row), []).append(row)
    for g in groups.values():
        g.sort(key=lambda r: -(r.get("wb_per_day", 0) + r.get("oz_per_day", 0) + r.get("ym_per_day", 0)) / 3)

    wb_x = Workbook()
    ws = wb_x.active
    ws.title = "Остатки"

    now_msk = datetime.utcnow() + timedelta(hours=3)
    N_COLS = 12
    ws["A1"] = (f"Остатки по площадкам на {now_msk.strftime('%d.%m.%Y %H:%M')} МСК "
                "(скорость продаж — среднее за 28 дней)")
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)

    headers = ["Артикул", "Название", "Бренд",
               "WB остаток", "WB шт/день", "WB дней",
               "Ozon остаток", "Ozon шт/день", "Ozon дней",
               "ЯМ остаток", "ЯМ шт/день", "ЯМ дней"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="374151")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    DAYS_FILL = {"red": "FFC7CE", "yellow": "FFEB9C", "green": "C6EFCE"}
    DAYS_FONT = {"red": "9C0006", "yellow": "9C6500", "green": "006100"}
    GROUP_FILL = PatternFill("solid", start_color="D9D9D9")

    def _days_val(d):
        return "—" if d is None or d >= 999 else d

    def _days_status(d):
        if d is None or d >= 999:
            return None
        return "red" if d <= 20 else "yellow" if d <= 45 else "green"

    r = 3
    for gname in _STOCK_GROUP_ORDER:
        grp_rows = groups.get(gname)
        if not grp_rows:
            continue
        # строка-заголовок группы
        gc = ws.cell(row=r, column=1, value=f"{gname} ({len(grp_rows)} арт.)")
        gc.font = Font(bold=True, size=11)
        for col in range(1, N_COLS + 1):
            ws.cell(row=r, column=col).fill = GROUP_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N_COLS)
        r += 1

        for row in grp_rows:
            vals = [row.get("supplierArticle", ""), row.get("name", ""), row.get("brand", ""),
                    row.get("wb_qty", 0), row.get("wb_per_day", 0), _days_val(row.get("wb_days")),
                    row.get("oz_qty", 0), row.get("oz_per_day", 0), _days_val(row.get("oz_days")),
                    row.get("ym_qty", 0), row.get("ym_per_day", 0), _days_val(row.get("ym_days"))]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                if col in (5, 8, 11):
                    c.number_format = "0.00"
            # подсветка «дней до OOS» цветом статуса — вместо отдельного столбца
            for col, dkey in ((6, "wb_days"), (9, "oz_days"), (12, "ym_days")):
                st = _days_status(row.get(dkey))
                if st:
                    dc = ws.cell(row=r, column=col)
                    dc.fill = PatternFill("solid", start_color=DAYS_FILL[st])
                    dc.font = Font(color=DAYS_FONT[st], bold=(st == "red"))
            r += 1

    widths = [14, 44, 15, 11, 11, 9, 12, 12, 10, 11, 11, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C3"          # шапка + артикул/название всегда видны

    buf = io.BytesIO()
    wb_x.save(buf)
    fname = f"ostatki_{now_msk.strftime('%Y-%m-%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


_reco_cache: dict = {}
_reco_cache_ts: float = 0.0
_RECO_TTL = 3600  # 1 час


@router.get("/supply-recommendations")
async def get_supply_recommendations():
    """Анализ остатков через Claude — рекомендации к поставке."""
    from config import ANTHROPIC_API_KEY
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY не настроен")

    global _reco_cache, _reco_cache_ts
    if _reco_cache and _time.monotonic() - _reco_cache_ts < _RECO_TTL:
        return _reco_cache

    # Собираем данные остатков
    dt_to   = datetime.utcnow()
    dt_from = dt_to - timedelta(days=28)
    (wb_sales, _, wb_stocks), oz_stocks, oz_sales, ym_stocks, ym_sales = await asyncio.gather(
        _fetch(dt_from, dt_to, None, None),
        ozon_client.get_stocks(),
        ozon_client.get_sales_28d(),
        ym_client.get_stocks(),
        ym_client.get_sales_28d(),
    )
    names = cost_store.get_names()
    rows = analytics.stocks_table_multi(
        wb_sales=wb_sales, wb_stocks=wb_stocks,
        oz_stocks=oz_stocks, oz_sales=oz_sales,
        ym_stocks=ym_stocks, ym_sales=ym_sales,
        names=names, days=28,
    )

    # Отдельная таблица на каждый кабинет: остаток, скорость, дней до OOS
    def _mp_table(prefix: str) -> str:
        out = ["Артикул | Название | Остаток | Шт/день | Дней до нуля"]
        for r in rows:
            qty = r.get(f"{prefix}_qty", 0)
            pd  = r.get(f"{prefix}_per_day", 0)
            if not qty and not pd:
                continue  # на этой площадке товара нет вовсе
            days = r.get(f"{prefix}_days", 999)
            days_s = "∞ (нет продаж)" if days >= 999 else str(days)
            out.append(f"{r.get('supplierArticle','')} | {r.get('name','')} | {qty} | {pd} | {days_s}")
        return "\n".join(out)

    prompt = f"""Ты аналитик маркетплейсов. Проанализируй остатки по каждому кабинету ОТДЕЛЬНО и дай рекомендации к поставке. Скорость продаж — среднее за 28 дней.

=== КАБИНЕТ WB (Wildberries) ===
{_mp_table('wb')}

=== КАБИНЕТ OZON ===
{_mp_table('oz')}

=== КАБИНЕТ ЯНДЕКС МАРКЕТ ===
{_mp_table('ym')}

Сделай отчёт строго в таком порядке — сначала WB, потом OZON, потом ЯМ. Для каждого кабинета:

**🟣 WB / 🔵 OZON / 🟡 ЯМ** (по разделу на кабинет):
- 🔴 Срочно (< 20 дней до нуля): артикул — сколько дней осталось, сколько штук отгрузить (считай запас на 60 дней: шт/день × 60 − остаток, округляй вверх до десятков).
- 🟡 Спланировать (20-45 дней): артикул — дней, рекомендуемая отгрузка.
- Если на площадке всё в порядке — одной строкой «Критичных позиций нет».
- Не перечисляй green-позиции.

В конце — **📋 СВОДКА** (3-5 предложений): общее состояние запасов, какой кабинет требует внимания в первую очередь, топ-3 артикула по срочности среди всех кабинетов с учётом объёма продаж.

Отвечай на русском, кратко и по делу, без воды."""

    import anthropic as _anthropic
    import logging as _logging
    _log = _logging.getLogger(__name__)
    try:
        client = _anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
        message = await client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}],
        )
    except Exception as exc:
        _log.error("Claude API error: %s", exc)
        raise HTTPException(status_code=500, detail=f"Claude API error: {exc}")
    text = message.content[0].text

    result = {"text": text, "generated_at": dt_to.strftime("%d.%m.%Y %H:%M UTC")}
    _reco_cache = result
    _reco_cache_ts = _time.monotonic()
    return result


@router.post("/supply-recommendations/invalidate", include_in_schema=False)
async def invalidate_reco_cache():
    global _reco_cache, _reco_cache_ts
    _reco_cache = {}
    _reco_cache_ts = 0.0
    return {"status": "ok"}


# ── Анализ текущих продаж через Claude ────────────────────────────────────────

_sales_analysis_cache: dict = {}
_sales_analysis_ts: float = 0.0
_SALES_ANALYSIS_TTL = 3600  # 1 час


@router.get("/sales_analysis")
async def get_sales_analysis(refresh: bool = False):
    """Анализ недельных продаж через Claude."""
    from config import ANTHROPIC_API_KEY
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY не настроен")

    global _sales_analysis_cache, _sales_analysis_ts
    if not refresh and _sales_analysis_cache and _time.monotonic() - _sales_analysis_ts < _SALES_ANALYSIS_TTL:
        return _sales_analysis_cache

    # Берём данные из weekly_orders кэша
    data = _wo_cache
    if not data:
        data = await get_weekly_orders()

    weeks = data.get("weeks", [])
    lines = ["Площадка | Артикул | Группа | " + " | ".join(weeks)]
    for mp_key in ("WB", "OZON", "YM"):
        block = data.get(mp_key, {})
        for s in (block.get("skus") or []):
            total = sum(s.get("rub") or [])
            if total < 100:
                continue
            rub_vals = " | ".join(f"{round(v):,}" for v in (s.get("rub") or []))
            lines.append(f"{mp_key} | {s['sku']} | {s.get('group','—')} | {rub_vals}")

    table_text = "\n".join(lines[:120])  # ограничение по размеру промпта
    now_str = datetime.utcnow().strftime("%d.%m.%Y")

    prompt = f"""Ты аналитик маркетплейсов. Дата анализа: {now_str}.

Ниже недельные продажи (₽) по артикулам за последние 8 недель (WB, OZON, YM):

{table_text}

Сформируй краткий структурированный анализ на русском:

1. **ТРЕНДЫ** — какие группы/артикулы растут, какие падают. Выдели топ-3 роста и топ-3 падения.

2. **ЛИДЕРЫ ПРОДАЖ** — топ-5 артикулов суммарно по всем площадкам за последние 2 недели.

3. **АНОМАЛИИ** — резкие скачки или провалы (>50%) между неделями. Что могло стать причиной?

4. **РЕКОМЕНДАЦИИ** — 3-5 конкретных действий для роста продаж на следующей неделе.

Будь краток, по делу. Используй эмодзи."""

    import anthropic as _anthropic
    import logging as _salog
    _salog = _salog.getLogger(__name__)
    try:
        client = _anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
        message = await client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}],
        )
    except Exception as exc:
        _salog.error("Claude sales_analysis error: %s", exc)
        raise HTTPException(status_code=500, detail=f"Claude API error: {exc}")

    text = message.content[0].text
    result = {"text": text, "generated_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}
    _sales_analysis_cache = result
    _sales_analysis_ts = _time.monotonic()
    return result



async def get_sales_analytics(
    date_from:   Annotated[Optional[str], _DATE_FROM] = None,
    date_to:     Annotated[Optional[str], _DATE_TO]   = None,
    days:        Annotated[int, _DAYS]                = 30,
    marketplace: str = "all",   # wb / ozon / ym / all
    brand:       str = "all",   # all / Джага / Satisfucktion / Aloe
):
    from collections import defaultdict

    dt_from, dt_to = _range(date_from, date_to, days)
    df_str = dt_from.strftime("%Y-%m-%d")
    dt_str = dt_to.strftime("%Y-%m-%d")

    (wb_sales_all, _, _), oz_rows, ym_rows = await asyncio.gather(
        _fetch(dt_from, dt_to, None, None),
        ozon_client.get_sales_detail(df_str, dt_str),
        ym_client.get_sales_detail(df_str, dt_str),
    )

    # Brand/name map: WB sales → catalog → cost_store
    uploaded_names = cost_store.get_names()

    def _art_brand(art: str, wb_brand: str = "") -> str:
        if wb_brand:
            return wb_brand
        cat = _catalog.CATALOG.get(art)
        if cat:
            return cat["brand"]
        return uploaded_names.get(art, "")

    def _art_name(art: str) -> str:
        if art in uploaded_names:
            return uploaded_names[art]
        cat = _catalog.CATALOG.get(art)
        if cat:
            return cat["name"]
        return ""

    brand_map: dict[str, str] = {}
    for s in wb_sales_all:
        art = s.get("supplierArticle", "")
        if art and art not in brand_map:
            brand_map[art] = _art_brand(art, s.get("brand", ""))
    # Fill from catalog for arts not in WB sales
    for art, info in _catalog.CATALOG.items():
        if art not in brand_map:
            brand_map[art] = info["brand"]

    brand_filter = brand if brand != "all" else None
    if brand_filter:
        wb_sales = [s for s in wb_sales_all if _art_brand(s.get("supplierArticle", ""), s.get("brand", "")) == brand_filter]
        oz_rows  = [r for r in oz_rows if brand_map.get(r["offer_id"]) == brand_filter]
        ym_rows  = [r for r in ym_rows if brand_map.get(r["shop_sku"]) == brand_filter]
    else:
        wb_sales = wb_sales_all

    # Date list for dynamics
    all_dates: list[str] = []
    cur = dt_from.date()
    end = dt_to.date()
    while cur <= end:
        all_dates.append(str(cur))
        from datetime import date as _date
        cur = (datetime.combine(cur, datetime.min.time()) + timedelta(days=1)).date()

    art_data: dict = defaultdict(lambda: {
        "wb_qty": 0, "wb_rev": 0.0,
        "oz_qty": 0, "oz_rev": 0.0,
        "ym_qty": 0, "ym_rev": 0.0,
    })
    date_data: dict = {d: {"wb_qty": 0, "wb_rev": 0.0, "oz_qty": 0, "oz_rev": 0.0, "ym_qty": 0, "ym_rev": 0.0}
                       for d in all_dates}

    _wb_sales_filtered = [s for s in wb_sales if analytics._is_sale(s)]

    if marketplace in ("wb", "all"):
        for s in _wb_sales_filtered:
            art  = s.get("supplierArticle", "")
            date = (s.get("date") or "")[:10]
            rev  = analytics._gross(s)
            art_data[art]["wb_qty"] += 1
            art_data[art]["wb_rev"] += rev
            if date in date_data:
                date_data[date]["wb_qty"] += 1
                date_data[date]["wb_rev"] += rev

    if marketplace in ("ozon", "all"):
        for r in oz_rows:
            art = r["offer_id"]
            art_data[art]["oz_qty"] += r["qty"]
            art_data[art]["oz_rev"] += r["revenue"]
            if r["date"] in date_data:
                date_data[r["date"]]["oz_qty"] += r["qty"]
                date_data[r["date"]]["oz_rev"] += r["revenue"]

    if marketplace in ("ym", "all"):
        for r in ym_rows:
            art = r["shop_sku"]
            art_data[art]["ym_qty"] += r["qty"]
            art_data[art]["ym_rev"] += r["revenue"]
            if r["date"] in date_data:
                date_data[r["date"]]["ym_qty"] += r["qty"]
                date_data[r["date"]]["ym_rev"] += r["revenue"]

    table = []
    for art, d in art_data.items():
        tq = d["wb_qty"] + d["oz_qty"] + d["ym_qty"]
        tr = d["wb_rev"] + d["oz_rev"] + d["ym_rev"]
        if tq == 0:
            continue
        table.append({
            "article": art,
            "name": _art_name(art),
            "brand": brand_map.get(art, ""),
            "wb_qty": d["wb_qty"], "wb_rev": round(d["wb_rev"], 2),
            "oz_qty": d["oz_qty"], "oz_rev": round(d["oz_rev"], 2),
            "ym_qty": d["ym_qty"], "ym_rev": round(d["ym_rev"], 2),
            "total_qty": tq, "total_rev": round(tr, 2),
        })
    table.sort(key=lambda x: x["total_rev"], reverse=True)

    total_qty = sum(r["total_qty"] for r in table)
    total_rev = sum(r["total_rev"] for r in table)
    period_days = max(1, _days(dt_from, dt_to))
    best = table[0] if table else None

    return {
        "kpi": {
            "total_qty": total_qty,
            "total_rev": round(total_rev, 2),
            "avg_per_day": round(total_qty / period_days, 1),
            "best": {"article": best["article"], "name": best["name"], "rev": best["total_rev"]} if best else None,
        },
        "dynamics": [{"date": d, **v} for d, v in sorted(date_data.items())],
        "top10": table[:10],
        "table": table,
    }


@router.get("/supplies")
async def get_supplies(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
    brand:     Annotated[Optional[str], _BRAND]     = None,
    category:  Annotated[Optional[str], _CATEGORY]  = None,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, _, stocks = await _fetch(dt_from, dt_to, brand, category)
    return analytics.supplies(sales, stocks, _days(dt_from, dt_to))


@router.get("/kpi")
async def get_kpi(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
    brand:     Annotated[Optional[str], _BRAND]     = None,
    category:  Annotated[Optional[str], _CATEGORY]  = None,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, orders, stocks = await _fetch(dt_from, dt_to, brand, category)
    return analytics.kpi_summary(sales, orders, stocks)


@router.get("/top-skus")
async def get_top_skus(
    date_from: Annotated[Optional[str], _DATE_FROM]       = None,
    date_to:   Annotated[Optional[str], _DATE_TO]         = None,
    days:      Annotated[int, _DAYS]                      = 30,
    n:         Annotated[int, Query(ge=1, le=50)]         = 20,
    brand:     Annotated[Optional[str], _BRAND]           = None,
    category:  Annotated[Optional[str], _CATEGORY]        = None,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, _, _ = await _fetch(dt_from, dt_to, brand, category)
    return analytics.top_skus(sales, _days(dt_from, dt_to), n)


@router.get("/abc-revenue")
async def get_abc_revenue(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, _, _ = await _fetch(dt_from, dt_to, None, None)
    return analytics.abc_by_revenue(sales, _days(dt_from, dt_to))


@router.get("/abc-turnover")
async def get_abc_turnover(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, _, stocks = await _fetch(dt_from, dt_to, None, None)
    return analytics.abc_by_turnover(sales, stocks, _days(dt_from, dt_to))


@router.get("/reorder")
async def get_reorder(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    sales, _, stocks = await _fetch(dt_from, dt_to, None, None)
    return analytics.reorder_forecast(sales, stocks, _days(dt_from, dt_to))


@router.get("/unit-economics")
async def get_unit_economics(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
    brand:     Annotated[Optional[str], _BRAND]     = None,
    category:  Annotated[Optional[str], _CATEGORY]  = None,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    costs = cost_store.get_costs()
    names = cost_store.get_names()

    # Try real reportDetailByPeriod first
    try:
        report = await cache.get_report_data(dt_from, dt_to)
        if report:
            if brand or category:
                report = [r for r in report
                          if (not brand    or r.get("brand_name") == brand)
                          and (not category or r.get("subject_name") == category)]
            rows = analytics.unit_economics_real(
                report, _days(dt_from, dt_to), costs or None, names or None)
            return {"rows": rows, "costs_loaded": cost_store.count(), "source": "report"}
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning("report API failed, falling back: %s", exc)

    # Fallback: estimated from sales/orders
    sales, orders, _ = await _fetch(dt_from, dt_to, brand, category)
    return {
        "rows": analytics.unit_economics(sales, orders, _days(dt_from, dt_to), costs or None),
        "costs_loaded": cost_store.count(),
        "source": "estimated",
    }


@router.get("/monthly-pivot")
async def get_monthly_pivot(
    date_from: Annotated[Optional[str], _DATE_FROM] = None,
    date_to:   Annotated[Optional[str], _DATE_TO]   = None,
    days:      Annotated[int, _DAYS]                = 30,
):
    dt_from, dt_to = _range(date_from, date_to, days)
    try:
        report = await cache.get_report_data(dt_from, dt_to)
        rows = analytics.monthly_pivot(report)
        return {"rows": rows}
    except Exception as exc:
        raise HTTPException(502, f"Report API: {exc}")


@router.post("/cache/invalidate", include_in_schema=False)
async def invalidate_cache():
    cache.invalidate()
    cache.invalidate_report()
    return {"status": "ok"}


def _week_ranges(n_weeks: int = 8) -> list[tuple]:
    """Last n_weeks Mon–Sun ranges (date objects), oldest first, including current week."""
    today = (datetime.utcnow() + timedelta(hours=3)).date()  # Moscow time
    monday_this_week = today - timedelta(days=today.weekday())
    weeks = []
    for i in range(n_weeks):
        start = monday_this_week - timedelta(weeks=(n_weeks - 1 - i))
        end = start + timedelta(days=6)
        weeks.append((start, end))
    return weeks


def _week_label(start, end) -> str:
    return f"{start.strftime('%d.%m')} - {end.strftime('%d.%m')}"


import time as _wtime
_weekly_cache: dict = {}
_weekly_cache_ts: float = 0.0
_WEEKLY_TTL = 1800  # 30 минут
_weekly_fetch_lock = asyncio.Lock()  # single-flight: только один фетч одновременно


@router.post("/weekly_summary/invalidate", include_in_schema=False)
async def invalidate_weekly_cache():
    global _weekly_cache, _weekly_cache_ts
    _weekly_cache = {}
    _weekly_cache_ts = 0.0
    return {"status": "ok"}


@router.get("/weekly_summary")
async def get_weekly_summary():
    """Сводка Продажи/Выкупы по неделям (Пн-Вс) за последние 8 недель, все МП."""
    global _weekly_cache, _weekly_cache_ts
    if _weekly_cache and _wtime.monotonic() - _weekly_cache_ts < _WEEKLY_TTL:
        return _weekly_cache

    # холодный старт: последний снапшот из БД отдаём сразу, пересборка — фоном
    # (WB-воронка строится минутами и в request-path упиралась в 100с лимит Render)
    if _weekly_cache_ts == 0.0:
        import snapshot as _snapmod
        snap = await asyncio.to_thread(_snapmod.load, "weekly_summary", None)
        if snap:
            _weekly_cache = snap
            _weekly_cache_ts = _wtime.monotonic()
            _dspawn(_weekly_refresh_bg())
            return snap

    async with _weekly_fetch_lock:
        # double-check: пока ждали лок, кто-то уже наполнил кеш
        if _weekly_cache and _wtime.monotonic() - _weekly_cache_ts < _WEEKLY_TTL:
            return _weekly_cache
        return await _build_weekly_summary()


async def _weekly_refresh_bg():
    async with _weekly_fetch_lock:
        await _build_weekly_summary()


async def _build_weekly_summary():
    global _weekly_cache, _weekly_cache_ts
    weeks = _week_ranges(8)
    n = len(weeks)
    dt_from = datetime.combine(weeks[0][0], datetime.min.time())
    dt_to   = datetime.combine(weeks[-1][1], datetime.min.time())

    import logging as _wslog
    _wslog = _wslog.getLogger("weekly_summary")

    date_from_str = weeks[0][0].strftime("%Y-%m-%d")
    date_to_str   = weeks[-1][1].strftime("%Y-%m-%d")
    week_str_ranges = [(s.strftime("%Y-%m-%d"), e.strftime("%Y-%m-%d")) for s, e in weeks]

    # OZON/YM — быстрые, отдельный короткий таймаут
    try:
        oz_rows, ym_rows = await asyncio.wait_for(
            asyncio.gather(
                ozon_client.get_sales_detail(date_from_str, date_to_str),
                ym_client.get_sales_detail(date_from_str, date_to_str),
            ),
            timeout=90,
        )
    except Exception as _exc:
        _wslog.error("[WS] OZON/YM fetch failed: %s", _exc)
        oz_rows, ym_rows = [], []

    # WB воронка — медленная (8 запросов с лимитом), длинный таймаут
    try:
        wb_funnel_weeks = await asyncio.wait_for(
            wb_client.get_nm_report_weeks(week_str_ranges), timeout=300,
        )
    except Exception as _exc:
        _wslog.error("[WS] WB funnel fetch failed: %s", _exc)
        wb_funnel_weeks = []
    wb_ok = any(f.get("orders_rub") or f.get("orders_qty") for f in wb_funnel_weeks)

    def week_index(date_str: str):
        try:
            d = datetime.strptime((date_str or "")[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
        for i, (s, e) in enumerate(weeks):
            if s <= d <= e:
                return i
        return None

    rub = {mp: {"sales": [0.0] * n, "buyout": [0.0] * n} for mp in ("WB", "OZON", "YM")}
    qty = {mp: {"sales": [0] * n, "buyout": [0] * n} for mp in ("WB", "OZON", "YM")}

    for i, fw in enumerate(wb_funnel_weeks):
        if i >= n:
            break
        rub["WB"]["sales"][i]  = fw.get("orders_rub", 0.0)
        qty["WB"]["sales"][i]  = fw.get("orders_qty", 0)
        rub["WB"]["buyout"][i] = fw.get("buyouts_rub", 0.0)
        qty["WB"]["buyout"][i] = fw.get("buyouts_qty", 0)

    for r in oz_rows:
        idx = week_index(r.get("date"))
        if idx is None:
            continue
        rub["OZON"]["sales"][idx] += r["revenue"]
        qty["OZON"]["sales"][idx] += r["qty"]
        if (r.get("status") or "").lower() == "delivered":
            # выкупы по дате доставки (как в Excel-запросе "Выкупы ШТ OZON")
            idx_d = week_index(r.get("delivered_date") or r.get("date"))
            if idx_d is not None:
                rub["OZON"]["buyout"][idx_d] += r["revenue"]
                qty["OZON"]["buyout"][idx_d] += r["qty"]

    for r in ym_rows:
        idx = week_index(r.get("date"))
        if idx is not None:
            rub["YM"]["sales"][idx] += r["revenue"]
            qty["YM"]["sales"][idx] += r["qty"]
        if (r.get("status") or "") == "DELIVERED":
            idx_d = week_index(r.get("update_date"))
            if idx_d is not None:
                rub["YM"]["buyout"][idx_d] += r["revenue"]
                qty["YM"]["buyout"][idx_d] += r["qty"]

    def block(d: dict, mp: str) -> dict:
        return {"sales": [round(v, 2) for v in d[mp]["sales"]],
                "buyout": [round(v, 2) for v in d[mp]["buyout"]]}

    def total(d: dict) -> dict:
        return {
            "sales":  [round(sum(d[mp]["sales"][i]  for mp in ("WB", "OZON", "YM")), 2) for i in range(n)],
            "buyout": [round(sum(d[mp]["buyout"][i] for mp in ("WB", "OZON", "YM")), 2) for i in range(n)],
        }

    result = {
        "weeks": [_week_label(s, e) for s, e in weeks],
        "rub": {"OZON": block(rub, "OZON"), "WB": block(rub, "WB"), "YM": block(rub, "YM"), "total": total(rub)},
        "qty": {"OZON": block(qty, "OZON"), "WB": block(qty, "WB"), "YM": block(qty, "YM"), "total": total(qty)},
    }
    if wb_ok:                      # кешируем только полный результат с данными WB
        _weekly_cache = result
        _weekly_cache_ts = _wtime.monotonic()
        import snapshot as _snapmod
        await asyncio.to_thread(_snapmod.save, "weekly_summary", result)
    return result


# ── Заказы по неделям с разбивкой по SKU ──────────────────────────────────────

import catalog as _cat

_wo_cache: dict = {}
_wo_cache_ts: float = 0.0
_WO_TTL = 1800
_wo_lock = asyncio.Lock()


@router.get("/weekly_orders")
async def get_weekly_orders():
    """Заказы по неделям (Пн-Вс) с разбивкой по SKU, все МП."""
    global _wo_cache, _wo_cache_ts
    if _wo_cache and _wtime.monotonic() - _wo_cache_ts < _WO_TTL:
        return _wo_cache

    # холодный старт: последний снапшот из БД (свежесть 60с — пересборка
    # придёт со следующим запросом или фоновым прогревом)
    if _wo_cache_ts == 0.0:
        import snapshot as _snapmod
        snap = await asyncio.to_thread(_snapmod.load, "weekly_orders", None)
        if snap:
            _wo_cache = snap
            _wo_cache_ts = _wtime.monotonic() - _WO_TTL + 60
            return snap

    async with _wo_lock:
        if _wo_cache and _wtime.monotonic() - _wo_cache_ts < _WO_TTL:
            return _wo_cache

        weeks = _week_ranges(8)
        n = len(weeks)
        date_from_str = weeks[0][0].strftime("%Y-%m-%d")
        date_to_str   = weeks[-1][1].strftime("%Y-%m-%d")
        dt_from = datetime.combine(weeks[0][0], datetime.min.time())
        dt_to   = datetime.combine(weeks[-1][1], datetime.min.time())

        def week_idx(date_str: str):
            try:
                d = datetime.strptime((date_str or "")[:10], "%Y-%m-%d").date()
            except ValueError:
                return None
            for i, (s, e) in enumerate(weeks):
                if s <= d <= e:
                    return i
            return None

        # ── WB. Первичный источник — sales_daily, куда воронка ЛК пишет
        # WB_ORDERS («Заказали», 1в1 с ЛК) и WB_CANCELS («Отменили и
        # вернули» по дате события). statistics /orders занижает заказы
        # (по спеке не отдаёт неподтверждённые оплаты — рассрочка, оплата
        # при получении), поэтому кеш /orders — только фолбэк, пока
        # воронка ещё ни разу не выгружалась. ──
        wb_by_sku: dict = {}   # sku → {rub, qty, cancel_rub, cancel_qty, name}
        wb_total_rub        = [0.0] * n
        wb_total_qty        = [0]   * n
        wb_total_cancel_rub = [0.0] * n
        wb_total_cancel_qty = [0]   * n

        def _wb_slot(sku, name):
            if sku not in wb_by_sku:
                wb_by_sku[sku] = {"rub": [0.0]*n, "qty": [0]*n,
                                  "cancel_rub": [0.0]*n, "cancel_qty": [0]*n,
                                  "name": name}
            return wb_by_sku[sku]

        wb_from_funnel = False
        try:
            sd_rows = await asyncio.to_thread(
                sales_history.get_history, date_from_str, date_to_str)
            if sales_history.funnel_fresh(26 * 14):   # свежее двух недель
                for r in sd_rows:
                    plat = str(r.get("platform") or "").upper()
                    if plat not in ("WB_ORDERS", "WB_CANCELS"):
                        continue
                    idx = week_idx(r.get("date"))
                    if idx is None:
                        continue
                    sku = r.get("sku") or ""
                    rev = float(r.get("revenue") or 0)
                    qty = int(r.get("qty") or 0)
                    e = _wb_slot(sku, sku)
                    if plat == "WB_ORDERS":
                        e["rub"][idx] += rev
                        e["qty"][idx] += qty
                        wb_total_rub[idx] += rev
                        wb_total_qty[idx] += qty
                        wb_from_funnel = True
                    else:
                        e["cancel_rub"][idx] += rev
                        e["cancel_qty"][idx] += qty
                        wb_total_cancel_rub[idx] += rev
                        wb_total_cancel_qty[idx] += qty
        except Exception as e:
            import logging; logging.getLogger(__name__).warning("weekly_orders WB(sd): %s", e)
        if not wb_from_funnel:
            wb_by_sku.clear()
            wb_total_rub        = [0.0] * n
            wb_total_qty        = [0]   * n
            wb_total_cancel_rub = [0.0] * n
            wb_total_cancel_qty = [0]   * n
            try:
                _, wb_orders, _ = await cache.get_raw_data(dt_from, dt_to)
                for o in wb_orders:
                    idx = week_idx(o.get("date") or o.get("lastChangeDate"))
                    if idx is None:
                        continue
                    raw = o.get("nmId") or o.get("supplierArticle") or ""
                    sku = _cat.resolve_wb(raw) if raw else str(raw)
                    name = o.get("subject") or o.get("category") or sku
                    price = float(o.get("priceWithDisc") or o.get("totalPrice") or 0)
                    e = _wb_slot(sku, name)
                    e["rub"][idx] += price
                    e["qty"][idx] += 1
                    wb_total_rub[idx] += price
                    wb_total_qty[idx] += 1
                    if bool(o.get("isCancel")):
                        e["cancel_rub"][idx] += price
                        e["cancel_qty"][idx] += 1
                        wb_total_cancel_rub[idx] += price
                        wb_total_cancel_qty[idx] += 1
            except Exception as e:
                import logging; logging.getLogger(__name__).warning("weekly_orders WB: %s", e)

        # ── OZON / YM ──
        oz_ok = ym_ok = True
        try:
            oz_rows, ym_rows = await asyncio.wait_for(
                asyncio.gather(
                    ozon_client.get_sales_detail(date_from_str, date_to_str),
                    ym_client.get_sales_detail(date_from_str, date_to_str),
                ),
                timeout=25,
            )
        except Exception as _exc:
            import logging; logging.getLogger(__name__).warning("weekly_orders OZON/YM: %s", _exc)
            oz_rows, ym_rows = [], []
            oz_ok = ym_ok = False

        def agg_rows(rows, sku_field, resolve):
            by_sku: dict = {}
            total_rub = [0.0] * n
            total_qty = [0]   * n
            for r in rows:
                idx = week_idx(r.get("date"))
                if idx is None:
                    continue
                raw = r.get(sku_field) or ""
                sku = resolve(raw) if raw else ""
                rev = float(r.get("revenue") or 0)
                qty = int(r.get("qty") or 0)
                if sku not in by_sku:
                    by_sku[sku] = {"rub": [0.0]*n, "qty": [0]*n, "name": sku}
                by_sku[sku]["rub"][idx] = round(by_sku[sku]["rub"][idx] + rev, 2)
                by_sku[sku]["qty"][idx] += qty
                total_rub[idx] = round(total_rub[idx] + rev, 2)
                total_qty[idx] += qty
            return by_sku, total_rub, total_qty

        oz_by_sku, oz_rub, oz_qty = agg_rows(oz_rows, "offer_id", _cat.resolve_ozon)
        ym_by_sku, ym_rub, ym_qty = agg_rows(ym_rows, "shop_sku", _cat.resolve_ym)

        def clean_sku(by_sku, has_cancel=False):
            """Sort SKUs by total revenue desc, round numbers."""
            result = []
            for sku, d in by_sku.items():
                cat = _cat.CATALOG.get(sku) or {}
                item = {
                    "sku": sku,
                    "name": cat.get("name") or d["name"] or sku,
                    "brand": cat.get("brand", ""),
                    "group": cat.get("group", ""),
                    "rub": [round(v, 2) for v in d["rub"]],
                    "qty": d["qty"],
                }
                if has_cancel:
                    item["cancel_rub"] = [round(v, 2) for v in d.get("cancel_rub", [0.0]*len(d["rub"]))]
                    item["cancel_qty"] = d.get("cancel_qty", [0]*len(d["qty"]))
                result.append(item)
            result.sort(key=lambda x: sum(x["rub"]), reverse=True)
            return result

        result = {
            "weeks": [_week_label(s, e) for s, e in weeks],
            "WB":   {"total_rub": [round(v,2) for v in wb_total_rub], "total_qty": wb_total_qty,
                     "total_cancel_rub": [round(v,2) for v in wb_total_cancel_rub],
                     "total_cancel_qty": wb_total_cancel_qty,
                     "skus": clean_sku(wb_by_sku, has_cancel=True)},
            "OZON": {"total_rub": oz_rub, "total_qty": oz_qty,
                     "skus": clean_sku(oz_by_sku)},
            "YM":   {"total_rub": ym_rub, "total_qty": ym_qty,
                     "skus": clean_sku(ym_by_sku)},
        }
        # Кешируем всегда. Если Ozon/YM упали — TTL короткий (5 мин),
        # чтобы при следующем заходе попробовать снова; если всё ок — 30 мин.
        _wo_cache = result
        _wo_cache_ts = _wtime.monotonic() if (oz_ok and ym_ok) else (_wtime.monotonic() - _WO_TTL + 300)
        if oz_ok and ym_ok:
            import snapshot as _snapmod
            await asyncio.to_thread(_snapmod.save, "weekly_orders", result)
        return result


@router.post("/weekly_orders/invalidate", include_in_schema=False)
async def invalidate_weekly_orders():
    global _wo_cache, _wo_cache_ts
    _wo_cache = {}; _wo_cache_ts = 0.0
    _om_cache.clear()
    return {"status": "ok"}


_om_cache: dict = {}      # period → (ts, result)
_OM_TTL = 600


@router.get("/orders_matrix")
async def get_orders_matrix(period: str = Query(default="day")):
    """Заказы той же матрицей, что weekly_orders, но по дням (14) или
    месяцам (6). Источник — вечная sales_daily (WB_ORDERS пишет воронка
    ЛК, «сегодня» дописывает statistics /orders), поэтому месяцы не
    упираются в 90-дневное окно WB API. Пилюли отмен — только в неделях."""
    period = period if period in ("day", "month") else "day"
    cached = _om_cache.get(period)
    if cached and _wtime.monotonic() - cached[0] < _OM_TTL:
        return cached[1]

    from datetime import date as _date
    import calendar as _cal
    import catalog as _cat
    today = (datetime.utcnow() + timedelta(hours=3)).date()
    if period == "day":
        days = [today - timedelta(days=i) for i in range(13, -1, -1)]
        DOW = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]
        labels = [f"{d.strftime('%d.%m')} {DOW[d.weekday()]}" for d in days]
        d_from, d_to = days[0], days[-1]
        pos = {d.strftime("%Y-%m-%d"): i for i, d in enumerate(days)}
        idx = lambda ds: pos.get((ds or "")[:10])
    else:
        y, m = today.year, today.month
        firsts = []
        for _ in range(6):
            firsts.append(_date(y, m, 1))
            m -= 1
            if m == 0:
                y, m = y - 1, 12
        firsts.reverse()
        MON = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
               "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
        labels = [f"{MON[f.month - 1]} {f.year}" for f in firsts]
        d_from = firsts[0]
        d_to = _date(firsts[-1].year, firsts[-1].month,
                     _cal.monthrange(firsts[-1].year, firsts[-1].month)[1])
        pos = {(f.year, f.month): i for i, f in enumerate(firsts)}
        def idx(ds):
            try:
                d = datetime.strptime((ds or "")[:10], "%Y-%m-%d")
            except ValueError:
                return None
            return pos.get((d.year, d.month))

    n = len(labels)
    rows = await asyncio.to_thread(
        sales_history.get_history,
        d_from.strftime("%Y-%m-%d"), d_to.strftime("%Y-%m-%d"))

    blocks = {p: {"by": {}, "rub": [0.0] * n, "qty": [0] * n}
              for p in ("WB", "OZON", "YM")}
    # WB: основным потоком идут заказы (WB_ORDERS — все заказанные, как
    # «Заказали» воронки ЛК; отмены не вычитаются);
    # для колонок старше их окна фолбэк — продажи (поток WB)
    wb_sales_fb = {"by": {}, "rub": [0.0] * n, "qty": [0] * n}
    for r in rows:
        plat = str(r.get("platform") or "").upper()
        if plat == "WB_ORDERS":
            b = blocks["WB"]
        elif plat == "WB":
            b = wb_sales_fb
        else:
            b = blocks.get(plat)
        i = idx(r.get("date"))
        if not b or i is None:
            continue
        sku = r.get("sku") or ""
        s = b["by"].setdefault(sku, {"rub": [0.0] * n, "qty": [0] * n})
        rev, qty = float(r.get("revenue") or 0), int(r.get("qty") or 0)
        s["rub"][i] += rev; s["qty"][i] += qty
        b["rub"][i] += rev; b["qty"][i] += qty
    wbb = blocks["WB"]
    for i in range(n):
        if wbb["qty"][i] == 0 and wb_sales_fb["qty"][i] > 0:
            wbb["rub"][i] = wb_sales_fb["rub"][i]
            wbb["qty"][i] = wb_sales_fb["qty"][i]
            for sku, s in wb_sales_fb["by"].items():
                if s["qty"][i] or s["rub"][i]:
                    t = wbb["by"].setdefault(sku, {"rub": [0.0] * n, "qty": [0] * n})
                    t["rub"][i] = s["rub"][i]; t["qty"][i] = s["qty"][i]

    def clean(by):
        out = []
        for sku, d in by.items():
            cat = _cat.CATALOG.get(sku) or {}
            out.append({"sku": sku, "name": cat.get("name") or sku,
                        "brand": cat.get("brand", ""), "group": cat.get("group", ""),
                        "rub": [round(v, 2) for v in d["rub"]], "qty": d["qty"]})
        out.sort(key=lambda x: sum(x["rub"]), reverse=True)
        return out

    result = {"weeks": labels, "period": period}
    for p in ("WB", "OZON", "YM"):
        result[p] = {"total_rub": [round(v, 2) for v in blocks[p]["rub"]],
                     "total_qty": blocks[p]["qty"],
                     "skus": clean(blocks[p]["by"])}
    _om_cache[period] = (_wtime.monotonic(), result)
    return result


@router.get("/ozon_debug", include_in_schema=False)
async def debug_ozon_fbo():
    """Живой тест v3/posting/fbo/list — показывает точную ошибку Ozon."""
    import httpx as _httpx
    import config as _cfg
    since = (datetime.utcnow() - timedelta(days=7)).strftime("%Y-%m-%dT00:00:00.000Z")
    to    = datetime.utcnow().strftime("%Y-%m-%dT23:59:59.000Z")
    out = {}
    for limit in (1000, 100):
        try:
            async with _httpx.AsyncClient(timeout=30) as c:
                r = await c.post("https://api-seller.ozon.ru/v3/posting/fbo/list",
                    headers={"Client-Id": _cfg.OZON_CLIENT_ID, "Api-Key": _cfg.OZON_API_KEY,
                             "Content-Type": "application/json"},
                    json={"cursor": "", "filter": {"since": since, "to": to},
                          "limit": limit, "sort_dir": "ASC",
                          "with": {"analytics_data": False, "financial_data": False, "legal_info": False}})
            body = r.json() if "json" in r.headers.get("content-type", "") else r.text[:400]
            n = len(body.get("postings") or []) if isinstance(body, dict) else None
            out[f"v3_limit_{limit}"] = {"status": r.status_code,
                                        "postings": n,
                                        "body": str(body)[:400] if r.status_code != 200 else f"OK, {n} postings"}
        except Exception as e:
            out[f"v3_limit_{limit}"] = {"error": str(e)}
    return out


@router.get("/weekly_orders/debug", include_in_schema=False)
async def debug_weekly_orders():
    """Диагностика: кеш WB + живой тест WB Statistics и Analytics API."""
    import cache as _cache_mod
    import config as _cfg
    import httpx as _httpx

    wb_orders_count = len(_cache_mod._store.orders) if hasattr(_cache_mod._store, 'orders') else -1
    wb_cache_age = round(_wtime.monotonic() - _cache_mod._store.fetched_at, 1) if hasattr(_cache_mod._store, 'fetched_at') else -1

    # Живой тест WB Statistics API (последние 7 дней, limit 1)
    wb_stats_status, wb_stats_body = None, None
    if _cfg.WB_API_KEY:
        date_from = (datetime.utcnow() - timedelta(days=7)).strftime("%Y-%m-%dT00:00:00")
        try:
            async with _httpx.AsyncClient(timeout=15) as _c:
                r = await _c.get(
                    "https://statistics-api.wildberries.ru/api/v1/supplier/orders",
                    headers={"Authorization": _cfg.WB_API_KEY},
                    params={"dateFrom": date_from, "flag": 0},
                )
            wb_stats_status = r.status_code
            body = r.json() if r.headers.get("content-type", "").startswith("application/json") else r.text[:300]
            if isinstance(body, list):
                wb_stats_body = f"list[{len(body)}]"
            else:
                wb_stats_body = str(body)[:300]
        except Exception as _e:
            wb_stats_status = "error"
            wb_stats_body = str(_e)

    return {
        "use_mock": _cfg.USE_MOCK,
        "wb_key_set": bool(_cfg.WB_API_KEY),
        "wb_key_suffix": _cfg.WB_API_KEY[-6:] if _cfg.WB_API_KEY else "",
        "wb_orders_in_cache": wb_orders_count,
        "wb_cache_age_sec": wb_cache_age,
        "wo_cache_filled": bool(_wo_cache),
        "wo_cache_age_sec": round(_wtime.monotonic() - _wo_cache_ts, 1) if _wo_cache_ts else -1,
        "wb_stats_api_status": wb_stats_status,
        "wb_stats_api_body": wb_stats_body,
    }


# ── Помесячная сводка Продажи/Выкупы ──────────────────────────────────────────

_RU_MONTHS = ["", "Янв", "Фев", "Мар", "Апр", "Май", "Июн",
              "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]


def _month_ranges(n_months: int = 6) -> list[tuple]:
    """Последние n_months календарных месяцев (date-диапазоны), старые первыми, вкл. текущий."""
    today = (datetime.utcnow() + timedelta(hours=3)).date()  # Moscow time
    ranges = []
    y, m = today.year, today.month
    months_back = []
    for _ in range(n_months):
        months_back.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    for (yy, mm) in reversed(months_back):
        start = datetime(yy, mm, 1).date()
        if mm == 12:
            end = datetime(yy, 12, 31).date()
        else:
            end = (datetime(yy, mm + 1, 1) - timedelta(days=1)).date()
        ranges.append((start, end))
    return ranges


def _month_label(start) -> str:
    return f"{_RU_MONTHS[start.month]} {start.year}"


_monthly_cache: dict = {}
_monthly_cache_ts: float = 0.0
_MONTHLY_TTL = 1800  # 30 минут
_monthly_fetch_lock = asyncio.Lock()


@router.post("/monthly_summary/invalidate", include_in_schema=False)
async def invalidate_monthly_cache():
    global _monthly_cache, _monthly_cache_ts
    _monthly_cache = {}
    _monthly_cache_ts = 0.0
    return {"status": "ok"}


@router.get("/monthly_summary")
async def get_monthly_summary():
    """Сводка Продажи/Выкупы по месяцам за последние 6 месяцев, все МП."""
    global _monthly_cache, _monthly_cache_ts
    if _monthly_cache and _wtime.monotonic() - _monthly_cache_ts < _MONTHLY_TTL:
        return _monthly_cache

    # холодный старт: снапшот из БД сразу, пересборка — фоном (см. weekly_summary)
    if _monthly_cache_ts == 0.0:
        import snapshot as _snapmod
        snap = await asyncio.to_thread(_snapmod.load, "monthly_summary", None)
        if snap:
            _monthly_cache = snap
            _monthly_cache_ts = _wtime.monotonic()
            _dspawn(_monthly_refresh_bg())
            return snap

    async with _monthly_fetch_lock:
        if _monthly_cache and _wtime.monotonic() - _monthly_cache_ts < _MONTHLY_TTL:
            return _monthly_cache
        return await _build_monthly_summary()


async def _monthly_refresh_bg():
    async with _monthly_fetch_lock:
        await _build_monthly_summary()


async def _build_monthly_summary():
    global _monthly_cache, _monthly_cache_ts
    months = _month_ranges(6)
    n = len(months)

    import logging as _mslog
    _mslog = _mslog.getLogger("monthly_summary")

    date_from_str = months[0][0].strftime("%Y-%m-%d")
    date_to_str   = months[-1][1].strftime("%Y-%m-%d")
    month_str_ranges = [(s.strftime("%Y-%m-%d"), e.strftime("%Y-%m-%d")) for s, e in months]

    try:
        oz_rows, ym_rows = await asyncio.wait_for(
            asyncio.gather(
                ozon_client.get_sales_detail(date_from_str, date_to_str),
                ym_client.get_sales_detail(date_from_str, date_to_str),
            ),
            timeout=90,
        )
    except Exception as _exc:
        _mslog.error("[MS] OZON/YM fetch failed: %s", _exc)
        oz_rows, ym_rows = [], []

    try:
        wb_funnel_months = await asyncio.wait_for(
            wb_client.get_nm_report_weeks(month_str_ranges), timeout=300,
        )
    except Exception as _exc:
        _mslog.error("[MS] WB funnel fetch failed: %s", _exc)
        wb_funnel_months = []
    wb_ok = any(f.get("orders_rub") or f.get("orders_qty") for f in wb_funnel_months)

    def month_index(date_str: str):
        try:
            d = datetime.strptime((date_str or "")[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
        for i, (s, e) in enumerate(months):
            if s <= d <= e:
                return i
        return None

    rub = {mp: {"sales": [0.0] * n, "buyout": [0.0] * n} for mp in ("WB", "OZON", "YM")}
    qty = {mp: {"sales": [0] * n, "buyout": [0] * n} for mp in ("WB", "OZON", "YM")}

    for i, fw in enumerate(wb_funnel_months):
        if i >= n:
            break
        rub["WB"]["sales"][i]  = fw.get("orders_rub", 0.0)
        qty["WB"]["sales"][i]  = fw.get("orders_qty", 0)
        rub["WB"]["buyout"][i] = fw.get("buyouts_rub", 0.0)
        qty["WB"]["buyout"][i] = fw.get("buyouts_qty", 0)

    for r in oz_rows:
        idx = month_index(r.get("date"))
        if idx is None:
            continue
        rub["OZON"]["sales"][idx] += r["revenue"]
        qty["OZON"]["sales"][idx] += r["qty"]
        if (r.get("status") or "").lower() == "delivered":
            idx_d = month_index(r.get("delivered_date") or r.get("date"))
            if idx_d is not None:
                rub["OZON"]["buyout"][idx_d] += r["revenue"]
                qty["OZON"]["buyout"][idx_d] += r["qty"]

    for r in ym_rows:
        idx = month_index(r.get("date"))
        if idx is not None:
            rub["YM"]["sales"][idx] += r["revenue"]
            qty["YM"]["sales"][idx] += r["qty"]
        if (r.get("status") or "") == "DELIVERED":
            idx_d = month_index(r.get("update_date"))
            if idx_d is not None:
                rub["YM"]["buyout"][idx_d] += r["revenue"]
                qty["YM"]["buyout"][idx_d] += r["qty"]

    def block(d: dict, mp: str) -> dict:
        return {"sales": [round(v, 2) for v in d[mp]["sales"]],
                "buyout": [round(v, 2) for v in d[mp]["buyout"]]}

    def total(d: dict) -> dict:
        return {
            "sales":  [round(sum(d[mp]["sales"][i]  for mp in ("WB", "OZON", "YM")), 2) for i in range(n)],
            "buyout": [round(sum(d[mp]["buyout"][i] for mp in ("WB", "OZON", "YM")), 2) for i in range(n)],
        }

    result = {
        "months": [_month_label(s) for s, e in months],
        "rub": {"OZON": block(rub, "OZON"), "WB": block(rub, "WB"), "YM": block(rub, "YM"), "total": total(rub)},
        "qty": {"OZON": block(qty, "OZON"), "WB": block(qty, "WB"), "YM": block(qty, "YM"), "total": total(qty)},
    }
    if wb_ok:
        _monthly_cache = result
        _monthly_cache_ts = _wtime.monotonic()
        import snapshot as _snapmod
        await asyncio.to_thread(_snapmod.save, "monthly_summary", result)
    return result


