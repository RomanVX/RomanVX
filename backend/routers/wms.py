"""WMS Market Partners — фулфилмент Чехов. Фаза 1 (MVP).

Отдельное приложение: /wms (фронт) + /api/wms/* (бэк). Своя авторизация
(cookie wms_session), свои пользователи: staff (склад) и client (клиент ФФ,
видит только своё). Дашборд-аналитика не затрагивается.

Несущие принципы (см. docs/wms_spec.md):
  1. Остаток = SUM(qty) по append-only журналу wms_moves. Правки — сторно.
  2. Каждая операция сразу пишет wms_billing (снапшот цены в момент скана).
  3. client_id в каждой таблице; клиентская роль жёстко отфильтрована.
"""
import asyncio
import hashlib
import json
import logging
import secrets
import threading
from datetime import date, datetime, timedelta

from fastapi import APIRouter, HTTPException, Request, Response
from fastapi.responses import HTMLResponse

import db

_log = logging.getLogger("wms")

# Один процесс (uvicorn, 1 воркер): межпоточный лок сериализует товарно-денежные
# операции — закрывает гонки «проверка остатка → списание», двойные сабмиты
# приёмки и погоню за id (ревью wf_77fbc29b).
_OP_LOCK = threading.Lock()

# id-колонка по диалекту: SQLite не понимает SERIAL (колонка теряет
# автоинкремент и молча пишет NULL в PK)
_ID = "id SERIAL PRIMARY KEY" if db.IS_PG else \
    "id INTEGER PRIMARY KEY AUTOINCREMENT"


def _insert_id(sql: str, params: tuple) -> int:
    """INSERT + надёжный id той же транзакцией (RETURNING / lastrowid) —
    вместо гоночного SELECT ... ORDER BY id DESC."""
    with db._conn() as con:
        cur = con.cursor()
        if db.IS_PG:
            cur.execute(db._tr(sql + " RETURNING id"), params)
            new_id = cur.fetchone()[0]
        else:
            cur.execute(sql, params)
            new_id = cur.lastrowid
        con.commit()
        return int(new_id)
router = APIRouter(prefix="/api/wms", tags=["wms"])

COOKIE = "wms_session"
SESSION_DAYS = 30

# Тариф по умолчанию для нового клиента — наш прайс (правится по клиенту).
# Ступени сборки по объёму упаковки, л (эквивалент суммы сторон/веса).
DEFAULT_TARIFF = {
    "assembly_tiers": [
        {"max_l": 1.7, "price": 65},   # до 45 см / 1.5 кг
        {"max_l": 4.0, "price": 85},   # до 60 см / 5 кг
        {"max_l": 10.0, "price": 110},  # до 90 см / 12 кг
    ],
    "extra_unit": 15,
    "receive_pallet": 250,     # приёмка паллетой (моно-короба)
    "receive_unit": 5,         # приёмка поштучно, маркированный
    "receive_unit_sorted": 15,  # поштучно с сортировкой (россыпь/микс)
    "return_process": 35,
    "kiz": 10,
    "storage": {
        "free_days": 30,
        "box_litres": 77,      # эквивалент короба 60×40×40
        "bands": [
            {"from_day": 31, "to_day": 90, "box_day": 15},
            {"from_day": 91, "to_day": 100000, "box_day": 30},
        ],
    },
}

SERVICE_NAMES = {
    "RECEIVE_PALLET": "Приёмка (паллета)",
    "RECEIVE_UNIT": "Приёмка (шт, маркированный)",
    "RECEIVE_UNIT_SORTED": "Приёмка (шт, с сортировкой)",
    "PICK_ORDER": "Сборка заказа",
    "PICK_ITEM_ADD": "Доп. единица в заказе",
    "STORAGE_DAY": "Хранение (сутки)",
    "RETURN_PROCESS": "Обработка возврата",
    "KIZ": "Маркировка «Честный знак»",
    "ADJUST": "Корректировка",
}


# ── Схема БД ────────────────────────────────────────────────────────────────

def _init():
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_clients (
        {_ID}, code TEXT UNIQUE, name TEXT, inn TEXT,
        contact TEXT, tg_chat_id TEXT, tariff TEXT, settings TEXT,
        created_at TEXT)""")
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_users (
        login TEXT PRIMARY KEY, salt TEXT, pass_hash TEXT,
        role TEXT, client_id INTEGER)""")
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_sessions (
        token TEXT PRIMARY KEY, login TEXT, role TEXT, client_id INTEGER,
        expires TEXT)""")
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_skus (
        {_ID}, client_id INTEGER, code TEXT, name TEXT,
        length_cm REAL, width_cm REAL, height_cm REAL,
        volume_l REAL, weight_g REAL, value_rub REAL,
        requires_expiry INTEGER DEFAULT 0, created_at TEXT)""")
    for col in ("length_cm", "width_cm", "height_cm"):
        try:
            db.execute(f"ALTER TABLE wms_skus ADD COLUMN {col} REAL")
        except Exception:
            pass
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_barcodes (
        barcode TEXT PRIMARY KEY, sku_id INTEGER)""")
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_inbounds (
        {_ID}, client_id INTEGER, status TEXT,
        expected_date TEXT, note TEXT, act_no TEXT,
        created_at TEXT, closed_at TEXT, created_by TEXT)""")
    try:
        db.execute("ALTER TABLE wms_inbounds ADD COLUMN pack_type TEXT")
    except Exception:
        pass
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_inbound_lines (
        {_ID}, inbound_id INTEGER, sku_id INTEGER,
        qty_expected INTEGER, qty_received INTEGER,
        batch_no TEXT, expiry_date TEXT,
        discrepancy_type TEXT, discrepancy_qty INTEGER, note TEXT)""")
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_inbound_boxes (
        {_ID}, inbound_id INTEGER, code TEXT UNIQUE, note TEXT,
        received INTEGER DEFAULT 0, created_at TEXT)""")
    try:
        db.execute("ALTER TABLE wms_inbound_boxes ADD COLUMN pallet_id INTEGER")
    except Exception:
        pass
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_inbound_pallets (
        {_ID}, inbound_id INTEGER, code TEXT UNIQUE, created_at TEXT)""")
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_inbound_box_items (
        {_ID}, box_id INTEGER, sku_id INTEGER, qty INTEGER,
        expiry_date TEXT, batch_no TEXT)""")
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_batches (
        {_ID}, client_id INTEGER, sku_id INTEGER,
        batch_no TEXT, expiry_date TEXT, received_at TEXT,
        inbound_id INTEGER)""")
    # append-only журнал движений: остаток = SUM(qty)
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_moves (
        {_ID}, client_id INTEGER, sku_id INTEGER,
        batch_id INTEGER, qty INTEGER, status TEXT,
        doc_type TEXT, doc_ref TEXT, note TEXT,
        user_login TEXT, created_at TEXT)""")
    # append-only начисления: цена зафиксирована в момент операции
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_billing (
        {_ID}, client_id INTEGER, service_code TEXT,
        qty REAL, price REAL, amount REAL, occurred_at TEXT,
        source_type TEXT, source_ref TEXT, sku_id INTEGER, note TEXT)""")
    db.execute(f"""CREATE TABLE IF NOT EXISTS wms_snapshots (
        day TEXT, client_id INTEGER, sku_id INTEGER, batch_id INTEGER,
        qty INTEGER, volume_l REAL, age_days INTEGER,
        PRIMARY KEY (day, batch_id))""")
    for t in ("wms_clients", "wms_skus", "wms_inbounds", "wms_inbound_lines",
              "wms_inbound_boxes", "wms_inbound_box_items",
              "wms_inbound_pallets",
              "wms_batches", "wms_moves", "wms_billing"):
        try:
            db.ensure_serial(t)
        except Exception:
            pass


def ensure_bootstrap():
    """Первый запуск: таблицы + стартовый сотрудник склада."""
    try:
        _init()
        n = db.fetchone("SELECT COUNT(*) FROM wms_users")[0]
        if n == 0:
            import os
            login = os.getenv("WMS_ADMIN_LOGIN", "sklad").strip()
            password = os.getenv("WMS_ADMIN_PASSWORD", "sklad2026").strip()
            salt = secrets.token_hex(8)
            db.execute("INSERT INTO wms_users (login, salt, pass_hash, role, client_id) "
                       "VALUES (?,?,?,?,NULL)",
                       (login, salt, _hash(salt, password), "staff"))
            _log.info("WMS: создан сотрудник «%s» (смени пароль!)", login)
    except Exception as e:
        _log.warning("WMS bootstrap: %s", e)


# ── Авторизация ─────────────────────────────────────────────────────────────

def _hash(salt: str, password: str) -> str:
    return hashlib.sha256((salt + password).encode()).hexdigest()


def _session_of(request: Request) -> dict | None:
    token = request.cookies.get(COOKIE)
    if not token:
        return None
    try:
        row = db.fetchone("SELECT login, role, client_id, expires "
                          "FROM wms_sessions WHERE token = ?", (token,))
    except Exception:
        return None
    if not row:
        return None
    login, role, client_id, expires = row
    try:
        if datetime.fromisoformat(expires) < datetime.utcnow():
            return None
    except ValueError:
        return None
    return {"login": login, "role": role, "client_id": client_id}


def _require(request: Request, staff_only: bool = False) -> dict:
    sess = _session_of(request)
    if not sess:
        raise HTTPException(status_code=401, detail="Не авторизован")
    if staff_only and sess["role"] != "staff":
        raise HTTPException(status_code=403, detail="Только для сотрудников склада")
    return sess


def _client_scope(sess: dict, client_id) -> int:
    """Клиент видит только себя; сотрудник — кого укажет."""
    if sess["role"] == "client":
        return int(sess["client_id"])
    if client_id in (None, "", 0):
        raise HTTPException(status_code=400, detail="Укажи client_id")
    return int(client_id)


@router.post("/auth/login")
async def wms_login(payload: dict, response: Response):
    login_ = str(payload.get("login") or "").strip()
    password = str(payload.get("password") or "")
    if not login_ or not password:
        raise HTTPException(status_code=400, detail="Введите логин и пароль")

    def _check():
        _init()
        return db.fetchone("SELECT salt, pass_hash, role, client_id "
                           "FROM wms_users WHERE login = ?", (login_,))
    row = await asyncio.to_thread(_check)
    if not row or _hash(row[0], password) != row[1]:
        raise HTTPException(status_code=401, detail="Неверный логин или пароль")
    token = secrets.token_hex(32)
    expires = (datetime.utcnow() + timedelta(days=SESSION_DAYS)).isoformat()

    def _save():
        db.execute("INSERT INTO wms_sessions (token, login, role, client_id, expires) "
                   "VALUES (?,?,?,?,?)", (token, login_, row[2], row[3], expires))
        db.execute("DELETE FROM wms_sessions WHERE expires < ?",
                   (datetime.utcnow().isoformat(),))
    await asyncio.to_thread(_save)
    response.set_cookie(COOKIE, token, max_age=SESSION_DAYS * 86400,
                        httponly=True, samesite="lax")
    out = {"login": login_, "role": row[2], "client_id": row[3]}
    if row[2] == "client":
        name = await asyncio.to_thread(
            db.fetchone, "SELECT name FROM wms_clients WHERE id = ?", (row[3],))
        out["client_name"] = name[0] if name else ""
    return out


@router.post("/auth/logout")
async def wms_logout(request: Request, response: Response):
    token = request.cookies.get(COOKIE)
    if token:
        await asyncio.to_thread(
            db.execute, "DELETE FROM wms_sessions WHERE token = ?", (token,))
    response.delete_cookie(COOKIE)
    return {"ok": True}


@router.get("/auth/me")
async def wms_me(request: Request):
    sess = _require(request)
    if sess["role"] == "client":
        name = await asyncio.to_thread(
            db.fetchone, "SELECT name FROM wms_clients WHERE id = ?",
            (sess["client_id"],))
        sess["client_name"] = name[0] if name else ""
    return sess


# ── Клиенты ─────────────────────────────────────────────────────────────────

@router.get("/clients")
async def clients_list(request: Request):
    _require(request, staff_only=True)

    def _load():
        rows = db.fetchall(
            "SELECT id, code, name, inn, contact, tg_chat_id FROM wms_clients "
            "ORDER BY id")
        out = []
        for r in rows:
            cid = r[0]
            stock = db.fetchone(
                "SELECT COALESCE(SUM(qty),0) FROM wms_moves "
                "WHERE client_id = ? AND status = 'available'", (cid,))[0]
            month = datetime.utcnow().strftime("%Y-%m")
            billed = db.fetchone(
                "SELECT COALESCE(SUM(amount),0) FROM wms_billing "
                "WHERE client_id = ? AND occurred_at >= ?", (cid, month + "-01"))[0]
            out.append({"id": cid, "code": r[1], "name": r[2], "inn": r[3],
                        "contact": r[4], "tg_chat_id": r[5],
                        "stock_units": int(stock or 0),
                        "billed_month": round(float(billed or 0))})
        return out
    return {"clients": await asyncio.to_thread(_load)}


@router.post("/clients")
async def clients_create(payload: dict, request: Request):
    _require(request, staff_only=True)
    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Нужно название клиента")
    code = str(payload.get("code") or "").strip() or \
        "CL" + secrets.token_hex(2).upper()

    def _save():
      with _OP_LOCK:
        if db.fetchone("SELECT 1 FROM wms_clients WHERE code = ?", (code,)):
            raise HTTPException(status_code=400,
                                detail=f"Код клиента {code} уже занят")
        return _insert_id(
            "INSERT INTO wms_clients (code, name, inn, contact, tg_chat_id, "
            "tariff, settings, created_at) VALUES (?,?,?,?,?,?,?,?)",
            (code, name, str(payload.get("inn") or ""),
             str(payload.get("contact") or ""),
             str(payload.get("tg_chat_id") or ""),
             json.dumps(DEFAULT_TARIFF, ensure_ascii=False), "{}",
             datetime.utcnow().isoformat()))
    cid = await asyncio.to_thread(_save)
    # логин клиента (если задан)
    login_ = str(payload.get("login") or "").strip()
    password = str(payload.get("password") or "").strip()
    if login_ and password:
        salt = secrets.token_hex(8)
        await asyncio.to_thread(
            db.execute,
            "INSERT INTO wms_users (login, salt, pass_hash, role, client_id) "
            "VALUES (?,?,?,?,?)",
            (login_, salt, _hash(salt, password), "client", cid))
    return {"id": cid, "code": code}


@router.get("/clients/{cid}/tariff")
async def tariff_get(cid: int, request: Request):
    sess = _require(request)
    cid = _client_scope(sess, cid)
    row = await asyncio.to_thread(
        db.fetchone, "SELECT tariff FROM wms_clients WHERE id = ?", (cid,))
    if not row:
        raise HTTPException(status_code=404, detail="Клиент не найден")
    return {"tariff": json.loads(row[0] or "{}")}


@router.post("/clients/{cid}/tariff")
async def tariff_set(cid: int, payload: dict, request: Request):
    _require(request, staff_only=True)
    tariff = payload.get("tariff")
    if not isinstance(tariff, dict):
        raise HTTPException(status_code=400, detail="tariff должен быть объектом")
    await asyncio.to_thread(
        db.execute, "UPDATE wms_clients SET tariff = ? WHERE id = ?",
        (json.dumps(tariff, ensure_ascii=False), cid))
    return {"ok": True}


def _tariff_of(cid: int) -> dict:
    row = db.fetchone("SELECT tariff FROM wms_clients WHERE id = ?", (cid,))
    t = json.loads(row[0]) if row and row[0] else {}
    out = json.loads(json.dumps(DEFAULT_TARIFF))
    for k, v in (t or {}).items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k].update(v)     # storage и будущие вложенные секции
        else:
            out[k] = v
    return out


# ── Товары клиента ──────────────────────────────────────────────────────────

@router.get("/skus")
async def skus_list(request: Request, client_id: int | None = None):
    sess = _require(request)
    cid = _client_scope(sess, client_id)

    def _load():
        rows = db.fetchall(
            "SELECT id, code, name, volume_l, weight_g, value_rub, "
            "requires_expiry, length_cm, width_cm, height_cm "
            "FROM wms_skus WHERE client_id = ? ORDER BY code", (cid,))
        bcs = db.fetchall(
            "SELECT b.sku_id, b.barcode FROM wms_barcodes b "
            "JOIN wms_skus s ON s.id = b.sku_id WHERE s.client_id = ?", (cid,))
        bmap: dict = {}
        for sid, bc in bcs:
            bmap.setdefault(sid, []).append(bc)
        return [{"id": r[0], "code": r[1], "name": r[2], "volume_l": r[3],
                 "weight_g": r[4], "value_rub": r[5],
                 "requires_expiry": bool(r[6]),
                 "length_cm": r[7], "width_cm": r[8], "height_cm": r[9],
                 "barcodes": bmap.get(r[0], [])} for r in rows]
    return {"skus": await asyncio.to_thread(_load), "client_id": cid}


@router.post("/skus")
async def skus_upsert(payload: dict, request: Request):
    _require(request, staff_only=True)
    cid = int(payload.get("client_id") or 0)
    items = payload.get("items") or []
    if not cid or not items:
        raise HTTPException(status_code=400, detail="Нужны client_id и items")

    def _save():
      with _OP_LOCK:
        n = 0
        for it in items:
            code = str(it.get("code") or "").strip().upper()
            if not code:
                continue
            L = float(it.get("length_cm") or 0)
            Wd = float(it.get("width_cm") or 0)
            H = float(it.get("height_cm") or 0)
            vol = float(it.get("volume_l") or 0)
            if L > 0 and Wd > 0 and H > 0:
                vol = round(L * Wd * H / 1000, 3)   # объём из габаритов
            row = db.fetchone(
                "SELECT id FROM wms_skus WHERE client_id = ? AND code = ?",
                (cid, code))
            if row:
                sid = row[0]
                db.execute(
                    "UPDATE wms_skus SET name=?, volume_l=?, weight_g=?, "
                    "value_rub=?, requires_expiry=?, length_cm=?, "
                    "width_cm=?, height_cm=? WHERE id=?",
                    (str(it.get("name") or ""), vol,
                     float(it.get("weight_g") or 0),
                     float(it.get("value_rub") or 0),
                     1 if it.get("requires_expiry") else 0,
                     L or None, Wd or None, H or None, sid))
            else:
                sid = _insert_id(
                    "INSERT INTO wms_skus (client_id, code, name, volume_l, "
                    "weight_g, value_rub, requires_expiry, length_cm, "
                    "width_cm, height_cm, created_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (cid, code, str(it.get("name") or ""), vol,
                     float(it.get("weight_g") or 0),
                     float(it.get("value_rub") or 0),
                     1 if it.get("requires_expiry") else 0,
                     L or None, Wd or None, H or None,
                     datetime.utcnow().isoformat()))
            for bc in (it.get("barcodes") or []):
                bc = str(bc).strip()
                if bc:
                    db.execute(
                        "INSERT INTO wms_barcodes (barcode, sku_id) VALUES (?,?) "
                        "ON CONFLICT (barcode) DO UPDATE SET sku_id = excluded.sku_id"
                        if db.IS_PG else
                        "INSERT OR REPLACE INTO wms_barcodes VALUES (?,?)",
                        (bc, sid))
            n += 1
        return n
    n = await asyncio.to_thread(_save)
    return {"saved": n}


@router.get("/skus/template")
async def skus_template(request: Request):
    """Шаблон Excel для загрузки товаров: шапка + примеры + подсказки."""
    _require(request)

    def _build():
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Товары"
        head = ["Артикул", "Название", "Длина, см", "Ширина, см",
                "Высота, см", "Объём, л", "Вес, г",
                "Ценность, ₽", "Срок годности (да/нет)", "Штрихкод"]
        ws.append(head)
        for ci, c in enumerate(ws[1], start=1):
            c.font = Font(bold=True, color="FFFFFF")
            # авто-поле (объём) — серым: заполнять не нужно
            c.fill = PatternFill("solid",
                                 fgColor="9aa5a0" if ci == 6 else "1a7f5a")
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        samples = [
            ["ABC-001", "Крем для рук 75 мл", 12, 4, 3, None, 90, 250,
             "да", "4601234567890"],
            ["ABC-002", "Шампунь 400 мл", 7, 7, 19, None, 460, 380,
             "да", "4601234567891"],
            ["ABC-003", "Расчёска", 18, 5, 2, None, 40, 150, "нет", ""],
        ]
        for ri, row in enumerate(samples, start=2):
            row[5] = f"=C{ri}*D{ri}*E{ri}/1000"   # объём считается сам
            ws.append(row)
        # формула объёма на 500 строк вперёд — клиент просто заполняет размеры
        grey = PatternFill("solid", fgColor="eef1ef")
        for ri in range(2, 502):
            cell = ws.cell(row=ri, column=6)
            if ri >= 5:
                cell.value = f'=IF(C{ri}*D{ri}*E{ri}=0,"",C{ri}*D{ri}*E{ri}/1000)'
            cell.fill = grey
        ws.cell(row=1, column=6).comment = None
        for col, w in zip("ABCDEFGHIJ",
                          (14, 32, 10, 10, 10, 12, 10, 12, 11, 18)):
            ws.column_dimensions[col].width = w
        ws2 = wb.create_sheet("Как заполнять")
        tips = [
            "Артикул — обязательное поле, ваш код товара (латиница/цифры).",
            "Название — как товар называется у вас (видно на приёмке и в остатках).",
            "Длина / Ширина / Высота, см — габариты единицы В УПАКОВКЕ.",
            "СЕРАЯ колонка «Объём, л» заполняется автоматически из габаритов —",
            "   в неё ничего не вписывайте. По ней ступень тарифа сборки и хранение.",
            "Вес, г — вес единицы в упаковке.",
            "Ценность, ₽ — закупочная стоимость единицы: предел ответственности",
            "   склада по договору ответственного хранения.",
            "СГ — «да», если у товара есть срок годности: тогда при приёмке",
            "   каждой партии срок обязателен (учёт FEFO).",
            "Штрихкод — EAN с упаковки. Если штрихкодов несколько — укажите",
            "   основной, остальные добавим отдельно.",
            "",
            "Примеры в первых трёх строках листа «Товары» — замените своими.",
            "Порядок колонок менять можно, лишние колонки игнорируются.",
        ]
        for t in tips:
            ws2.append([t])
        ws2.column_dimensions["A"].width = 78
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
    data = await asyncio.to_thread(_build)
    from urllib.parse import quote
    from fastapi.responses import Response as _Resp
    fname = quote("Шаблон товаров WMS.xlsx")
    return _Resp(content=data,
                 media_type="application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet",
                 headers={"Content-Disposition":
                          f"attachment; filename*=UTF-8''{fname}"})


@router.post("/skus/import")
async def skus_import_file(request: Request):
    """Загрузка товаров файлом (.xlsx/.csv). Шапка распознаётся по словам:
    артикул/sku, название, объём(л), вес(г), ценность/стоимость, сг/срок,
    штрихкод/barcode. Лишние колонки игнорируются."""
    _require(request, staff_only=True)
    form = await request.form()
    cid = int(form.get("client_id") or 0)
    up = form.get("file")
    if not cid or up is None:
        raise HTTPException(status_code=400, detail="Нужны client_id и file")
    raw = await up.read()
    fname = (getattr(up, "filename", "") or "").lower()

    def _parse():
        rows: list[list] = []
        if fname.endswith(".csv") or fname.endswith(".txt"):
            import csv
            import io
            text = raw.decode("utf-8-sig", errors="replace")
            delim = ";" if text.count(";") >= text.count(",") else ","
            rows = [r for r in csv.reader(io.StringIO(text), delimiter=delim)]
        else:
            import io
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
        if not rows:
            raise HTTPException(status_code=400, detail="Файл пустой")
        # найти шапку в первых 5 строках
        def _low(r):
            return [str(v or "").strip().lower() for v in r]
        idx = {"code": 0, "name": 1, "volume_l": None, "weight_g": None,
               "value_rub": None, "expiry": None, "barcode": None,
               "length": None, "width": None, "height": None}
        start = 0
        for ri, r in enumerate(rows[:5]):
            cells = _low(r)
            if any("артикул" in c or c == "sku" for c in cells):
                for ci, c in enumerate(cells):
                    if "артикул" in c or c == "sku":
                        idx["code"] = ci
                    elif "назван" in c or "name" in c:
                        idx["name"] = ci
                    elif "длин" in c or "length" in c:
                        idx["length"] = ci
                    elif "ширин" in c or "width" in c:
                        idx["width"] = ci
                    elif "высот" in c or "height" in c:
                        idx["height"] = ci
                    elif "объ" in c or "литр" in c or "volume" in c:
                        idx["volume_l"] = ci
                    elif "вес" in c or "weight" in c:
                        idx["weight_g"] = ci
                    elif "ценност" in c or "стоимост" in c or "value" in c:
                        idx["value_rub"] = ci
                    elif c.startswith("сг") or "срок" in c or "expiry" in c:
                        idx["expiry"] = ci
                    elif "штрих" in c or "barcode" in c or "шк" == c:
                        idx["barcode"] = ci
                start = ri + 1
                break
        items = []
        def _n(r, i):
            if i is None or i >= len(r):
                return 0.0
            try:
                return float(str(r[i]).replace(",", "."))
            except (TypeError, ValueError):
                return 0.0
        def _t(r, i):
            return str(r[i] or "").strip() if i is not None and i < len(r) else ""
        for r in rows[start:]:
            if not r:
                continue
            code = _t(r, idx["code"]).upper()
            if not code or code.startswith("ИТОГО"):
                continue
            exp = _t(r, idx["expiry"]).lower()
            bc = _t(r, idx["barcode"])
            # штрихкоды из Excel часто приходят как 2.04e+12
            if bc.endswith(".0"):
                bc = bc[:-2]
            items.append({
                "code": code, "name": _t(r, idx["name"]),
                "length_cm": _n(r, idx["length"]),
                "width_cm": _n(r, idx["width"]),
                "height_cm": _n(r, idx["height"]),
                "volume_l": _n(r, idx["volume_l"]),
                "weight_g": _n(r, idx["weight_g"]),
                "value_rub": _n(r, idx["value_rub"]),
                "requires_expiry": exp in ("1", "да", "yes", "true", "+"),
                "barcodes": [bc] if bc else []})
        return items
    items = await asyncio.to_thread(_parse)
    if not items:
        raise HTTPException(status_code=400,
                            detail="Не нашёл ни одной строки с артикулом")
    res = await skus_upsert({"client_id": cid, "items": items}, request)
    return {"saved": res["saved"], "parsed": len(items)}


# ── Приёмка (ASN) ───────────────────────────────────────────────────────────

@router.get("/inbounds")
async def inbounds_list(request: Request, client_id: int | None = None,
                        status: str | None = None):
    sess = _require(request)
    cid = _client_scope(sess, client_id) if (sess["role"] == "client"
                                             or client_id) else None

    def _load():
        q = ("SELECT i.id, i.client_id, c.name, i.status, i.expected_date, "
             "i.act_no, i.created_at, i.closed_at, i.note, i.pack_type "
             "FROM wms_inbounds i "
             "JOIN wms_clients c ON c.id = i.client_id WHERE 1=1")
        params: list = []
        if cid:
            q += " AND i.client_id = ?"
            params.append(cid)
        if status:
            q += " AND i.status = ?"
            params.append(status)
        q += " ORDER BY i.id DESC LIMIT 100"
        rows = db.fetchall(q, tuple(params))
        # короба и палеты всех заявок — тремя запросами, не N×3
        boxes_by_ib: dict = {}
        pallets_by_ib: dict = {}
        ids = [r[0] for r in rows]
        if ids:
            ph = ",".join("?" * len(ids))
            for p in db.fetchall(
                    f"SELECT id, inbound_id, code FROM wms_inbound_pallets "
                    f"WHERE inbound_id IN ({ph}) ORDER BY id", tuple(ids)):
                pallets_by_ib.setdefault(p[1], []).append(
                    {"id": p[0], "code": p[2]})
            brows = db.fetchall(
                f"SELECT id, inbound_id, code, received, note, pallet_id "
                f"FROM wms_inbound_boxes WHERE inbound_id IN ({ph}) "
                f"ORDER BY id", tuple(ids))
            items_by_box: dict = {}
            bids = [b[0] for b in brows]
            if bids:
                ph2 = ",".join("?" * len(bids))
                irows = db.fetchall(
                    f"SELECT x.id, x.box_id, s.code, s.name, x.qty, "
                    f"x.expiry_date, s.requires_expiry "
                    f"FROM wms_inbound_box_items x "
                    f"JOIN wms_skus s ON s.id = x.sku_id "
                    f"WHERE x.box_id IN ({ph2}) ORDER BY s.code", tuple(bids))
                for ir in irows:
                    items_by_box.setdefault(ir[1], []).append({
                        "id": ir[0], "sku": ir[2], "name": ir[3],
                        "qty": ir[4], "expiry_date": ir[5],
                        "requires_expiry": ir[6]})
            for b in brows:
                boxes_by_ib.setdefault(b[1], []).append({
                    "id": b[0], "code": b[2], "received": b[3],
                    "note": b[4], "pallet_id": b[5],
                    "items": items_by_box.get(b[0], [])})
        out = []
        for r in rows:
            lines = db.fetchall(
                "SELECT l.id, s.code, s.name, l.qty_expected, l.qty_received, "
                "l.batch_no, l.expiry_date, l.discrepancy_type, "
                "l.discrepancy_qty FROM wms_inbound_lines l "
                "JOIN wms_skus s ON s.id = l.sku_id WHERE l.inbound_id = ? "
                "ORDER BY s.code", (r[0],))
            out.append({
                "id": r[0], "client_id": r[1], "client": r[2], "status": r[3],
                "expected_date": r[4], "act_no": r[5], "created_at": r[6],
                "closed_at": r[7], "note": r[8], "pack_type": r[9],
                "pallets": pallets_by_ib.get(r[0], []),
                "boxes": boxes_by_ib.get(r[0], []),
                "lines": [{"id": x[0], "sku": x[1], "name": x[2],
                           "qty_expected": x[3], "qty_received": x[4],
                           "batch_no": x[5], "expiry_date": x[6],
                           "discrepancy_type": x[7], "discrepancy_qty": x[8]}
                          for x in lines]})
        return out
    return {"inbounds": await asyncio.to_thread(_load)}


@router.post("/inbounds")
async def inbound_create(payload: dict, request: Request):
    """Создание поставки. Помимо строк может сразу создать структуру
    упаковки (мастер WB-стиля): pack_type box|mono, boxes N — короба;
    pallets M × boxes_per_pallet K — монопаллеты с коробами."""
    sess = _require(request)
    cid = _client_scope(sess, payload.get("client_id"))
    lines = payload.get("lines") or []   # можно пустую: наполняется коробами
    pack_type = str(payload.get("pack_type") or "box")
    n_boxes = min(max(int(payload.get("boxes") or 0), 0), 200)
    n_pallets = min(max(int(payload.get("pallets") or 0), 0), 50)
    per_pallet = min(max(int(payload.get("boxes_per_pallet") or 0), 0), 100)

    def _save():
      with _OP_LOCK:
        now = datetime.utcnow().isoformat()
        iid = _insert_id(
            "INSERT INTO wms_inbounds (client_id, status, expected_date, note, "
            "act_no, created_at, created_by, pack_type) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (cid, "expected", str(payload.get("expected_date") or ""),
             str(payload.get("note") or ""), "",
             now, sess["login"], pack_type))
        skipped = []
        for ln in lines:
            code = str(ln.get("sku") or "").strip().upper()
            row = db.fetchone(
                "SELECT id FROM wms_skus WHERE client_id = ? AND code = ?",
                (cid, code))
            if not row:
                skipped.append(code)
                continue
            db.execute(
                "INSERT INTO wms_inbound_lines (inbound_id, sku_id, "
                "qty_expected, qty_received) VALUES (?,?,?,0)",
                (iid, row[0], int(ln.get("qty") or 0)))
        if pack_type == "mono" and n_pallets > 0:
            for _ in range(n_pallets):
                seq = _next_seq("wms_inbound_pallets", iid) + 1
                pcode = f"MP-P-{iid:05d}-{seq:02d}"
                pid = _insert_id(
                    "INSERT INTO wms_inbound_pallets (inbound_id, code, "
                    "created_at) VALUES (?,?,?)", (iid, pcode, now))
                _make_boxes(iid, per_pallet, pid, now)
        elif n_boxes > 0:
            _make_boxes(iid, n_boxes, None, now)
        return iid, skipped
    iid, skipped = await asyncio.to_thread(_save)
    return {"id": iid, "skipped_unknown_skus": skipped}


# ── Короба поставки (WB ФБО-стиль): клиент сам пакует и клеит ШК ────────────

# Code 128: 107 паттернов ширин (бар/пробел попеременно), значения 0–106.
_C128 = (
    "212222 222122 222221 121223 121322 131222 122213 122312 132212 221213 "
    "221312 231212 112232 122132 122231 113222 123122 123221 223211 221132 "
    "221231 213212 223112 312131 311222 321122 321221 312212 322112 322211 "
    "212123 212321 232121 111323 131123 131321 112313 132113 132311 211313 "
    "231113 231311 112133 112331 132131 113123 113321 133121 313121 211331 "
    "231131 213113 213311 213131 311123 311321 331121 312113 312311 332111 "
    "314111 221411 431111 111224 111422 121124 121421 141122 141221 112214 "
    "112412 122114 122411 142112 142211 241211 221114 413111 241112 134111 "
    "111242 121142 121241 114212 124112 124211 411212 421112 421211 212141 "
    "214121 412121 111143 111341 131141 114113 114311 411113 411311 113141 "
    "114131 311141 411131 211412 211214 211232 2331112").split()


def _code128_svg(text: str, module: int = 2, height: int = 56) -> str:
    """Штрихкод Code 128 B как SVG — без внешних библиотек."""
    vals = [104] + [max(0, min(94, ord(c) - 32)) for c in text]
    vals.append((vals[0] + sum(i * v for i, v in enumerate(vals[1:], 1))) % 103)
    vals.append(106)
    x = 10 * module  # тихая зона
    rects = []
    for v in vals:
        for i, w in enumerate(_C128[v]):
            wpx = int(w) * module
            if i % 2 == 0:
                rects.append(f'<rect x="{x}" y="0" width="{wpx}" height="{height}"/>')
            x += wpx
    total = x + 10 * module
    return (f'<svg xmlns="http://www.w3.org/2000/svg" '
            f'viewBox="0 0 {total} {height}" preserveAspectRatio="none">'
            + "".join(rects) + "</svg>")


def _box_head(bid: int, sess: dict):
    """Короб + заявка; клиенту чужой короб отдаём как 404 (не светим)."""
    row = db.fetchone(
        "SELECT b.id, b.inbound_id, b.code, b.received, i.client_id, i.status "
        "FROM wms_inbound_boxes b JOIN wms_inbounds i ON i.id = b.inbound_id "
        "WHERE b.id = ?", (bid,))
    if not row or (sess["role"] == "client"
                   and int(row[4]) != int(sess["client_id"])):
        raise HTTPException(status_code=404, detail="Короб не найден")
    return row


def _inbound_open(iid: int, sess: dict) -> tuple:
    """Заявка, доступная на редактирование; чужому клиенту — 404."""
    head = db.fetchone(
        "SELECT client_id, status FROM wms_inbounds WHERE id = ?", (iid,))
    if not head or (sess["role"] == "client"
                    and int(head[0]) != int(sess["client_id"])):
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    if head[1] == "done":
        raise HTTPException(status_code=400, detail="Заявка уже принята")
    return head


def _next_seq(table: str, iid: int) -> int:
    """Следующий суффикс кода — от максимального (короба могли удалять)."""
    seq = 0
    for (code,) in db.fetchall(
            f"SELECT code FROM {table} WHERE inbound_id = ?", (iid,)):
        try:
            seq = max(seq, int(str(code).rsplit("-", 1)[1]))
        except (ValueError, IndexError):
            pass
    return seq


def _make_boxes(iid: int, count: int, pallet_id, now: str) -> list:
    seq = _next_seq("wms_inbound_boxes", iid)
    created = []
    for _ in range(count):
        seq += 1
        code = f"MP-{iid:05d}-{seq:02d}"
        bid = _insert_id(
            "INSERT INTO wms_inbound_boxes (inbound_id, code, note, "
            "received, created_at, pallet_id) VALUES (?,?,?,0,?,?)",
            (iid, code, "", now, pallet_id))
        created.append({"id": bid, "code": code, "pallet_id": pallet_id})
    return created


@router.post("/inbounds/{iid}/boxes")
async def box_create(iid: int, payload: dict, request: Request):
    sess = _require(request)
    count = min(max(int(payload.get("count") or 1), 1), 100)
    pallet_id = payload.get("pallet_id")

    def _save():
      with _OP_LOCK:
        _inbound_open(iid, sess)
        if pallet_id:
            p = db.fetchone(
                "SELECT id FROM wms_inbound_pallets WHERE id = ? "
                "AND inbound_id = ?", (int(pallet_id), iid))
            if not p:
                raise HTTPException(status_code=404, detail="Палета не найдена")
        return _make_boxes(iid, count,
                           int(pallet_id) if pallet_id else None,
                           datetime.utcnow().isoformat())
    return {"boxes": await asyncio.to_thread(_save)}


@router.post("/inbounds/{iid}/pallets")
async def pallet_create(iid: int, payload: dict, request: Request):
    """Палета + сразу N коробов в ней (WB-стиль «монопаллета/короба»)."""
    sess = _require(request)
    n_boxes = min(max(int(payload.get("boxes") or 0), 0), 100)

    def _save():
      with _OP_LOCK:
        _inbound_open(iid, sess)
        now = datetime.utcnow().isoformat()
        seq = _next_seq("wms_inbound_pallets", iid) + 1
        code = f"MP-P-{iid:05d}-{seq:02d}"
        pid = _insert_id(
            "INSERT INTO wms_inbound_pallets (inbound_id, code, created_at) "
            "VALUES (?,?,?)", (iid, code, now))
        boxes = _make_boxes(iid, n_boxes, pid, now)
        return {"id": pid, "code": code, "boxes": boxes}
    return {"pallet": await asyncio.to_thread(_save)}


@router.delete("/pallets/{pid}")
async def pallet_delete(pid: int, request: Request):
    """Удаляет палету вместе с её коробами и их составом."""
    sess = _require(request)

    def _del():
      with _OP_LOCK:
        row = db.fetchone(
            "SELECT inbound_id FROM wms_inbound_pallets WHERE id = ?", (pid,))
        if not row:
            raise HTTPException(status_code=404, detail="Палета не найдена")
        _inbound_open(int(row[0]), sess)
        for (bid,) in db.fetchall(
                "SELECT id FROM wms_inbound_boxes WHERE pallet_id = ?", (pid,)):
            db.execute("DELETE FROM wms_inbound_box_items WHERE box_id = ?",
                       (bid,))
        db.execute("DELETE FROM wms_inbound_boxes WHERE pallet_id = ?", (pid,))
        db.execute("DELETE FROM wms_inbound_pallets WHERE id = ?", (pid,))
    await asyncio.to_thread(_del)
    return {"ok": True}


@router.post("/boxes/{bid}/items")
async def box_items_set(bid: int, payload: dict, request: Request):
    """Полная замена состава короба: [{sku, qty, expiry_date, batch_no?}]."""
    sess = _require(request)
    items = payload.get("items") or []

    def _save():
      with _OP_LOCK:
        b = _box_head(bid, sess)
        if b[5] == "done":
            raise HTTPException(status_code=400, detail="Заявка уже принята")
        checked = []
        for it in items:
            code = str(it.get("sku") or "").strip().upper()
            qty = int(it.get("qty") or 0)
            if not code or qty <= 0:
                continue
            if qty > 1_000_000:
                raise HTTPException(status_code=400,
                                    detail=f"Неправдоподобное кол-во {qty} "
                                           f"для {code}")
            row = db.fetchone(
                "SELECT id, requires_expiry, code FROM wms_skus "
                "WHERE client_id = ? AND code = ?", (b[4], code))
            if not row:
                raise HTTPException(status_code=400,
                                    detail=f"Артикула {code} нет в справочнике")
            exp = str(it.get("expiry_date") or "").strip()
            if row[1] and not exp:
                raise HTTPException(
                    status_code=400,
                    detail=f"У {row[2]} обязателен срок годности")
            checked.append((row[0], qty, exp, str(it.get("batch_no") or "")))
        db.execute("DELETE FROM wms_inbound_box_items WHERE box_id = ?", (bid,))
        for sku_id, qty, exp, bno in checked:
            db.execute(
                "INSERT INTO wms_inbound_box_items (box_id, sku_id, qty, "
                "expiry_date, batch_no) VALUES (?,?,?,?,?)",
                (bid, sku_id, qty, exp, bno))
        return len(checked)
    return {"saved": await asyncio.to_thread(_save)}


@router.delete("/boxes/{bid}")
async def box_delete(bid: int, request: Request):
    sess = _require(request)

    def _del():
      with _OP_LOCK:
        b = _box_head(bid, sess)
        if b[5] == "done":
            raise HTTPException(status_code=400, detail="Заявка уже принята")
        db.execute("DELETE FROM wms_inbound_box_items WHERE box_id = ?", (bid,))
        db.execute("DELETE FROM wms_inbound_boxes WHERE id = ?", (bid,))
    await asyncio.to_thread(_del)
    return {"ok": True}


def _label_block(code: str, meta: str, rows_html: str, head_html: str) -> str:
    import html as _h
    return f"""<div class="lbl">
  {_code128_svg(code)}
  <div class="code">{_h.escape(code)}</div>
  <div class="meta">{meta}</div>
  <table><tr>{head_html}</tr>
  {rows_html}</table>
</div>"""


def _label_page(title: str, blocks: list[str]) -> HTMLResponse:
    import html as _h
    body = "\n".join(blocks)
    page = f"""<!DOCTYPE html><html lang="ru"><head><meta charset="utf-8">
<title>{_h.escape(title)}</title><style>
body {{ font: 14px/1.4 Arial, sans-serif; margin: 0; padding: 16px; color: #111; }}
.lbl {{ width: 100mm; border: 1px dashed #999; padding: 6mm; border-radius: 4px;
  margin-bottom: 10mm; page-break-after: always; }}
.lbl:last-child {{ page-break-after: auto; }}
svg {{ display: block; margin: 0 auto; width: 88mm; height: 20mm; }}
.code {{ font-size: 22px; font-weight: 800; letter-spacing: 3px; text-align: center; margin-top: 2mm; }}
.meta {{ color: #444; font-size: 12px; text-align: center; margin-top: 1mm; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 4mm; font-size: 12px; }}
td, th {{ border-bottom: 1px solid #ddd; padding: 3px 4px; text-align: left; }}
td.n, th.n {{ text-align: right; }}
.btn {{ margin: 0 0 12px; padding: 10px 18px; font-size: 15px; cursor: pointer; }}
@media print {{ .btn {{ display: none; }} .lbl {{ border: none; margin-bottom: 0; }} body {{ padding: 0; }} }}
</style></head><body>
<button class="btn" onclick="print()">🖨 Печать</button>
{body}</body></html>"""
    return HTMLResponse(page, headers={"Cache-Control": "no-store"})


def _box_label_html(code: str, client_name: str, iid: int, items) -> str:
    """items: [(sku_code, name, qty, expiry)]"""
    import html as _h
    rows = "".join(
        f"<tr><td>{_h.escape(str(r[0]))}</td><td>{_h.escape(str(r[1] or ''))}</td>"
        f"<td class='n'>{r[2]}</td><td>{_h.escape(str(r[3] or '—'))}</td></tr>"
        for r in items) or "<tr><td colspan='4'>Короб пуст</td></tr>"
    total = sum(r[2] or 0 for r in items)
    meta = (f"{_h.escape(client_name)} · заявка №{iid} · {total} шт · "
            f"Market Partners, Чехов")
    return _label_block(code, meta, rows,
                        "<th>Артикул</th><th>Название</th>"
                        "<th class='n'>шт</th><th>Срок годности</th>")


def _pallet_label_html(code: str, client_name: str, iid: int, boxes) -> str:
    """boxes: [(box_code, units)]"""
    import html as _h
    rows = "".join(
        f"<tr><td>{_h.escape(str(b[0]))}</td><td class='n'>{b[1]}</td></tr>"
        for b in boxes) or "<tr><td colspan='2'>Палета пуста</td></tr>"
    total = sum(b[1] or 0 for b in boxes)
    meta = (f"{_h.escape(client_name)} · заявка №{iid} · {len(boxes)} кор. · "
            f"{total} шт · Market Partners, Чехов")
    return _label_block(code, meta, rows,
                        "<th>Короб</th><th class='n'>шт</th>")


def _box_items_for_label(bid: int):
    return db.fetchall(
        "SELECT s.code, s.name, x.qty, x.expiry_date "
        "FROM wms_inbound_box_items x JOIN wms_skus s ON s.id = x.sku_id "
        "WHERE x.box_id = ? ORDER BY s.code", (bid,))


def _client_name(cid) -> str:
    row = db.fetchone("SELECT name FROM wms_clients WHERE id = ?", (cid,))
    return row[0] if row else ""


@router.get("/boxes/{bid}/label")
async def box_label(bid: int, request: Request):
    """Печатная этикетка короба: ШК Code128 + состав. Открывается в новой
    вкладке (cookie той же сессии), кнопка печати скрывается при печати."""
    sess = _require(request)

    def _load():
        b = _box_head(bid, sess)
        return b, _client_name(b[4]), _box_items_for_label(bid)
    b, client_name, items = await asyncio.to_thread(_load)
    code = str(b[2])
    return _label_page(code, [_box_label_html(code, client_name, b[1], items)])


@router.get("/pallets/{pid}/label")
async def pallet_label(pid: int, request: Request):
    sess = _require(request)

    def _load():
        row = db.fetchone(
            "SELECT p.id, p.inbound_id, p.code, i.client_id "
            "FROM wms_inbound_pallets p "
            "JOIN wms_inbounds i ON i.id = p.inbound_id WHERE p.id = ?", (pid,))
        if not row or (sess["role"] == "client"
                       and int(row[3]) != int(sess["client_id"])):
            raise HTTPException(status_code=404, detail="Палета не найдена")
        boxes = db.fetchall(
            "SELECT b.code, COALESCE((SELECT SUM(x.qty) "
            "FROM wms_inbound_box_items x WHERE x.box_id = b.id), 0) "
            "FROM wms_inbound_boxes b WHERE b.pallet_id = ? ORDER BY b.id",
            (pid,))
        return row, _client_name(row[3]), boxes
    row, client_name, boxes = await asyncio.to_thread(_load)
    code = str(row[2])
    return _label_page(code,
                       [_pallet_label_html(code, client_name, row[1], boxes)])


@router.get("/inbounds/{iid}/labels")
async def inbound_labels(iid: int, request: Request):
    """Все этикетки заявки одной страницей: палеты + короба, по одной на
    лист печати — как «Печать всех ШК» в WB."""
    sess = _require(request)

    def _load():
        head = db.fetchone(
            "SELECT client_id FROM wms_inbounds WHERE id = ?", (iid,))
        if not head or (sess["role"] == "client"
                        and int(head[0]) != int(sess["client_id"])):
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        cname = _client_name(head[0])
        blocks = []
        for pid, pcode in db.fetchall(
                "SELECT id, code FROM wms_inbound_pallets "
                "WHERE inbound_id = ? ORDER BY id", (iid,)):
            pboxes = db.fetchall(
                "SELECT b.code, COALESCE((SELECT SUM(x.qty) "
                "FROM wms_inbound_box_items x WHERE x.box_id = b.id), 0) "
                "FROM wms_inbound_boxes b WHERE b.pallet_id = ? ORDER BY b.id",
                (pid,))
            blocks.append(_pallet_label_html(str(pcode), cname, iid, pboxes))
        for bid, bcode in db.fetchall(
                "SELECT id, code FROM wms_inbound_boxes "
                "WHERE inbound_id = ? ORDER BY id", (iid,)):
            blocks.append(_box_label_html(
                str(bcode), cname, iid, _box_items_for_label(bid)))
        return blocks
    blocks = await asyncio.to_thread(_load)
    if not blocks:
        raise HTTPException(status_code=400,
                            detail="В заявке нет коробов и палет")
    return _label_page(f"ШК заявки №{iid}", blocks)


def _norm_date(v) -> str:
    """Дата из Excel/строки → YYYY-MM-DD (принимает ДД.ММ.ГГГГ и ISO)."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return ""
    for sep in (".", "/"):
        p = s.split(sep)
        if len(p) == 3 and len(p[2]) == 4:
            try:
                return f"{int(p[2]):04d}-{int(p[1]):02d}-{int(p[0]):02d}"
            except ValueError:
                return s[:10]
    return s[:10]


@router.get("/inbounds/{iid}/boxes/template")
async def boxes_template(iid: int, request: Request):
    """Шаблон Excel раскладки по коробам — со штрихкодами коробов, как у WB.
    Одна строка = один товар в коробе; несколько товаров в коробе —
    несколько строк с одним ШК короба."""
    sess = _require(request)

    def _build():
        head = db.fetchone(
            "SELECT client_id FROM wms_inbounds WHERE id = ?", (iid,))
        if not head or (sess["role"] == "client"
                        and int(head[0]) != int(sess["client_id"])):
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        cid = head[0]
        boxes = db.fetchall(
            "SELECT id, code FROM wms_inbound_boxes "
            "WHERE inbound_id = ? ORDER BY id", (iid,))
        if not boxes:
            raise HTTPException(
                status_code=400,
                detail="Сначала добавьте короба (или палету с коробами) — "
                       "шаблон выгружается с их штрихкодами")
        items_by_box: dict = {}
        for bid, sku, qty, exp in db.fetchall(
                "SELECT x.box_id, s.code, x.qty, x.expiry_date "
                "FROM wms_inbound_box_items x "
                "JOIN wms_skus s ON s.id = x.sku_id "
                "JOIN wms_inbound_boxes b ON b.id = x.box_id "
                "WHERE b.inbound_id = ? ORDER BY x.id", (iid,)):
            items_by_box.setdefault(bid, []).append((sku, qty, exp))
        skus = db.fetchall(
            "SELECT code, name, requires_expiry FROM wms_skus "
            "WHERE client_id = ? ORDER BY code", (cid,))

        import io
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Раскладка"
        ws.append(["ШК короба", "Артикул или баркод", "Кол-во",
                   "Срок годности"])
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1a7f5a")
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        grey_font = Font(color="6b7671")
        for _bid, code in boxes:
            rows = items_by_box.get(_bid) or [("", None, "")]
            for sku, qty, exp in rows:
                ws.append([code, sku, qty, exp])
                ws.cell(row=ws.max_row, column=1).font = grey_font
        for col, w in zip("ABCD", (16, 22, 9, 14)):
            ws.column_dimensions[col].width = w
        ws2 = wb.create_sheet("Товары")
        ws2.append(["Артикул", "Название", "Срок годности обязателен"])
        for c in ws2[1]:
            c.font = Font(bold=True)
        for r in skus:
            ws2.append([r[0], r[1], "да" if r[2] else ""])
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 40
        ws2.column_dimensions["C"].width = 24
        ws3 = wb.create_sheet("Как заполнять")
        for t in (
            "Одна строка = один товар в одном коробе.",
            "Несколько товаров в коробе — несколько строк с одним ШК короба.",
            "ШК короба уже проставлены — копируйте код в новые строки.",
            "Вместо полного кода можно писать номер короба: 1, 2, 3…",
            "Артикул — из листа «Товары»; можно указать баркод с упаковки.",
            "Срок годности — в формате ДД.ММ.ГГГГ (для товаров с пометкой",
            "   «срок годности обязателен» — обязательно).",
            "Загрузка файла ПОЛНОСТЬЮ заменяет состав упомянутых коробов.",
            "Короба, которых нет в файле, не меняются.",
        ):
            ws3.append([t])
        ws3.column_dimensions["A"].width = 72
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
    data = await asyncio.to_thread(_build)
    from urllib.parse import quote
    from fastapi.responses import Response as _Resp
    fname = quote(f"Раскладка по коробам — заявка {iid}.xlsx")
    return _Resp(content=data,
                 media_type="application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet",
                 headers={"Content-Disposition":
                          f"attachment; filename*=UTF-8''{fname}"})


@router.post("/inbounds/{iid}/boxes/import")
async def boxes_import(iid: int, request: Request):
    """Раскладка по коробам файлом (.xlsx/.csv). Понимает и наш шаблон,
    и шаблон WB («Баркод товара | Кол-во товаров | ШК короба | Срок
    годности»). Всё или ничего: при любой ошибке файл не применяется."""
    sess = _require(request)
    form = await request.form()
    up = form.get("file")
    if up is None:
        raise HTTPException(status_code=400, detail="Нужен файл")
    raw = await up.read()
    fname = (getattr(up, "filename", "") or "").lower()

    def _rows():
        if fname.endswith(".csv") or fname.endswith(".txt"):
            import csv
            import io
            text = raw.decode("utf-8-sig", errors="replace")
            delim = ";" if text.count(";") >= text.count(",") else ","
            return [r for r in csv.reader(io.StringIO(text), delimiter=delim)]
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        return [list(r) for r in wb.active.iter_rows(values_only=True)]

    def _save():
      with _OP_LOCK:
        cid, _status = _inbound_open(iid, sess)
        rows = _rows()
        if not rows:
            raise HTTPException(status_code=400, detail="Файл пустой")
        # шапка: наш шаблон и шаблон WB распознаются по словам
        idx = {"box": None, "sku": None, "qty": None, "expiry": None}
        start = 0
        for ri, r in enumerate(rows[:5]):
            cells = [str(v or "").strip().lower() for v in r]
            if any("короб" in c for c in cells):
                for ci, c in enumerate(cells):
                    if "короб" in c and idx["box"] is None:
                        idx["box"] = ci
                    elif "артикул" in c or "баркод" in c or "sku" in c:
                        idx["sku"] = ci
                    elif "кол" in c or "qty" in c:
                        idx["qty"] = ci
                    elif "срок" in c or "годн" in c or "expiry" in c:
                        idx["expiry"] = ci
                start = ri + 1
                break
        if idx["box"] is None or idx["sku"] is None or idx["qty"] is None:
            raise HTTPException(
                status_code=400,
                detail="Не нашёл колонки: нужны «ШК короба», "
                       "«Артикул/баркод» и «Кол-во»")
        boxes = {str(c): b for b, c in db.fetchall(
            "SELECT id, code FROM wms_inbound_boxes "
            "WHERE inbound_id = ? ORDER BY id", (iid,))}
        by_suffix = {}
        for code, bid in boxes.items():
            try:
                by_suffix[int(code.rsplit("-", 1)[1])] = bid
            except (ValueError, IndexError):
                pass
        sku_by_code = {}
        sku_meta = {}
        for sid, code, req in db.fetchall(
                "SELECT id, code, requires_expiry FROM wms_skus "
                "WHERE client_id = ?", (cid,)):
            sku_by_code[str(code).upper()] = sid
            sku_meta[sid] = (str(code), bool(req))
        for bc, sid in db.fetchall(
                "SELECT bc.barcode, bc.sku_id FROM wms_barcodes bc "
                "JOIN wms_skus s ON s.id = bc.sku_id WHERE s.client_id = ?",
                (cid,)):
            sku_by_code.setdefault(str(bc).upper(), sid)

        # фаза 1: разобрать и проверить ВСЁ, ошибки собрать разом
        plan: dict = {}     # box_id -> [(sku_id, qty, expiry)]
        errors: list = []
        units = 0
        for ri, r in enumerate(rows[start:], start=start + 1):
            def _cell(i):
                return r[i] if i is not None and i < len(r) else None
            box_raw = str(_cell(idx["box"]) or "").strip()
            sku_raw = str(_cell(idx["sku"]) or "").strip()
            if sku_raw.endswith(".0"):
                sku_raw = sku_raw[:-2]
            if not box_raw and not sku_raw:
                continue
            try:
                qty = int(float(str(_cell(idx["qty"]) or 0).replace(",", ".")))
            except ValueError:
                qty = 0
            if not sku_raw and not qty:
                continue    # строка пустого короба из шаблона
            bid = boxes.get(box_raw)
            if bid is None and box_raw.isdigit():
                bid = by_suffix.get(int(box_raw))
            if bid is None:
                errors.append(f"стр. {ri}: короб «{box_raw}» не найден")
                continue
            if not sku_raw:
                errors.append(f"стр. {ri}: есть кол-во, но не указан артикул")
                continue
            sid = sku_by_code.get(sku_raw.upper())
            if sid is None:
                errors.append(f"стр. {ri}: артикул/баркод «{sku_raw}» "
                              f"не найден в справочнике")
                continue
            if qty <= 0 or qty > 1_000_000:
                errors.append(f"стр. {ri}: неверное кол-во для «{sku_raw}»")
                continue
            exp = _norm_date(_cell(idx["expiry"]))
            code_h, req = sku_meta[sid]
            if req and not exp:
                errors.append(f"стр. {ri}: у {code_h} обязателен срок "
                              f"годности")
                continue
            plan.setdefault(bid, []).append((sid, qty, exp))
            units += qty
        if errors:
            raise HTTPException(status_code=400,
                                detail="Файл не применён:\n" +
                                       "\n".join(errors[:12]))
        if not plan:
            raise HTTPException(status_code=400,
                                detail="В файле нет ни одной строки раскладки")
        # фаза 2: заменить состав упомянутых коробов
        for bid, items in plan.items():
            db.execute("DELETE FROM wms_inbound_box_items WHERE box_id = ?",
                       (bid,))
            for sid, qty, exp in items:
                db.execute(
                    "INSERT INTO wms_inbound_box_items (box_id, sku_id, qty, "
                    "expiry_date, batch_no) VALUES (?,?,?,?,?)",
                    (bid, sid, qty, exp, ""))
        return {"boxes": len(plan), "units": units}
    return await asyncio.to_thread(_save)


@router.post("/inbounds/{iid}/receive_boxes")
async def inbound_receive_boxes(iid: int, payload: dict, request: Request):
    """Приёмка по коробам (WB-стиль): склад отмечает принятые короба,
    факт и сроки годности предзаполнены раскладкой клиента."""
    sess = _require(request, staff_only=True)
    boxes = payload.get("boxes") or []
    mode = str(payload.get("receive_mode") or "pallet")
    pallets = int(payload.get("pallets") or 0)

    def _save():
      with _OP_LOCK:
        head = db.fetchone(
            "SELECT client_id, status FROM wms_inbounds WHERE id = ?", (iid,))
        if not head:
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        cid, status = head
        if status == "done":
            raise HTTPException(status_code=400, detail="Заявка уже принята")
        box_rows = db.fetchall(
            "SELECT id FROM wms_inbound_boxes WHERE inbound_id = ?", (iid,))
        if not box_rows:
            raise HTTPException(status_code=400, detail="В заявке нет коробов")
        t = _tariff_of(cid)
        now = datetime.utcnow().isoformat()
        items_by_box: dict = {}
        declared: dict = {}   # sku_id -> заявлено раскладкой по коробам
        for xid, bx, sku_id, qty, exp, bno in db.fetchall(
                "SELECT x.id, x.box_id, x.sku_id, x.qty, x.expiry_date, "
                "x.batch_no FROM wms_inbound_box_items x "
                "JOIN wms_inbound_boxes b ON b.id = x.box_id "
                "WHERE b.inbound_id = ?", (iid,)):
            items_by_box.setdefault(bx, []).append((xid, sku_id, qty, exp, bno))
            declared[sku_id] = declared.get(sku_id, 0) + (qty or 0)
        payload_map = {int(x.get("box_id") or 0): x for x in boxes}
        # фаза 1: валидация всего до первой записи
        by_key: dict = {}     # (sku_id, expiry, batch_no) -> принятое кол-во
        got_boxes = []
        total_units = 0
        for (bx,) in box_rows:
            pb = payload_map.get(bx) or {}
            received = bool(pb.get("received", True))
            overrides = {int(x.get("item_id") or 0): x
                         for x in (pb.get("items") or [])}
            for xid, sku_id, qty, exp, bno in items_by_box.get(bx, []):
                ov = overrides.get(xid) or {}
                q = int(ov.get("qty", qty) or 0) if received else 0
                e = str(ov.get("expiry_date", exp) or "").strip()
                if q < 0 or q > 1_000_000:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Неправдоподобное кол-во {q} в коробе — "
                               f"проверь поле «Факт» (не отсканирован ли ШК)")
                if q > 0:
                    req = db.fetchone(
                        "SELECT requires_expiry, code FROM wms_skus "
                        "WHERE id = ?", (sku_id,))
                    if req and req[0] and not e:
                        raise HTTPException(
                            status_code=400,
                            detail=f"У {req[1]} обязателен срок годности")
                    key = (sku_id, e, str(bno or ""))
                    by_key[key] = by_key.get(key, 0) + q
                    total_units += q
            got_boxes.append((bx, 1 if received else 0))
        # фаза 2: партии по (SKU, СГ) + движения
        for (sku_id, exp, bno), q in by_key.items():
            batch_id = _insert_id(
                "INSERT INTO wms_batches (client_id, sku_id, batch_no, "
                "expiry_date, received_at, inbound_id) VALUES (?,?,?,?,?,?)",
                (cid, sku_id, bno, exp, now, iid))
            db.execute(
                "INSERT INTO wms_moves (client_id, sku_id, batch_id, qty, "
                "status, doc_type, doc_ref, note, user_login, created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (cid, sku_id, batch_id, q, "available", "receipt",
                 f"IN-{iid}", "", sess["login"], now))
        for bx, rec in got_boxes:
            db.execute("UPDATE wms_inbound_boxes SET received = ? WHERE id = ?",
                       (rec, bx))
        # строки заявки: qty_received по факту; чего не было в строках —
        # добавляем с ожиданием из раскладки (клиент заявил её коробами)
        recv_by_sku: dict = {}
        for (sku_id, _e, _b), q in by_key.items():
            recv_by_sku[sku_id] = recv_by_sku.get(sku_id, 0) + q
        line_rows = db.fetchall(
            "SELECT id, sku_id, qty_expected FROM wms_inbound_lines "
            "WHERE inbound_id = ?", (iid,))
        line_by_sku = {r[1]: (r[0], r[2]) for r in line_rows}
        for sku_id in set(declared) | set(recv_by_sku) | set(line_by_sku):
            got = recv_by_sku.get(sku_id, 0)
            if sku_id in line_by_sku:
                lid, exp_q = line_by_sku[sku_id]
            else:
                exp_q = declared.get(sku_id, 0)
                lid = _insert_id(
                    "INSERT INTO wms_inbound_lines (inbound_id, sku_id, "
                    "qty_expected, qty_received) VALUES (?,?,?,0)",
                    (iid, sku_id, exp_q))
            disc_type = None
            disc_qty = 0
            if got != exp_q:
                disc_type = "short" if got < exp_q else "over"
                disc_qty = got - exp_q
            db.execute(
                "UPDATE wms_inbound_lines SET qty_received=?, "
                "discrepancy_type=?, discrepancy_qty=? WHERE id=?",
                (got, disc_type, disc_qty, lid))
        act_no = f"ПР-{iid:05d}"
        db.execute(
            "UPDATE wms_inbounds SET status='done', closed_at=?, act_no=? "
            "WHERE id=?", (now, act_no, iid))
        if mode == "pallet" and pallets > 0:
            price = float(t.get("receive_pallet") or 0)
            db.execute(
                "INSERT INTO wms_billing (client_id, service_code, qty, price, "
                "amount, occurred_at, source_type, source_ref, sku_id, note) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (cid, "RECEIVE_PALLET", pallets, price, pallets * price,
                 now, "inbound", act_no, None, ""))
        elif total_units > 0:
            code = ("RECEIVE_UNIT_SORTED" if mode == "unit_sorted"
                    else "RECEIVE_UNIT")
            price = float(t.get("receive_unit_sorted" if mode == "unit_sorted"
                                else "receive_unit") or 0)
            db.execute(
                "INSERT INTO wms_billing (client_id, service_code, qty, price, "
                "amount, occurred_at, source_type, source_ref, sku_id, note) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (cid, code, total_units, price, total_units * price,
                 now, "inbound", act_no, None, ""))
        boxes_ok = sum(1 for _, rec in got_boxes if rec)
        return {"act_no": act_no, "units": total_units,
                "boxes_received": boxes_ok, "boxes_total": len(got_boxes)}
    return await asyncio.to_thread(_save)


@router.post("/inbounds/{iid}/receive")
async def inbound_receive(iid: int, payload: dict, request: Request):
    """Завершение приёмки: факт по строкам → партии + движения + начисления."""
    sess = _require(request, staff_only=True)
    lines = payload.get("lines") or []
    mode = str(payload.get("receive_mode") or "pallet")  # pallet|unit|unit_sorted
    pallets = int(payload.get("pallets") or 0)

    def _save():
      with _OP_LOCK:
        head = db.fetchone(
            "SELECT client_id, status FROM wms_inbounds WHERE id = ?", (iid,))
        if not head:
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        cid, status = head
        if status == "done":
            raise HTTPException(status_code=400, detail="Заявка уже принята")
        t = _tariff_of(cid)
        now = datetime.utcnow().isoformat()
        # фаза 1: валидация ВСЕХ строк до первой записи — чтобы ошибка
        # в середине не оставляла в базе половину приёмки
        checked = []
        for ln in lines:
            lid = int(ln.get("line_id") or 0)
            qty = int(ln.get("qty_received") or 0)
            lrow = db.fetchone(
                "SELECT sku_id, qty_expected FROM wms_inbound_lines "
                "WHERE id = ? AND inbound_id = ?", (lid, iid))
            if not lrow:
                continue
            sku_id, qty_exp = lrow
            req = db.fetchone(
                "SELECT requires_expiry, code FROM wms_skus WHERE id = ?",
                (sku_id,))
            if req and req[0] and not str(ln.get("expiry_date") or "").strip():
                raise HTTPException(
                    status_code=400,
                    detail=f"У {req[1]} обязателен срок годности")
            checked.append((ln, lid, qty, sku_id, qty_exp))
        if not checked:
            raise HTTPException(status_code=400,
                                detail="Ни одна строка не относится к заявке")
        total_units = 0
        for ln, lid, qty, sku_id, qty_exp in checked:
            disc_type = None
            disc_qty = 0
            if qty != qty_exp:
                disc_type = "short" if qty < qty_exp else "over"
                disc_qty = qty - qty_exp
            if str(ln.get("discrepancy_type") or "").strip():
                disc_type = str(ln["discrepancy_type"]).strip()
                disc_qty = int(ln.get("discrepancy_qty") or disc_qty)
            db.execute(
                "UPDATE wms_inbound_lines SET qty_received=?, batch_no=?, "
                "expiry_date=?, discrepancy_type=?, discrepancy_qty=?, note=? "
                "WHERE id=?",
                (qty, str(ln.get("batch_no") or ""),
                 str(ln.get("expiry_date") or ""), disc_type, disc_qty,
                 str(ln.get("note") or ""), lid))
            if qty <= 0:
                continue
            bid = _insert_id(
                "INSERT INTO wms_batches (client_id, sku_id, batch_no, "
                "expiry_date, received_at, inbound_id) VALUES (?,?,?,?,?,?)",
                (cid, sku_id, str(ln.get("batch_no") or ""),
                 str(ln.get("expiry_date") or ""), now, iid))
            db.execute(
                "INSERT INTO wms_moves (client_id, sku_id, batch_id, qty, "
                "status, doc_type, doc_ref, note, user_login, created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (cid, sku_id, bid, qty, "available", "receipt",
                 f"IN-{iid}", "", sess["login"], now))
            total_units += qty
        act_no = f"ПР-{iid:05d}"
        db.execute(
            "UPDATE wms_inbounds SET status='done', closed_at=?, act_no=? "
            "WHERE id=?", (now, act_no, iid))
        # начисление приёмки
        if mode == "pallet" and pallets > 0:
            price = float(t.get("receive_pallet") or 0)
            db.execute(
                "INSERT INTO wms_billing (client_id, service_code, qty, price, "
                "amount, occurred_at, source_type, source_ref, sku_id, note) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (cid, "RECEIVE_PALLET", pallets, price, pallets * price,
                 now, "inbound", act_no, None, ""))
        elif total_units > 0:
            code = "RECEIVE_UNIT_SORTED" if mode == "unit_sorted" else "RECEIVE_UNIT"
            price = float(t.get("receive_unit_sorted" if mode == "unit_sorted"
                                else "receive_unit") or 0)
            db.execute(
                "INSERT INTO wms_billing (client_id, service_code, qty, price, "
                "amount, occurred_at, source_type, source_ref, sku_id, note) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (cid, code, total_units, price, total_units * price,
                 now, "inbound", act_no, None, ""))
        return {"act_no": act_no, "units": total_units}
    res = await asyncio.to_thread(_save)
    return res


# ── Остатки ─────────────────────────────────────────────────────────────────

def _stock_rows(cid: int):
    rows = db.fetchall(
        "SELECT m.sku_id, s.code, s.name, s.volume_l, m.batch_id, "
        "b.batch_no, b.expiry_date, b.received_at, m.status, "
        "COALESCE(SUM(m.qty),0) "
        "FROM wms_moves m JOIN wms_skus s ON s.id = m.sku_id "
        "LEFT JOIN wms_batches b ON b.id = m.batch_id "
        "WHERE m.client_id = ? "
        "GROUP BY m.sku_id, s.code, s.name, s.volume_l, m.batch_id, "
        "b.batch_no, b.expiry_date, b.received_at, m.status "
        "HAVING COALESCE(SUM(m.qty),0) != 0 ORDER BY s.code", (cid,))
    return rows


@router.get("/stock")
async def stock(request: Request, client_id: int | None = None):
    sess = _require(request)
    cid = _client_scope(sess, client_id)

    def _load():
        t = _tariff_of(cid)
        free_days = int(t["storage"].get("free_days") or 30)
        today = datetime.utcnow().date()
        by_sku: dict = {}
        for r in _stock_rows(cid):
            sku_id, code, name, vol, bid, bno, expiry, received, status, qty = r
            e = by_sku.setdefault(code, {
                "sku": code, "name": name, "volume_l": vol,
                "available": 0, "quarantine": 0, "batches": []})
            if status == "available":
                e["available"] += qty
            else:
                e["quarantine"] += qty
            age = None
            free_left = None
            if received:
                age = (today - datetime.fromisoformat(received).date()).days
                free_left = max(free_days - age, 0)
            e["batches"].append({
                "batch_id": bid, "batch_no": bno, "expiry": expiry,
                "received_at": (received or "")[:10], "age_days": age,
                "free_days_left": free_left, "qty": qty, "status": status})
        return sorted(by_sku.values(), key=lambda x: x["sku"])
    return {"stock": await asyncio.to_thread(_load), "client_id": cid}


@router.get("/moves")
async def moves(request: Request, client_id: int | None = None,
                limit: int = 200):
    """Журнал движений (что и когда списалось/пришло) — виден и клиенту:
    его требование «видно, как остатки списываются» закрывается этой лентой."""
    sess = _require(request)
    cid = _client_scope(sess, client_id)

    def _load():
        return db.fetchall(
            "SELECT m.created_at, s.code, s.name, m.qty, m.status, "
            "m.doc_type, m.doc_ref, b.batch_no, m.user_login, m.note "
            "FROM wms_moves m JOIN wms_skus s ON s.id = m.sku_id "
            "LEFT JOIN wms_batches b ON b.id = m.batch_id "
            "WHERE m.client_id = ? ORDER BY m.id DESC LIMIT ?",
            (cid, min(int(limit or 200), 1000)))
    rows = await asyncio.to_thread(_load)
    DOC = {"receipt": "Приёмка", "ship": "Отгрузка заказа",
           "return": "Возврат", "writeoff": "Списание",
           "adjust": "Корректировка"}
    return {"client_id": cid, "moves": [{
        "at": (r[0] or "")[:16].replace("T", " "),
        "sku": r[1], "name": r[2], "qty": r[3],
        "status": r[4], "doc": DOC.get(r[5], r[5]), "ref": r[6],
        "batch_no": r[7], "user": r[8], "note": r[9]} for r in rows]}


# ── Операции: сборка/отгрузка, возврат, корректировка ───────────────────────

def _assembly_price(t: dict, volume_l: float) -> float:
    for tier in t.get("assembly_tiers") or []:
        if volume_l <= float(tier.get("max_l") or 0):
            return float(tier.get("price") or 0)
    tiers = t.get("assembly_tiers") or []
    return float(tiers[-1]["price"]) if tiers else 0.0


def _ensure_batch(cid: int, sku_id: int, ref: str, now: str) -> int:
    """Служебная партия, когда партий SKU не существует (возврат/плюс-
    корректировка в пустой сток): без неё товар выпадает из возраста
    хранения и FIFO."""
    return _insert_id(
        "INSERT INTO wms_batches (client_id, sku_id, batch_no, expiry_date, "
        "received_at, inbound_id) VALUES (?,?,?,?,?,NULL)",
        (cid, sku_id, f"SRV-{ref[:16]}", "", now))


def _fifo_consume(cid: int, sku_id: int, qty: int, doc_type: str,
                  doc_ref: str, user: str, now: str):
    """Списание FIFO по партиям (старые первыми). Возвращает списанное."""
    batches = db.fetchall(
        "SELECT m.batch_id, COALESCE(SUM(m.qty),0), b.received_at "
        "FROM wms_moves m LEFT JOIN wms_batches b ON b.id = m.batch_id "
        "WHERE m.client_id = ? AND m.sku_id = ? AND m.status='available' "
        "GROUP BY m.batch_id, b.received_at "
        "HAVING COALESCE(SUM(m.qty),0) > 0 ORDER BY b.received_at NULLS LAST"
        if db.IS_PG else
        "SELECT m.batch_id, COALESCE(SUM(m.qty),0), b.received_at "
        "FROM wms_moves m LEFT JOIN wms_batches b ON b.id = m.batch_id "
        "WHERE m.client_id = ? AND m.sku_id = ? AND m.status='available' "
        "GROUP BY m.batch_id, b.received_at "
        "HAVING COALESCE(SUM(m.qty),0) > 0 ORDER BY b.received_at",
        (cid, sku_id))
    left = qty
    for bid, avail, _recv in batches:
        if left <= 0:
            break
        take = min(left, int(avail))
        db.execute(
            "INSERT INTO wms_moves (client_id, sku_id, batch_id, qty, status, "
            "doc_type, doc_ref, note, user_login, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (cid, sku_id, bid, -take, "available", doc_type, doc_ref, "",
             user, now))
        left -= take
    return qty - left


@router.post("/ops/ship")
async def ops_ship(payload: dict, request: Request):
    """Сборка и отгрузка заказов FBS: [{ref, items:[{sku, qty}]}].
    Списание FIFO + начисление по ступени габарита (объём самой крупной
    единицы заказа) + доп. единицы."""
    sess = _require(request, staff_only=True)
    cid = _client_scope(sess, payload.get("client_id"))
    orders = payload.get("orders") or []
    if not orders:
        raise HTTPException(status_code=400, detail="Нет заказов")

    def _save():
      with _OP_LOCK:
        t = _tariff_of(cid)
        now = datetime.utcnow().isoformat()
        shipped, errors = 0, []
        for o in orders:
            ref = str(o.get("ref") or "").strip() or f"SHIP-{secrets.token_hex(3)}"
            agg: dict = {}
            for it in (o.get("items") or []):
                code = str(it.get("sku") or "").strip().upper()
                q = int(it.get("qty") or 0)
                if code and q > 0:
                    agg[code] = agg.get(code, 0) + q
            units = 0
            max_vol = 0.0
            plan = []
            for code, qty in agg.items():
                row = db.fetchone(
                    "SELECT id, volume_l FROM wms_skus "
                    "WHERE client_id = ? AND code = ?", (cid, code))
                if not row:
                    errors.append(f"{ref}: SKU {code} не найден")
                    plan = []
                    break
                avail = db.fetchone(
                    "SELECT COALESCE(SUM(qty),0) FROM wms_moves "
                    "WHERE client_id=? AND sku_id=? AND status='available'",
                    (cid, row[0]))[0]
                if int(avail or 0) < qty:
                    errors.append(f"{ref}: {code} не хватает "
                                  f"(есть {avail}, надо {qty})")
                    plan = []
                    break
                plan.append((row[0], code, qty, float(row[1] or 0)))
            if not plan:
                continue
            for sku_id, code, qty, vol in plan:
                done = _fifo_consume(cid, sku_id, qty, "ship", ref,
                                     sess["login"], now)
                if done != qty:   # под локом невозможно, но не молчим
                    errors.append(f"{ref}: {code} списано {done} из {qty}")
                units += done
                max_vol = max(max_vol, vol)
            price = _assembly_price(t, max_vol)
            db.execute(
                "INSERT INTO wms_billing (client_id, service_code, qty, price, "
                "amount, occurred_at, source_type, source_ref, sku_id, note) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (cid, "PICK_ORDER", 1, price, price, now, "order", ref, None,
                 f"{units} ед."))
            if units > 1:
                extra = float(t.get("extra_unit") or 0)
                db.execute(
                    "INSERT INTO wms_billing (client_id, service_code, qty, "
                    "price, amount, occurred_at, source_type, source_ref, "
                    "sku_id, note) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (cid, "PICK_ITEM_ADD", units - 1, extra,
                     (units - 1) * extra, now, "order", ref, None, ""))
            shipped += 1
        return {"shipped": shipped, "errors": errors}
    return await asyncio.to_thread(_save)


@router.post("/ops/return")
async def ops_return(payload: dict, request: Request):
    sess = _require(request, staff_only=True)
    cid = _client_scope(sess, payload.get("client_id"))
    code = str(payload.get("sku") or "").strip().upper()
    qty = int(payload.get("qty") or 0)
    verdict = str(payload.get("verdict") or "to_stock")
    if verdict not in ("to_stock", "damaged", "dispose"):
        raise HTTPException(status_code=400, detail="Вердикт: to_stock/damaged/dispose")
    if not code or qty <= 0:
        raise HTTPException(status_code=400, detail="Нужны sku и qty")

    def _save():
      with _OP_LOCK:
        row = db.fetchone(
            "SELECT id FROM wms_skus WHERE client_id = ? AND code = ?",
            (cid, code))
        if not row:
            raise HTTPException(status_code=404, detail=f"SKU {code} не найден")
        sku_id = row[0]
        now = datetime.utcnow().isoformat()
        ref = str(payload.get("ref") or "") or f"RET-{secrets.token_hex(3)}"
        # возврат кладём в самую старую живую партию (возраст не сбрасывается)
        b = db.fetchone(
            "SELECT id FROM wms_batches WHERE client_id=? AND sku_id=? "
            "ORDER BY received_at LIMIT 1", (cid, sku_id))
        bid = b[0] if b else _ensure_batch(cid, sku_id, ref, now)
        status = "available" if verdict == "to_stock" else "quarantine"
        if verdict != "dispose":
            db.execute(
                "INSERT INTO wms_moves (client_id, sku_id, batch_id, qty, "
                "status, doc_type, doc_ref, note, user_login, created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (cid, sku_id, bid, qty, status, "return", ref, verdict,
                 sess["login"], now))
        t = _tariff_of(cid)
        price = float(t.get("return_process") or 0)
        db.execute(
            "INSERT INTO wms_billing (client_id, service_code, qty, price, "
            "amount, occurred_at, source_type, source_ref, sku_id, note) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (cid, "RETURN_PROCESS", qty, price, qty * price, now, "return",
             ref, sku_id, verdict))
        return {"ok": True, "ref": ref}
    return await asyncio.to_thread(_save)


@router.post("/ops/adjust")
async def ops_adjust(payload: dict, request: Request):
    """Корректировка (сторно-стиль): qty_delta ±, причина обязательна."""
    sess = _require(request, staff_only=True)
    cid = _client_scope(sess, payload.get("client_id"))
    code = str(payload.get("sku") or "").strip().upper()
    delta = int(payload.get("qty_delta") or 0)
    reason = str(payload.get("reason") or "").strip()
    if not code or not delta or not reason:
        raise HTTPException(status_code=400,
                            detail="Нужны sku, qty_delta и причина")

    def _save():
      with _OP_LOCK:
        row = db.fetchone(
            "SELECT id FROM wms_skus WHERE client_id = ? AND code = ?",
            (cid, code))
        if not row:
            raise HTTPException(status_code=404, detail=f"SKU {code} не найден")
        now = datetime.utcnow().isoformat()
        if delta < 0:
            done = _fifo_consume(cid, row[0], -delta, "adjust",
                                 f"ADJ:{reason[:40]}", sess["login"], now)
            return {"adjusted": -done}
        b = db.fetchone(
            "SELECT id FROM wms_batches WHERE client_id=? AND sku_id=? "
            "ORDER BY received_at LIMIT 1", (cid, row[0]))
        bid = b[0] if b else _ensure_batch(cid, row[0], "ADJ", now)
        db.execute(
            "INSERT INTO wms_moves (client_id, sku_id, batch_id, qty, status, "
            "doc_type, doc_ref, note, user_login, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (cid, row[0], bid, delta, "available", "adjust",
             f"ADJ:{reason[:40]}", "", sess["login"], now))
        return {"adjusted": delta}
    return await asyncio.to_thread(_save)


# ── Начисления ──────────────────────────────────────────────────────────────

@router.get("/billing")
async def billing(request: Request, client_id: int | None = None,
                  date_from: str | None = None, date_to: str | None = None):
    sess = _require(request)
    cid = _client_scope(sess, client_id)
    f = (date_from or datetime.utcnow().strftime("%Y-%m-01"))
    t_ = (date_to or datetime.utcnow().strftime("%Y-%m-%d")) + "T23:59:59"

    def _load():
        rows = db.fetchall(
            "SELECT id, service_code, qty, price, amount, occurred_at, "
            "source_type, source_ref, note FROM wms_billing "
            "WHERE client_id = ? AND occurred_at >= ? AND occurred_at <= ? "
            "ORDER BY id DESC LIMIT 500", (cid, f, t_))
        totals = db.fetchall(
            "SELECT service_code, COALESCE(SUM(qty),0), COALESCE(SUM(amount),0) "
            "FROM wms_billing WHERE client_id = ? AND occurred_at >= ? "
            "AND occurred_at <= ? GROUP BY service_code", (cid, f, t_))
        return rows, totals
    rows, totals = await asyncio.to_thread(_load)
    return {
        "client_id": cid, "date_from": f, "date_to": t_[:10],
        "events": [{"id": r[0], "service": r[1],
                    "service_name": SERVICE_NAMES.get(r[1], r[1]),
                    "qty": r[2], "price": r[3], "amount": r[4],
                    "at": (r[5] or "")[:16].replace("T", " "),
                    "source": f"{r[6]} {r[7]}".strip(), "note": r[8]}
                   for r in rows],
        "totals": [{"service": r[0],
                    "service_name": SERVICE_NAMES.get(r[0], r[0]),
                    "qty": r[1], "amount": round(float(r[2] or 0), 2)}
                   for r in totals],
        "total": round(sum(float(r[2] or 0) for r in totals), 2),
    }


# ── Хранение: ежедневное начисление ─────────────────────────────────────────

def storage_accrue(day: str | None = None) -> dict:
    """Раз в день: снапшот остатков по партиям + начисление хранения
    партиям старше бесплатного периода. Идемпотентно по (day, batch):
    снапшот и начисление пишутся одной транзакцией."""
    with _OP_LOCK:
        return _storage_accrue_locked(day)


def _storage_accrue_locked(day: str | None = None) -> dict:
    _init()
    d = day or (datetime.utcnow() + timedelta(hours=3)).strftime("%Y-%m-%d")
    today = date.fromisoformat(d)
    clients = db.fetchall("SELECT id FROM wms_clients")
    written = 0
    billed = 0.0
    for (cid,) in clients:
        t = _tariff_of(cid)
        st = t.get("storage") or {}
        free_days = int(st.get("free_days") or 30)
        box_l = float(st.get("box_litres") or 77)
        bands = st.get("bands") or []
        rows = db.fetchall(
            "SELECT m.batch_id, m.sku_id, s.volume_l, b.received_at, "
            "COALESCE(SUM(m.qty),0) FROM wms_moves m "
            "JOIN wms_skus s ON s.id = m.sku_id "
            "LEFT JOIN wms_batches b ON b.id = m.batch_id "
            "WHERE m.client_id = ? AND m.status = 'available' "
            "GROUP BY m.batch_id, m.sku_id, s.volume_l, b.received_at "
            "HAVING COALESCE(SUM(m.qty),0) > 0", (cid,))
        for bid, sku_id, vol, received, qty in rows:
            if not bid or not received:
                continue
            age = (today - datetime.fromisoformat(received).date()).days
            volume = float(vol or 0) * int(qty)
            exists = db.fetchone(
                "SELECT 1 FROM wms_snapshots WHERE day = ? AND batch_id = ?",
                (d, bid))
            if exists:
                continue
            rate = 0.0
            if age > free_days:
                for band in bands:
                    if int(band.get("from_day") or 0) <= age \
                            <= int(band.get("to_day") or 0):
                        rate = float(band.get("box_day") or 0)
                        break
            boxes = volume / box_l if box_l > 0 else 0.0
            amount = round(boxes * rate, 2) if rate > 0 else 0.0
            # пара снапшот+начисление — одна транзакция: сбой между ними
            # не потеряет списание хранения (маркер идемпотентности = снапшот)
            with db._conn() as con:
                cur = con.cursor()
                cur.execute(db._tr(
                    "INSERT INTO wms_snapshots (day, client_id, sku_id, "
                    "batch_id, qty, volume_l, age_days) VALUES (?,?,?,?,?,?,?)"),
                    (d, cid, sku_id, bid, int(qty), volume, age))
                if amount > 0:
                    cur.execute(db._tr(
                        "INSERT INTO wms_billing (client_id, service_code, "
                        "qty, price, amount, occurred_at, source_type, "
                        "source_ref, sku_id, note) VALUES (?,?,?,?,?,?,?,?,?,?)"),
                        (cid, "STORAGE_DAY", round(boxes, 3), rate, amount,
                         d + "T03:00:00", "storage", f"batch-{bid}", sku_id,
                         f"возраст {age} дн"))
                con.commit()
            written += 1
            billed += amount
    return {"day": d, "snapshots": written, "billed": round(billed, 2)}


@router.post("/storage/accrue")
async def storage_accrue_now(request: Request):
    _require(request, staff_only=True)
    return await asyncio.to_thread(storage_accrue)


# ── Пользователи WMS (staff) ────────────────────────────────────────────────

@router.post("/users")
async def wms_user_upsert(payload: dict, request: Request):
    _require(request, staff_only=True)
    login_ = str(payload.get("login") or "").strip()
    password = str(payload.get("password") or "")
    role = str(payload.get("role") or "staff")
    client_id = payload.get("client_id")
    if not login_ or not password or role not in ("staff", "client"):
        raise HTTPException(status_code=400,
                            detail="Нужны login, password, role=staff|client")
    if role == "client" and not client_id:
        raise HTTPException(status_code=400,
                            detail="Для клиента нужен client_id")
    salt = secrets.token_hex(8)

    def _save():
        db.execute("DELETE FROM wms_users WHERE login = ?", (login_,))
        db.execute(
            "INSERT INTO wms_users (login, salt, pass_hash, role, client_id) "
            "VALUES (?,?,?,?,?)",
            (login_, salt, _hash(salt, password), role,
             int(client_id) if client_id else None))
    await asyncio.to_thread(_save)
    return {"ok": True}
