"""Reviews endpoints — auto-fetch from WB / Ozon / YM APIs."""
import asyncio
import logging

from fastapi import APIRouter, BackgroundTasks, Body, Query

import reviews_client as rc
import review_ai

router = APIRouter(prefix="/api/reviews", tags=["reviews"])
_log = logging.getLogger(__name__)

_bg_tasks: set = set()  # ссылки на фоновые задачи, чтобы их не собрал GC


@router.get("/data")
async def get_reviews_data(
    background_tasks: BackgroundTasks,
    platform: str = Query("all"),
    limit: int = Query(5000),
):
    """Return reviews, ratings, dynamics, stats. Triggers a background refresh if stale."""
    background_tasks.add_task(rc.refresh_all)
    return {
        "reviews":  rc.get_all_reviews(platform=platform, limit=limit),
        "ratings":  rc.get_rating_table(),
        "dynamics": rc.get_rating_dynamics(),
        "stats":    rc.get_stats(),
        "drafts":   rc.get_drafts(),
        "auto":     auto_status,
    }


@router.get("/rules")
async def get_reply_rules():
    """Правила стиля ИИ-ответов («так не пиши — пиши так»), редактируются с фронта."""
    return {"rules": await asyncio.to_thread(review_ai.get_rules)}


@router.post("/rules")
async def add_reply_rule(body: dict = Body(...)):
    text = str((body or {}).get("text") or "").strip()
    if not text:
        return {"error": "Пустое правило"}
    await asyncio.to_thread(review_ai.add_rule, text)
    return {"rules": await asyncio.to_thread(review_ai.get_rules)}


@router.delete("/rules/{rule_id}")
async def delete_reply_rule(rule_id: int):
    await asyncio.to_thread(review_ai.delete_rule, rule_id)
    return {"rules": await asyncio.to_thread(review_ai.get_rules)}


@router.get("/export")
async def export_reviews(
    platform: str = Query("all"),
    only_text: bool = Query(False),
    skus: str = Query(""),
    art: str = Query(""),
):
    """Выгрузка отзывов в Excel с теми же фильтрами, что на дашборде.

    skus — список выбранных артикулов через запятую (мультиселект);
    art — текстовый поиск (совместимость)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    rows = rc.get_all_reviews(platform=platform, limit=100000)
    if only_text:
        rows = [r for r in rows if r.get("text")]
    sku_set = {s.strip() for s in skus.split(",") if s.strip()}
    if sku_set:
        rows = [r for r in rows if (r.get("sku") or "") in sku_set]
    elif art:
        q = art.strip().lower()
        rows = [r for r in rows
                if q in (r.get("sku") or "").lower()
                or q in (r.get("name") or "").lower()
                or q in str(r.get("nm") or "")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отзывы"
    ws.append(["Площадка", "Артикул", "Название", "Группа", "Дата",
               "Рейтинг", "Текст отзыва", "Ответ"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4F46E5")
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append([
            r.get("platform", ""), r.get("sku", ""), r.get("name", ""),
            r.get("group", ""), r.get("date", ""), r.get("rating", ""),
            r.get("text", ""), r.get("answer", ""),
        ])
    for i, w in enumerate([10, 14, 34, 16, 12, 9, 70, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")
        row[7].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="reviews.xlsx"'})


@router.post("/refresh")
async def force_refresh(force: bool = Query(True)):
    """Запускает обновление с площадок в фоне и сразу отвечает.

    Полная синхронизация (WB+Ozon+YM с пагинацией) занимает десятки секунд —
    Render обрывает запросы на ~30с, поэтому ждать её в запросе нельзя.
    Фронт поллит /data и подхватывает данные по мере готовности.
    """
    t = asyncio.create_task(rc.refresh_all(force=force))
    _bg_tasks.add(t)
    t.add_done_callback(_bg_tasks.discard)
    return {"ok": True, "scheduled": True, "stats": rc.get_stats()}


@router.get("/answer-stats")
async def answer_stats():
    """How many reviews already have our saved answer, per platform."""
    return rc.get_answer_stats()


@router.post("/analyze-style")
async def analyze_style(platform: str = Query("WB"), sample: int = Query(300)):
    """Analyze our past answers and return a reusable style guide."""
    return await review_ai.analyze_style(platform=platform, sample=sample)


@router.post("/draft")
async def make_draft(id: str = Query(...)):
    """Generate (or regenerate) an AI draft reply for one review."""
    review = rc.get_review(id)
    if not review:
        return {"error": "Отзыв не найден"}
    try:
        draft = await review_ai.generate_reply(review, platform=review["platform"])
    except Exception as e:      # ошибка API — понятным текстом, а не 500
        _log.warning("draft: %s", e)
        return {"error": f"Генерация не удалась: {str(e)[:200]}"}
    if not draft:
        return {"error": "Не удалось сгенерировать (проверьте ANTHROPIC_API_KEY)"}
    rc.save_draft(id, draft, status="pending")
    return {"id": id, "draft": draft, "status": "pending"}


@router.post("/draft-batch")
async def make_drafts(platform: str = Query("all"), limit: int = Query(20)):
    """Generate drafts for up to `limit` unanswered reviews without a draft."""
    platforms = ["WB", "Ozon", "YM"] if platform == "all" else [platform]
    todo = []
    for p in platforms:
        todo += rc.get_unanswered(platform=p, limit=limit)
    made = 0
    for review in todo[:limit]:
        try:
            draft = await review_ai.generate_reply(review, platform=review["platform"])
        except Exception as e:
            _log.warning("draft-batch: %s", e)
            break
        if draft:
            rc.save_draft(review["id"], draft, status="pending")
            made += 1
    return {"generated": made, "requested": len(todo[:limit])}


@router.post("/approve")
async def approve_draft(
    id: str = Query(...),
    publish: bool = Query(True),
    body: dict = Body(default=None),
):
    """Approve a draft (optionally edited). If publish and WB review — post to WB."""
    d = rc.get_draft(id)
    if not d:
        return {"error": "Черновик не найден"}
    text = ((body or {}).get("text") or d["draft"] or "").strip()
    if not text:
        return {"error": "Текст ответа пустой"}
    published, msg = False, ""
    if publish:
        published, msg = await rc.post_answer(id, text)
        if not published:
            low = (msg or "").lower()
            # отзыв скрыт/удалён площадкой (Ozon 404 GetReviewByUUID и т.п.) —
            # ответить на него нельзя никогда, убираем из очереди
            if "content not found" in low or "getreviewbyuuid" in low \
                    or "review not found" in low:
                rc.set_draft_status(id, "gone")
                return {"id": id, "status": "gone", "published": False,
                        "gone": True,
                        "error": "Отзыв уже скрыт или удалён площадкой — "
                                 "ответить на него нельзя. Убрал из очереди."}
            return {"id": id, "status": "pending", "published": False, "error": msg}
    rc.save_draft(id, text, status="approved")
    return {"id": id, "status": "approved", "published": published, "message": msg}


@router.post("/decline")
async def decline_draft(id: str = Query(...)):
    """Mark a draft as declined."""
    rc.set_draft_status(id, "declined")
    return {"id": id, "status": "declined"}


# ── Автоответы: 3-5★ отвечаем и публикуем сами, 1-2★ остаются людям ─────────

AUTO_MIN_RATING = 3

# телеметрия автоответов — видна на фронте (когда прогонялся, что сделал)
auto_status: dict = {"last_run": "", "result": None, "error": ""}


def _is_gone(msg: str) -> bool:
    low = (msg or "").lower()
    return ("content not found" in low or "getreviewbyuuid" in low
            or "review not found" in low)


def _is_empty_ban(msg: str) -> bool:
    """Ozon: «cannot comment on empty review» — на отзыв без текста ответить
    нельзя в принципе. Помечаем skip, чтобы не долбить API каждые 20 минут."""
    return "empty review" in (msg or "").lower()


async def auto_reply_pass(limit: int = 60) -> dict:
    """Один проход автоответов: сгенерировать и опубликовать ответы на
    неотвеченные отзывы с рейтингом >= AUTO_MIN_RATING (включая беститекстовые).
    Негатив (1-2★) не трогаем — его смотрят люди. Старые pending-черновики
    на позитив публикуются без повторной генерации."""
    from datetime import datetime, timedelta
    auto_status["last_run"] = (datetime.utcnow()
                               + timedelta(hours=3)).strftime("%H:%M")
    auto_status["error"] = ""
    published = generated = failed = gone = 0
    errors: list = []          # первые сообщения площадок — для диагностики

    def _note(rid, msg):
        if len(errors) < 3:
            errors.append(f"{rid}: {str(msg)[:140]}")
    skipped = 0
    # 1) готовые черновики на позитив — просто публикуем
    for d in rc.get_pending_autoable(AUTO_MIN_RATING, limit):
        ok, msg = await rc.post_answer(d["id"], d["draft"])
        if ok:
            rc.save_draft(d["id"], d["draft"], status="auto")
            published += 1
        elif _is_gone(msg):
            rc.set_draft_status(d["id"], "gone")
            gone += 1
        elif _is_empty_ban(msg):
            rc.set_draft_status(d["id"], "skip")
            skipped += 1
        else:
            failed += 1
            _note(d["id"], msg)
            _log.warning("auto publish %s: %s", d["id"], msg)
    # 2) свежие без черновика — генерим и публикуем
    todo = []
    for p in ("WB", "Ozon", "YM"):
        todo += [r for r in rc.get_unanswered(platform=p, limit=limit)
                 if int(r.get("rating") or 0) >= AUTO_MIN_RATING]
    for review in todo[:max(limit - published, 0)]:
        # Ozon не принимает комментарии к отзывам без текста — не тратим
        # генерацию, помечаем «ответ не требуется»
        if review["platform"] == "Ozon" and not (review.get("text") or "").strip():
            rc.save_draft(review["id"], "", status="skip")
            skipped += 1
            continue
        try:
            draft = await review_ai.generate_reply(
                review, platform=review["platform"])
        except Exception as e:
            _log.warning("auto generate: %s", e)
            auto_status["error"] = f"генерация: {str(e)[:180]}"
            break   # скорее всего лимит/ключ API — прерываем проход целиком
        if not draft:
            failed += 1
            continue
        generated += 1
        ok, msg = await rc.post_answer(review["id"], draft)
        if ok:
            rc.save_draft(review["id"], draft, status="auto")
            published += 1
        elif _is_gone(msg):
            rc.save_draft(review["id"], draft, status="gone")
            gone += 1
        elif _is_empty_ban(msg):
            rc.save_draft(review["id"], draft, status="skip")
            skipped += 1
        else:
            # площадка не приняла — оставляем как pending-черновик человеку
            rc.save_draft(review["id"], draft, status="pending")
            failed += 1
            _note(review["id"], msg)
            _log.warning("auto publish %s: %s", review["id"], msg)
    if published or failed or gone:
        _log.info("auto-reviews: published=%d generated=%d gone=%d failed=%d",
                  published, generated, gone, failed)
    res = {"published": published, "generated": generated, "gone": gone,
           "skipped": skipped, "failed": failed, "errors": errors}
    auto_status["result"] = res
    return res


@router.post("/auto-pass")
async def run_auto_pass(limit: int = Query(60)):
    """Ручной запуск прохода автоответов в фоне (Render обрывает запросы
    ~30с, а проход с генерацией может идти минуты). Результат — в поле
    auto ответа /data (строка статуса на фронте)."""
    auto_status["error"] = ""
    auto_status["result"] = None
    t = asyncio.create_task(auto_reply_pass(min(int(limit or 60), 200)))
    _bg_tasks.add(t)
    t.add_done_callback(_bg_tasks.discard)
    return {"scheduled": True}
