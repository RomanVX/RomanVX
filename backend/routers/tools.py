"""Инструменты WB. Первый: Продуктолог — анализ отзывов по каждому артикулу.

Из всех собранных отзывов WB по SKU строится сводка: сильные и слабые
стороны товара с частотой упоминаний (%), и рекомендация к доработке.
Анализ делает Claude, результат кэшируется в БД и пересобирается в фоне
только для артикулов, где появились новые отзывы.
"""
import asyncio
import base64
import os
import json
import logging
import re
import time as _time
from datetime import datetime, timedelta

from fastapi import APIRouter, File, HTTPException, Query, Request, UploadFile

import db
from config import ANTHROPIC_API_KEY

_log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/tools", tags=["tools"])

_building = False
_error = ""
_progress = ""

_MODEL = "claude-opus-4-8"
_MAX_REVIEWS_PER_SKU = 120
_MIN_TEXTS = 3          # меньше 3 текстовых отзывов — анализировать нечего


def _init_table():
    db.execute("CREATE TABLE IF NOT EXISTS productolog "
               "(sku TEXT PRIMARY KEY, data TEXT, reviews_count INTEGER, built_at TEXT)")


def _wb_reviews_by_sku() -> dict[str, list]:
    """{sku: [(rating, text, date)]} — все отзывы WB из БД."""
    rows = db.fetchall(
        "SELECT sku, rating, text, created_at FROM reviews "
        "WHERE platform = 'WB' AND sku IS NOT NULL AND sku != '' "
        "ORDER BY created_at DESC")
    out: dict[str, list] = {}
    for sku, rating, text, dt in rows:
        out.setdefault(sku, []).append((int(rating or 0), (text or "").strip(), dt))
    return out


def _stats(reviews: list) -> dict:
    n = len(reviews)
    pos = sum(1 for r, _, _ in reviews if r >= 4)
    neu = sum(1 for r, _, _ in reviews if r == 3)
    neg = n - pos - neu
    avg = round(sum(r for r, _, _ in reviews) / n, 2) if n else 0
    return {"count": n, "avg": avg,
            "pos": round(pos / n * 100) if n else 0,
            "neu": round(neu / n * 100) if n else 0,
            "neg": round(neg / n * 100) if n else 0}


async def _analyze_sku(sku: str, name: str, reviews: list) -> dict | None:
    """LLM-анализ отзывов одного артикула → {pluses, minuses, recommendation}."""
    texts = [(r, t) for r, t, _ in reviews if t][:_MAX_REVIEWS_PER_SKU]
    if len(texts) < _MIN_TEXTS:
        return None
    body = "\n".join(f"[{r}★] {t[:400]}" for r, t in texts)
    prompt = f"""Ты — продуктолог маркетплейс-селлера. Проанализируй отзывы покупателей о товаре «{name}» (артикул {sku}).

ОТЗЫВЫ ({len(texts)} шт., в скобках оценка):
{body}

Верни СТРОГО JSON без пояснений и без markdown:
{{
 "pluses":  [{{"tag": "короткая формулировка плюса (2-4 слова)", "pct": число % отзывов с этим плюсом}}],
 "minuses": [{{"tag": "короткая формулировка минуса", "pct": число}}],
 "recommendation": "1-3 предложения: что конкретно доработать в продукте/упаковке/карточке"
}}
Правила: до 6 плюсов и до 6 минусов, сортируй по частоте, pct — целое число (доля отзывов, где тема упомянута), формулировки конкретные («Течёт крышка», а не «Плохое качество»). Если минусов нет — пустой список и рекомендация «Существенных проблем в отзывах нет»."""
    import anthropic
    client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
    msg = await client.messages.create(
        model=_MODEL, max_tokens=1200,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = msg.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.strip("`")
        raw = raw[raw.find("{"):raw.rfind("}") + 1]
    try:
        data = json.loads(raw[raw.find("{"):raw.rfind("}") + 1])
    except (ValueError, TypeError):
        _log.warning("productolog %s: не распарсился ответ LLM", sku)
        return None
    return {"pluses": data.get("pluses") or [],
            "minuses": data.get("minuses") or [],
            "recommendation": str(data.get("recommendation") or "")}


def _set_progress(msg: str) -> None:
    global _progress
    _progress = msg


async def _build_bg(force: bool = False) -> None:
    global _building, _error, _progress
    if _building:
        return
    _building = True
    _error = ""
    try:
        if not ANTHROPIC_API_KEY:
            _error = "ANTHROPIC_API_KEY не настроен"
            return
        import catalog as _cat
        _init_table()
        by_sku = await asyncio.to_thread(_wb_reviews_by_sku)
        cached = {r[0]: r[2] for r in await asyncio.to_thread(
            db.fetchall, "SELECT sku, data, reviews_count FROM productolog")}
        # пересобираем только SKU, где отзывов стало заметно больше
        todo = []
        for sku, revs in by_sku.items():
            n_texts = sum(1 for _, t, _ in revs if t)
            if n_texts < _MIN_TEXTS:
                continue
            if not force and sku in cached and n_texts <= (cached[sku] or 0) + 2:
                continue
            todo.append((sku, revs, n_texts))
        todo.sort(key=lambda x: -x[2])
        total = len(todo)
        done = 0
        errors = 0
        sem = asyncio.Semaphore(4)   # 4 товара анализируем параллельно

        async def _one(sku, revs, n_texts):
            nonlocal done, errors
            name = _cat.lookup(sku).get("name", sku)
            async with sem:
                try:
                    data = await _analyze_sku(sku, name, revs)
                except Exception as e:
                    errors += 1
                    _log.warning("productolog %s: %s", sku, e)
                    return f"{sku}: {str(e)[:150]}"
                if data:
                    await asyncio.to_thread(
                        db.execute,
                        "INSERT INTO productolog (sku, data, reviews_count, built_at) VALUES (?,?,?,?) "
                        "ON CONFLICT (sku) DO UPDATE SET data = excluded.data, "
                        "reviews_count = excluded.reviews_count, built_at = excluded.built_at",
                        (sku, json.dumps(data, ensure_ascii=False), n_texts,
                         datetime.utcnow().isoformat()))
            return None

        async def _tracked(sku, revs, n_texts):
            nonlocal done
            err = await _one(sku, revs, n_texts)
            done += 1
            _set_progress(f"проанализировано {done} из {total}"
                          + (f", ошибок {errors}" if errors else ""))
            return err

        errs = await asyncio.gather(*[_tracked(s, r, n) for s, r, n in todo])
        errs = [e for e in errs if e]
        if errs:
            _error = f"{len(errs)} арт. с ошибкой (напр. {errs[0]})"
        _log.info("productolog: обновлено %d SKU, ошибок %d", total - len(errs), len(errs))
    except Exception as e:
        _error = str(e)[:300]
        _log.error("productolog build: %s", e)
    finally:
        _building = False
        _progress = ""


# держим ссылки на фоновые задачи (иначе GC их убивает)
_bg: set = set()


def _spawn(coro):
    # тяжёлый шлюз тут сознательно НЕ используется: сборки инструментов
    # (LLM-анализ, fullstats 1 req/мин) длятся десятки минут, но памяти
    # почти не потребляют — под heavy.guard они заблокировали бы финансы
    t = asyncio.get_event_loop().create_task(coro)
    _bg.add(t)
    t.add_done_callback(_bg.discard)


@router.get("/productolog")
async def get_productolog(refresh: bool = Query(default=False)):
    """Продуктолог: анализ отзывов WB по артикулам."""
    from config import USE_MOCK
    if USE_MOCK:
        import mock_data
        return mock_data.generate_productolog()
    import catalog as _cat
    _init_table()
    if refresh and not _building:
        _spawn(_build_bg(force=False))
    by_sku = await asyncio.to_thread(_wb_reviews_by_sku)
    rows = await asyncio.to_thread(
        db.fetchall, "SELECT sku, data, built_at FROM productolog")
    analyzed = {r[0]: (json.loads(r[1] or "{}"), r[2]) for r in rows}
    items = []
    for sku, revs in by_sku.items():
        st = _stats(revs)
        if st["count"] < _MIN_TEXTS:
            continue
        info = _cat.lookup(sku)
        a, built = analyzed.get(sku, ({}, None))
        # анализировать можно только по ТЕКСТОВЫМ отзывам; если их мало —
        # товар не «ждёт анализа», а просто не подлежит ему (звёзды без слов)
        n_texts = sum(1 for _, t, _ in revs if t)
        analyzable = n_texts >= _MIN_TEXTS
        items.append({
            "sku": sku, "name": info.get("name", sku), "group": info.get("brand", ""),
            **st,
            "pluses": a.get("pluses") or [],
            "minuses": a.get("minuses") or [],
            "recommendation": a.get("recommendation") or "",
            "analyzed": bool(a), "analyzable": analyzable,
            "text_reviews": n_texts, "built_at": (built or "")[:10],
            "wb_link": _wb_link(sku),
        })
    # проблемные сверху: доля негатива, затем количество отзывов
    items.sort(key=lambda x: (-x["neg"], -x["count"]))
    # «ждут анализа» — только те, что реально можно проанализировать
    pending = sum(1 for it in items if not it["analyzed"] and it["analyzable"])
    if pending and not _building:
        _spawn(_build_bg(force=False))
    return {"items": items, "building": _building, "progress": _progress,
            "error": _error, "pending": pending}


_ART_TO_NM: dict = {}


def _wb_link(sku: str) -> str:
    """Ссылка на карточку WB по артикулу (через nmId из каталога)."""
    global _ART_TO_NM
    if not _ART_TO_NM:
        try:
            import catalog as _cat
            m = getattr(_cat, "WB_ID_TO_ART", {}) or {}
            # nmId→art  ⇒  art(upper)→nmId (первый выигрывает)
            for nm, art in m.items():
                key = str(art).strip().upper()
                if key and key not in _ART_TO_NM:
                    _ART_TO_NM[key] = str(nm)
        except Exception:
            pass
    nm = _ART_TO_NM.get(str(sku).strip().upper())
    return f"https://www.wildberries.ru/catalog/{nm}/detail.aspx" if nm else ""


def _prod_reason(it: dict) -> str:
    """Сухое обоснование по фактам: рейтинг, отзывы, доля оценок."""
    avg = float(it.get("avg") or 0)
    n = int(it.get("count") or 0)
    return (f"Рейтинг {avg:.2f}★, отзывов {n}. "
            f"Положительных {it.get('pos', 0)}%, нейтральных {it.get('neu', 0)}%, "
            f"негативных {it.get('neg', 0)}%.")


@router.get("/productolog/export")
async def export_productolog():
    """Выгрузка продуктолога в Excel: лист 1 — сводка с обоснованием,
    лист 2 — подробные плюсы/минусы/рекомендации."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    data = await get_productolog(refresh=False)
    items = data.get("items", [])

    # нейтральный порядок — по числу отзывов (сортировку делает сам пользователь)
    verdict_items = sorted(items, key=lambda x: -int(x.get("count") or 0))

    wb = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="4F46E5")

    def _style_header(ws):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hfill
            c.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Лист 1: Сводка ──
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.append(["Артикул", "Наименование", "Группа", "Кол-во отзывов",
                "Рейтинг", "% негатива", "Обоснование", "Ссылка WB"])
    _style_header(ws1)
    link_font = Font(color="0563C1", underline="single")
    for it in verdict_items:
        if not it.get("analyzed") and it.get("count", 0) < _MIN_TEXTS:
            continue
        ws1.append([
            it.get("sku", ""), it.get("name", ""), it.get("group", ""),
            it.get("count", 0),
            round(it.get("avg", 0), 2) if it.get("avg") is not None else "",
            f'{it.get("neg", 0)}%', _prod_reason(it), "",
        ])
        link = _wb_link(it.get("sku", ""))
        if link:
            cell = ws1.cell(ws1.max_row, 8)
            cell.value = "Открыть на WB"
            cell.hyperlink = link
            cell.font = link_font
    for i, w in enumerate([12, 34, 14, 13, 10, 11, 55, 16], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    for row in ws1.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws1.freeze_panes = "A2"

    # ── Лист 2: Плюсы / Минусы / Рекомендации ──
    ws2 = wb.create_sheet("Плюсы-минусы")
    ws2.append(["Артикул", "Наименование", "Кол-во отзывов", "Рейтинг",
                "Плюсы (тема · %)", "Минусы (тема · %)", "Рекомендация к доработке",
                "Ссылка WB"])
    _style_header(ws2)

    def _chips(lst):
        return "\n".join(f"{p.get('tag','')} · {p.get('pct','')}%" for p in (lst or []))

    for it in sorted(items, key=lambda x: (-int(x.get("neg") or 0), -int(x.get("count") or 0))):
        ws2.append([
            it.get("sku", ""), it.get("name", ""), it.get("count", 0),
            round(it.get("avg", 0), 2) if it.get("avg") is not None else "",
            _chips(it.get("pluses")) if it.get("analyzed")
            else ("⏳ анализируется" if it.get("analyzable", True)
                  else f"мало текстовых отзывов ({it.get('text_reviews', 0)})"),
            _chips(it.get("minuses")) if it.get("analyzed") else "",
            it.get("recommendation", "") if it.get("analyzed") else "", "",
        ])
        link = _wb_link(it.get("sku", ""))
        if link:
            cell = ws2.cell(ws2.max_row, 8)
            cell.value = "Открыть на WB"
            cell.hyperlink = link
            cell.font = link_font
    for i, w in enumerate([12, 34, 13, 10, 34, 34, 55, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="productolog.xlsx"'})


# ══ Остатки по кластерам (федеральным округам) ════════════════════════════════

_CLUSTER_KEYS = [
    ("Центральный",      ["колед", "подольск", "электрост", "тула", "белые столбы",
                          "радумл", "вёшки", "вешки", "обухово", "чехов", "рязан",
                          "котовск", "голицын", "пушкино", "внуково", "москв",
                          "иваново", "ярослав", "воронеж", "белгород", "домодедово",
                          "сабурово", "черная грязь", "чёрная грязь", "щербинка"]),
    ("Северо-Западный",  ["санкт", "уткина", "шушары", "псков", "спб", "новосаратовка",
                          "петербург", "калининград", "мурманс"]),
    ("Южный",            ["краснодар", "тихорецк", "невинномыс", "волгоград", "ростов",
                          "астрахан", "крым", "симферопол", "ставропол"]),
    ("Приволжский",      ["казан", "самара", "новосемейкино", "сарапул", "уфа",
                          "нижний новгород", "пенза", "саратов", "чебоксар", "оренбург",
                          "киров", "перм", "ижевск"]),
    ("Уральский",        ["екатеринбург", "испытателей", "перспективн", "челябинск",
                          "тюмен", "курган", "сургут"]),
    ("Сибирский",        ["новосибир", "омск", "красноярск", "кемеров", "иркутск",
                          "барнаул", "томск"]),
    ("Дальневосточный",  ["хабаровск", "владивосток", "артём", "артем", "благовещенск"]),
]
_CLUSTER_ORDER = [c for c, _ in _CLUSTER_KEYS] + ["Прочее"]


def _cluster_of(name: str) -> str:
    low = (name or "").lower()
    for cluster, keys in _CLUSTER_KEYS:
        if any(k in low for k in keys):
            return cluster
    return "Прочее"


def _okrug_cluster(okrug: str) -> str:
    """«Центральный федеральный округ» → «Центральный»."""
    low = (okrug or "").lower()
    for cluster, _ in _CLUSTER_KEYS:
        if low.startswith(cluster.lower().split("-")[0][:6]):
            return cluster
    for cluster, _ in _CLUSTER_KEYS:
        if cluster.lower() in low:
            return cluster
    return "Прочее"


_clusters_cache: dict = {}
_clusters_ts: float = 0.0


@router.get("/clusters")
async def get_clusters(refresh: bool = Query(default=False)):
    """Остатки и продажи WB по кластерам (федеральным округам складов)."""
    import time as _time
    global _clusters_cache, _clusters_ts
    if not refresh and _clusters_cache and _time.monotonic() - _clusters_ts < 1800:
        return _clusters_cache

    from datetime import timedelta
    import analytics
    import cache
    DAYS = 28
    dt_to = datetime.utcnow()
    sales, _orders, stocks = await cache.get_raw_data(dt_to - timedelta(days=DAYS), dt_to)

    agg: dict[str, dict] = {c: {"stock": 0, "sold": 0, "loc_hit": 0, "loc_total": 0}
                            for c in _CLUSTER_ORDER}
    # SKU-разрез: остаток на складах кластера и СПРОС покупателей кластера
    sku_stock: dict = {c: {} for c in _CLUSTER_ORDER}   # кластер → sku → шт на складах
    sku_demand: dict = {c: {} for c in _CLUSTER_ORDER}  # кластер → sku → выкупов покупателями округа
    for s in stocks or []:
        cl = _cluster_of(s.get("warehouseName"))
        agg[cl]["stock"] += int(s.get("quantity") or 0)
        sku = (s.get("supplierArticle") or "").strip()
        if sku:
            sku_stock[cl][sku] = sku_stock[cl].get(sku, 0) + int(s.get("quantity") or 0)
    for r in sales or []:
        if not analytics._is_sale(r):
            continue
        wh_cl = _cluster_of(r.get("warehouseName"))
        agg[wh_cl]["sold"] += 1
        buyer = _okrug_cluster(r.get("oblastOkrugName") or r.get("regionName") or "")
        if buyer != "Прочее":
            agg[buyer]["loc_total"] += 1
            if buyer == wh_cl:
                agg[buyer]["loc_hit"] += 1
            sku = (r.get("supplierArticle") or "").strip()
            if sku:
                sku_demand[buyer][sku] = sku_demand[buyer].get(sku, 0) + 1

    TARGET_DAYS = 30
    items = []
    for c in _CLUSTER_ORDER:
        a = agg[c]
        if a["stock"] == 0 and a["sold"] == 0:
            continue
        spd = round(a["sold"] / DAYS, 2)
        coverage = round(a["stock"] / spd, 1) if spd > 0 else None
        need = max(0, round((TARGET_DAYS - (coverage or 0)) * spd)) if spd > 0 else 0
        if spd == 0:
            status = "no_sales"
        elif coverage < 7:
            status = "urgent"
        elif coverage < 15:
            status = "warn"
        elif coverage > 90:
            status = "over"
        else:
            status = "ok"
        loc = round(a["loc_hit"] / a["loc_total"] * 100) if a["loc_total"] else None
        # что именно везти: покрытие по SKU от СПРОСА округа (не от отгрузок)
        import catalog as _cat
        skus = []
        for sku, dem in sku_demand[c].items():
            d_spd = dem / DAYS
            if d_spd <= 0:
                continue
            st_here = sku_stock[c].get(sku, 0)
            cov = round(st_here / d_spd, 1)
            sku_need = max(0, round((TARGET_DAYS - cov) * d_spd))
            if sku_need > 0:
                skus.append({"sku": _cat.canon(sku), "name": _cat.lookup(sku).get("name", sku),
                             "demand_spd": round(d_spd, 2), "stock": st_here,
                             "coverage": cov, "need": sku_need})
        skus.sort(key=lambda x: -x["need"])
        # «Остальное»: прочие артикулы округа (в остатках или спросе), которых
        # нет в «что везти» — лежат без спроса, либо покрытие в норме
        in_need = {s["sku"] for s in skus}
        other = []
        all_sku = set(sku_stock[c].keys()) | set(sku_demand[c].keys())
        for sku in all_sku:
            canon = _cat.canon(sku)
            if canon in in_need:
                continue
            dem = sku_demand[c].get(sku, 0)
            d_spd = round(dem / DAYS, 2)
            st_here = sku_stock[c].get(sku, 0)
            cov = round(st_here / d_spd, 1) if d_spd > 0 else None
            other.append({"sku": canon, "name": _cat.lookup(sku).get("name", sku),
                          "demand_spd": d_spd, "stock": st_here,
                          "coverage": cov if cov is not None else "∞", "need": 0})
        other.sort(key=lambda x: -x["stock"])
        items.append({"cluster": c, "stock": a["stock"], "spd": spd,
                      "coverage": coverage, "need": need, "status": status,
                      "localization": loc, "demand": a["loc_total"],
                      "need_by_demand": sum(s["need"] for s in skus),
                      "skus": skus[:15], "other_skus": other})

    total_hit = sum(a["loc_hit"] for a in agg.values())
    total_dem = sum(a["loc_total"] for a in agg.values())
    result = {
        "items": items,
        "days": DAYS,
        "target_days": TARGET_DAYS,
        "localization_total": round(total_hit / total_dem * 100, 1) if total_dem else None,
        "weak": sum(1 for it in items if it["status"] in ("urgent", "warn")),
        "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC"),
    }
    _clusters_cache = result
    _clusters_ts = _time.monotonic()
    return result


@router.get("/clusters/export")
async def export_clusters():
    """Выгрузка остатков по кластерам в Excel: сводка по округам + листы
    «Что везти» и «Остальное» с артикулами."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    data = await get_clusters(refresh=False)
    items = data.get("items", [])
    wb = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="38bdf8")

    def _hdr(ws):
        for c in ws[1]:
            c.font = Font(bold=True, color="0b2a3a")
            c.fill = hfill
            c.alignment = Alignment(vertical="center", wrap_text=True)

    # Лист 1: сводка по округам
    ws1 = wb.active
    ws1.title = "Кластеры"
    ws1.append(["Округ", "Продаж/день", "Остаток, шт", "Покрытие, дн",
                "К заказу, шт", "Локализация, %", "Спрос (заказов)", "Статус"])
    _hdr(ws1)
    ST = {"urgent": "СРОЧНО", "warn": "мало", "ok": "ок", "over": "избыток",
          "no_sales": "нет продаж"}
    for it in items:
        ws1.append([it.get("cluster"), it.get("spd"), it.get("stock"),
                    it.get("coverage"), it.get("need"), it.get("localization"),
                    it.get("demand"), ST.get(it.get("status"), it.get("status"))])
    for i, w in enumerate([20, 13, 13, 13, 12, 15, 16, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # Листы 2 и 3: что везти / остальное
    def _sku_sheet(title, key):
        ws = wb.create_sheet(title)
        ws.append(["Округ", "Артикул", "Название", "Спрос/день", "Здесь, шт",
                   "Покрытие, дн", "Везти, шт"])
        _hdr(ws)
        for it in items:
            for s in it.get(key, []):
                ws.append([it.get("cluster"), s.get("sku"), s.get("name"),
                           s.get("demand_spd"), s.get("stock"), s.get("coverage"),
                           s.get("need")])
        for i, w in enumerate([18, 14, 34, 12, 11, 13, 11], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    _sku_sheet("Что везти", "skus")
    _sku_sheet("Остальное", "other_skus")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="clusters.xlsx"'})


# ══ Контроль рекламы WB ═══════════════════════════════════════════════════════

_ADV_TYPES = {4: "Каталог", 5: "Карточка", 6: "Поиск", 7: "Рекомендации",
              8: "Авто", 9: "Аукцион"}
_ADV_STATUS = {4: "готова", 7: "завершена", 8: "отказ", 9: "🟢 активна", 11: "⏸ пауза"}

_adv_building = False
_adv_error = ""
_adv_progress = ""
_ADV_DAYS = 28


def _adv_init():
    db.execute("CREATE TABLE IF NOT EXISTS adv_tool "
               "(campaign_id BIGINT PRIMARY KEY, data TEXT, built_at TEXT)")
    db.execute("CREATE TABLE IF NOT EXISTS kv_cache (k TEXT PRIMARY KEY, v TEXT)")


def _adv_verdict(c: dict) -> tuple[str, str]:
    """Правило-вердикт по кампании: (код, пояснение)."""
    spend, revenue, clicks = c["spend"], c["revenue"], c["clicks"]
    if spend < 100:
        return "idle", "почти нет расхода"
    if revenue <= 0:
        return "waste", f"потрачено {round(spend)} ₽ без заказов"
    drr = spend / revenue * 100
    if drr > 30:
        return "bad", f"ДРР {round(drr)}% — реклама дороже, чем приносит"
    if drr > 15:
        return "warn", f"ДРР {round(drr)}% — на грани"
    return "good", f"ДРР {round(drr)}% — эффективна"


async def _adv_build_bg(force: bool = False) -> None:
    global _adv_building, _adv_error, _adv_progress
    if _adv_building:
        return
    _adv_building = True
    _adv_error = ""
    try:
        import advert_client as ac
        _adv_init()
        _adv_progress = "список кампаний"
        ids = await ac.get_all_campaign_ids_ext()
        if not ids:
            _adv_error = "Кампании не найдены (проверьте права токена «Продвижение»)"
            return
        meta = await ac.get_campaigns_meta(ids)
        # фолбэк: тип/статус из promotion/count, если детали кампаний пусты
        cnt_meta = ac.get_count_meta()
        for aid, cm in cnt_meta.items():
            m = meta.setdefault(aid, {"name": f"Кампания {aid}", "nms": [], "bids": {}})
            if m.get("type") is None:
                m["type"] = cm.get("type")
            if m.get("status") is None:
                m["status"] = cm.get("status")
        end = (datetime.utcnow() + timedelta(hours=3)).date()
        begin = end - timedelta(days=_ADV_DAYS)
        _adv_progress = "статистика кампаний (fullstats, ~минута)"
        stats = await ac.get_fullstats_campaigns(ids, begin.isoformat(), end.isoformat())

        import catalog as _cat
        campaigns = []
        for aid in ids:
            m = meta.get(aid) or {}
            s = stats.get(aid) or {}
            spend = round(float(s.get("sum") or 0), 2)
            status = m.get("status")
            if spend <= 0 and status not in (9, 11):
                continue   # старые пустые кампании не показываем
            # какие SKU продвигает: nm-разбивка расходов из fullstats,
            # фолбэк — товары из настроек кампании
            sku_spend: dict[str, dict] = {}
            for nid, cell in (s.get("nms") or {}).items():
                art = _cat.resolve_wb(nid)
                sc = sku_spend.setdefault(art, {"sum": 0.0, "orders": 0})
                sc["sum"] += cell.get("sum", 0)
                sc["orders"] += cell.get("orders", 0)
            if not sku_spend:
                for nid in m.get("nms") or []:
                    sku_spend.setdefault(_cat.resolve_wb(nid), {"sum": 0.0, "orders": 0})
            skus = sorted(
                [{"sku": k, "spend": round(v["sum"]), "orders": v["orders"]}
                 for k, v in sku_spend.items()], key=lambda x: -x["spend"])[:12]
            campaigns.append({
                "id": aid, "name": m.get("name") or str(aid),
                "type": _ADV_TYPES.get(m.get("type"), str(m.get("type") or "")),
                "type_id": m.get("type"),
                "status": _ADV_STATUS.get(status, str(status or "")),
                "active": status == 9,
                "spend": spend,
                "views": s.get("views") or 0,
                "clicks": s.get("clicks") or 0,
                "ctr": round((s.get("clicks") or 0) / s["views"] * 100, 2) if s.get("views") else 0,
                "cpc": round(spend / s["clicks"], 1) if s.get("clicks") else None,
                "orders": s.get("orders") or 0,
                "revenue": round(float(s.get("sum_price") or 0)),
                "atbs": s.get("atbs") or 0,
                "mode": "Автоматическая" if m.get("type") == 8 else "Ручная",
                "bids": m.get("bids") or {},
                "skus": skus,
            })
        for c in campaigns:
            c["cpo"] = round(c["spend"] / c["orders"]) if c["orders"] else None
            c["drr"] = round(c["spend"] / c["revenue"] * 100, 1) if c["revenue"] else None
            v, why = _adv_verdict(c)
            c["verdict"], c["verdict_why"] = v, why

        def _save(c: dict):
            db.execute(
                "INSERT INTO adv_tool (campaign_id, data, built_at) VALUES (?,?,?) "
                "ON CONFLICT (campaign_id) DO UPDATE SET data = excluded.data, built_at = excluded.built_at",
                (c["id"], json.dumps(c, ensure_ascii=False), datetime.utcnow().isoformat()))

        # сохраняем сразу (без фраз) — рестарты сервера не теряют прогресс
        for c in campaigns:
            await asyncio.to_thread(_save, c)
        _log.info("adv tool: %d кампаний сохранено, собираем фразы", len(campaigns))

        # ключевые фразы — по кампаниям с расходом (щадяще, с паузами),
        # каждая кампания дозаписывается в БД по мере готовности
        with_spend = sorted([c for c in campaigns if c["spend"] > 100],
                            key=lambda x: -x["spend"])[:15]
        for i, c in enumerate(with_spend):
            _adv_progress = f"фразы {i + 1}/{len(with_spend)}: {c['name'][:30]}"
            try:
                words = await ac.get_campaign_words(c["id"], c.get("type_id"))
            except Exception:
                words = []
            for w in words:
                # кандидат в минус: тратит, почти не кликают, заказов не видно
                w["flag"] = ("minus" if w.get("sum", 0) > 300 and w.get("ctr", 0) < 1.5
                             else "hot" if w.get("clicks", 0) >= 20 and w.get("ctr", 0) >= 4
                             else "")
            words.sort(key=lambda w: -(w.get("sum") or w.get("views") or 0))
            c["words"] = words[:40]
            await asyncio.to_thread(_save, c)
            await asyncio.sleep(2)

        # LLM-совет по оптимизации (одним вызовом)
        try:
            if ANTHROPIC_API_KEY and campaigns:
                _adv_progress = "советы Claude"
                table = "\n".join(
                    f"- {c['name']} [{c['type']}, {c['status']}]: расход {c['spend']}₽, "
                    f"показы {c['views']}, клики {c['clicks']} (CTR {c['ctr']}%), "
                    f"заказы {c['orders']} на {c['revenue']}₽, ДРР {c['drr'] or '—'}%"
                    for c in sorted(campaigns, key=lambda x: -x["spend"])[:20])
                waste_words = []
                for c in campaigns:
                    for w in c.get("words") or []:
                        if w.get("flag") == "minus":
                            waste_words.append(f"«{w['phrase']}» ({c['name'][:25]}): {round(w['sum'])}₽, CTR {w['ctr']}%")
                prompt = f"""Ты — специалист по рекламе Wildberries. Данные кампаний селлера за {_ADV_DAYS} дней:
{table}

Фразы-кандидаты в минус (тратят, не кликаются):
{chr(10).join(waste_words[:25]) or 'нет'}

Дай 3-6 конкретных рекомендаций по оптимизации: что остановить/минусовать, куда перераспределить бюджет, где потенциал масштабирования. Формат: маркированный список, каждая рекомендация 1-2 предложения с цифрами. Без воды и вступлений."""
                import anthropic
                client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
                msg = await client.messages.create(model=_MODEL, max_tokens=900,
                                                   messages=[{"role": "user", "content": prompt}])
                advice = msg.content[0].text.strip()
                await asyncio.to_thread(
                    db.execute,
                    "INSERT INTO kv_cache (k, v) VALUES ('adv_advice', ?) "
                    "ON CONFLICT (k) DO UPDATE SET v = excluded.v",
                    (json.dumps({"text": advice, "at": datetime.utcnow().isoformat()},
                                ensure_ascii=False),))
        except Exception as e:
            _log.warning("adv advice: %s", e)
        _log.info("adv tool built: %d campaigns", len(campaigns))
    except Exception as e:
        _adv_error = str(e)[:300]
        _log.error("adv tool build: %s", e)
    finally:
        _adv_building = False
        _adv_progress = ""


@router.get("/adv")
async def get_adv(refresh: bool = Query(default=False)):
    """Контроль рекламы: кампании, ДРР, ключевые фразы, советы."""
    _adv_init()
    rows = await asyncio.to_thread(db.fetchall, "SELECT data, built_at FROM adv_tool")
    campaigns = []
    newest = ""
    for data, built in rows:
        try:
            campaigns.append(json.loads(data))
            newest = max(newest, built or "")
        except ValueError:
            pass
    stale = True
    if newest:
        try:
            stale = (datetime.utcnow() - datetime.fromisoformat(newest)).total_seconds() > 6 * 3600
        except ValueError:
            pass
    if (refresh or stale or not campaigns) and not _adv_building:
        _spawn(_adv_build_bg())
    advice = None
    row = await asyncio.to_thread(db.fetchone, "SELECT v FROM kv_cache WHERE k = 'adv_advice'")
    if row:
        try:
            advice = json.loads(row[0])
        except ValueError:
            pass
    campaigns.sort(key=lambda c: (-c.get("active", False), -(c.get("spend") or 0)))
    total_spend = round(sum(c.get("spend") or 0 for c in campaigns))
    total_rev = round(sum(c.get("revenue") or 0 for c in campaigns))
    waste = round(sum(c.get("spend") or 0 for c in campaigns if c.get("verdict") == "waste"))
    return {"campaigns": campaigns, "days": _ADV_DAYS,
            "total_spend": total_spend, "total_revenue": total_rev,
            "total_drr": round(total_spend / total_rev * 100, 1) if total_rev else None,
            "waste": waste,
            "advice": advice,
            "building": _adv_building, "progress": _adv_progress, "error": _adv_error,
            "built_at": newest[:16].replace("T", " ")}


# ══ Калькулятор ниши: выходить с товаром или нет ═══════════════════════════════

_NICHE_REVIEW_RATE = 0.04   # ~4% покупателей оставляют отзыв (отраслевая оценка)


def _niche_init():
    db.execute("CREATE TABLE IF NOT EXISTS niche_snapshots "
               "(query TEXT, nm_id BIGINT, feedbacks INTEGER, price REAL, "
               "snap_date TEXT, PRIMARY KEY (query, nm_id, snap_date))")
    db.execute("CREATE TABLE IF NOT EXISTS niche_history "
               "(query TEXT PRIMARY KEY, data TEXT, built_at TEXT)")


_niche_last_body = ""   # сырое тело последнего ответа поиска (для диагностики)

# WB жёстко банит IP за очереди запросов; после серии 429 бан «липкий»
# и держится часами. Темп по умолчанию — 1 запрос в 45 сек (можно менять
# переменной окружения WB_SEARCH_INTERVAL без деплоя).
_wb_search_lock = asyncio.Lock()
_wb_search_last: float = 0.0
# 45с было под сожжённый серверный IP; на резидентском прокси хватает 15с
_WB_SEARCH_INTERVAL = float(os.getenv("WB_SEARCH_INTERVAL", "15") or 15)


async def _wb_throttle():
    global _wb_search_last
    import time as _t
    async with _wb_search_lock:
        wait = _WB_SEARCH_INTERVAL - (_t.monotonic() - _wb_search_last)
        if wait > 0:
            await asyncio.sleep(wait)
        _wb_search_last = _t.monotonic()


def _parse_loose(text: str) -> dict | None:
    """Максимально терпеливый разбор ответа WB search.

    Тело бывает испорчено: прокси оставляет маркеры chunked-передачи,
    а анти-бот WB дописывает к metadata-JSON хвост «Not Found». Пробуем
    по очереди: чистый json → чистка чанков → raw_decode (берёт первый
    валидный JSON-объект и игнорирует мусорный хвост)."""
    import re as _re
    for candidate in (text, _re.sub(r"\r?\n[0-9a-fA-F]{1,6}\r?\n", "", text)):
        try:
            return json.loads(candidate)
        except ValueError:
            pass
        try:
            obj, _ = json.JSONDecoder().raw_decode(candidate.lstrip())
            if isinstance(obj, dict):
                return obj
        except ValueError:
            pass
    return None


async def _wb_search_preset(client, params: dict, meta_text: str, ver: str = "v13") -> dict | None:
    """Preset-ответ WB: metadata несёт catalog_value=preset=N — сама выдача
    отдаётся тем же эндпоинтом при повторе с параметром preset."""
    import re as _re
    global _niche_last_body
    m = _re.search(r'"catalog_value":"preset=(\d+)"', meta_text)
    if not m:
        return None
    preset_id = m.group(1)
    for drop_query in (False, True):
        p2 = dict(params)
        p2["preset"] = preset_id
        if drop_query:
            p2.pop("query", None)
        # без троттлинга: браузер шлёт follow-up мгновенно, пауза только
        # протухшивает токены и провоцирует анти-бот на приманку
        await asyncio.sleep(1.0)
        try:
            r = await client.get(
                f"https://search.wb.ru/exactmatch/ru/common/{ver}/search", params=p2)
            _niche_last_body += (f"\n\n=== PRESET {preset_id}{' (без query)' if drop_query else ''}"
                                 f" → HTTP {r.status_code} ===\n" + r.text[:1500])
            if not r.is_success:
                continue
            payload = _parse_loose(r.text)
            if payload and ((payload.get("data") or {}).get("products")):
                return payload
        except Exception as e:
            _niche_last_body += f"\n\n=== PRESET → EXC {str(e)[:200]} ==="
    return None


async def _wb_search_stage2(client, params: dict, meta_text: str, ver: str = "v13") -> dict | None:
    """Анти-бот WB: metadata-ответ несёт токены qv/kcl — повторяем запрос
    с ними. Иногда и второй ответ metadata-только (с новыми токенами),
    поэтому идём до 3 переходов, каждый раз со свежими qv/kcl."""
    import re as _re
    global _niche_last_body
    body = meta_text
    for hop in range(1, 4):
        qv = _re.search(r'"qv":"([^"]+)"', body)
        if not qv:
            return None
        p2 = dict(params)
        p2["qv"] = qv.group(1)
        kcl = _re.search(r'"kcl":"([^"]+)"', body)
        if kcl:
            p2["kcl"] = kcl.group(1)
        # qv-токен короткоживущий: повтор должен уйти сразу, как в браузере
        await asyncio.sleep(0.7)
        try:
            r = await client.get(
                f"https://search.wb.ru/exactmatch/ru/common/{ver}/search", params=p2)
            _niche_last_body += f"\n\n=== STAGE2 hop{hop} qv → HTTP {r.status_code} ===\n" + r.text[:1500]
            if not r.is_success:
                return None
            payload = _parse_loose(r.text)
            if payload and ((payload.get("data") or {}).get("products")):
                return payload
            # metadata снова — берём свежие токены из нового тела и повторяем
            new_body = r.text
            if new_body == body:
                return None
            body = new_body
        except Exception as e:
            _niche_last_body += f"\n\n=== STAGE2 hop{hop} qv → EXC {str(e)[:200]} ==="
            return None
    return None


async def _wb_public_search(query: str, limit: int = 60) -> tuple[list[dict], int]:
    """Публичная выдача WB по запросу → (товары, всего найдено).

    Тот же эндпоинт, что использует сайт wildberries.ru (и все сервисы
    аналитики). Формат периодически меняется — пробуем v5 и v4."""
    import httpx
    global _niche_last_err, _niche_last_body
    _niche_last_err = ""
    import os
    proxy = os.getenv("WB_SEARCH_PROXY", "").strip() or None

    # Сайт WB (июль 2026) получает выдачу через v18 c appType=64 — старый
    # v13/appType=1 отдаёт всем preset-заглушку с qv (снято с DevTools).
    # Пробуем варианты от нового к старому; для старого остаётся qv-цепочка.
    variants = [("v18", 64), ("v13", 64), ("v13", 1)]
    env_ver = os.getenv("WB_SEARCH_VER", "").strip()
    if env_ver:
        variants.insert(0, (env_ver, int(os.getenv("WB_SEARCH_APPTYPE", "64") or 64)))

    def _mk_params(ver: str, apptype: int) -> dict:
        p = {"ab_testing": "false", "appType": apptype, "curr": "rub",
             "dest": -1257786, "sort": "popular", "resultset": "catalog",
             "page": 1, "spp": 30, "lang": "ru", "locale": "ru",
             "suppressSpellcheck": "false", "query": query}
        if ver >= "v18":
            p["hide_dtype"] = 15   # как шлёт сайт
        else:
            p["reg"] = 1
            p["regions"] = "80,38,83,4,64,33,68,70,30,40,86,75,69,1,66,110,22,31,48,71,114"
        return p

    async with httpx.AsyncClient(timeout=30, proxy=proxy, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36",
            "Accept": "*/*", "Origin": "https://www.wildberries.ru",
            "Referer": "https://www.wildberries.ru/catalog/0/search.aspx",
            "x-requested-with": "XMLHttpRequest",
            "Accept-Encoding": "identity"}) as client:
        _niche_last_body = ""
        for ver, apptype in variants:
            params = _mk_params(ver, apptype)
            tag = f"{ver}/appType={apptype}"
            try:
                await _wb_throttle()
                r = await client.get(
                    f"https://search.wb.ru/exactmatch/ru/common/{ver}/search", params=params)
            except Exception as e:
                _niche_last_err = f"{tag}: {str(e)[:120]}"
                continue
            _niche_last_body += f"\n\n=== {tag} → HTTP {r.status_code} ===\n" + r.text[:4000]
            if r.status_code == 429:
                _niche_last_err = ("HTTP 429 — WB ограничивает IP; дайте инструменту "
                                   "пару минут отдыха")
                break   # с этого IP сейчас всё будет 429 — варианты не помогут
            if not r.is_success:
                _niche_last_err = f"{tag}: HTTP {r.status_code}"
                continue
            try:
                txt = r.text
                payload = _parse_loose(txt)
                data = (payload.get("data") or {}) if payload else {}
                products = data.get("products") or []
                if not products and '"qv"' in txt:
                    p2 = await _wb_search_stage2(client, params, txt, ver)
                    if p2:
                        data = p2.get("data") or {}
                        products = data.get("products") or []
                if not products and '"catalog_value":"preset=' in txt:
                    p3 = await _wb_search_preset(client, params, txt, ver)
                    if p3:
                        data = p3.get("data") or {}
                        products = data.get("products") or []
                if products:
                    total = int(data.get("total") or len(products))
                    out = []
                    for p in products[:limit]:
                        price = None
                        for sz in p.get("sizes") or []:
                            pr = (sz.get("price") or {}).get("product")
                            if pr:
                                price = pr / 100
                                break
                        if price is None and p.get("salePriceU"):
                            price = p["salePriceU"] / 100
                        out.append({
                            "nm": p.get("id"),
                            "name": p.get("name") or "",
                            "brand": p.get("brand") or "",
                            "price": round(price) if price else None,
                            "rating": p.get("reviewRating") or p.get("rating") or 0,
                            "feedbacks": int(p.get("feedbacks") or 0),
                            "subject_id": p.get("subjectId"),
                            "supplier": p.get("supplier") or "",
                        })
                    if out:
                        return out, total
                    _niche_last_err = f"{tag}: пустая выдача"
                elif '"qv"' in txt or '"preset=' in txt:
                    _niche_last_err = (f"{tag}: только анти-бот metadata — пробуем "
                                       "следующий вариант")
                else:
                    _niche_last_err = f"{tag}: пустая выдача"
            except Exception as e:
                _log.warning("wb public search %s: %s", tag, e)
                _niche_last_err = f"{tag}: {str(e)[:120]}"
    return [], 0


_niche_last_err = ""


# ══ Калькулятор маржи: затраты на единицу + ценообразование ═══════════════════
_wb_comm_cache: dict = {}     # nmId → {pct, subject}
_wb_comm_ts: float = 0.0
_WB_COMM_TTL = 12 * 3600


async def _wb_commission_by_nm() -> dict:
    """Официальная комиссия WB по каждому nmId кабинета: тарифы по категориям
    (tariffs/commission) × категория карточки (cards/list). Кеш 12ч."""
    global _wb_comm_cache, _wb_comm_ts
    if _wb_comm_cache and _time.monotonic() - _wb_comm_ts < _WB_COMM_TTL:
        return _wb_comm_cache
    import wb_client
    try:
        tariffs, subjects = await asyncio.gather(
            wb_client.get_commission_tariffs(), wb_client.get_card_subjects())
    except Exception as e:
        _log.warning("WB commission tariffs failed: %s", e)
        return _wb_comm_cache
    if not tariffs or not subjects:
        return _wb_comm_cache
    out = {}
    for nm, s in subjects.items():
        t = tariffs.get(int(s.get("subjectID") or 0))
        if t and t.get("fbo") is not None:
            out[nm] = {"pct": float(t["fbo"]), "subject": t.get("subjectName") or s.get("subjectName", "")}
    if out:
        _wb_comm_cache, _wb_comm_ts = out, _time.monotonic()
        try:
            await asyncio.to_thread(_tariff_watch, out)
        except Exception as e:
            _log.warning("tariff watch failed: %s", e)
    return out


def _tariff_watch(comm_by_nm: dict) -> None:
    """Следит за изменением комиссии WB по НАШИМ категориям: сравнивает с
    последним виденным тарифом и при изменении пишет алерт (виден в шапке)."""
    import snapshot as _snap
    cur = {}
    for v in comm_by_nm.values():
        if v.get("subject"):
            cur[v["subject"]] = v["pct"]
    seen = _snap.load("wb_tariffs_seen", None)
    if seen:
        changes = [{"subject": s, "old": seen[s], "new": p}
                   for s, p in cur.items() if s in seen and abs(seen[s] - p) >= 0.01]
        if changes:
            _snap.save("wb_tariff_alert", {
                "date": datetime.utcnow().strftime("%Y-%m-%d"),
                "changes": changes})
            _log.info("WB тарифы изменились: %s", changes)
    _snap.save("wb_tariffs_seen", cur)


@router.post("/agent_review")
async def agent_review_now(mp: str = Query(default="WB")):
    """Ручной запуск разбора агента (кнопка/тест): собрать и отправить в TG."""
    import agent_review
    return await agent_review.send_review(mp)


@router.get("/tariff_alert")
async def tariff_alert():
    """Алерт об изменении комиссии WB по категориям кабинета (14 дней)."""
    import snapshot as _snap
    alert = await asyncio.to_thread(_snap.load, "wb_tariff_alert", None)
    if not alert:
        return {}
    try:
        age = (datetime.utcnow() - datetime.strptime(alert["date"], "%Y-%m-%d")).days
    except (KeyError, ValueError):
        return {}
    return alert if age <= 14 else {}


def _sales_forecast(platform: str) -> dict:
    """Прогноз продаж на месяц по каждому SKU — модель Холта с затуханием.

    По 12 неделям истории sales_daily: экспоненциальное сглаживание уровня
    и тренда (α=0.5, β=0.3, затухание φ=0.85). Недельные корзины гасят
    сезонность дней недели; затухание не даёт тренду улетать при экстраполяции.
    Возвращает {SKU_UPPER: {"qty_f": шт/мес, "trend": %/мес}}."""
    from datetime import date, timedelta
    yesterday = date.today() - timedelta(days=1)
    since = (yesterday - timedelta(days=12 * 7 - 1)).isoformat()
    try:
        rows = db.fetchall(
            "SELECT sale_date, sku, qty FROM sales_daily "
            "WHERE platform=? AND sale_date>=? AND sale_date<=?",
            (platform, since, yesterday.isoformat()))
    except Exception as e:
        _log.warning("forecast: sales_daily недоступна: %s", e)
        return {}

    daily: dict[str, dict[str, int]] = {}
    for d, sku, q in rows:
        k = str(sku or "").strip().upper()
        if k:
            daily.setdefault(k, {})[str(d)[:10]] = int(q or 0)

    out = {}
    for sku, dd in daily.items():
        # недельные корзины, от старых к свежим (неделя 0 заканчивается вчера)
        weeks = []
        for i in range(11, -1, -1):
            end = yesterday - timedelta(days=7 * i)
            s = sum(dd.get((end - timedelta(days=j)).isoformat(), 0) for j in range(7))
            weeks.append(s)
        # обрезаем пустой префикс (товар ещё не продавался)
        first = next((i for i, w in enumerate(weeks) if w > 0), None)
        if first is None:
            continue
        weeks = weeks[first:]
        if len(weeks) < 3:
            continue  # мало истории — калькулятор возьмёт среднее по окну

        # Холт с затуханием
        a, b_, phi = 0.5, 0.3, 0.85
        level, trend = float(weeks[0]), 0.0
        for w in weeks[1:]:
            prev = level
            level = a * w + (1 - a) * (level + phi * trend)
            trend = b_ * (level - prev) + (1 - b_) * phi * trend
        fc_weeks = [max(level + trend * sum(phi ** k for k in range(1, h + 1)), 0.0)
                    for h in range(1, 5)]
        month = sum(fc_weeks) * (30.44 / 28)
        # здравый смысл: не дальше ×[0.3..2.5] от факта последних 4 недель
        recent = sum(weeks[-4:]) * (30.44 / 28) if len(weeks) >= 4 else month
        if recent > 0:
            month = max(min(month, recent * 2.5), recent * 0.3)
        out[sku] = {"qty_f": round(month, 1),
                    "trend": round((month / recent - 1) * 100) if recent > 0 else 0}
    return out


@router.get("/margin")
async def get_margin(mp: str = Query(default="WB")):
    """Затраты на единицу товара из фактической юнитки — для калькулятора
    цены/маржи. По каждому SKU: цена (средняя), % комиссии/эквайринга,
    логистика/хранение/продвижение/себес на штуку."""
    from routers import finance as _fin
    try:
        if mp == "OZON":
            data = await _fin.get_ozon_unit(months=6)
        elif mp == "YM":
            data = await _fin.get_ym_unit(months=6)
        else:
            data = await _fin.get_wb_unit(months=6)
    except Exception as e:
        return {"items": [], "error": str(e)[:200]}

    skus = data.get("skus", [])
    if not skus:
        return {"items": [], "message": data.get("message") or "Нет данных юнитки"}

    # официальные тарифы комиссии WB по категориям (nmId → pct/subject)
    comm_tariff = await _wb_commission_by_nm() if mp == "WB" else {}
    # прогнозная модель продаж (Холт по 12 нед истории)
    forecast = await asyncio.to_thread(
        _sales_forecast, {"WB": "WB", "OZON": "Ozon", "YM": "YM"}.get(mp, "WB"))

    # Окна усреднения — у статей разная «скорость устаревания»:
    #   свежие (цена, комиссия %, логистика, ДРР) — текущий + прошлый месяц;
    #   сглаженные (хранение, штрафы/удержания) — последние 4 месяца, чтобы
    #   разовая поставка или штраф не искажали штуку;
    #   мало продаж в свежем окне (<10 шт) — окно расширяется автоматически.
    all_mks = sorted({mk for r in skus for mk in (r.get("months") or {})})
    W_RECENT, W_SMOOTH = all_mks[-2:], all_mks[-4:]

    def _mk_label(keys):
        if not keys:
            return ""
        f = lambda mk: f"{mk[5:7]}.{mk[2:4]}"
        return f(keys[0]) if len(keys) == 1 else f"{f(keys[0])}–{f(keys[-1])}"

    items = []
    for r in skus:
        months = r.get("months") or {}

        def agg(keys):
            tot = {}
            for mk in keys:
                for k, v in (months.get(mk) or {}).items():
                    if isinstance(v, (int, float)):
                        tot[k] = tot.get(k, 0) + v
            return tot

        # свежее окно с автоматическим расширением при малых продажах
        for keys in (W_RECENT, W_SMOOTH, all_mks):
            recent, recent_keys = agg(keys), keys
            if recent.get("qty", 0) >= 10:
                break
        smooth, smooth_keys = agg(W_SMOOTH), W_SMOOTH
        if smooth.get("qty", 0) < 10 or len(recent_keys) > len(W_SMOOTH):
            smooth, smooth_keys = agg(all_mks), all_mks

        qty, rev = recent.get("qty", 0), recent.get("revenue", 0)
        if qty <= 0 or rev <= 0:
            continue
        per = lambda k: recent.get(k, 0) / qty
        sqty = smooth.get("qty", 0) or 1
        sper = lambda k: smooth.get(k, 0) / sqty
        price0 = round(rev / qty)
        paid = recent.get("paid", 0)
        total_qty = sum((c or {}).get("qty", 0) for c in months.values())
        # дней в свежем окне: текущий месяц учитываем по фактическим дням,
        # иначе среднее «шт/мес» занижается в полтора раза в середине месяца
        _now = datetime.utcnow() + timedelta(hours=3)
        days_cov = 0
        for mk in recent_keys:
            y_, m_ = int(mk[:4]), int(mk[5:7])
            if (y_, m_) == (_now.year, _now.month):
                days_cov += max(_now.day - 1, 1)
            else:
                import calendar as _cal
                days_cov += _cal.monthrange(y_, m_)[1]
        nm = r.get("nmId")
        tariff = comm_tariff.get(int(nm)) if nm and str(nm).isdigit() else None
        items.append({
            "sku": r.get("sku"), "nmId": r.get("nmId"),
            "name": r.get("name", ""), "group": r.get("brand", ""),
            "qty": round(total_qty),
            "qty_m": round(qty / max(days_cov, 1) * 30.44, 1),  # продажи в месяц (по фактическим дням окна)
            "qty_f": (forecast.get(str(r.get("sku") or "").strip().upper()) or {}).get("qty_f"),
            "trend": (forecast.get(str(r.get("sku") or "").strip().upper()) or {}).get("trend"),
            "window": _mk_label(recent_keys),               # окно свежих статей
            "price0": price0,                              # средняя цена (до СПП)
            "buyer0": round(paid / qty) if paid > 0 else None,  # цена для покупателя (после СПП)
            # комиссия WB: официальный тариф категории (tariffs/commission),
            # фолбэки — commission_percent из последней продажи, затем средняя
            "comm_pct": (round(tariff["pct"], 2) if tariff
                         else round(float(r["commNow"]), 2) if r.get("commNow")
                         else round(recent.get("commission", 0) / rev * 100, 2)),
            "comm_exact": bool(tariff or r.get("commNow")),
            "subject": (tariff or {}).get("subject", ""),
            "acq_pct": round(recent.get("acquiring", 0) / rev * 100, 2),
            "logist": round(per("delivery")),              # логистика на штуку (фикс)
            "storage": round(sper("storage") + sper("acceptance")),  # сглажено
            "other": round(max(sper("penalty") + sper("deductions"), 0)),  # сглажено
            "advert": round(max(per("advert"), 0)),        # продвижение на штуку
            "cogs": round(r.get("unitCost") or sper("cogs")),  # себестоимость
        })
    items.sort(key=lambda x: -x["qty"])
    return {"items": items, "mp": mp,
            "window_recent": _mk_label(W_RECENT), "window_smooth": _mk_label(W_SMOOTH),
            "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}


@router.post("/margin/export")
async def margin_export(body: dict):
    """Экспорт калькулятора маржи в Excel — с учётом правок пользователя
    (цена/себес/ДРР), налога и целевых марж из шапки."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    mp = body.get("mp") or "WB"
    tax = float(body.get("tax") or 0)
    adv_on = bool(body.get("adv_on", True))
    targets = [float(t) for t in (body.get("targets") or [15, 25, 35])][:3]
    ov = body.get("overrides") or {}

    data = await get_margin(mp=mp)
    items = data.get("items") or []
    if not items:
        raise HTTPException(400, data.get("message") or "Нет данных")

    taxf = 1 - tax / 100

    def calc(b):
        o = ov.get(b["sku"]) or {}
        price = float(o.get("price", b["price0"]) or 0)
        cogs = float(o.get("cogs", b["cogs"]) or 0)
        drr0 = round(b["advert"] / b["price0"] * 100, 1) if b["price0"] else 0
        drr = float(o.get("drr", drr0) or 0) if adv_on else 0.0
        pct = (b["comm_pct"] + b["acq_pct"] + drr) / 100
        fixed = b["logist"] + b["storage"] + b["other"] + cogs
        gross = price * (1 - pct) - fixed
        profit = gross - (max(gross, 0) * tax / 100)
        margin = profit / price * 100 if price > 0 else 0
        be = fixed / (1 - pct) if pct < 1 else None
        tgt_prices = []
        for t in targets:
            denom = (1 - pct) - (t / 100) / taxf if taxf > 0 else 0
            tgt_prices.append(fixed / denom if denom > 0.001 else None)
        buyer_k = (b["buyer0"] / b["price0"]) if b.get("buyer0") and b["price0"] else None
        return price, cogs, drr, profit, margin, be, tgt_prices, buyer_k

    wb_x = openpyxl.Workbook()
    ws = wb_x.active
    ws.title = "Калькулятор маржи"
    head = ["Артикул", "Название", "Категория WB", "Комиссия %", "Эквайринг %",
            "Логистика ₽", "Хранение ₽", "Прочее ₽", "ДРР %", "Себес ₽",
            "Цена ₽ (до СПП)", "Цена покупателя ₽", "Прибыль/ед ₽", "Маржа %",
            "Безубыток ₽"]
    for t in targets:
        head += [f"Цена для {t:g}% ₽", f"Цена клиента при {t:g}% ₽"]
    ws.append(head)
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4A5568")
    for c in ws[1]:
        c.font, c.fill = hf, fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for b in items:
        price, cogs, drr, profit, margin, be, tgts, bk = calc(b)
        row = [b["sku"], b["name"], b.get("subject") or "", b["comm_pct"], b["acq_pct"],
               b["logist"], b["storage"], b["other"], drr, round(cogs),
               round(price), round(price * bk) if bk else None,
               round(profit), round(margin, 1), round(be) if be else None]
        for tp in tgts:
            row += [round(tp) if tp else None,
                    round(tp * bk) if (tp and bk) else None]
        ws.append(row)
    for i in range(1, len(head) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14 if i > 2 else (26 if i == 2 else 12)
    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb_x.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="margin.xlsx"'})


# ══ Спрос WB (Джем): поисковые запросы по своим товарам ═══════════════════════
_demand_cache: dict = {}
_demand_ts: float = 0.0
_demand_building: bool = False
_demand_error: str = ""
_DEMAND_TTL = 6 * 3600


def _num(v):
    """{current,dynamics} → (current, dynamics) безопасно."""
    if isinstance(v, dict):
        return v.get("current") or 0, v.get("dynamics") or 0
    return (v or 0), 0


async def _build_demand_bg() -> None:
    global _demand_cache, _demand_ts, _demand_building, _demand_error
    if _demand_building:
        return
    _demand_building = True
    _demand_error = ""
    try:
        import wb_client
        import catalog as _cat
        from datetime import date
        nmids = [int(x) for x in getattr(_cat, "WB_ID_TO_ART", {}).keys() if str(x).isdigit()]
        if not nmids:
            _demand_error = "нет nmId в каталоге кабинета"
            return
        cur_e = date.today() - timedelta(days=1)
        cur_s = cur_e - timedelta(days=6)
        pw_e = cur_s - timedelta(days=1)
        pw_s = pw_e - timedelta(days=6)
        per = lambda s, e: {"start": s.isoformat(), "end": e.isoformat()}

        items = []
        # nmIds батчами по 50 — топ запросов по каждому товару
        for i in range(0, len(nmids), 50):
            batch = nmids[i:i + 50]
            body = {"currentPeriod": per(cur_s, cur_e), "pastPeriod": per(pw_s, pw_e),
                    "nmIds": batch, "topOrderBy": "openToCart",
                    "orderBy": {"field": "avgPosition", "mode": "asc"}, "limit": 30}
            st, resp = await wb_client.analytics_post(
                "/api/v2/search-report/product/search-texts", body)
            if st != 200 or not isinstance(resp, dict):
                _demand_error = f"search-texts HTTP {st}: {str(resp)[:200]}"
                continue
            for it in (resp.get("data") or {}).get("items", []):
                text = str(it.get("text") or "")
                if text.startswith("#"):    # артикул вместо фразы — пропускаем
                    continue
                freq_c, freq_d = _num(it.get("frequency"))
                pos_c, pos_d = _num(it.get("avgPosition"))
                ord_c, ord_d = _num(it.get("orders"))
                o2c_c, _ = _num(it.get("openToCart"))
                c2o_c, _ = _num(it.get("cartToOrder"))
                art = _cat.resolve_wb(it.get("nmId")) if it.get("nmId") else ""
                items.append({
                    "query": text, "nm": it.get("nmId"),
                    "sku": art if art and not str(art).isdigit() else it.get("vendorCode", ""),
                    "name": it.get("name") or "", "group": it.get("subjectName") or "",
                    "freq": int(freq_c), "freq_dyn": int(freq_d),
                    "week_freq": int(it.get("weekFrequency") or 0),
                    "position": round(pos_c, 1), "pos_dyn": round(pos_d, 1),
                    "orders": int(ord_c), "orders_dyn": int(ord_d),
                    "open_to_cart": round(o2c_c, 1), "cart_to_order": round(c2o_c, 1),
                })
            await asyncio.sleep(1)
        # точка роста: высокий спрос + низкая позиция (глубоко в выдаче)
        for it in items:
            it["opportunity"] = it["freq"] >= 100 and it["position"] >= 20
        items.sort(key=lambda x: -x["freq"])
        _demand_cache = {"items": items,
                         "period": f"{cur_s.isoformat()} — {cur_e.isoformat()}",
                         "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}
        _demand_ts = _t_mono()
        import snapshot as _snap
        _snap.save("demand_jam", _demand_cache)
    except Exception as e:
        _demand_error = str(e)[:300]
        _log.warning("demand build: %s", e)
    finally:
        _demand_building = False


@router.get("/demand")
async def get_demand(refresh: bool = Query(default=False)):
    """Спрос по своим товарам (Джем): поисковые запросы, частотность,
    позиция, заказы, точки роста."""
    global _demand_cache, _demand_ts
    fresh = _demand_cache and _t_mono() - _demand_ts < _DEMAND_TTL
    if refresh or not fresh:
        if not _demand_cache:
            import snapshot as _snap
            snap = await asyncio.to_thread(_snap.load, "demand_jam", None)
            if snap:
                _demand_cache = snap
                _demand_ts = _t_mono() - _DEMAND_TTL + 300
        if not _demand_building:
            _spawn(_build_demand_bg())
    if _demand_cache:
        return {**_demand_cache, "building": _demand_building, "error": _demand_error}
    if _demand_error and not _demand_building:
        return {"items": [], "message": f"⚠ Джем: {_demand_error}"}
    return {"items": [], "building": True,
            "message": "⏳ Собираем спрос по товарам из Джем — обновится само"}


@router.get("/jam/probe", include_in_schema=False)
async def jam_probe(query: str = Query(default="шторы блэкаут")):
    """Диагностика Джем-методов на живом ключе: какие эндпоинты доступны
    и что возвращают. Проверяем и «по своим товарам» (nmIds из каталога),
    и кандидатов на отчёт «Поисковые запросы на WB» по свободному слову."""
    import wb_client
    import catalog as _cat
    from datetime import date, timedelta

    nmids = [int(x) for x in list(getattr(_cat, "WB_ID_TO_ART", {}).keys())[:20] if str(x).isdigit()]
    cur_e = date.today() - timedelta(days=1)
    cur_s = cur_e - timedelta(days=6)
    pw_e = cur_s - timedelta(days=1)
    pw_s = pw_e - timedelta(days=6)
    per = lambda s, e: {"start": s.isoformat(), "end": e.isoformat()}

    tests = [
        ("report (свои)", "/api/v2/search-report/report",
         {"currentPeriod": per(cur_s, cur_e), "pastPeriod": per(pw_s, pw_e),
          "nmIds": nmids, "positionCluster": "all", "offset": 0,
          "orderBy": {"field": "avgPosition", "mode": "asc"},
          "includeSubstitutedSKUs": True, "includeSearchTexts": False, "limit": 30}),
        ("table/details (свои — тексты запросов+метрики)", "/api/v2/search-report/table/details",
         {"currentPeriod": per(cur_s, cur_e), "pastPeriod": per(pw_s, pw_e),
          "nmIds": nmids, "positionCluster": "all", "includeSubstitutedSKUs": True,
          "includeSearchTexts": True, "orderBy": {"field": "avgPosition", "mode": "asc"},
          "limit": 30, "offset": 0}),
        ("product/search-texts (топ запросов товара)", "/api/v2/search-report/product/search-texts",
         {"currentPeriod": per(cur_s, cur_e), "pastPeriod": per(pw_s, pw_e),
          "nmIds": [nmids[0]] if nmids else [0], "topOrderBy": "openToCart",
          "orderBy": {"field": "avgPosition", "mode": "asc"}, "limit": 30}),
        ("product/orders (заказы по запросам)", "/api/v2/search-report/product/orders",
         {"period": per(cur_s, cur_e), "nmId": nmids[0] if nmids else 0,
          "searchTexts": []}),
    ]
    out = []
    for name, path, body in tests:
        try:
            st, resp = await wb_client.analytics_post(path, body)
            head = json.dumps(resp, ensure_ascii=False)[:500] if isinstance(resp, (dict, list)) else str(resp)[:500]
            out.append({"метод": name, "path": path, "status": st, "ответ": head})
        except Exception as e:
            out.append({"метод": name, "path": path, "error": str(e)[:300]})
    return {"nmids_проверено": len(nmids), "результаты": out}


@router.get("/niche/last_response", include_in_schema=False)
async def niche_last_response():
    """Сырое тело последнего ответа поиска WB — без нового запроса."""
    from fastapi.responses import PlainTextResponse
    return PlainTextResponse(_niche_last_body or "ещё не было запросов")


@router.get("/niche/proxycheck", include_in_schema=False)
async def niche_proxycheck():
    """Каким выходным IP ходим в WB: маскированный прокси из env + egress-IP.

    Проверка идёт на нейтральный ipify (не WB), лимиты не жжёт."""
    import httpx
    import os
    proxy = os.getenv("WB_SEARCH_PROXY", "").strip() or None
    masked = None
    if proxy:
        # прячем пароль: http://user:pass@host:port → http://user:***@host:port
        import re as _re
        masked = _re.sub(r"(://[^:/@]+):[^@]*@", r"\1:***@", proxy)
    out = {"proxy_env": masked or "не задан (прямое соединение)"}
    try:
        async with httpx.AsyncClient(timeout=20, proxy=proxy) as client:
            r = await client.get("https://api.ipify.org?format=json")
            out["egress_ip"] = r.json().get("ip")
        try:
            async with httpx.AsyncClient(timeout=20, proxy=proxy) as client:
                g = await client.get(f"http://ip-api.com/json/{out['egress_ip']}?fields=country,isp")
                out.update(g.json())
        except Exception:
            pass
    except Exception as e:
        out["error"] = f"через прокси не удалось выйти в сеть: {str(e)[:200]}"
    return out


@router.get("/niche/debug", include_in_schema=False)
async def niche_debug(query: str = Query(default="крем для лица"),
                      full: bool = Query(default=False)):
    """Диагностика публичного поиска WB: статусы по версиям API."""
    import httpx
    import os
    proxy = os.getenv("WB_SEARCH_PROXY", "").strip() or None
    out = {"proxy": "настроен" if proxy else "нет"}
    params = {"ab_testing": "false", "appType": 1, "curr": "rub", "dest": -1257786,
              "sort": "popular", "resultset": "catalog", "page": 1, "spp": 30,
              "lang": "ru", "locale": "ru", "reg": 1,
              "regions": "80,38,83,4,64,33,68,70,30,40,86,75,69,1,66,110,22,31,48,71,114",
              "query": query}
    async with httpx.AsyncClient(timeout=20, proxy=proxy, headers={
            "User-Agent": "Mozilla/5.0", "Referer": "https://www.wildberries.ru/",
            "Accept-Encoding": "gzip, deflate"}) as client:
        for ver in ("v13",):   # только один запрос — очереди выжигают лимит WB
            try:
                await _wb_throttle()
                r = await client.get(f"https://search.wb.ru/exactmatch/ru/common/{ver}/search",
                                     params=params)
                body = r.text[:300]
                n = 0
                try:
                    n = len((r.json().get("data") or {}).get("products") or [])
                except Exception:
                    pass
                out[ver] = {"status": r.status_code, "products": n,
                            "body_head": r.text[:4000] if full else body}
            except Exception as e:
                out[ver] = {"error": str(e)[:200]}
    return out


_commission_cache: dict = {}


async def _wb_commission(subject_id: int | None) -> float | None:
    """Комиссия WB (FBW, %) по предмету — официальный API тарифов."""
    global _commission_cache
    if not subject_id:
        return None
    if not _commission_cache:
        try:
            import httpx
            from config import WB_API_KEY
            async with httpx.AsyncClient(timeout=30) as client:
                r = await client.get(
                    "https://common-api.wildberries.ru/api/v1/tariffs/commission",
                    headers={"Authorization": WB_API_KEY}, params={"locale": "ru"})
                for rep in (r.json().get("report") or []):
                    sid = rep.get("subjectID")
                    if sid:
                        _commission_cache[int(sid)] = float(
                            rep.get("kgvpMarketplace") or rep.get("paidStorageKgvp") or 0)
        except Exception as e:
            _log.warning("wb tariffs: %s", e)
            _commission_cache = {0: 0.0}
    return _commission_cache.get(int(subject_id))


_niche_job: dict = {"status": "idle"}

# ── Локальный агент (сбор выдачи WB с домашнего IP пользователя) ──────────────
# WB банит IP дата-центров/прокси, но домашний IP пользователя пускает.
# Агент на ПК пользователя опрашивает /niche/pending, тянет выдачу и
# отдаёт её в /niche/ingest. Очередь и «почтовый ящик» — в памяти.
_agent_pending: dict = {}     # query -> запрошено (monotonic)
_agent_inbox: dict = {}       # query -> {products, total} или {error}


@router.get("/niche/pending")
async def niche_pending(token: str = ""):
    """Список запросов, ждущих сбора локальным агентом."""
    if os.getenv("WB_AGENT_TOKEN", "") and token != os.getenv("WB_AGENT_TOKEN", ""):
        raise HTTPException(status_code=401, detail="bad token")
    now = _t_mono()
    # чистим протухшие (агент офлайн > 5 мин)
    for q in [q for q, ts in _agent_pending.items() if now - ts > 300]:
        _agent_pending.pop(q, None)
    return {"queries": list(_agent_pending.keys())}


@router.post("/niche/ingest")
async def niche_ingest(payload: dict, token: str = ""):
    """Агент отдаёт собранную выдачу: {query, products:[...], total} или {query, error}."""
    if os.getenv("WB_AGENT_TOKEN", "") and token != os.getenv("WB_AGENT_TOKEN", ""):
        raise HTTPException(status_code=401, detail="bad token")
    query = str(payload.get("query") or "").strip().lower()
    if not query:
        raise HTTPException(status_code=400, detail="query required")
    _agent_pending.pop(query, None)
    if payload.get("error"):
        _agent_inbox[query] = {"error": str(payload["error"])[:300]}
    else:
        products = payload.get("products") or []
        # починка парсинга рейтинга: у карточек с целым рейтингом DOM-регулярка
        # склеивает его со счётчиком оценок («5» + «448» → 5448) — дробные
        # (4.73) приходят корректно. Всё, что больше 5 — берём первую цифру.
        for p in products:
            r = p.get("rating")
            if isinstance(r, (int, float)) and r > 5:
                p["rating"] = float(str(int(r))[0])
        _agent_inbox[query] = {"products": products,
                               "total": int(payload.get("total") or 0)}
    return {"ok": True}


def _t_mono() -> float:
    import time as _t
    return _t.monotonic()


async def _wb_agent_search(query: str) -> tuple[list, int]:
    """Ставит запрос в очередь агента и ждёт результат (агент на ПК юзера)."""
    global _niche_last_err
    _agent_inbox.pop(query, None)
    _agent_pending[query] = _t_mono()
    for _ in range(60):            # ~3 мин ожидания результата от агента
        await asyncio.sleep(3)
        res = _agent_inbox.pop(query, None)
        if res is None:
            continue
        if res.get("error"):
            _niche_last_err = f"агент: {res['error']}"
            return [], 0
        return res.get("products") or [], int(res.get("total") or 0)
    _agent_pending.pop(query, None)
    _niche_last_err = ("локальный агент не ответил за 3 мин — запущен ли он на ПК "
                       "и открыт ли доступ к WB?")
    return [], 0


def _niche_relevant(products: list, query: str) -> bool:
    """Анти-бот WB иногда отдаёт «приманку» — 1-2 случайных товара (MacBook
    на запрос про крем). Хоть какие-то товары должны содержать слова запроса."""
    q_tokens = [w[:max(4, len(w) - 2)].lower() for w in query.split() if len(w) > 3]
    if not q_tokens:
        return True
    return any(any(t in (p.get("name") or "").lower() for t in q_tokens)
               for p in products)


async def _wb_fetch_search(query: str, limit: int = 60) -> tuple[list, int]:
    """Выдача через сервис wb-fetch (headless-Chrome проходит анти-бот WB).

    Холодный старт браузера на free-CPU занимает до 2-3 минут — таймаут
    щедрый, прогресс виден в стадии задачи."""
    global _niche_last_err, _niche_last_body
    import os
    fetch_url = os.getenv("WB_FETCH_URL", "").strip()
    try:
        import httpx
        async with httpx.AsyncClient(timeout=300) as fc:
            fr = await fc.get(fetch_url.rstrip("/") + "/search",
                              params={"query": query, "limit": limit,
                                      "token": os.getenv("WB_FETCH_TOKEN", "")})
            fj = fr.json()
            if fj.get("ok") and fj.get("products"):
                _niche_last_body = (f"WB-FETCH (headless): {len(fj['products'])} "
                                    f"товаров, total={fj.get('total')}")
                return fj["products"], int(fj.get("total") or len(fj["products"]))
            _niche_last_err = f"wb-fetch: {fj.get('error') or f'HTTP {fr.status_code}, пусто'}"
            _niche_last_body = f"WB-FETCH ответ: {str(fj)[:1200]}"
    except Exception as e:
        _niche_last_err = f"wb-fetch недоступен: {str(e)[:150]}"
        _niche_last_body = f"WB-FETCH exc: {str(e)[:400]}"
    return [], 0


async def _niche_job_run(payload: dict, query: str) -> None:
    """Фоновый анализ ниши с автоповторами.

    Многоступенчатая схема (qv/preset, троттлинг, ретраи при приманке)
    занимает больше 100с — лимита запроса на Render, поэтому анализ идёт
    фоном, а фронт опрашивает /niche/status."""
    global _niche_job, _niche_last_err
    import os
    try:
        # Локальный агент (домашний IP пользователя) — приоритетный путь:
        # WB банит IP серверов/прокси, но домашний адрес пускает.
        if os.getenv("WB_AGENT_MODE", "").strip() == "1":
            for attempt in (1, 2):
                _niche_job["stage"] = (f"ждём локальный агент — сбор выдачи WB "
                                       f"с вашего ПК (попытка {attempt}/2)")
                products, total = await _wb_agent_search(query)
                if products and not _niche_relevant(products, query):
                    _niche_last_err = "агент получил нерелевантную выдачу"
                    products = []
                if products:
                    _niche_job["stage"] = "выдача получена — считаем метрики и вердикт"
                    await _analyze_niche_impl(payload, query, products, total)
                    _niche_job = {"status": "done", "query": query}
                    return
                if attempt == 1:
                    await asyncio.sleep(5)
            _niche_job = {"status": "error", "query": query,
                          "error": _niche_last_err or "агент не дал выдачу"}
            return

        # wb-fetch настроен → он единственный путь: прямые запросы всё равно
        # получают 429/заглушку, а перебор только затягивает и жжёт лимиты
        if os.getenv("WB_FETCH_URL", "").strip():
            for attempt in (1, 2):
                _niche_job["stage"] = (f"wb-fetch: браузер открывает WB "
                                       f"(попытка {attempt}/2; холодный старт — до 2-3 мин)")
                products, total = await _wb_fetch_search(query)
                if products and not _niche_relevant(products, query):
                    _niche_last_err = "браузер получил нерелевантную выдачу"
                    products = []
                if products:
                    _niche_job["stage"] = "выдача получена — считаем метрики и вердикт"
                    await _analyze_niche_impl(payload, query, products, total)
                    _niche_job = {"status": "done", "query": query}
                    return
                if attempt == 1:
                    _niche_job["stage"] = f"{_niche_last_err} — повтор через 45с"
                    await asyncio.sleep(45)
            _niche_job = {"status": "error", "query": query,
                          "error": (_niche_last_err or "wb-fetch не дал выдачу")
                          + ". Смотрите лог сервиса wb-fetch в Render."}
            return

        last_reason = ""
        for attempt in range(1, 4):
            _niche_job["stage"] = f"запрашиваем выдачу WB (попытка {attempt}/3)"
            products, total = await _wb_public_search(query)
            if products and not _niche_relevant(products, query):
                last_reason = "WB подсовывает нерелевантную приманку вместо выдачи"
                products = []
            elif not products:
                last_reason = _niche_last_err or "пустая выдача"
            if products:
                _niche_job["stage"] = "выдача получена — считаем метрики и вердикт"
                await _analyze_niche_impl(payload, query, products, total)
                _niche_job = {"status": "done", "query": query}
                return
            if "429" in last_reason:
                break   # бан по IP держится часами — повторы его только продлевают
            if attempt < 3:
                _niche_job["stage"] = f"{last_reason} — пауза и новая попытка ({attempt}/3)"
                await asyncio.sleep(30)
        hint = (" WB ограничивает IP: подождите час-другой или смените IP прокси "
                "в ЛК proxy6." if "429" in last_reason else
                " Если повторяется на каждой попытке — WB пометил IP прокси как бота; "
                "надёжное лечение — резидентский прокси.")
        _niche_job = {"status": "error", "query": query, "error": last_reason + hint}
    except Exception as e:
        _log.warning("niche job: %s", e)
        _niche_job = {"status": "error", "query": query, "error": str(e)[:300]}


@router.post("/niche")
async def analyze_niche(payload: dict):
    """Скоринг ниши по поисковому запросу WB.

    {query, price?, cost?, logistics?} → запускает фоновый анализ
    (конкуренты, метрики ниши, юнит-прикидка, вердикт Claude); статус —
    в /niche/status, результат — в /niche/get."""
    query = str(payload.get("query") or "").strip().lower()
    if not query:
        raise HTTPException(status_code=400, detail="Введите поисковый запрос")
    _niche_init()
    global _niche_job
    if _niche_job.get("status") == "running":
        return {"building": True, "query": _niche_job.get("query"),
                "stage": _niche_job.get("stage", "")}
    _niche_job = {"status": "running", "query": query, "stage": "старт"}
    _spawn(_niche_job_run(payload, query))
    return {"building": True, "query": query, "stage": "старт"}


# ══ Расчёт ущерба по складу (пожар/утрата): срез из платного хранения ═════════
async def _fire_report_data(warehouse: str, on_date: str) -> dict:
    """warehouse — один или несколько складов через запятую."""
    import wb_client
    import cost_store
    import catalog as _cat

    end = on_date
    start = (datetime.fromisoformat(on_date) - timedelta(days=3)).strftime("%Y-%m-%d")
    rows = await wb_client.get_paid_storage(start, end)
    wanted = [w.strip().lower() for w in warehouse.split(",") if w.strip()]

    # последний доступный день ≤ даты события по каждому (склад, nmId)
    per_key: dict = {}
    for r in rows:
        wh_raw = str(r.get("warehouse") or "")
        wl = wh_raw.lower()
        match = next((w for w in wanted if w in wl), None)
        if not match:
            continue
        d = str(r.get("date") or "")[:10]
        nm = r.get("nmId")
        if not nm:
            continue
        cur = per_key.setdefault((wh_raw, nm), {"date": "", "qty": 0,
                                                "vendorCode": r.get("vendorCode") or ""})
        if d > cur["date"]:
            cur["date"] = d
            cur["qty"] = 0
        if d == cur["date"]:
            cur["qty"] += int(r.get("barcodesCount") or 0)

    costs = cost_store.get_costs()
    names = cost_store.get_names()
    try:
        prices = await wb_client.get_current_prices()
    except Exception:
        prices = {}

    by_wh: dict = {}
    for (wh_raw, nm), v in per_key.items():
        if v["qty"] <= 0:
            continue
        art = _cat.canon(v["vendorCode"]) or str(nm)
        cost = costs.get(art) or 0
        retail = (prices.get(art) or {}).get("discounted") or 0
        by_wh.setdefault(wh_raw, []).append({
            "sku": art, "nmId": nm, "name": names.get(art) or _cat.lookup(art).get("name", ""),
            "qty": v["qty"], "snapshot_date": v["date"],
            "cost": cost, "cost_sum": round(v["qty"] * cost),
            "retail": retail, "retail_sum": round(v["qty"] * retail),
        })

    warehouses = []
    for wh_raw, items in sorted(by_wh.items()):
        items.sort(key=lambda x: -x["retail_sum"])
        warehouses.append({"warehouse": wh_raw, "items": items,
                           "totals": {"qty": sum(i["qty"] for i in items),
                                      "cost_sum": sum(i["cost_sum"] for i in items),
                                      "retail_sum": sum(i["retail_sum"] for i in items)}})
    grand = {"qty": sum(w["totals"]["qty"] for w in warehouses),
             "cost_sum": sum(w["totals"]["cost_sum"] for w in warehouses),
             "retail_sum": sum(w["totals"]["retail_sum"] for w in warehouses)}
    return {"warehouse": warehouse, "on_date": on_date,
            "warehouses": warehouses, "totals": grand,
            "rows_fetched": len(rows)}


@router.get("/fire_report")
async def fire_report(warehouse: str = Query(default="Электросталь,Котовск"),
                      date: str = Query(default="")):
    """Ущерб по складу на дату (по умолчанию — последние доступные данные):
    остатки из отчёта платного хранения × себес и × розница."""
    on_date = date or (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
    try:
        return await _fire_report_data(warehouse, on_date)
    except Exception as e:
        _log.error("fire_report: %s", e)
        return {"error": str(e)[:300]}


@router.get("/fire_report/export")
async def fire_report_export(warehouse: str = Query(default="Электросталь,Котовск"),
                             date: str = Query(default="")):
    """Excel-расчёт ущерба для претензии WB."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    on_date = date or (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
    d = await _fire_report_data(warehouse, on_date)
    wb_x = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="B45309")
    head = ["Артикул", "nmId", "Название", "Остаток, шт", "Дата среза",
            "Себестоимость, ₽", "Ущерб по себестоимости, ₽",
            "Розничная цена, ₽", "Стоимость по рознице, ₽"]

    def _sheet(ws, title, items, totals):
        ws.append([title])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])
        ws.append(head)
        for c in ws[3]:
            c.font, c.fill = hf, fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for i in items:
            ws.append([i["sku"], i["nmId"], i["name"], i["qty"], i["snapshot_date"],
                       i["cost"] or None, i["cost_sum"], i["retail"] or None,
                       i["retail_sum"]])
        ws.append(["ИТОГО", "", "", totals["qty"], "", "", totals["cost_sum"],
                   "", totals["retail_sum"]])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        for idx in range(1, len(head) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 30 if idx == 3 else 15
        ws.freeze_panes = "A4"

    first = True
    for w in d["warehouses"]:
        ws = wb_x.active if first else wb_x.create_sheet()
        ws.title = w["warehouse"][:28]
        _sheet(ws, f"Склад {w['warehouse']}: остатки на {on_date} "
                   f"(отчёт платного хранения WB)", w["items"], w["totals"])
        first = False
    # сводный лист
    ws = wb_x.active if first else wb_x.create_sheet()
    ws.title = "Итого"
    ws.append([f"Сводный расчёт ущерба на {on_date}"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["Склад", "Остаток, шт", "Ущерб по себестоимости, ₽", "Стоимость по рознице, ₽"])
    for c in ws[3]:
        c.font, c.fill = hf, fill
    for w in d["warehouses"]:
        ws.append([w["warehouse"], w["totals"]["qty"], w["totals"]["cost_sum"],
                   w["totals"]["retail_sum"]])
    t = d["totals"]
    ws.append(["ИТОГО", t["qty"], t["cost_sum"], t["retail_sum"]])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for idx in range(1, 5):
        ws.column_dimensions[get_column_letter(idx)].width = 26

    buf = io.BytesIO()
    wb_x.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="fire_damage.xlsx"'})


# ══ Остатки по складам WB + стоимость по себестоимости ════════════════════════
async def _wh_stocks_data() -> dict:
    import cache as _c
    import cost_store
    import catalog as _cat
    _, _, stocks = await _c.get_raw_data(
        datetime.utcnow() - timedelta(days=1), datetime.utcnow())
    costs = cost_store.get_costs()
    names = cost_store.get_names()

    by_sku: dict = {}
    wh_totals: dict = {}
    for r in stocks:
        art = str(r.get("supplierArticle") or "").strip()
        if not art:
            continue
        wh = (r.get("warehouseName") or "?").strip()
        qty = int(r.get("quantity") or 0)
        s = by_sku.setdefault(art, {"sku": art, "qty": 0, "to_client": 0,
                                    "from_client": 0, "warehouses": {}})
        s["qty"] += qty
        s["to_client"] += int(r.get("inWayToClient") or 0)
        s["from_client"] += int(r.get("inWayFromClient") or 0)
        if qty > 0:
            s["warehouses"][wh] = s["warehouses"].get(wh, 0) + qty
            wh_totals[wh] = wh_totals.get(wh, 0) + qty

    items = []
    for art, s in by_sku.items():
        if s["qty"] <= 0 and s["to_client"] <= 0:
            continue
        cost = costs.get(art) or 0
        s["name"] = names.get(art) or _cat.lookup(art).get("name", "")
        s["group"] = _cat.lookup(art).get("brand", "")
        s["cost"] = cost
        s["value"] = round(s["qty"] * cost)
        s["value_with_way"] = round((s["qty"] + s["to_client"]) * cost)
        items.append(s)
    items.sort(key=lambda x: -x["value"])
    warehouses = sorted(wh_totals.items(), key=lambda kv: -kv[1])
    return {"items": items,
            "warehouses": [{"name": w, "qty": q} for w, q in warehouses],
            "totals": {"qty": sum(i["qty"] for i in items),
                       "to_client": sum(i["to_client"] for i in items),
                       "value": sum(i["value"] for i in items),
                       "value_with_way": sum(i["value_with_way"] for i in items),
                       "no_cost": sum(1 for i in items if not i["cost"])},
            "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}


@router.get("/whstocks")
async def get_wh_stocks():
    """Остатки WB по складам + стоимость остатков по себестоимости."""
    try:
        return await _wh_stocks_data()
    except Exception as e:
        _log.error("whstocks: %s", e)
        return {"items": [], "error": str(e)[:200]}


@router.get("/whstocks/export")
async def wh_stocks_export():
    """Excel: лист «По товарам» (сток×склады×стоимость) + «По складам»."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    d = await _wh_stocks_data()
    items = d["items"]
    whs = [w["name"] for w in d["warehouses"]]

    wb_x = openpyxl.Workbook()
    ws = wb_x.active
    ws.title = "По товарам"
    head = ["Артикул", "Название", "Группа", "Остаток, шт", "К клиенту, шт",
            "От клиента, шт", "Себес, ₽", "Стоимость остатка, ₽",
            "Стоимость с учётом в пути, ₽"] + whs
    ws.append(head)
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4A5568")
    for c in ws[1]:
        c.font, c.fill = hf, fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i in items:
        ws.append([i["sku"], i["name"], i["group"], i["qty"], i["to_client"],
                   i["from_client"], i["cost"] or None, i["value"],
                   i["value_with_way"]] + [i["warehouses"].get(w) for w in whs])
    t = d["totals"]
    ws.append(["ИТОГО", "", "", t["qty"], t["to_client"], "", "", t["value"],
               t["value_with_way"]] + [w2["qty"] for w2 in d["warehouses"]])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    ws.freeze_panes = "D2"
    for idx in range(1, len(head) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 26 if idx == 2 else 13

    ws2 = wb_x.create_sheet("По складам")
    ws2.append(["Склад", "Остаток, шт"])
    for c in ws2[1]:
        c.font, c.fill = hf, fill
    for w in d["warehouses"]:
        ws2.append([w["name"], w["qty"]])
    ws2.column_dimensions["A"].width = 32

    buf = io.BytesIO()
    wb_x.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="wh_stocks.xlsx"'})


# ══ Дашборд конкурентов: ежедневный срез выдачи WB по нашим запросам ══════════
def _comp_init():
    db.execute("""CREATE TABLE IF NOT EXISTS competitor_daily (
        day TEXT, query TEXT, position INTEGER, nm TEXT, brand TEXT,
        name TEXT, price REAL, rating REAL, feedbacks INTEGER, is_ours INTEGER,
        PRIMARY KEY (day, query, nm))""")


def competitor_queries() -> list[str]:
    """Список запросов: свой (задан пользователем, kv_cache) или автогенерация
    из названий SKU — убираем объёмы/фасовку: «Крем для лица ночной 50 мл» →
    «крем для лица ночной»."""
    import snapshot as _snap
    custom = _snap.load("competitor_queries_custom", None)
    if custom:
        return [str(q).strip().lower() for q in custom if str(q).strip()][:20]
    import re
    import catalog as _cat
    import cost_store
    names = cost_store.get_names()
    seen, out = set(), []
    for art in getattr(_cat, "WB_ID_TO_ART", {}).values():
        nm_name = names.get(art) or _cat.lookup(art).get("name", "")
        if not nm_name:
            continue
        q = nm_name.lower()
        q = re.sub(r"\b\d+[.,]?\d*\s*(мл|ml|г|гр|мг|л|шт)\b\.?", " ", q)
        q = re.sub(r"[,()«»\"]| - ", " ", q)
        q = re.sub(r"\s+", " ", q).strip()
        q = " ".join(q.split()[:7])          # длинные хвосты режем
        if len(q) >= 8 and q not in seen:
            seen.add(q)
            out.append(q)
    return out[:15]


async def competitors_collect_daily() -> dict:
    """Суточный сбор: каждый запрос → домашний агент → топ-30 в competitor_daily."""
    import catalog as _cat
    _comp_init()
    ours = {str(nm) for nm in getattr(_cat, "WB_ID_TO_ART", {})}
    day = datetime.utcnow().strftime("%Y-%m-%d")
    queries = competitor_queries()
    ok, fail = 0, 0
    for q in queries:
        try:
            products, _total = await _wb_agent_search(q)
            if not products or not _niche_relevant(products, q):
                fail += 1
                continue
            rows = []
            for pos, p in enumerate(products[:30], 1):
                nm = str(p.get("nm") or "")
                if not nm:
                    continue
                rows.append((day, q, pos, nm, (p.get("brand") or "")[:60],
                             (p.get("name") or "")[:120], p.get("price"),
                             p.get("rating"), int(p.get("feedbacks") or 0),
                             1 if nm in ours else 0))
            if rows:
                await asyncio.to_thread(
                    db.executemany,
                    "INSERT INTO competitor_daily VALUES (?,?,?,?,?,?,?,?,?,?) "
                    "ON CONFLICT(day, query, nm) DO UPDATE SET position=excluded.position, "
                    "price=excluded.price, rating=excluded.rating, feedbacks=excluded.feedbacks"
                    if db.IS_PG else
                    "INSERT OR REPLACE INTO competitor_daily VALUES (?,?,?,?,?,?,?,?,?,?)",
                    rows)
                ok += 1
        except Exception as e:
            _log.warning("competitors %s: %s", q, e)
            fail += 1
        await asyncio.sleep(8)   # щадим домашний агент и WB
    _log.info("competitors: собрано %d/%d запросов за %s", ok, len(queries), day)
    return {"day": day, "ok": ok, "fail": fail, "queries": len(queries)}


@router.get("/competitors/queries")
async def competitors_queries_get():
    """Текущий список запросов + признак «свой/автогенерация»."""
    import snapshot as _snap
    custom = await asyncio.to_thread(_snap.load, "competitor_queries_custom", None)
    return {"queries": competitor_queries(), "custom": bool(custom)}


@router.post("/competitors/queries")
async def competitors_queries_set(body: dict):
    """Сохранить свой список запросов (queries: [...]); пустой = вернуть автогенерацию."""
    import snapshot as _snap
    queries = [str(q).strip().lower() for q in (body.get("queries") or []) if str(q).strip()]
    queries = list(dict.fromkeys(queries))[:20]   # дедуп, максимум 20
    if queries:
        await asyncio.to_thread(_snap.save, "competitor_queries_custom", queries)
    else:
        await asyncio.to_thread(_snap.save, "competitor_queries_custom", [])
    return {"queries": competitor_queries(), "custom": bool(queries)}


@router.post("/competitors/collect")
async def competitors_collect_now():
    """Ручной запуск сбора (кнопка/тест). Требует работающего домашнего агента."""
    return await competitors_collect_daily()


@router.get("/competitors")
async def competitors_get():
    """Последний срез по конкурентам + изменения цен к прошлому срезу."""
    _comp_init()
    days = await asyncio.to_thread(
        db.fetchall, "SELECT DISTINCT day FROM competitor_daily ORDER BY day DESC")
    days = [str(d[0])[:10] for d in days[:2]]
    if not days:
        return {"queries": [], "message": "Данных ещё нет — запустите сбор "
                "(нужен работающий домашний агент)"}
    cur_day = days[0]
    prev_day = days[1] if len(days) > 1 else None
    rows = await asyncio.to_thread(
        db.fetchall,
        "SELECT day, query, position, nm, brand, name, price, rating, feedbacks, is_ours "
        "FROM competitor_daily WHERE day IN (?, ?) ORDER BY query, position",
        (cur_day, prev_day or cur_day))
    prev_price = {}
    prev_pos = {}
    prev_fb = {}
    by_q: dict = {}
    for d, q, pos, nm, brand, name, price, rating, fb, is_ours in rows:
        if str(d)[:10] != cur_day:
            prev_price[(q, nm)] = price
            prev_pos[(q, nm)] = pos
            prev_fb[(q, nm)] = fb
            continue
        fb_delta = (fb - prev_fb[(q, nm)]) if (q, nm) in prev_fb else None
        by_q.setdefault(q, []).append({
            "position": pos, "nm": nm, "brand": brand, "name": name,
            "price": price, "rating": rating, "feedbacks": fb,
            "is_ours": bool(is_ours),
            "price_prev": prev_price.get((q, nm)),
            "position_prev": prev_pos.get((q, nm)),
            # прирост отзывов со вчера и оценка продаж (~1 отзыв на 40 покупок)
            "fb_delta": fb_delta,
            "sales_est": round(fb_delta * 40) if fb_delta and fb_delta > 0 else None,
        })
    return {"day": cur_day, "prev_day": prev_day,
            "queries": [{"query": q, "items": items} for q, items in by_q.items()]}


@router.get("/niche/status")
async def niche_status():
    """Состояние фонового анализа ниши."""
    return dict(_niche_job)


async def _analyze_niche_impl(payload: dict, query: str,
                              products: list, total: int) -> dict:
    today = datetime.utcnow().date().isoformat()
    # прошлые замеры — для оценки продаж по приросту отзывов
    prev_rows = await asyncio.to_thread(
        db.fetchall,
        "SELECT nm_id, feedbacks, snap_date FROM niche_snapshots "
        "WHERE query = ? AND snap_date < ? ORDER BY snap_date", (query, today))
    prev: dict[int, tuple] = {}
    for nm, fb, dt_ in prev_rows:
        prev[int(nm)] = (int(fb or 0), dt_)   # берём самый старый→последняя запись перезапишет? нет: оставляем первый
    # (первый замер по nm — самая старая точка)
    first_seen: dict[int, tuple] = {}
    for nm, fb, dt_ in prev_rows:
        if int(nm) not in first_seen:
            first_seen[int(nm)] = (int(fb or 0), dt_)

    sales_est_available = False
    for idx, p in enumerate(products):
        p["position"] = idx + 1          # позиция в выдаче WB (порядок)
        p["wb_url"] = f"https://www.wildberries.ru/catalog/{p['nm']}/detail.aspx" if p.get("nm") else ""
        p["sales_month_est"] = None
        base = first_seen.get(int(p["nm"] or 0))
        if base:
            fb0, dt0 = base
            try:
                days = (datetime.utcnow().date() - datetime.fromisoformat(dt0).date()).days
            except ValueError:
                days = 0
            if days >= 3:
                delta = max(0, p["feedbacks"] - fb0)
                p["sales_month_est"] = round(delta / days * 30 / _NICHE_REVIEW_RATE)
                sales_est_available = True

    # сохраняем сегодняшний снимок
    snap_rows = [(query, int(p["nm"]), p["feedbacks"], float(p["price"] or 0), today)
                 for p in products if p.get("nm")]
    await asyncio.to_thread(
        db.executemany,
        "INSERT INTO niche_snapshots (query, nm_id, feedbacks, price, snap_date) "
        "VALUES (?,?,?,?,?) ON CONFLICT (query, nm_id, snap_date) DO UPDATE "
        "SET feedbacks = excluded.feedbacks, price = excluded.price", snap_rows)

    # ── метрики ниши ──
    prices = sorted(p["price"] for p in products if p["price"])
    med_price = prices[len(prices) // 2] if prices else 0
    top30 = products[:30]
    fb_total = sum(p["feedbacks"] for p in top30)
    brand_fb: dict[str, int] = {}
    for p in top30:
        brand_fb[p["brand"]] = brand_fb.get(p["brand"], 0) + p["feedbacks"]
    top3_share = round(sum(sorted(brand_fb.values(), reverse=True)[:3]) / fb_total * 100) if fb_total else 0
    newcomers = sum(1 for p in top30 if p["feedbacks"] < 50)
    avg_rating = round(sum(float(p["rating"] or 0) for p in top30) / len(top30), 2) if top30 else 0

    # ── юнит-прикидка ──
    price = float(payload.get("price") or 0) or med_price
    cost = float(payload.get("cost") or 0)
    logistics = float(payload.get("logistics") or 70)
    subj = next((p["subject_id"] for p in top30 if p.get("subject_id")), None)
    commission_pct = await _wb_commission(subj)
    unit = None
    if price:
        comm = price * (commission_pct or 19) / 100
        profit = price - comm - logistics - cost
        unit = {"price": round(price), "commission_pct": commission_pct or 19,
                "commission": round(comm), "logistics": round(logistics),
                "cost": round(cost),
                "profit": round(profit),
                "margin": round(profit / price * 100) if price else 0}

    # ── вердикт Claude ──
    verdict = None
    try:
        if ANTHROPIC_API_KEY:
            comp = "\n".join(
                f"- {p['brand']} «{p['name'][:50]}»: {p['price']}₽, ★{p['rating']}, отзывов {p['feedbacks']}"
                + (f", ~{p['sales_month_est']} прод/мес" if p.get("sales_month_est") else "")
                for p in top30[:20])
            prompt = f"""Ты — аналитик маркетплейсов. Оцени, стоит ли селлеру выходить на Wildberries с товаром по запросу «{query}».

НИША: найдено {total} товаров. Медианная цена {med_price}₽. Средний рейтинг топ-30: {avg_rating}. Суммарно отзывов у топ-30: {fb_total}. Доля топ-3 брендов: {top3_share}% (монополизация). Карточек с <50 отзывов в топ-30: {newcomers} (шанс новичку).
{'ЮНИТ-ПРИКИДКА: цена ' + str(unit['price']) + '₽, комиссия ' + str(unit['commission_pct']) + '%, логистика ' + str(unit['logistics']) + '₽, себес ' + str(unit['cost']) + '₽ → прибыль ' + str(unit['profit']) + '₽ (' + str(unit['margin']) + '%)' if unit and cost else ''}

ТОП КОНКУРЕНТОВ:
{comp}

Дай вердикт: 1) итог одной строкой (ИДТИ / ИДТИ ОСТОРОЖНО / НЕ ИДТИ + почему); 2) 3-4 пункта обоснования с цифрами; 3) при каких условиях заходить (цена, на что давить). Кратко, без воды."""
            import anthropic
            client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
            msg = await client.messages.create(model=_MODEL, max_tokens=2000,
                                               messages=[{"role": "user", "content": prompt}])
            verdict = msg.content[0].text.strip()
    except Exception as e:
        _log.warning("niche verdict: %s", e)

    result = {
        "query": query, "total": total,
        "median_price": med_price,
        "price_min": prices[0] if prices else None,
        "price_max": prices[-1] if prices else None,
        "avg_rating": avg_rating,
        "feedbacks_top30": fb_total,
        "top3_brand_share": top3_share,
        "newcomers_top30": newcomers,
        "sales_est_available": sales_est_available,
        "unit": unit,
        "verdict": verdict,
        "products": products[:30],
        "analyzed_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC"),
    }
    await asyncio.to_thread(
        db.execute,
        "INSERT INTO niche_history (query, data, built_at) VALUES (?,?,?) "
        "ON CONFLICT (query) DO UPDATE SET data = excluded.data, built_at = excluded.built_at",
        (query, json.dumps(result, ensure_ascii=False), datetime.utcnow().isoformat()))
    return result


@router.get("/niche/history")
async def niche_history():
    """Последние проанализированные ниши."""
    _niche_init()
    rows = await asyncio.to_thread(
        db.fetchall, "SELECT query, built_at FROM niche_history ORDER BY built_at DESC")
    return {"items": [{"query": r[0], "built_at": (r[1] or "")[:16].replace("T", " ")}
                      for r in rows[:20]]}


@router.get("/niche/get")
async def niche_get(query: str = Query(...)):
    """Сохранённый анализ ниши (без пересчёта)."""
    _niche_init()
    row = await asyncio.to_thread(
        db.fetchone, "SELECT data FROM niche_history WHERE query = ?", (query.strip().lower(),))
    if not row:
        raise HTTPException(status_code=404, detail="Нет сохранённого анализа")
    return json.loads(row[0])


_WB_BASKETS = [143, 287, 431, 719, 1007, 1061, 1115, 1169, 1313, 1601, 1655,
               1919, 2045, 2189, 2405, 2621, 2837, 3053, 3269, 3485, 3701,
               3917, 4133, 4349, 4565, 4877, 5189, 5501, 5813, 6125, 6437]


def _wb_photo(nm, size: str = "c516x688") -> str:
    """URL заглавного фото карточки WB (basket-XX по vol)."""
    try:
        nm = int(nm)
    except (TypeError, ValueError):
        return ""
    if not nm:
        return ""
    vol, part = nm // 100000, nm // 1000
    b = len(_WB_BASKETS) + 1
    for i, r in enumerate(_WB_BASKETS):
        if vol <= r:
            b = i + 1
            break
    return (f"https://basket-{b:02d}.wbbasket.ru/vol{vol}/part{part}/{nm}"
            f"/images/{size}/1.webp")


@router.get("/visuals")
async def get_visuals(query: str = Query(...), refresh: bool = Query(default=False)):
    """Анализ визуалов топ-20: заглавные фото карточек конкурентов +
    сравнение через Claude (общие приёмы, что работает, как выделиться).
    Берёт товары из последнего анализа ниши по этому запросу."""
    _niche_init()
    q = query.strip().lower()
    db.execute("CREATE TABLE IF NOT EXISTS visuals_history "
               "(query TEXT PRIMARY KEY, data TEXT, built_at TEXT)")
    if not refresh:
        row = await asyncio.to_thread(
            db.fetchone, "SELECT data FROM visuals_history WHERE query = ?", (q,))
        if row:
            return json.loads(row[0])

    nrow = await asyncio.to_thread(
        db.fetchone, "SELECT data FROM niche_history WHERE query = ?", (q,))
    if not nrow:
        return {"items": [], "message": "Сначала соберите нишу в «Калькуляторе ниши» "
                "по этому запросу — визуалы берутся из неё."}
    nd = json.loads(nrow[0])
    # только реальные товары (у рекламных/пустых вставок нет ни имени, ни данных)
    products = [p for p in (nd.get("products") or [])
                if p.get("nm") and (p.get("name") or p.get("feedbacks") or p.get("price"))][:20]
    items = [{
        "position": p.get("position") or (i + 1), "nm": p.get("nm"),
        "name": p.get("name", ""), "brand": p.get("brand", ""),
        "price": p.get("price"), "rating": p.get("rating"),
        "feedbacks": p.get("feedbacks"),
        # реальное фото из карточки (агент) в приоритете, иначе — по формуле
        "photo": p.get("photo") or _wb_photo(p.get("nm"), "c246x328"),
        "photo_big": p.get("photo") or _wb_photo(p.get("nm"), "c516x688"),
        "wb_url": p.get("wb_url") or (f"https://www.wildberries.ru/catalog/{p['nm']}/detail.aspx" if p.get("nm") else ""),
    } for i, p in enumerate(products)]

    # фото качаем на СЕРВЕРЕ и кодируем base64 — Anthropic не может тянуть
    # их сам (WB CDN отдаёт браузеру, но не серверу-фетчеру Anthropic)
    async def _fetch_img(it):
        import httpx
        nm = it.get("nm")
        urls = []
        # реальный URL из карточки (агент) — самый надёжный, в т.ч. для видео
        real = it.get("photo") or ""
        if real.startswith("http"):
            urls.append(real)
        # запас: перебор размеров/кадров по формуле basket
        for size in ("c516x688", "big", "c246x328"):
            for frame in (1, 2, 3):
                u = _wb_photo(nm, size)
                if u:
                    urls.append(u.replace("/1.webp", f"/{frame}.webp"))
        seen_u = set()
        async with httpx.AsyncClient(timeout=12) as c:
            for url in urls:
                if url in seen_u:
                    continue
                seen_u.add(url)
                try:
                    r = await c.get(url, headers={"Referer": "https://www.wildberries.ru/"})
                    if r.status_code == 200 and r.content and len(r.content) > 800:
                        return it, base64.b64encode(r.content).decode(), "image/webp"
                except Exception:
                    continue
        return it, None, None

    analysis = None
    try:
        if ANTHROPIC_API_KEY and items:
            fetched = await asyncio.gather(*[_fetch_img(it) for it in items])
            got = [(it, b64, mt) for it, b64, mt in fetched if b64]
            if not got:
                analysis = ("⚠ Не удалось скачать фото карточек для анализа "
                            "(WB CDN недоступен серверу). Фото видно в сетке ниже.")
            else:
                content = [{"type": "text", "text":
                    f"""Ты — арт-директор и маркетолог маркетплейсов. Перед тобой заглавные фото
топ-{len(got)} карточек WB по запросу «{query}» (в порядке позиций в выдаче).
Проанализируй их ВИЗУАЛ как единую выборку и дай практику:

1) ОБЩИЕ ПРИЁМЫ топа: композиция (продукт крупно / модель / флэтлей), фон и цвета,
   текст на фото (что пишут, где), инфографика, бейджи, эмоции.
2) ЧТО ДЕЛАЮТ ЛИДЕРЫ (первые позиции) иначе, чем нижняя часть.
3) ПАТТЕРНЫ, которые повторяются у большинства (значит, работают в нише).
4) КАК ВЫДЕЛИТЬСЯ: 3-5 конкретных рекомендаций для НОВОЙ карточки, чтобы
   попасть в стандарт ниши, но зацепить взгляд.
Кратко, по делу, с привязкой к тому, что видишь на фото."""}]
                for it, b64, mt in got:
                    content.append({"type": "text", "text": f"Позиция {it['position']}: {it['brand']} — {it['name'][:60]}, {it['price']}₽"})
                    content.append({"type": "image", "source": {
                        "type": "base64", "media_type": mt, "data": b64}})
                import anthropic
                client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
                msg = await client.messages.create(model=_MODEL, max_tokens=3000,
                                                   messages=[{"role": "user", "content": content}])
                analysis = msg.content[0].text.strip()
    except Exception as e:
        _log.warning("visuals analysis: %s", e)
        analysis = f"⚠ Не удалось получить анализ: {str(e)[:200]}"

    result = {"query": query, "items": items, "analysis": analysis,
              "analyzed_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}
    await asyncio.to_thread(
        db.execute,
        "INSERT INTO visuals_history (query, data, built_at) VALUES (?,?,?) "
        "ON CONFLICT (query) DO UPDATE SET data = excluded.data, built_at = excluded.built_at",
        (q, json.dumps(result, ensure_ascii=False), datetime.utcnow().isoformat()))
    return result


@router.post("/visuals/prompt")
async def visuals_prompt(payload: dict):
    """Генерирует промт для nano banana (Gemini image) — заглавная карточка
    WB на основе разбора топ-20 визуалов + комментарии пользователя."""
    query = str(payload.get("query") or "").strip().lower()
    notes = str(payload.get("notes") or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="Нужен запрос")
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY не настроен")
    _niche_init()
    db.execute("CREATE TABLE IF NOT EXISTS visuals_history "
               "(query TEXT PRIMARY KEY, data TEXT, built_at TEXT)")
    row = await asyncio.to_thread(
        db.fetchone, "SELECT data FROM visuals_history WHERE query = ?", (query,))
    if not row:
        raise HTTPException(status_code=404, detail="Сначала соберите визуалы топ-20 по этому запросу")
    vd = json.loads(row[0])
    analysis = vd.get("analysis") or ""
    items = vd.get("items") or []
    comp = "\n".join(f"- поз.{it.get('position')}: {it.get('brand','')}, {it.get('price')}₽, "
                     f"★{it.get('rating')}, отзывов {it.get('feedbacks')}" for it in items[:20])

    sys_prompt = f"""Ты — промт-инженер для генератора изображений nano banana (Google Gemini image).
Задача: составить ГОТОВЫЙ промт для генерации ЗАГЛАВНОГО фото карточки Wildberries
по нише «{query}», опираясь на разбор визуалов топ-20 конкурентов и пожелания селлера.

РАЗБОР ВИЗУАЛОВ ТОП-20:
{analysis}

ТОП-20 (позиции/цены/рейтинги):
{comp}

ПОЖЕЛАНИЯ СЕЛЛЕРА (обязательно учесть, приоритет над общим стандартом):
{notes or '(нет — ориентируйся на стандарт ниши и на то, как выделиться)'}

Верни ответ СТРОГО в таком виде, без лишних пояснений:

ПРОМТ (для nano banana, на английском):
<детальный промт: тип кадра/композиция, продукт (что на флаконе/упаковке), фон и цвета,
текст и инфографика на карточке (крупная цифра-крючок, бейджи, мини-иконки свойств),
стиль освещения, соотношение сторон 3:4 (вертикаль под WB), фотореализм/студийность.
Промт должен и попадать в стандарт ниши, и содержать элемент-отстройку.>

ЧТО ОТСТРАИВАЕТ (на русском, 2-3 пункта):
<чем эта карточка зацепит на фоне топа>"""

    try:
        import anthropic
        client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
        msg = await client.messages.create(model=_MODEL, max_tokens=2200,
                                           messages=[{"role": "user", "content": sys_prompt}])
        return {"prompt": msg.content[0].text.strip()}
    except Exception as e:
        _log.warning("visuals prompt: %s", e)
        raise HTTPException(status_code=500, detail=f"Не удалось сгенерировать промт: {str(e)[:200]}")


@router.get("/niche/export")
async def niche_export(query: str = Query(...)):
    """Выгрузка анализа ниши в Excel: лист 1 — сводка+вердикт, лист 2 —
    топ-30 конкурентов (позиция, цена, рейтинг, отзывы, оценка продаж, ссылка)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    _niche_init()
    row = await asyncio.to_thread(
        db.fetchone, "SELECT data FROM niche_history WHERE query = ?", (query.strip().lower(),))
    if not row:
        raise HTTPException(status_code=404, detail="Нет сохранённого анализа")
    d = json.loads(row[0])
    products = d.get("products", [])

    wb = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="4F46E5")
    link_font = Font(color="0563C1", underline="single")

    def _hdr(ws):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hfill
            c.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Лист 1: Сводка ──
    ws1 = wb.active
    ws1.title = "Сводка ниши"
    ws1.append(["Показатель", "Значение"])
    _hdr(ws1)
    unit = d.get("unit") or {}
    rows1 = [
        ("Запрос", d.get("query", "")),
        ("Товаров найдено", d.get("total", "")),
        ("Медианная цена, ₽", d.get("median_price", "")),
        ("Мин. цена, ₽", d.get("price_min", "")),
        ("Макс. цена, ₽", d.get("price_max", "")),
        ("Средний рейтинг топ-30", d.get("avg_rating", "")),
        ("Отзывов у топ-30", d.get("feedbacks_top30", "")),
        ("Монополизация (доля топ-3 брендов), %", d.get("top3_brand_share", "")),
        ("Карточек с <50 отзывов (шанс новичку)", d.get("newcomers_top30", "")),
    ]
    if unit:
        rows1 += [
            ("— Юнит-прикидка —", ""),
            ("Цена, ₽", unit.get("price", "")),
            ("Комиссия WB, %", unit.get("commission_pct", "")),
            ("Логистика, ₽", unit.get("logistics", "")),
            ("Себестоимость, ₽", unit.get("cost", "")),
            ("Прибыль с единицы, ₽", unit.get("profit", "")),
            ("Маржа, %", unit.get("margin", "")),
        ]
    for k, v in rows1:
        ws1.append([k, v])
    if d.get("verdict"):
        ws1.append(["", ""])
        ws1.append(["Вердикт (Claude)", d["verdict"]])
        ws1.cell(ws1.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws1.column_dimensions["A"].width = 42
    ws1.column_dimensions["B"].width = 80

    # ── Лист 2: Топ-30 конкурентов ──
    ws2 = wb.create_sheet("Конкуренты топ-30")
    ws2.append(["Позиция", "Артикул WB", "Название", "Бренд", "Продавец",
                "Цена, ₽", "Рейтинг", "Отзывы", "~Продаж/мес", "Ссылка WB"])
    _hdr(ws2)
    for p in products:
        ws2.append([
            p.get("position", ""), p.get("nm", ""), p.get("name", ""),
            p.get("brand", ""), p.get("supplier", ""),
            p.get("price", ""), p.get("rating", ""), p.get("feedbacks", ""),
            p.get("sales_month_est") if p.get("sales_month_est") is not None else "—", "",
        ])
        link = p.get("wb_url") or (f"https://www.wildberries.ru/catalog/{p['nm']}/detail.aspx"
                                   if p.get("nm") else "")
        if link:
            cell = ws2.cell(ws2.max_row, 10)
            cell.value = "Открыть"
            cell.hyperlink = link
            cell.font = link_font
    for i, w in enumerate([9, 13, 40, 18, 22, 10, 9, 10, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in ws2.iter_rows(min_row=2):
        r[2].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = "".join(ch for ch in query if ch.isalnum() or ch in " -_")[:40].strip() or "niche"
    from urllib.parse import quote as _q
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=niche.xlsx; filename*=UTF-8''{_q(safe)}.xlsx"})


# ══ Воронка Ozon (Premium-аналитика) ═══════════════════════════════════════════

_funnel_cache: dict = {}
_funnel_ts: float = 0.0

# метрики Premium-аналитики Ozon: воронка от показа до заказа
_FUNNEL_METRICS = ["hits_view_search", "hits_view_pdp", "hits_tocart_pdp",
                   "ordered_units", "revenue", "position_category", "session_view"]


def _funnel_bottleneck(m: dict) -> tuple[str, str]:
    """Где товар теряет продажи: (код, пояснение с цифрами)."""
    search = m.get("hits_view_search") or 0
    pdp = m.get("hits_view_pdp") or 0
    tocart = m.get("hits_tocart_pdp") or 0
    orders = m.get("ordered_units") or 0
    pos = m.get("position_category") or 0
    ctr = pdp / search * 100 if search else 0
    cart = tocart / pdp * 100 if pdp else 0
    buy = orders / tocart * 100 if tocart else 0

    if search < 300:
        return "visibility", f"мало показов ({int(search)} за месяц) — товар почти не видят в поиске. Нужны реклама/трафик и SEO-заголовок."
    # средняя позиция глубокая → показы есть, но далеко в выдаче
    if pos and pos > 80 and ctr < 12:
        return "ctr", (f"позиция в категории ~{int(pos)}, из показов в карточку заходят {ctr:.0f}% "
                       f"— товар глубоко в выдаче. Поднимать ставками/органикой, усилить главное фото и цену в выдаче.")
    if ctr < 8:
        return "ctr", (f"из {int(search)} показов в карточку заходят {ctr:.0f}% "
                       f"— слабый отклик в выдаче. Главное фото, заголовок, цена/скидка, рейтинг.")
    # основное горлышко ниши — карточка → корзина
    if cart < 6:
        return "cart", (f"карточку открыли {int(pdp)} раз, в корзину положили {cart:.1f}% "
                        f"— карточка не убеждает. Контент (фото/видео/инфографика), отзывы, состав, цена.")
    if buy < 30:
        return "checkout", (f"из корзины выкупают {buy:.0f}% — теряете на финише. "
                            f"Сроки доставки, наличие на ближних складах, цена против корзины конкурентов.")
    return "ok", f"воронка здоровая: {ctr:.0f}% в карточку, {cart:.1f}% в корзину, {buy:.0f}% выкуп."


_ozc_cache: dict = {}
_ozc_ts: float = 0.0
_OZ_TURNOVER = {"DEFICIT": "🔴 дефицит", "POPULAR": "🟢 ходовой",
                "AVERAGE": "средний", "SURPLUS": "избыток",
                "NO_SALES": "нет продаж", "UNKNOWN": "—"}
_OZ_TARGET_DAYS = 30


@router.get("/ozon/clusters")
async def get_ozon_clusters(refresh: bool = Query(default=False)):
    """Остатки Ozon по кластерам: покрытие, скорость, что везти / остальное.
    Всё считает сам Ozon (ads/idc/оборачиваемость) — группируем по кластерам."""
    import time as _t
    global _ozc_cache, _ozc_ts
    if not refresh and _ozc_cache and _t.monotonic() - _ozc_ts < 1800:
        return _ozc_cache

    import ozon_client
    import catalog as _cat
    rows = await ozon_client.get_stocks_by_cluster()
    if not rows:
        return {"items": [], "message": "Ozon не вернул аналитику остатков (нужен ключ с доступом к аналитике)"}

    # строки приходят ПО СКЛАДАМ (SKU × склад): ads_cluster/idc_cluster
    # одинаковы у всех складов кластера — сначала склеиваем в (кластер × SKU),
    # суммируя только складские остатки, иначе продажи/день задваиваются
    per_sku: dict = {}
    for r in rows:
        cl = r.get("cluster_name") or "—"
        offer = r.get("offer_id") or ""
        sku = _cat.canon(offer) if offer else str(r.get("sku") or "")
        p = per_sku.setdefault((cl, sku), {
            "cluster": cl, "sku": sku, "name": r.get("name") or "",
            "ads": float(r.get("ads_cluster") or 0),
            "idc": r.get("idc_cluster"),
            "grade": r.get("turnover_grade_cluster") or r.get("turnover_grade") or "UNKNOWN",
            "stock": 0, "excess": 0})
        p["stock"] += int(r.get("available_stock_count") or 0)
        p["excess"] += int(r.get("excess_stock_count") or 0)

    clusters: dict = {}
    for p in per_sku.values():
        ads, idc = p["ads"], p["idc"]
        c = clusters.setdefault(p["cluster"], {"cluster": p["cluster"], "stock": 0,
                                               "ads": 0.0, "excess": 0, "skus": []})
        c["stock"] += p["stock"]
        c["ads"] += ads
        c["excess"] += p["excess"]
        need = 0
        if ads > 0 and idc is not None and idc < _OZ_TARGET_DAYS:
            need = max(0, round((_OZ_TARGET_DAYS - idc) * ads))
        c["skus"].append({"sku": p["sku"], "name": p["name"], "ads": round(ads, 2),
                          "stock": p["stock"], "idc": idc if idc is not None else "∞",
                          "grade": _OZ_TURNOVER.get(p["grade"], p["grade"]),
                          "need": need, "excess": p["excess"]})

    items = []
    for c in clusters.values():
        spd = round(c["ads"], 2)
        cov = round(c["stock"] / spd, 1) if spd > 0 else None
        need_total = sum(s["need"] for s in c["skus"])
        status = ("no_sales" if spd == 0 else "urgent" if (cov or 99) < 7
                  else "warn" if (cov or 99) < 15 else "over" if (cov or 0) > 90 else "ok")
        to_bring = sorted([s for s in c["skus"] if s["need"] > 0], key=lambda x: -x["need"])
        other = sorted([s for s in c["skus"] if s["need"] == 0], key=lambda x: -x["stock"])
        items.append({"cluster": c["cluster"], "stock": c["stock"], "spd": spd,
                      "coverage": cov, "need": need_total, "status": status,
                      "excess": c["excess"],
                      "skus": to_bring[:20], "other_skus": other})
    items.sort(key=lambda x: -x["stock"])

    result = {"items": items, "target_days": _OZ_TARGET_DAYS,
              "weak": sum(1 for it in items if it["status"] in ("urgent", "warn")),
              "excess_total": sum(it["excess"] for it in items),
              "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}
    _ozc_cache = result
    _ozc_ts = _t.monotonic()
    return result


@router.get("/ozon/clusters/export")
async def export_ozon_clusters():
    """Выгрузка остатков Ozon по кластерам в Excel."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    data = await get_ozon_clusters(refresh=False)
    items = data.get("items", [])
    wb = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="3b82f6")

    def _hdr(ws):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hfill
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "Кластеры Ozon"
    ws1.append(["Кластер", "Продаж/день", "Остаток, шт", "Покрытие, дн",
                "К заказу, шт", "Излишки, шт", "Статус"])
    _hdr(ws1)
    ST = {"urgent": "СРОЧНО", "warn": "мало", "ok": "ок", "over": "избыток", "no_sales": "нет продаж"}
    for it in items:
        ws1.append([it["cluster"], it["spd"], it["stock"], it["coverage"],
                    it["need"], it["excess"], ST.get(it["status"], it["status"])])
    for i, w in enumerate([32, 13, 13, 13, 12, 12, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    def _sheet(title, key):
        ws = wb.create_sheet(title)
        ws.append(["Кластер", "Артикул", "Название", "Продаж/день", "Остаток, шт",
                   "Покрытие, дн", "Оборачиваемость", "Везти, шт", "Излишки, шт"])
        _hdr(ws)
        for it in items:
            for s in it.get(key, []):
                ws.append([it["cluster"], s["sku"], s["name"], s["ads"], s["stock"],
                           s["idc"], s["grade"], s["need"], s["excess"]])
        for i, w in enumerate([30, 14, 34, 12, 11, 12, 15, 10, 11], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    _sheet("Что везти", "skus")
    _sheet("Остальное", "other_skus")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ozon_clusters.xlsx"'})


@router.get("/ozon/stocks_probe", include_in_schema=False)
async def ozon_stocks_probe():
    """Диагностика: реальная структура ответа Ozon по остаткам с кластерами.
    Пробуем /v1/analytics/stocks и /v2/analytics/stock_on_warehouses."""
    import ozon_client
    out = {}
    # v1/analytics/stocks — берём первые SKU
    try:
        skus = await ozon_client._get_all_skus()
        batch = [s["sku"] for s in skus if s["sku"]][:20]
        data = await ozon_client._post("/v1/analytics/stocks", {"skus": batch})
        items = data.get("items") or []
        out["v1_stocks"] = {"count": len(items),
                            "keys": sorted(items[0].keys()) if items else [],
                            "sample": items[0] if items else None}
    except Exception as e:
        out["v1_stocks_error"] = str(e)[:300]
    # v2/stock_on_warehouses — есть ли warehouse/cluster
    try:
        data = await ozon_client._post("/v2/analytics/stock_on_warehouses",
                                       {"limit": 20, "offset": 0, "warehouse_type": "ALL"})
        rows = (data.get("result") or {}).get("rows") or []
        out["v2_warehouses"] = {"count": len(rows),
                                "keys": sorted(rows[0].keys()) if rows else [],
                                "sample": rows[0] if rows else None}
    except Exception as e:
        out["v2_warehouses_error"] = str(e)[:300]
    return out


_ozads_cache: dict = {}
_ozads_ts: float = 0.0
_OZADS_TTL = 3 * 3600
_ADV_TYPE_RU = {
    "SEARCH_PROMO": "Продвижение в поиске", "SKU": "Трафареты",
    "REF_BLOGGER": "Блогеры", "BANNER": "Баннер", "BRAND_SHELF": "Брендовая полка",
}


def _ozads_advice_stub(camps: list, total: dict) -> str:
    return ""


@router.get("/ozads")
async def get_ozads(refresh: bool = Query(default=False), days: int = Query(default=28)):
    """Мега-аналитика рекламы Ozon: кампании, ДРР, воронка, куда уходят деньги."""
    import time as _t
    global _ozads_cache, _ozads_ts
    if not refresh and _ozads_cache and _t.monotonic() - _ozads_ts < _OZADS_TTL:
        return _ozads_cache

    import ozon_perf_client as pc
    if not pc.configured():
        return {"items": [], "error": "Не заданы OZON_PERF_CLIENT_ID / OZON_PERF_CLIENT_SECRET "
                "в переменных окружения."}

    # холодный старт: снапшот из БД
    if not refresh and not _ozads_cache:
        import snapshot as _snap
        snap = await asyncio.to_thread(_snap.load, "ozads", None)
        if snap:
            _ozads_cache = snap
            _ozads_ts = _t.monotonic() - _OZADS_TTL + 300
            return snap

    from datetime import date
    d_end = date.today() - timedelta(days=1)     # Ozon отдаёт с задержкой
    d_start = d_end - timedelta(days=days - 1)
    d_from, d_to = d_start.isoformat(), d_end.isoformat()

    try:
        daily = await pc.get_daily(d_from, d_to)
        camps_meta = {str(c.get("id")): c for c in await pc.get_campaigns()}
    except Exception as e:
        return {"items": [], "error": f"Ozon Performance: {str(e)[:200]}"}

    # агрегируем по кампании
    agg: dict[str, dict] = {}
    by_day: dict[str, dict] = {}
    for r in daily:
        a = agg.setdefault(r["id"], {"id": r["id"], "title": r["title"],
                                     "views": 0, "clicks": 0, "spent": 0.0,
                                     "orders": 0, "orders_money": 0.0})
        for k in ("views", "clicks", "orders"):
            a[k] += r[k]
        a["spent"] += r["spent"]
        a["orders_money"] += r["orders_money"]
        if r["title"] and not a["title"]:
            a["title"] = r["title"]
        d = by_day.setdefault(r["date"], {"spent": 0.0, "orders_money": 0.0, "orders": 0})
        d["spent"] += r["spent"]
        d["orders_money"] += r["orders_money"]
        d["orders"] += r["orders"]

    def _metrics(a: dict) -> dict:
        views, clicks, spent = a["views"], a["clicks"], a["spent"]
        orders, om = a["orders"], a["orders_money"]
        meta = camps_meta.get(a["id"], {})
        typ = meta.get("advObjectType", "")
        return {
            **a,
            "type": _ADV_TYPE_RU.get(typ, typ or "—"),
            "state": meta.get("state", "").replace("CAMPAIGN_STATE_", ""),
            "daily_budget": pc._num(meta.get("dailyBudget")),
            "ctr": round(clicks / views * 100, 2) if views else 0,
            "cpc": round(spent / clicks) if clicks else 0,
            "cr": round(orders / clicks * 100, 1) if clicks else 0,   # заказ/клик
            "cpo": round(spent / orders) if orders else 0,            # цена заказа
            "drr": round(spent / om * 100, 1) if om else None,        # ДРР
            "roas": round(om / spent, 1) if spent else None,          # окупаемость
        }

    items = [_metrics(a) for a in agg.values() if a["spent"] > 0 or a["orders"] > 0]
    items.sort(key=lambda x: -x["spent"])

    T = {"views": sum(i["views"] for i in items), "clicks": sum(i["clicks"] for i in items),
         "spent": sum(i["spent"] for i in items), "orders": sum(i["orders"] for i in items),
         "orders_money": sum(i["orders_money"] for i in items)}
    T["ctr"] = round(T["clicks"] / T["views"] * 100, 2) if T["views"] else 0
    T["cpo"] = round(T["spent"] / T["orders"]) if T["orders"] else 0
    T["drr"] = round(T["spent"] / T["orders_money"] * 100, 1) if T["orders_money"] else None
    T["roas"] = round(T["orders_money"] / T["spent"], 1) if T["spent"] else None

    # советы Claude — по агрегатам
    advice = None
    try:
        if ANTHROPIC_API_KEY and items:
            top = "\n".join(
                f"- «{i['title']}» ({i['type']}): расход {round(i['spent'])}₽, {i['orders']} заказов, "
                f"выручка {round(i['orders_money'])}₽, ДРР {i['drr']}%, CTR {i['ctr']}%, цена заказа {i['cpo']}₽"
                for i in items[:15])
            prompt = f"""Ты — специалист по рекламе на Ozon. Вот кампании за {days} дней ({d_from}—{d_to}).
ИТОГО: расход {round(T['spent'])}₽, выручка с рекламы {round(T['orders_money'])}₽, ДРР {T['drr']}%, заказов {T['orders']}.

КАМПАНИИ:
{top}

Дай практику кратко: 1) куда утекают деньги (кампании с высоким ДРР / без заказов); 2) что масштабировать (низкий ДРР, хороший ROAS); 3) 3-4 конкретных действия. С цифрами, без воды."""
            import anthropic
            client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
            msg = await client.messages.create(model=_MODEL, max_tokens=1200,
                                               messages=[{"role": "user", "content": prompt}])
            advice = msg.content[0].text.strip()
    except Exception as e:
        _log.warning("ozads advice: %s", e)

    chart = [{"date": d, **v} for d, v in sorted(by_day.items())]
    result = {"items": items, "total": T, "chart": chart, "advice": advice,
              "period": f"{'.'.join(reversed(d_from.split('-')))} — {'.'.join(reversed(d_to.split('-')))}",
              "days": days, "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}
    _ozads_cache = result
    _ozads_ts = _t.monotonic()
    import snapshot as _snap
    await asyncio.to_thread(_snap.save, "ozads", result)
    return result


@router.get("/ozads/phrases_probe", include_in_schema=False)
async def ozads_phrases_probe():
    """Диагностика: отчёт по поисковым фразам Ozon Performance (async-цикл).
    Ищем search-promo кампании, заказываем отчёт, ждём, скачиваем."""
    import ozon_perf_client as pc
    import asyncio as _a
    from datetime import date
    if not pc.configured():
        return {"error": "not configured"}
    out = {}
    camps = await pc.get_campaigns()
    # кампании продвижения в поиске / трафаретов
    search = [c for c in camps if any(t in str(c.get("advObjectType", "")).upper()
              for t in ("SEARCH", "SKU")) and c.get("state") == "CAMPAIGN_STATE_RUNNING"]
    out["search_campaigns"] = [{"id": c.get("id"), "type": c.get("advObjectType"),
                                "title": c.get("title")} for c in search[:10]]
    if not search:
        out["note"] = "нет активных search-promo кампаний"
        return out
    ids = [str(c.get("id")) for c in search[:5]]
    d_to = (date.today() - timedelta(days=1)).isoformat()
    d_from = (date.today() - timedelta(days=14)).isoformat()

    st, data = await pc.api_post("/api/client/statistics/phrases",
                                 {"campaigns": ids, "dateFrom": d_from, "dateTo": d_to})
    uuid = data.get("UUID") if isinstance(data, dict) else None
    out["request"] = {"status": st, "uuid": uuid}
    if not uuid:
        return out
    for _ in range(40):
        await _a.sleep(3)
        s2, meta = await pc.api_get(f"/api/client/statistics/{uuid}")
        state = meta.get("state") if isinstance(meta, dict) else None
        out["state"] = state
        if state in ("OK", "ERROR"):
            break
    # скачиваем ZIP напрямую и показываем шапку CSV
    import httpx as _hx, io as _io, zipfile as _zip
    async with _hx.AsyncClient(timeout=120) as c:
        r = await c.get(f"{pc.BASE}/api/client/statistics/report",
                        headers=await pc._headers(), params={"UUID": uuid})
    out["download_status"] = r.status_code
    try:
        zf = _zip.ZipFile(_io.BytesIO(r.content))
        out["files"] = zf.namelist()[:5]
        first = zf.read(zf.namelist()[0]).decode("utf-8-sig", errors="replace")
        out["csv_head"] = first.splitlines()[:8]      # шапка + первые строки
    except Exception as e:
        out["unzip_error"] = str(e)[:200]
    return out


_ozphr_cache: dict = {}
_ozphr_ts: float = 0.0


@router.get("/ozphrases")
async def get_ozphrases(refresh: bool = Query(default=False), days: int = Query(default=14)):
    """Поисковые запросы Ozon (Performance): по каким фразам показываются и
    кликают товары в рекламе. Агрегат по фразе с показами/кликами/CTR."""
    import time as _t
    global _ozphr_cache, _ozphr_ts
    if not refresh and _ozphr_cache and _t.monotonic() - _ozphr_ts < 6 * 3600:
        return _ozphr_cache

    import ozon_perf_client as pc
    if not pc.configured():
        return {"items": [], "error": "Performance API не настроен"}

    if not refresh and not _ozphr_cache:
        import snapshot as _snap
        snap = await asyncio.to_thread(_snap.load, "ozphrases_v2", None)
        if snap:
            _ozphr_cache = snap
            _ozphr_ts = _t.monotonic() - 6 * 3600 + 300
            return snap

    import catalog as _cat
    from datetime import date
    d_to = (date.today() - timedelta(days=1)).isoformat()
    d_from = (date.today() - timedelta(days=days)).isoformat()
    try:
        camps = await pc.get_campaigns()
        sku_ids = [str(c.get("id")) for c in camps
                   if "SKU" in str(c.get("advObjectType", "")).upper()
                   and c.get("state") == "CAMPAIGN_STATE_RUNNING"]
        rows = await pc.get_phrases(sku_ids, d_from, d_to) if sku_ids else []
    except Exception as e:
        return {"items": [], "error": f"Ozon Performance: {str(e)[:200]}"}

    # агрегируем ПО ТОВАРУ (sku), внутри — фразы
    prods: dict[str, dict] = {}
    for r in rows:
        ph = r["phrase"].strip()
        skunum = str(r.get("sku") or "")
        if not ph or not skunum:
            continue
        art = _cat.resolve_ozon(skunum) if skunum.isdigit() else skunum
        p = prods.setdefault(skunum, {
            "sku": skunum, "art": art if art and not str(art).isdigit() else "",
            "name": r.get("product") or "", "views": 0, "clicks": 0, "phrases": {}})
        p["views"] += r["views"]
        p["clicks"] += r["clicks"]
        f = p["phrases"].setdefault(ph.lower(), {"phrase": ph, "views": 0, "clicks": 0})
        f["views"] += r["views"]
        f["clicks"] += r["clicks"]

    items = []
    for p in prods.values():
        phrases = []
        for f in p["phrases"].values():
            f["ctr"] = round(f["clicks"] / f["views"] * 100, 1) if f["views"] else 0
            phrases.append(f)
        phrases.sort(key=lambda x: -x["views"])
        items.append({
            "sku": p["sku"], "art": p["art"], "name": p["name"],
            "views": p["views"], "clicks": p["clicks"],
            "ctr": round(p["clicks"] / p["views"] * 100, 1) if p["views"] else 0,
            "phrase_count": len(phrases), "phrases": phrases,
            # для группировки на фронте
            "group": _cat.lookup(p["art"]).get("brand", "") if p["art"] else "",
        })
    items.sort(key=lambda x: -x["views"])

    result = {"items": items,
              "total_views": sum(i["views"] for i in items),
              "total_clicks": sum(i["clicks"] for i in items),
              "period": f"{'.'.join(reversed(d_from.split('-')))} — {'.'.join(reversed(d_to.split('-')))}",
              "days": days, "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}
    _ozphr_cache = result
    _ozphr_ts = _t.monotonic()
    import snapshot as _snap
    await asyncio.to_thread(_snap.save, "ozphrases_v2", result)
    return result


@router.get("/ozphrases/export")
async def ozphrases_export():
    """Выгрузка поисковых запросов Ozon в Excel."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    d = await get_ozphrases(refresh=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Запросы Ozon"
    ws.append(["Товар", "Артикул", "Поисковый запрос", "Показы", "Клики", "CTR %"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4F46E5")
    for it in d.get("items", []):
        # строка-итог по товару
        ws.append([it.get("name") or it.get("sku"), it.get("art") or "",
                   f'ИТОГО {it.get("phrase_count")} фраз', it["views"], it["clicks"], it["ctr"]])
        for f in it.get("phrases") or []:
            ws.append(["", "", f["phrase"], f["views"], f["clicks"], f["ctr"]])
    for i, w in enumerate([40, 12, 46, 10, 9, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ozon_phrases.xlsx"'})


@router.get("/ozads/export")
async def ozads_export():
    """Выгрузка рекламы Ozon в Excel: кампании + метрики + дни."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    d = await get_ozads(refresh=False)
    items = d.get("items", [])
    T = d.get("total", {})
    wb = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="4F46E5")

    def _hdr(ws):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hfill
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "Кампании"
    ws1.append(["Кампания", "Тип", "Статус", "Расход, ₽", "Заказы", "Выручка, ₽",
                "ДРР %", "ROAS", "Цена заказа, ₽", "Показы", "Клики", "CTR %", "CR %"])
    _hdr(ws1)
    ws1.append(["ИТОГО", "", "", round(T.get("spent", 0)), T.get("orders", 0),
                round(T.get("orders_money", 0)), T.get("drr"), T.get("roas"),
                T.get("cpo"), T.get("views"), T.get("clicks"), T.get("ctr"), ""])
    for i in items:
        ws1.append([i.get("title") or i.get("id"), i.get("type"), i.get("state"),
                    round(i.get("spent", 0)), i.get("orders"), round(i.get("orders_money", 0)),
                    i.get("drr"), i.get("roas"), i.get("cpo"), i.get("views"),
                    i.get("clicks"), i.get("ctr"), i.get("cr")])
    for idx, w in enumerate([34, 20, 12, 12, 9, 13, 9, 8, 13, 11, 10, 9, 8], 1):
        ws1.column_dimensions[get_column_letter(idx)].width = w
    ws1.freeze_panes = "A3"

    ws2 = wb.create_sheet("По дням")
    ws2.append(["Дата", "Расход, ₽", "Заказы", "Выручка, ₽", "ДРР %"])
    _hdr(ws2)
    for r in d.get("chart", []):
        drr = round(r["spent"] / r["orders_money"] * 100, 1) if r.get("orders_money") else None
        ws2.append([r["date"], round(r["spent"]), r["orders"], round(r["orders_money"]), drr])
    for idx, w in enumerate([14, 13, 9, 13, 9], 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ozon_ads.xlsx"'})


@router.get("/ozads/probe", include_in_schema=False)
async def ozads_probe():
    """Диагностика Ozon Performance API на живом токене: авторизация,
    список кампаний и кандидаты эндпоинтов статистики/расходов."""
    import ozon_perf_client as pc
    from datetime import date, timedelta
    out = {"configured": pc.configured()}
    if not pc.configured():
        out["hint"] = "не заданы OZON_PERF_CLIENT_ID / OZON_PERF_CLIENT_SECRET"
        return out
    # 1) токен
    try:
        tok = await pc._get_token()
        out["token"] = "ok" if tok else "пусто"
    except Exception as e:
        out["token_error"] = str(e)[:300]
        return out
    # 2) кампании
    try:
        st, data = await pc.api_get("/api/client/campaign")
        camps = (data.get("list") or data.get("campaigns") or []) if isinstance(data, dict) else []
        out["campaigns"] = {"status": st, "count": len(camps),
                            "keys": sorted(camps[0].keys()) if camps else [],
                            "sample": camps[0] if camps else None}
    except Exception as e:
        out["campaigns_error"] = str(e)[:300]
        camps = []
    cid = str((camps[0].get("id") if camps else "")) or ""
    d_to = (date.today() - timedelta(days=1)).isoformat()
    d_from = (date.today() - timedelta(days=8)).isoformat()
    # 3) кандидаты статистики/расходов — пробуем несколько путей
    tests = [
        ("expense/json GET", "GET", "/api/client/statistics/expense/json",
         {"dateFrom": d_from, "dateTo": d_to}),
        ("daily/json GET", "GET", "/api/client/statistics/daily/json",
         {"campaignId": cid, "dateFrom": d_from, "dateTo": d_to}),
        ("statistics POST (async→UUID)", "POST", "/api/client/statistics",
         {"campaigns": [cid] if cid else [], "dateFrom": d_from, "dateTo": d_to, "groupBy": "DATE"}),
    ]
    out["stats_probe"] = {}
    for name, method, path, arg in tests:
        try:
            st, data = await (pc.api_post(path, arg) if method == "POST" else pc.api_get(path, arg))
            body = data if isinstance(data, dict) else str(data)[:600]
            if isinstance(body, dict):
                # укоротим — только ключи и первую строку
                body = {"keys": sorted(body.keys()),
                        "head": {k: body[k] for k in list(body)[:4]}}
            out["stats_probe"][name] = {"status": st, "path": path, "body": body}
        except Exception as e:
            out["stats_probe"][name] = {"error": str(e)[:200]}
    return out


@router.get("/funnel/probe", include_in_schema=False)
async def funnel_probe():
    """Диагностика воронки Ozon: сырые метрики по первым SKU + период."""
    import ozon_client
    from datetime import date
    d_end = date.today() - timedelta(days=1)
    d_from = (d_end - timedelta(days=27)).isoformat()
    d_to = d_end.isoformat()
    rows = await ozon_client.get_analytics_data(d_from, d_to, _FUNNEL_METRICS)
    sample = []
    for r in rows[:8]:
        m = r["metrics"]
        sample.append({"sku": r["sku"], "name": (r.get("name") or "")[:40], **m})
    return {"period": f"{d_from} — {d_to}", "metrics_requested": _FUNNEL_METRICS,
            "rows": len(rows), "sample": sample}


@router.get("/funnel")
async def get_funnel(refresh: bool = Query(default=False)):
    """Воронка Ozon по SKU: показы → карточка → корзина → заказ (Premium)."""
    import time as _t
    global _funnel_cache, _funnel_ts
    if not refresh and _funnel_cache and _t.monotonic() - _funnel_ts < 6 * 3600:
        return _funnel_cache

    # холодный старт: последняя воронка из БД (свежесть 10 мин), потом обновим
    if not refresh and not _funnel_cache:
        import snapshot as _snapmod
        snap = await asyncio.to_thread(_snapmod.load, "oz_funnel", None)
        if snap:
            _funnel_cache = snap
            _funnel_ts = _t.monotonic() - 6 * 3600 + 600
            return snap

    import ozon_client
    import catalog as _cat
    from datetime import date
    # Ozon отдаёт аналитику с задержкой — последний ПОЛНЫЙ день это вчера.
    # Берём завершённое окно: с (вчера−27) по вчера = 28 полных дней.
    d_end = date.today() - timedelta(days=1)
    d_start = d_end - timedelta(days=27)
    d_to = d_end.isoformat()
    d_from = d_start.isoformat()
    try:
        rows = await ozon_client.get_analytics_data(d_from, d_to, _FUNNEL_METRICS)
    except Exception as e:
        msg = str(e)[:300]
        hint = (" Метрики воронки доступны с подпиской Ozon Premium — проверьте, "
                "что она активна и у API-ключа есть роль «Аналитика»."
                if "403" in msg or "premium" in msg.lower() else "")
        return {"items": [], "error": msg + hint}

    items = []
    for r in rows:
        m = r["metrics"]
        if not any(m.values()):
            continue
        sku = _cat.resolve_ozon(r["sku"]) if str(r["sku"]).isdigit() else str(r["sku"])
        search = m.get("hits_view_search") or 0
        pdp = m.get("hits_view_pdp") or 0
        tocart = m.get("hits_tocart_pdp") or 0
        orders = m.get("ordered_units") or 0
        code, why = _funnel_bottleneck(m)
        items.append({
            "sku": sku,
            "name": _cat.lookup(sku).get("name") or r["name"] or sku,
            "group": _cat.lookup(sku).get("brand", ""),
            "search": int(search), "pdp": int(pdp), "tocart": int(tocart),
            "orders": int(orders),
            "revenue": round(m.get("revenue") or 0),
            "sessions": int(m.get("session_view") or 0),
            "position": int(round(m.get("position_category") or 0)) or None,
            "ctr": round(pdp / search * 100, 1) if search else None,
            "cart_pct": round(tocart / pdp * 100, 1) if pdp else None,
            "buy_pct": round(orders / tocart * 100, 1) if tocart else None,
            "bottleneck": code, "bottleneck_why": why,
        })
    order = {"visibility": 0, "ctr": 1, "cart": 2, "checkout": 3, "ok": 4}
    items.sort(key=lambda x: (order.get(x["bottleneck"], 9), -x["revenue"]))
    _d = lambda s: ".".join(reversed(s.split("-")))   # 2026-07-14 → 14.07.2026
    result = {"items": items, "days": 28,
              "period": f"{_d(d_from)} — {_d(d_to)}",
              "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}
    _funnel_cache = result
    _funnel_ts = _t.monotonic()
    import snapshot as _snapmod
    await asyncio.to_thread(_snapmod.save, "oz_funnel", result)
    return result


# ══ Радар трендов: нарастающие поисковые запросы (Ozon API + выгрузки ЛК) ═════
def _trend_init():
    db.execute("""CREATE TABLE IF NOT EXISTS trend_weekly (
        week TEXT, source TEXT, query TEXT,
        cnt REAL, items_cnt REAL, price REAL, conv REAL, gmv REAL,
        d28 REAL, d7 REAL,
        PRIMARY KEY (week, source, query))""")
    for col in ("d28", "d7"):      # таблица могла быть создана без колонок динамики
        try:
            db.execute(f"ALTER TABLE trend_weekly ADD COLUMN {col} REAL")
        except Exception:
            pass


def _trend_monday(d) -> str:
    return (d - timedelta(days=d.weekday())).strftime("%Y-%m-%d")


def _trend_last_week() -> tuple[str, str]:
    """Последняя ЗАВЕРШЁННАЯ неделя, отстоящая от сегодня минимум на 3 дня
    (Ozon досчитывает данные ~3 дня)."""
    from datetime import date as _date
    safe_end = _date.today() - timedelta(days=3)
    monday = safe_end - timedelta(days=safe_end.weekday())     # пн текущей «безопасной» недели
    prev_mon = monday - timedelta(days=7)
    return prev_mon.isoformat(), (prev_mon + timedelta(days=6)).isoformat()


def _trend_pick(d: dict, keys, default=None):
    for k in keys:
        if k in d and d[k] not in (None, ""):
            return d[k]
    return default


_TREND_Q = ("query", "search_query", "query_text", "text", "phrase")
_TREND_CNT = ("queries_count", "query_count", "count", "unique_search_users",
              "search_count", "popularity")
_TREND_GMV = ("gmv", "revenue", "ordered_amount")
_TREND_CONV = ("view_conversion", "conversion", "conv_to_cart")


def _trend_num(v) -> float:
    try:
        return float(str(v).replace(",", ".").replace(" ", "").replace(" ", ""))
    except (TypeError, ValueError):
        return 0.0


def _trend_read_xlsx(raw: bytes) -> list[list]:
    """Читает XLSX; выгрузки Ozon несут битый styles.xml, который openpyxl
    не переваривает — при ошибке подменяем стили на валидные и повторяем."""
    import io
    import warnings
    import zipfile
    import openpyxl

    def _load(data: bytes):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb_x = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = max(wb_x.worksheets, key=lambda s: s.max_row or 0)
        return [list(r) for r in ws.iter_rows(values_only=True)]

    try:
        return _load(raw)
    except Exception:
        pass
    xf = b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>' * 1000
    styles = (b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
              b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
              b'<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
              b'<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
              b'<borders count="1"><border/></borders>'
              b'<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
              b'<cellXfs count="1000">' + xf + b'</cellXfs></styleSheet>')
    src = zipfile.ZipFile(io.BytesIO(raw))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for it in src.infolist():
            dst.writestr(it.filename,
                         styles if it.filename == "xl/styles.xml" else src.read(it.filename))
    return _load(buf.getvalue())


async def trends_collect_ozon() -> dict:
    """Снять тексты запросов своих товаров из Ozon Seller API за последнюю
    завершённую неделю и сложить в trend_weekly (source='ozon_my')."""
    import ozon_client
    import snapshot as _snap
    _trend_init()
    d_from, d_to = _trend_last_week()
    try:
        items = await ozon_client.get_query_details(d_from, d_to)
    except Exception:
        # частая причина 400 — метод требует явный список skus
        skus = [str(p.get("sku")) for p in await ozon_client._get_all_skus()
                if p.get("sku")]
        items = await ozon_client.get_query_details(d_from, d_to, skus=skus)
    if items:
        await asyncio.to_thread(_snap.save, "trends_probe_last",
                                {"period": [d_from, d_to], "sample": items[:3]})
    rows = []
    for it in items:
        q = str(_trend_pick(it, _TREND_Q, "") or "").strip().lower()
        if not q:
            continue
        rows.append((d_from, "ozon_my", q[:200],
                     _trend_num(_trend_pick(it, _TREND_CNT, 0)),
                     0.0, 0.0,
                     _trend_num(_trend_pick(it, _TREND_CONV, 0)),
                     _trend_num(_trend_pick(it, _TREND_GMV, 0))))
    # один запрос приходит по нескольким SKU; частота у запроса рыночная,
    # одна и та же — берём max (не сумму!), GMV же наш — суммируем
    merged: dict = {}
    for w, s, q, cnt, ic, pr, cv, gm in rows:
        m = merged.setdefault((w, s, q), [w, s, q, 0.0, 0.0, 0.0, 0.0, 0.0, None, None])
        m[3] = max(m[3], cnt)
        m[6] = max(m[6], cv)
        m[7] += gm
    rows = [tuple(v) for v in merged.values()]
    if rows:
        await asyncio.to_thread(
            db.executemany,
            "INSERT INTO trend_weekly VALUES (?,?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(week, source, query) DO UPDATE SET cnt=excluded.cnt, "
            "conv=excluded.conv, gmv=excluded.gmv"
            if db.IS_PG else
            "INSERT OR REPLACE INTO trend_weekly VALUES (?,?,?,?,?,?,?,?,?,?)",
            rows)
    _log.info("trends: ozon_my %s—%s → %d запросов", d_from, d_to, len(rows))
    return {"week": d_from, "period": [d_from, d_to], "rows": len(rows)}


def _http_err(e: Exception) -> str:
    """Ошибка httpx с телом ответа — Ozon в 400 пишет, что именно не так."""
    body = ""
    resp = getattr(e, "response", None)
    if resp is not None:
        try:
            body = " · " + resp.text[:250]
        except Exception:
            pass
    return str(e)[:200] + body


@router.post("/trends/collect")
async def trends_collect_now():
    """Ручной сбор поисковых запросов Ozon (кнопка на фронте)."""
    try:
        return await trends_collect_ozon()
    except Exception as e:
        _log.error("trends collect: %s", _http_err(e))
        return {"error": _http_err(e)}


@router.get("/trends/probe", include_in_schema=False)
async def trends_probe():
    """Диагностика: сырой ответ product-queries/details (первые 3 записи)."""
    import ozon_client
    d_from, d_to = _trend_last_week()
    out = {"period": [d_from, d_to]}
    try:
        items = await ozon_client.get_query_details(d_from, d_to)
        out["details"] = {"rows": len(items), "sample": items[:3]}
    except Exception as e:
        out["details_error"] = _http_err(e)
    try:
        skus = [str(p.get("sku")) for p in await ozon_client._get_all_skus() if p.get("sku")]
        out["skus"] = len(skus)
        items = await ozon_client.get_query_details(d_from, d_to, skus=skus[:1000])
        out["details_with_skus"] = {"rows": len(items), "sample": items[:3]}
    except Exception as e:
        out["details_with_skus_error"] = _http_err(e)
    try:
        base = await ozon_client.get_product_queries(d_from, d_to)
        out["base"] = {"rows": len(base), "sample": base[:3]}
    except Exception as e:
        out["base_error"] = _http_err(e)
    try:
        raw = await ozon_client._post("/v1/analytics/product-queries/details", {
            "date_from": ozon_client._ts(d_from),
            "date_to": ozon_client._ts(d_to, end=True),
            "skus": [str(p.get("sku")) for p in await ozon_client._get_all_skus()
                     if p.get("sku")][:1000],
            "page": 0, "page_size": 100, "limit_by_sku": 15})
        out["details_raw"] = {k: (v[:5] if isinstance(v, list) else v)
                              for k, v in raw.items()} if isinstance(raw, dict) else raw
    except Exception as e:
        out["details_raw_error"] = _http_err(e)
    return out


@router.post("/trends/upload")
async def trends_upload(file: UploadFile = File(...), week: str = Query(default="")):
    """Выгрузка «Поисковые запросы» из ЛК Ozon (Аналитика → Расширение
    ассортимента) — XLSX/CSV с РЫНОЧНЫМИ запросами. week = понедельник периода
    выгрузки (иначе — последняя завершённая неделя)."""
    import io
    _trend_init()
    wk = (week or _trend_last_week()[0])[:10]
    raw = await file.read()
    rows: list[list] = []
    name = (file.filename or "").lower()
    try:
        if name.endswith(".csv") or b";" in raw[:2000] and not name.endswith((".xlsx", ".xls")):
            import csv
            text = raw.decode("utf-8-sig", errors="replace")
            delim = ";" if text.count(";") >= text.count(",") else ","
            rows = list(csv.reader(io.StringIO(text), delimiter=delim))
        else:
            rows = _trend_read_xlsx(raw)
    except Exception as e:
        return {"error": f"Не смог прочитать файл: {str(e)[:200]}"}

    # ищем строку заголовков: есть «запрос» и это именно таблица (≥3 колонок),
    # а не строки «Период: …» / «Сортировка: по популярности запроса»
    hdr_i, hdr = -1, []
    for i, r in enumerate(rows[:20]):
        cells = [str(c or "").strip().lower() for c in r]
        filled = sum(1 for c in cells if c)
        if filled >= 3 and any("запрос" in c for c in cells):
            hdr_i, hdr = i, cells
            break
    if hdr_i < 0:
        return {"error": "Не нашёл колонку с запросами — это точно выгрузка "
                         "«Поисковые запросы» из ЛК Ozon?"}

    def _col(*words, exclude=()):
        for j, h in enumerate(hdr):
            if any(w in h for w in words) and not any(x in h for x in exclude):
                return j
        return -1

    c_q = _col("запрос")
    c_cnt = _col("количество запрос", "популярн", "частот")
    if c_cnt < 0:
        c_cnt = _col("запросов", exclude=("поисков",))
    c_items = _col("показано товар", "товаров", "количество товар")
    c_price = _col("цена", "стоимост", exclude=("заказано",))
    c_conv = _col("конверс")
    c_gmv = _col("заказано на сумму")
    c_d28 = _col("динамика за 28")
    c_d7 = _col("динамика за 7")

    def _cell(r, c):
        return _trend_num(r[c]) if 0 <= c < len(r) else 0.0

    out = []
    for r in rows[hdr_i + 1:]:
        if not r or c_q >= len(r):
            continue
        q = str(r[c_q] or "").strip().lower()
        if not q or len(q) < 2:              # пропускаем пустые и строку описаний «—»
            continue
        # динамика в выгрузке — доля (1.42 = +142%); храним в процентах
        out.append((wk, "ozon_market", q[:200],
                    _cell(r, c_cnt), _cell(r, c_items), _cell(r, c_price),
                    _cell(r, c_conv), _cell(r, c_gmv),
                    round(_cell(r, c_d28) * 100, 1) if 0 <= c_d28 < len(r) and r[c_d28] not in (None, "") else None,
                    round(_cell(r, c_d7) * 100, 1) if 0 <= c_d7 < len(r) and r[c_d7] not in (None, "") else None))
    if not out:
        return {"error": "Файл прочитан, но строк с запросами не нашлось"}
    await asyncio.to_thread(
        db.executemany,
        "INSERT INTO trend_weekly VALUES (?,?,?,?,?,?,?,?,?,?) "
        "ON CONFLICT(week, source, query) DO UPDATE SET cnt=excluded.cnt, "
        "items_cnt=excluded.items_cnt, price=excluded.price, conv=excluded.conv, "
        "gmv=excluded.gmv, d28=excluded.d28, d7=excluded.d7"
        if db.IS_PG else
        "INSERT OR REPLACE INTO trend_weekly VALUES (?,?,?,?,?,?,?,?,?,?)",
        out)
    _log.info("trends: upload %s → %d рыночных запросов (неделя %s)",
              file.filename, len(out), wk)
    return {"week": wk, "rows": len(out)}


def _trend_score(series: list[tuple[str, float]]) -> dict:
    """Метрики тренда по недельной серии [(week, cnt), …] (отсортирована).

    slope_pct — средний рост %/нед (лин. регрессия, нормированная на среднее),
    z — последняя точка против истории, accel — ускорение (наклон 2-й половины
    минус 1-й), score 0–100, stage — стадия жизни тренда."""
    import math
    pts = [(i, c) for i, (_w, c) in enumerate(series) if c is not None]
    n = len(pts)
    if n < 2:
        return {"slope_pct": None, "z": None, "accel": None, "score": None,
                "stage": "мало данных"}
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    mean_y = sum(ys) / n or 1.0

    def _slope(px, py):
        m = len(px)
        if m < 2:
            return 0.0
        mx = sum(px) / m
        my = sum(py) / m
        den = sum((x - mx) ** 2 for x in px) or 1.0
        return sum((x - mx) * (y - my) for x, y in zip(px, py)) / den

    slope_pct = _slope(xs, ys) / mean_y * 100
    half = n // 2
    accel = (_slope(xs[half:], ys[half:]) - _slope(xs[:half], ys[:half])) / mean_y * 100
    hist = ys[:-1]
    mh = sum(hist) / len(hist)
    sd = (sum((y - mh) ** 2 for y in hist) / len(hist)) ** 0.5 or max(mh * .1, 1.0)
    z = (ys[-1] - mh) / sd
    score = 50 + 25 * math.tanh(slope_pct / 20) + 15 * math.tanh(z / 2) \
        + 10 * math.tanh(accel / 15)
    if z >= 2.5 and slope_pct > 0:
        stage = "выстрел"
    elif slope_pct >= 15 and z >= 0.5:
        stage = "рост"
    elif slope_pct >= 5:
        stage = "оживление"
    elif slope_pct <= -10:
        stage = "спад"
    else:
        stage = "плато"
    return {"slope_pct": round(slope_pct, 1), "z": round(z, 2),
            "accel": round(accel, 1), "score": round(max(0, min(100, score))),
            "stage": stage}


@router.get("/trends")
async def trends_get(weeks: int = Query(default=12), min_cnt: float = Query(default=0),
                     q: str = Query(default="")):
    """Радар трендов: недельные серии запросов + скоринг «нарастания»."""
    from config import USE_MOCK
    if USE_MOCK:
        demo = [
            ("крем с муцином улитки", "ozon_market", [120, 180, 340, 610, 980, 1720], 210),
            ("пептидная сыворотка", "ozon_market", [850, 900, 1100, 1350, 1600, 2050], 480),
            ("гель для умывания pH", "ozon_my", [40, 44, 52, 66, 78, 95], 0),
            ("лубрикант на водной основе", "ozon_my", [210, 205, 215, 208, 220, 212], 0),
            ("тонер с центеллой", "ozon_market", [2400, 2300, 2100, 1850, 1600, 1350], 890),
        ]
        items = []
        from datetime import date as _date
        mon = _date.today() - timedelta(days=_date.today().weekday() + 7)
        wks = [(mon - timedelta(days=7 * (5 - i))).isoformat() for i in range(6)]
        for q, src, ser, ic in demo:
            sc = _trend_score(list(zip(wks, [float(x) for x in ser])))
            items.append({"query": q, "source": src, "series": ser, "weeks": wks,
                          "last": ser[-1], "items_cnt": ic, "price": 0, "conv": 0,
                          **sc})
        items.sort(key=lambda x: -(x["score"] or 0))
        return {"items": items, "weeks": wks, "sources": {"ozon_my": 2, "ozon_market": 3},
                "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}

    _trend_init()
    weeks = max(4, min(26, weeks))
    rows = await asyncio.to_thread(
        db.fetchall,
        "SELECT week, source, query, cnt, items_cnt, price, conv, gmv, d28, d7 "
        "FROM trend_weekly ORDER BY week")
    if not rows:
        return {"items": [], "weeks": [], "sources": {},
                "message": "Данных ещё нет: нажмите «Собрать из Ozon» и/или "
                           "загрузите выгрузку «Поисковые запросы» из ЛК"}
    words = [w.strip().lower() for w in (q or "").replace(",", " ").split() if w.strip()]
    all_weeks = sorted({str(r[0])[:10] for r in rows})[-weeks:]
    wk_ix = {w: i for i, w in enumerate(all_weeks)}
    by_key: dict = {}
    for w, src, qq, cnt, ic, pr, cv, gm, d28, d7 in rows:
        w = str(w)[:10]
        if w not in wk_ix:
            continue
        if words and not any(x in qq for x in words):
            continue
        it = by_key.setdefault((src, qq), {
            "query": qq, "source": src, "series": [None] * len(all_weeks),
            "items_cnt": 0, "price": 0, "conv": 0, "gmv": 0,
            "d28": None, "d7": None})
        it["series"][wk_ix[w]] = cnt or 0
        if ic:
            it["items_cnt"] = ic
        if pr:
            it["price"] = pr
        if cv:
            it["conv"] = cv
        if gm:
            it["gmv"] = (it["gmv"] or 0) + gm
        if d28 is not None:
            it["d28"] = d28
        if d7 is not None:
            it["d7"] = d7
    items = []
    src_cnt: dict = {}
    for (src, _qq), it in by_key.items():
        pts = [(w, c) for w, c in zip(all_weeks, it["series"])]
        last = next((c for c in reversed(it["series"]) if c is not None), 0)
        if min_cnt and (last or 0) < min_cnt:
            continue
        sc = _trend_score(pts)
        if sc["score"] is None and it["d28"] is not None:
            # серия из одной точки, но у выгрузки ЛК есть готовая динамика:
            # 28-дневный рост → средненедельный темп, ускорение — 7д против него
            import math
            g = it["d28"] / 100
            wk_pct = (abs(1 + g) ** 0.25 * (1 if g >= -1 else -1) - 1) * 100 if g > -1 else -20.0
            d7 = it["d7"] if it["d7"] is not None else 0.0
            accel = d7 - wk_pct
            score = 50 + 25 * math.tanh(wk_pct / 20) + 10 * math.tanh(accel / 15)
            if wk_pct >= 15 and d7 > 0:
                stage = "рост"
            elif wk_pct >= 5:
                stage = "оживление"
            elif wk_pct <= -10:
                stage = "спад"
            else:
                stage = "плато"
            if it["d28"] >= 100 and d7 > 0:
                stage = "выстрел"
            sc = {"slope_pct": round(wk_pct, 1), "z": None,
                  "accel": round(accel, 1), "score": round(max(0, min(100, score))),
                  "stage": stage}
        it.update(sc)
        it["last"] = last
        src_cnt[src] = src_cnt.get(src, 0) + 1
        items.append(it)
    items.sort(key=lambda x: (-(x["score"] or -1), -(x["last"] or 0)))
    # лимит по каждому источнику отдельно — иначе «свои» вытесняются рынком
    capped, per_src = [], {}
    for it in items:
        c = per_src.get(it["source"], 0)
        if c < 400:
            capped.append(it)
            per_src[it["source"]] = c + 1
    return {"items": capped, "weeks": all_weeks, "sources": src_cnt,
            "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}


# ══ Агент-стратег ═════════════════════════════════════════════════════════════
def _owner_only(request):
    import auth as _auth
    sess = _auth._session_of(request) or {}
    if sess.get("role") != "owner":
        raise HTTPException(403, "Только для владельца")


@router.get("/strategy")
async def strategy_get(request: Request):
    """Текущий план стратега + задачи с статусами (только владелец)."""
    _owner_only(request)
    import agent_strategist as st
    import snapshot as _snap
    plan = await asyncio.to_thread(_snap.load, "strategist_plan", None) or {}
    tasks = await asyncio.to_thread(st._tasks_load)
    return {"plan": plan.get("text", ""), "updated": plan.get("updated"),
            "tasks": tasks, "due": len(st.due_tasks()),
            "running": st._running}


@router.post("/strategy/run")
async def strategy_run(request: Request, body: dict | None = None):
    """Запустить стратегическую сессию (фоном). body.focus — фокус сессии."""
    _owner_only(request)
    import agent_strategist as st
    if st._running:
        return {"error": "Стратег уже работает — дождись окончания текущей сессии"}
    focus = str((body or {}).get("focus") or "")
    _spawn(st.run_session(trigger="кнопка на дашборде", focus=focus))
    return {"started": True}


# ══ Репрайсер (владелец) ══════════════════════════════════════════════════════
_repr_ov_cache: dict = {}
_repr_ov_building = False


async def _repr_refresh_bg():
    global _repr_ov_cache, _repr_ov_building
    import repricer as rp
    import snapshot as _snap
    try:
        ov = await rp.overview()
        ov["fetched_at"] = datetime.utcnow().strftime("%H:%M UTC")
        _repr_ov_cache = ov
        await asyncio.to_thread(_snap.save, "repricer_overview", ov)
    except Exception as e:
        _log.warning("repricer bg: %s", e)
    finally:
        _repr_ov_building = False


@router.get("/repricer")
async def repricer_get(request: Request):
    """Конфиг репрайсера × цены × маржа. Мгновенно из кеша + фон-обновление."""
    global _repr_ov_building, _repr_ov_cache
    _owner_only(request)
    import repricer as rp
    import snapshot as _snap
    if not _repr_ov_cache:      # рестарт: поднимаем последний снимок из БД
        snap = await asyncio.to_thread(_snap.load, "repricer_overview", None)
        if snap:
            _repr_ov_cache = snap
    if _repr_ov_cache:
        # конфиг и предложения всегда свежие (лёгкие), цены/маржа — из кеша
        cfg = {c["art"]: c for c in await asyncio.to_thread(rp.cfg_load)}
        for it in _repr_ov_cache.get("items", []):
            c = cfg.get(it["art"])
            if c:
                it.update({k: c[k] for k in ("target", "min_ozon", "min_wb", "active")})
        _repr_ov_cache["proposals"] = await asyncio.to_thread(rp.proposals_load)
        if not _repr_ov_building:
            _repr_ov_building = True
            _spawn(_repr_refresh_bg())
        return _repr_ov_cache
    # холодный старт: быстрый снимок без юнитки (секунды), полный — фоном
    ov = await rp.overview(include_margin=False)
    ov["fetched_at"] = datetime.utcnow().strftime("%H:%M UTC")
    ov["partial"] = True
    globals()["_repr_ov_cache"] = ov
    if not _repr_ov_building:
        globals()["_repr_ov_building"] = True
        _spawn(_repr_refresh_bg())
    return ov


@router.post("/repricer/set")
async def repricer_set(request: Request, body: dict):
    """Правка строки: {art, target?, min_wb?, min_ozon?, active?}."""
    _owner_only(request)
    import repricer as rp
    art = str(body.get("art") or "")
    ok = await asyncio.to_thread(
        rp.cfg_set, art, target=body.get("target"), min_wb=body.get("min_wb"),
        min_ozon=body.get("min_ozon"),
        active=None if body.get("active") is None else int(bool(body.get("active"))))
    return {"ok": ok}


@router.post("/repricer/preset")
async def repricer_preset(request: Request, body: dict):
    """Переключить конфиг на пресет low/mid/high."""
    _owner_only(request)
    import repricer as rp
    n = await asyncio.to_thread(rp.preset_apply, str(body.get("name") or ""))
    return {"applied": n}


@router.post("/repricer/proposal")
async def repricer_proposal(request: Request, body: dict):
    """Принять/отклонить предложение стратега: {art, action: apply|reject}."""
    _owner_only(request)
    import repricer as rp
    art = str(body.get("art") or "")
    if body.get("action") == "apply":
        return {"ok": await asyncio.to_thread(rp.proposal_apply, art)}
    return {"ok": await asyncio.to_thread(rp.proposal_reject, art)}


@router.get("/repricer/export")
async def repricer_export(request: Request):
    """XLSX в формате шаблона Ozon — загрузить в ЛК Ozon → Репрайсер."""
    _owner_only(request)
    import io
    import repricer as rp
    from fastapi.responses import StreamingResponse
    data = await asyncio.to_thread(rp.export_ozon_xlsx)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=repricer_ozon.xlsx"})


@router.post("/repricer/sync")
async def repricer_sync(request: Request):
    """Цели репрайсера = фактические цены для покупателя из кабинетов."""
    _owner_only(request)
    import repricer as rp
    n = await rp.sync_targets_to_fact()
    return {"synced": n}


@router.get("/repricer/probe", include_in_schema=False)
async def repricer_probe(request: Request):
    """Сырые ценовые поля Ozon по первым товарам — подбор правильных полей."""
    _owner_only(request)
    import ozon_client
    data = await ozon_client._post("/v5/product/info/prices",
                                   {"filter": {"visibility": "ALL"}, "limit": 5})
    items = data.get("items") or (data.get("result") or {}).get("items") or []
    return {"count": len(items), "raw": items[:4]}


# ══ Сравнение FBW → FBS: как изменятся затраты ════════════════════════════════
@router.get("/fbs_compare")
async def fbs_compare():
    """По каждому SKU: комиссия FBW против FBS (официальные тарифы), логистика
    FBW (факт из юнитки) против расчётной FBS (база + литры), хранение WB → 0.
    Скорость доставки и конверсия НЕ моделируются — только прямые затраты."""
    import wb_client

    data = await get_margin(mp="WB")
    items = data.get("items") or []
    if not items:
        return {"items": [], "message": data.get("message") or "юнитка собирается"}

    # nmId → (fbo pct, fbs pct) напрямую из тарифов × категорий карточек
    try:
        tariffs, subjects = await asyncio.gather(
            wb_client.get_commission_tariffs(), wb_client.get_card_subjects())
    except Exception as e:
        return {"items": [], "error": f"тарифы WB недоступны: {str(e)[:150]}"}
    comm_nm = {}
    for nm, sinfo in subjects.items():
        t = tariffs.get(int(sinfo.get("subjectID") or 0))
        if t:
            comm_nm[str(nm)] = {"fbo": t.get("fbo"), "fbs": t.get("fbs")}

    def _n(v):
        try:
            return float(str(v).replace(",", "."))
        except (TypeError, ValueError):
            return 0.0

    # объём единицы, л — из отчёта платного хранения (последние 2 дня)
    vols: dict = {}
    try:
        from datetime import date as _date
        d_to = (_date.today() - timedelta(days=1)).isoformat()
        d_from = (_date.today() - timedelta(days=2)).isoformat()
        for r in await wb_client.get_paid_storage(d_from, d_to):
            v = _n(r.get("volume"))
            if v > 0 and r.get("nmId"):
                vols[str(r["nmId"])] = v
    except Exception as e:
        _log.warning("fbs volumes: %s", e)

    # логистика FBS: медианный тариф Marketplace по складам
    fbs_base, fbs_liter = 50.0, 5.0     # дефолт, если API недоступен
    try:
        boxes = await wb_client.get_box_tariffs()
        bases = sorted(_n(w.get("boxDeliveryMarketplaceBase")) for w in boxes
                       if _n(w.get("boxDeliveryMarketplaceBase")) > 0)
        liters = sorted(_n(w.get("boxDeliveryMarketplaceLiter")) for w in boxes
                        if _n(w.get("boxDeliveryMarketplaceLiter")) > 0)
        if bases:
            fbs_base = bases[len(bases) // 2]
        if liters:
            fbs_liter = liters[len(liters) // 2]
    except Exception as e:
        _log.warning("fbs box tariffs: %s", e)

    out, t_fbo, t_fbs, t_qty = [], 0.0, 0.0, 0
    for b in items:
        sku = b.get("sku")
        price = b.get("price0") or 0
        qty = round(b.get("qty_f") if b.get("qty_f") is not None else b.get("qty_m", 0) or 0)
        import catalog as _cat
        nm_id = next((str(n) for n, a in getattr(_cat, "WB_ID_TO_ART", {}).items()
                      if a == sku), "")
        cinfo = comm_nm.get(nm_id) or {}
        fbo_pct = cinfo.get("fbo") or b.get("comm_pct") or 0
        fbs_pct = cinfo.get("fbs")
        if fbs_pct is None:
            fbs_pct = fbo_pct                     # нет тарифа — считаем равной
        vol = vols.get(nm_id, 0) or 1.0
        # прямая доставка FBS + 20% надбавка на невыкупы/возвраты (факт FBW
        # в юнитке невыкупы уже содержит — выравниваем сопоставимость)
        fbs_logist = round((fbs_base + fbs_liter * max(vol - 1, 0)) * 1.2, 1)
        fbo_logist = b.get("logist") or 0
        storage = b.get("storage") or 0
        d_unit = round((price * (fbs_pct - fbo_pct) / 100)
                       + (fbs_logist - fbo_logist) - storage, 1)
        out.append({
            "sku": sku, "price": price, "qty_month": qty, "volume_l": vol,
            "fbo_comm_pct": round(fbo_pct, 1), "fbs_comm_pct": round(fbs_pct, 1),
            "fbo_logist": round(fbo_logist, 1), "fbs_logist": fbs_logist,
            "fbo_storage": round(storage, 1),
            "delta_unit": d_unit,                  # >0 = FBS дороже на штуку
            "delta_month": round(d_unit * qty),
        })
        t_fbo += (price * fbo_pct / 100 + fbo_logist + storage) * qty
        t_fbs += (price * fbs_pct / 100 + fbs_logist) * qty
        t_qty += qty
    out.sort(key=lambda x: x["delta_month"])
    return {"items": out, "qty_month": t_qty,
            "fbo_month": round(t_fbo), "fbs_month": round(t_fbs),
            "delta_month": round(t_fbs - t_fbo),
            "fbs_tariff": {"base": fbs_base, "liter": fbs_liter},
            "note": "delta>0 — FBS дороже. Не учтены: своё хранение/сборка/дорога "
                    "до пункта приёма WB, падение конверсии из-за сроков доставки, "
                    "буст карточек FBW в выдаче."}


# ══ Ставка CPM: запросы и ставки действующих РК ═══════════════════════════════
_bidq_cache: dict = {}
_bidq_ts: float = 0.0
_bidq_building = [False]


async def _bidq_refresh_bg():
    global _bidq_cache, _bidq_ts
    import time as _t
    import snapshot as _snap
    try:
        out = await bid_queries(refresh=True)
        if out.get("rows"):
            await asyncio.to_thread(_snap.save, "bidq_default", out)
    except Exception as e:
        _log.warning("bidq bg: %s", e)
    finally:
        _bidq_building[0] = False


def _bidq_init():
    db.execute("""CREATE TABLE IF NOT EXISTS adv_cluster_daily (
        day TEXT, advert_id INTEGER, nm TEXT, cluster TEXT,
        views INTEGER, clicks INTEGER, orders_cnt INTEGER,
        spend REAL, bid REAL,
        PRIMARY KEY (day, advert_id, nm, cluster))""")


async def bid_collect_daily() -> dict:
    """Суточный срез кластерной статистики и ставок в БД — вечная история."""
    import advert_client as ac
    _bidq_init()
    day = (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
    ids = await ac.get_all_campaign_ids_ext()
    details = await ac.get_campaigns_info(ids[:50])
    act = [c for c in details if c.get("status") in (9, 11)]
    pairs = [(int(c["advertId"]), int(n["nm_id"]))
             for c in act for n in (c.get("nm_settings") or []) if n.get("nm_id")]
    stats = await ac.get_cluster_stats(pairs, day, day)
    bids = {}
    for r in await ac.get_cluster_bids(pairs):
        bids[(int(r.get("advert_id") or 0), str(r.get("cluster") or "").lower())] = r.get("bid")
    rows = [(day, int(st["advert_id"]), str(st["nm_id"]), st["cluster"],
             st.get("views") or 0, st.get("clicks") or 0, st.get("orders") or 0,
             st.get("spend") or 0,
             bids.get((int(st["advert_id"]), str(st["cluster"] or "").lower())))
            for st in stats if st.get("cluster")]
    if rows:
        await asyncio.to_thread(
            db.executemany,
            "INSERT INTO adv_cluster_daily VALUES (?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(day, advert_id, nm, cluster) DO UPDATE SET "
            "views=excluded.views, clicks=excluded.clicks, "
            "orders_cnt=excluded.orders_cnt, spend=excluded.spend, bid=excluded.bid"
            if db.IS_PG else
            "INSERT OR REPLACE INTO adv_cluster_daily VALUES (?,?,?,?,?,?,?,?,?)",
            rows)
    _log.info("bid history: %s — %d строк кластеров", day, len(rows))
    return {"day": day, "rows": len(rows)}


@router.get("/bid/queries")
async def bid_queries(refresh: bool = Query(default=False),
                      date_from: str = Query(default=""),
                      date_to: str = Query(default="")):
    """Активные кампании WB: ставка кампании + фразы со статистикой CTR,
    привязанные к SKU и его юнитке. Кеш 1 час (advert API лимитный)."""
    import time as _t
    import snapshot as _snap
    global _bidq_cache, _bidq_ts
    default_period = not date_from and not date_to
    if default_period and not _bidq_cache:
        snap = await asyncio.to_thread(_snap.load, "bidq_default", None)
        if snap:
            _bidq_cache = snap
            _bidq_ts = _t.monotonic() - 3600
    if not refresh and default_period and _bidq_cache:
        if _t.monotonic() - _bidq_ts >= 3600 and not _bidq_building[0]:
            _bidq_building[0] = True
            _spawn(_bidq_refresh_bg())
        return _bidq_cache
    import advert_client as ac
    import catalog as _cat

    data = await get_margin(mp="WB")
    marg = {}
    for b in data.get("items") or []:
        import agent_review as _ar
        m = _ar._margin_math(b)
        marg[str(b.get("sku", "")).upper()] = {
            "price": m["price_seller"], "be_drr": m["be_drr_pct"]}
    nm_to_art = {str(k): v for k, v in getattr(_cat, "WB_ID_TO_ART", {}).items()}

    try:
        ids = await asyncio.wait_for(ac.get_all_campaign_ids_ext(), timeout=45)
        _log.info("bid/queries: кампаний всего %d", len(ids))
        details = await asyncio.wait_for(ac.get_campaigns_info(ids[:50]), timeout=45)
    except asyncio.TimeoutError:
        return {"rows": [], "error": "advert API не ответил — попробуй позже"}
    except Exception as e:
        return {"rows": [], "error": f"advert API: {str(e)[:200]}"}
    # 9 = идут показы, 11 = пауза; завершённые не интересны
    act = [c for c in details if c.get("status") in (7, 9, 11)] or details
    _log.info("bid/queries: с деталями %d, активных/пауза %d", len(details), len(act))

    # пары (кампания, артикул) из настроек товаров кампаний (спека v2/adverts)
    pairs, camp_meta = [], {}
    for c in act:
        cid = int(c.get("advertId") or c.get("id") or 0)
        st = c.get("settings") or {}
        camp_meta[cid] = {"name": st.get("name") or c.get("name") or str(cid),
                          "bid_type": c.get("bid_type"), "status": c.get("status")}
        for nm_s in c.get("nm_settings") or []:
            nm = nm_s.get("nm_id")
            if not nm:
                continue
            bk = nm_s.get("bids_kopecks") or {}
            camp_meta[cid].setdefault("nm_bids", {})[int(nm)] = round(
                (bk.get("search") or bk.get("recommendations") or 0) / 100)
            pairs.append((cid, int(nm)))
    _log.info("bid/queries: пар кампания×артикул %d", len(pairs))

    from datetime import date as _date
    d_to = (date_to or (_date.today() - timedelta(days=1)).isoformat())[:10]
    d_from = (date_from or (_date.today() - timedelta(days=14)).isoformat())[:10]
    stats = await ac.get_cluster_stats(pairs, d_from, d_to)
    cbids = {}
    try:
        for r in await ac.get_cluster_bids(pairs):
            key = (int(r.get("advert_id") or 0),
                   str(r.get("cluster") or "").lower())
            if r.get("bid"):
                cbids[key] = r["bid"]
    except Exception as e:
        _log.warning("bid/queries cluster bids: %s", e)

    rows = []
    seen_clusters = set()
    for st in sorted(stats, key=lambda x: -(x.get("views") or 0))[:1000]:
        cid = int(st.get("advert_id") or 0)
        nm = st.get("nm_id")
        meta = camp_meta.get(cid, {})
        art = nm_to_art.get(str(nm)) or str(nm)
        m = marg.get(str(art).upper()) or {}
        clicks, orders = st.get("clicks") or 0, st.get("orders") or 0
        rows.append({
            "campaign": meta.get("name"), "camp_id": cid,
            "skus": [art], "phrase": st.get("cluster"),
            "views": st.get("views"), "clicks": clicks, "ctr": st.get("ctr"),
            "orders": orders,
            "cr": round(orders / clicks * 100, 1) if clicks else None,
            "spend": st.get("spend"), "avg_pos": st.get("avg_pos"),
            "bid": cbids.get((cid, str(st.get("cluster") or "").lower()))
                   or (meta.get("nm_bids") or {}).get(int(nm) if nm else 0),
            "price": m.get("price"), "be_drr": m.get("be_drr"),
            "cluster": False,
        })
        seen_clusters.add((cid, str(st.get("cluster") or "").lower()))
    # фразы со ставкой, но без показов за период — тоже показываем (Саша: «не все фразы»)
    for (cid, cl), bidv in cbids.items():
        if (cid, cl) in seen_clusters or not cl:
            continue
        meta = camp_meta.get(cid, {})
        nm = next(iter(meta.get("nm_bids") or {}), None)
        art = nm_to_art.get(str(nm)) or (str(nm) if nm else "?")
        m = marg.get(str(art).upper()) or {}
        rows.append({"campaign": meta.get("name"), "camp_id": cid,
                     "skus": [art], "phrase": cl, "views": 0, "clicks": 0,
                     "ctr": None, "orders": 0, "cr": None, "spend": None,
                     "avg_pos": None, "bid": bidv,
                     "price": m.get("price"), "be_drr": m.get("be_drr"),
                     "cluster": False, "no_stats": True})
    if not rows and act:
        for cid, meta in camp_meta.items():
            rows.append({"campaign": meta.get("name"), "camp_id": cid,
                         "skus": sorted({nm_to_art.get(str(n), str(n))
                                         for n in (meta.get("nm_bids") or {})}),
                         "bid": next(iter((meta.get("nm_bids") or {}).values()), None),
                         "phrase": None, "no_words": True})
    out = {"rows": rows, "campaigns": len(act), "active_ids": len(ids),
           "period": [d_from, d_to],
           "fetched_at": datetime.utcnow().strftime("%d.%m %H:%M UTC")}
    if default_period:
        _bidq_cache, _bidq_ts = out, _t.monotonic()
        await asyncio.to_thread(_snap.save, "bidq_default", out)
    return out


@router.get("/bid/probe", include_in_schema=False)
async def bid_probe(request: Request):
    """Сырые ответы кандидатов на детали кампаний — подбор полей."""
    _owner_only(request)
    import advert_client as ac
    out = {}
    try:
        ids = await ac.get_all_campaign_ids_ext()
        out["ids"] = ids[:10]
    except Exception as e:
        out["ids_error"] = str(e)[:200]
        ids = []
    for name, coro in (
        ("promotion_adverts", ac._post("/adv/v1/promotion/adverts", ids[:3])),
        ("auction_adverts", ac._get("/adv/v0/auction/adverts")),
        ("advert_v2", ac._get("/api/advert/v2/adverts")),
    ):
        try:
            data = await coro
            s = data if isinstance(data, list) else [data]
            out[name] = s[:2]
        except Exception as e:
            out[name + "_error"] = str(e)[:200]
    return out


@router.post("/adv/sandbox", include_in_schema=False)
async def adv_sandbox(request: Request, body: dict):
    """Owner-only прокси к advert-api (песочница или бой) для отладки методов.
    body: {path, method: GET|POST|PATCH|PUT, params?, json?, sandbox: true}."""
    _owner_only(request)
    import httpx
    import advert_client as ac
    base = ("https://advert-api-sandbox.wildberries.ru" if body.get("sandbox", True)
            else "https://advert-api.wildberries.ru")
    path = str(body.get("path") or "")
    if not path.startswith("/"):
        return {"error": "path должен начинаться с /"}
    method = str(body.get("method") or "GET").upper()
    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.request(method, base + path, headers=ac._headers(),
                                params=body.get("params"), json=body.get("json"))
        try:
            payload = r.json()
        except Exception:
            payload = r.text[:2000]
        return {"status": r.status_code, "body": payload}
    except Exception as e:
        return {"error": str(e)[:300]}


@router.get("/supply/volumes", include_in_schema=False)
async def supply_volumes(request: Request):
    """Точный литраж каждой штуки из габаритов карточек Ozon."""
    _owner_only(request)
    import ozon_client
    vols = await ozon_client.get_product_volumes()
    return {"count": len(vols),
            "volumes": {k: v for k, v in sorted(vols.items())}}


# ══ Ловец слотов поставки Ozon ════════════════════════════════════════════════
@router.get("/supply/orders")
async def supply_orders(request: Request):
    """Заявки на поставку Ozon (включая черновики DATA_FILLING)."""
    _owner_only(request)
    import ozon_client
    orders = await ozon_client.get_supply_orders()
    slim = []
    for o in orders:
        ts = o.get("timeslot") or {}
        slim.append({"order_id": o.get("order_id"),
                     "number": o.get("order_number") or o.get("supply_order_number"),
                     "state": o.get("state"),
                     "dropoff": (o.get("dropoff_warehouse") or {}).get("name"),
                     "created": (o.get("created_date") or "")[:16],
                     "timeslot": ts})
    return {"orders": slim}


@router.get("/supply/probe", include_in_schema=False)
async def supply_probe(request: Request, q: str = ""):
    """Сырой просмотр заявок в браузере: /api/tools/supply/probe?q=119613312.
    Отдаёт полный JSON заявки + доступные таймслоты по ней."""
    _owner_only(request)
    import ozon_client
    orders = await ozon_client.get_supply_orders()
    out = []
    for o in orders:
        num = str(o.get("order_number") or o.get("supply_order_number") or "")
        if q and q not in num and q != str(o.get("order_id")):
            continue
        item = {"order": o}
        try:
            item["timeslots"] = await ozon_client.get_order_timeslots(o.get("order_id"))
        except Exception as e:
            item["timeslots_error"] = str(e)[:400]
        out.append(item)
        await asyncio.sleep(0.5)
    return {"found": len(out), "total_orders": len(orders), "orders": out}


# ══ Короба (грузоместа) заявок Ozon из файла производства ════════════════════
_BOX_ALIAS = {"питер": "санкт петербург", "спб": "санкт петербург",
              "мск": "москва", "екб": "екатеринбург"}


def _box_norm(s) -> str:
    """Нормализация имени кластера: буквы+пробелы, нижний регистр."""
    s = re.sub(r"[^а-яёa-z ]", " ", str(s or "").lower().replace("-", " "))
    s = " ".join(s.split())
    return _BOX_ALIAS.get(s.replace(" ", ""), _BOX_ALIAS.get(s, s))


def _box_cluster_match(file_name: str, api_name: str) -> bool:
    """«Санк Петербург» ↔ «Санкт-Петербург и СЗО», «Москва» ↔ «Москва, МО и
    Дальние регионы»: каждый токен из файла находится в API-имени по префиксу."""
    ftoks = _box_norm(file_name).split()
    atoks = _box_norm(api_name).split()
    if not ftoks or not atoks:
        return False
    return all(any(a.startswith(f) or f.startswith(a) for a in atoks)
               for f in ftoks)


@router.post("/supply/boxes/upload")
async def supply_boxes_upload(request: Request, file: UploadFile = File(...)):
    """Файл производства (лист «отгрузка OZON»): строки SKU×короб×кластер.
    Короб целиком едет на один кластер («Новый кластер», иначе «Кластер»)."""
    _owner_only(request)
    import io
    import openpyxl
    import snapshot as _snap
    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        return {"error": f"Не смог открыть XLSX: {str(e)[:200]}"}
    ws = None
    for name in wb.sheetnames:
        hdr = [str(c.value or "").strip().lower() for c in wb[name][1]]
        if any("короб" in h for h in hdr) and any("кластер" in h for h in hdr) \
                and any("ozon" in h or "озон" in h for h in hdr):
            ws = wb[name]
            break
    if ws is None:
        return {"error": "Не нашёл лист с колонками «Кластер» и «Коробов OZON» "
                         "(обычно лист «отгрузка OZON»)"}
    hdr = [str(c.value or "").strip().lower() for c in ws[1]]

    def col(*words, exclude=()):
        for j, h in enumerate(hdr):
            if any(w in h for w in words) and not any(x in h for x in exclude):
                return j
        return -1

    c_art = col("артикул")
    c_bar = col("штрихкод")
    c_cl = col("кластер", exclude=("новый",))
    c_qty = col("шт", exclude=("штрихкод",))
    c_box = col("короб")
    c_exp = col("срок")
    c_wave = col("волна")
    c_new = col("новый")
    if -1 in (c_art, c_bar, c_cl, c_qty, c_box):
        return {"error": f"Не хватает колонок (артикул/штрихкод/кластер/шт/короб), заголовок: {hdr}"}

    boxes: dict[str, dict] = {}
    conflicts: list[str] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        art = str(r[c_art] or "").strip() if c_art < len(r) else ""
        if not art:
            continue
        try:
            box_no = int(float(str(r[c_box]).strip()))
            qty = int(float(str(r[c_qty]).strip()))
        except (TypeError, ValueError):
            continue
        if qty <= 0:
            continue
        src = str(r[c_cl] or "").strip()
        target = str(r[c_new] or "").strip() if (c_new >= 0 and c_new < len(r)
                                                 and r[c_new]) else src
        wave = ""
        if c_wave >= 0 and c_wave < len(r) and r[c_wave]:
            m = re.search(r"\d+", str(r[c_wave]))
            wave = m.group(0) if m else str(r[c_wave]).strip()
        exp = None
        if c_exp >= 0 and c_exp < len(r) and hasattr(r[c_exp], "strftime"):
            exp = r[c_exp].strftime("%Y-%m-%dT00:00:00Z")
        key = f"{src}|{box_no}"
        b = boxes.setdefault(key, {"src": src, "box": box_no, "target": target,
                                   "wave": wave, "items": []})
        if b["target"] != target or b["wave"] != wave:
            conflicts.append(f"{src} короб {box_no}: разные кластер/волна в строках")
        b["items"].append({"offer_id": art, "barcode": str(r[c_bar] or "").strip(),
                           "quantity": qty, "expires_at": exp})
    plan = sorted(boxes.values(), key=lambda b: (b["src"], b["box"]))
    if not plan:
        return {"error": "Не разобрал ни одного короба — проверь файл"}
    await asyncio.to_thread(_snap.save, "supply_boxes_plan", plan)
    waves: dict[str, int] = {}
    for b in plan:
        waves[b["wave"] or "?"] = waves.get(b["wave"] or "?", 0) + 1
    return {"boxes": len(plan), "items": sum(len(b["items"]) for b in plan),
            "qty": sum(i["quantity"] for b in plan for i in b["items"]),
            "waves": dict(sorted(waves.items())),
            "clusters": sorted({b["target"] for b in plan}),
            "conflicts": sorted(set(conflicts))[:10]}


async def _box_match_order(order_id: int, wave: str):
    """Заявка + её поставки-кластера + короба выбранной волны по кластерам."""
    import ozon_client
    import snapshot as _snap
    plan = await asyncio.to_thread(_snap.load, "supply_boxes_plan", None) or []
    if not plan:
        raise HTTPException(400, "Сначала загрузи файл с коробами")
    orders = await ozon_client.get_supply_orders()
    order = next((o for o in orders
                  if str(o.get("order_id")) == str(order_id)), None)
    if not order:
        raise HTTPException(404, "Заявка не найдена")
    names = await ozon_client.get_cluster_names()
    supplies = []
    for s in order.get("supplies") or []:
        cid = int(s.get("macrolocal_cluster_id") or 0)
        supplies.append({"supply_id": s.get("supply_id"),
                         "bundle_id": s.get("bundle_id"),
                         "cluster_id": cid,
                         "cluster": names.get(cid, f"кластер {cid}")})
    wave = str(wave or "").strip()
    # файл может быть без колонки «Волна» — тогда берём все короба
    waves_present = any(str(b.get("wave") or "").strip() for b in plan)
    matched: dict[int, list] = {sp["supply_id"]: [] for sp in supplies}
    unmatched = []
    for b in plan:
        if waves_present and wave and str(b.get("wave")) != wave:
            continue
        sp = next((sp for sp in supplies
                   if _box_cluster_match(b["target"], sp["cluster"])), None)
        if sp:
            matched[sp["supply_id"]].append(b)
        else:
            unmatched.append(b)
    return order, supplies, matched, unmatched


@router.get("/supply/boxes/match")
async def supply_boxes_match(request: Request, order_id: int, wave: str = "1"):
    """Сверка: короба из файла против состава заявки, по каждому кластеру."""
    _owner_only(request)
    import ozon_client
    order, supplies, matched, unmatched = await _box_match_order(order_id, wave)
    out = []
    for sp in supplies:
        boxes = matched.get(sp["supply_id"]) or []
        bundle = await ozon_client.get_bundle_items([sp["bundle_id"]]) \
            if sp["bundle_id"] else []
        need: dict[str, int] = {}
        for it in bundle:
            bc = str(it.get("barcode") or "")
            need[bc] = need.get(bc, 0) + int(it.get("quantity") or 0)
        got: dict[str, int] = {}
        for b in boxes:
            for it in b["items"]:
                got[it["barcode"]] = got.get(it["barcode"], 0) + it["quantity"]
        diffs = []
        for bc in sorted(set(need) | set(got)):
            if need.get(bc, 0) != got.get(bc, 0):
                art = next((i.get("offer_id") for i in bundle
                            if str(i.get("barcode")) == bc), None) \
                    or next((i["offer_id"] for b in boxes for i in b["items"]
                             if i["barcode"] == bc), bc)
                diffs.append({"offer_id": art, "in_order": need.get(bc, 0),
                              "in_file": got.get(bc, 0)})
        out.append({"cluster": sp["cluster"], "supply_id": sp["supply_id"],
                    "boxes": len(boxes), "qty": sum(got.values()),
                    "order_qty": sum(need.values()),
                    "ok": bool(boxes) and not diffs, "diffs": diffs[:15]})
    return {"order": order.get("order_number"), "wave": wave, "clusters": out,
            "unmatched": sorted({f"{b['target']} ({b['src']} короб {b['box']})"
                                 for b in unmatched})[:20]}


@router.post("/supply/boxes/apply")
async def supply_boxes_apply(request: Request, body: dict):
    """Заливка коробов в Ozon по каждой поставке заявки (cargoes/create,
    replace: старые грузоместа удаляются)."""
    _owner_only(request)
    import ozon_client
    order_id = int(body.get("order_id") or 0)
    wave = str(body.get("wave") or "1")
    order, supplies, matched, _ = await _box_match_order(order_id, wave)
    results = []
    for sp in supplies:
        boxes = matched.get(sp["supply_id"]) or []
        if not boxes:
            results.append({"cluster": sp["cluster"],
                            "status": "SKIP", "note": "в файле нет коробов"})
            continue
        bundle = await ozon_client.get_bundle_items([sp["bundle_id"]]) \
            if sp["bundle_id"] else []
        by_bar = {str(i.get("barcode") or ""): i for i in bundle}
        cargoes = []
        for b in sorted(boxes, key=lambda x: (x["src"], x["box"])):
            items = []
            for it in b["items"]:
                bi = by_bar.get(it["barcode"]) or {}
                item = {"barcode": it["barcode"],
                        "offer_id": bi.get("offer_id") or it["offer_id"],
                        "quantity": int(it["quantity"]),
                        "quant": int(bi.get("quant") or 1)}
                if it.get("expires_at"):
                    item["expires_at"] = it["expires_at"]
                items.append(item)
            cargoes.append({"key": f"{b['src']}-{b['box']}",
                            "value": {"type": "BOX", "items": items}})
        if len(cargoes) > 30:
            results.append({"cluster": sp["cluster"], "status": "FAILED",
                            "note": f"{len(cargoes)} коробок — лимит 30"})
            continue
        try:
            resp = await ozon_client.cargoes_create(
                sp["supply_id"], cargoes, delete_current=True)
        except Exception as e:
            results.append({"cluster": sp["cluster"], "status": "FAILED",
                            "note": _http_err(e)})
            continue
        op = resp.get("operation_id") or ""
        status, err = ("UNKNOWN", resp.get("errors"))
        for _i in range(20):
            if not op:
                break
            await asyncio.sleep(3)
            try:
                info = await ozon_client.cargoes_create_info(op)
            except Exception:
                continue
            status = info.get("status") or "UNKNOWN"
            if status in ("SUCCESS", "FAILED"):
                err = info.get("errors")
                break
        note = None
        if err and (err.get("error_reasons") or err.get("items_validation")):
            note = str(err)[:400]
        results.append({"cluster": sp["cluster"], "boxes": len(cargoes),
                        "status": status, "note": note})
    return {"order": order.get("order_number"), "wave": wave, "results": results}


@router.get("/supply/labels")
async def supply_labels(request: Request, order_id: int):
    """ZIP с PDF-этикетками грузомест: файл на каждый кластер заявки."""
    _owner_only(request)
    import io as _io
    import zipfile as _zf
    from urllib.parse import quote
    from fastapi.responses import StreamingResponse
    import ozon_client
    orders = await ozon_client.get_supply_orders()
    order = next((o for o in orders
                  if str(o.get("order_id")) == str(order_id)), None)
    if not order:
        raise HTTPException(404, "Заявка не найдена")
    names = await ozon_client.get_cluster_names()
    buf = _io.BytesIO()
    with _zf.ZipFile(buf, "w", _zf.ZIP_DEFLATED) as zf:
        for s in order.get("supplies") or []:
            cid = int(s.get("macrolocal_cluster_id") or 0)
            cluster = names.get(cid, f"кластер {cid}")
            try:
                resp = await ozon_client.cargoes_label_create(s.get("supply_id"))
                op = resp.get("operation_id")
                if not op:
                    raise RuntimeError(str(resp.get("errors") or resp)[:300])
                url = None
                for _i in range(30):
                    await asyncio.sleep(2)
                    st = await ozon_client.cargoes_label_get(op)
                    stat = st.get("status") or ""
                    if stat == "SUCCESS":
                        res = st.get("result") or {}
                        url = res.get("file_url") or res.get("file_guid")
                        break
                    if stat == "FAILED":
                        raise RuntimeError(str(st.get("errors") or "FAILED")[:300])
                if not url:
                    raise RuntimeError("этикетки не готовы за 60 секунд")
                pdf = await ozon_client.download_label_file(url)
                zf.writestr(f"{cluster}.pdf", pdf)
            except Exception as e:
                zf.writestr(f"{cluster}_ОШИБКА.txt", _http_err(e))
    buf.seek(0)
    num = str(order.get("order_number") or order_id).replace("/", "-")
    fname = quote(f"Этикетки_{num}.zip")
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition":
                 f"attachment; filename=labels_{num}.zip; filename*=UTF-8''{fname}"})


def _watch_list_load():
    import snapshot as _snap
    v = _snap.load("slot_watch", None)
    if not v:
        return []
    return v if isinstance(v, list) else [v]


@router.post("/supply/watch")
async def supply_watch(request: Request, body: dict):
    """Охоты (несколько параллельно): {draft_id}|{order_id} добавляет,
    {off: id} снимает одну, {off: true} — все."""
    _owner_only(request)
    import snapshot as _snap
    watches = await asyncio.to_thread(_watch_list_load)
    off = body.get("off")
    if off:
        if off is True:
            watches = []
        else:
            watches = [w for w in watches
                       if str(w.get("draft_id") or w.get("order_id")) != str(off)]
        await asyncio.to_thread(_snap.save, "slot_watch", watches)
        return {"watching": len(watches)}
    order_id = int(body.get("order_id") or 0)
    draft_id = int(body.get("draft_id") or 0)
    if order_id:
        watches = [w for w in watches if w.get("order_id") != order_id]
        watches.append({"mode": "order", "order_id": order_id,
                        "started": datetime.utcnow().isoformat(), "seen": []})
        await asyncio.to_thread(_snap.save, "slot_watch", watches[:6])
        return {"watching": len(watches), "added": order_id}
    if not draft_id:
        return {"error": "нужен draft_id (номер черновика) или order_id (заявка)"}
    clusters = body.get("clusters") or []
    if not clusters:
        import ozon_client
        rows = await ozon_client.get_stocks_by_cluster()
        clusters = sorted({int(r["macrolocal_cluster_id"]) for r in rows
                           if r.get("macrolocal_cluster_id")})
    watches = [w for w in watches if w.get("draft_id") != draft_id]
    watches.append({"mode": "draft", "draft_id": draft_id,
                    "clusters": clusters[:20],
                    "started": datetime.utcnow().isoformat(), "seen": []})
    await asyncio.to_thread(_snap.save, "slot_watch", watches[:6])
    return {"watching": len(watches), "added": draft_id}


@router.get("/supply/watch")
async def supply_watch_status(request: Request):
    _owner_only(request)
    return {"watches": await asyncio.to_thread(_watch_list_load)}


async def slot_watch_tick() -> None:
    """Один проход охоты по всем целям: новые слоты — в Telegram."""
    import snapshot as _snap
    import ozon_client
    import agent_review as _ar
    watches = await asyncio.to_thread(_watch_list_load)
    if not watches:
        return
    keep, dirty = [], False
    for cfg in watches:
        if (datetime.utcnow() - datetime.fromisoformat(cfg["started"])).days >= 7:
            await _ar.tg_send(f"Охота за слотами ({cfg.get('draft_id') or cfg.get('order_id')}): "
                              "неделя прошла, выключаю.")
            dirty = True
            continue
        keep.append(cfg)
        try:
            if cfg.get("mode") == "order":
                resp = await ozon_client.get_order_timeslots(cfg["order_id"])
            else:
                resp = await ozon_client.get_draft_timeslots(cfg["draft_id"], cfg["clusters"])
        except Exception as e:
            _log.warning("slot watch %s: %s",
                         cfg.get("draft_id") or cfg.get("order_id"), str(e)[:200])
            continue
        slots = ozon_client.extract_timeslots(resp)
        seen = set(cfg.get("seen") or [])
        fresh = [s for s in slots if f"{s['from']}|{s['to']}" not in seen]
        if fresh:
            tgt = cfg.get("draft_id") or cfg.get("order_id")
            kind = "черновик" if cfg.get("mode") != "order" else "заявка"
            lines = [f"СЛОТЫ ПОЯВИЛИСЬ — {kind} {tgt}, бронируй в ЛК:"]
            for s in fresh[:15]:
                lines.append(f"{s['from'][:16].replace('T', ' ')} — {s['to'][11:16]}")
            await _ar.tg_send("\n".join(lines))
            seen |= {f"{s['from']}|{s['to']}" for s in fresh}
            cfg["seen"] = sorted(seen)[-300:]
            dirty = True
        await asyncio.sleep(1)
    if dirty or len(keep) != len(watches):
        await asyncio.to_thread(_snap.save, "slot_watch", keep)
