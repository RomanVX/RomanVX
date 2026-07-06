"""Finance router — WB sales reports + P&L by month."""
import asyncio
import logging
import time as _time
from datetime import datetime, timedelta

from fastapi import APIRouter, Query, HTTPException

import wb_finance_client
import cost_store

router = APIRouter(prefix="/api/finance", tags=["finance"])
_log = logging.getLogger(__name__)

# Ссылки на фоновые задачи: asyncio держит task'и слабыми ссылками,
# и create_task без сохранения ссылки может быть убит GC ДО выполнения —
# из-за этого детальный отчёт WB не загружался вовсе.
_bg_tasks: set = set()


def _spawn(coro, light: bool = False) -> None:
    """light=True — долгие, но лёгкие по памяти сборки (fullstats 1 req/мин):
    их не пускаем через heavy.guard, иначе они на десятки минут заблокируют
    тяжёлые финансовые сборки."""
    import heavy
    t = asyncio.create_task(coro if light else heavy.guard(coro))
    _bg_tasks.add(t)
    t.add_done_callback(_bg_tasks.discard)

# ── кэш: 60 мин (rate limit WB Finance API — 1 req/min) ──────────────────────
_wb_cache: dict = {}
_wb_cache_ts: float = 0.0
_WB_TTL = 3600
_wb_lock = asyncio.Lock()

# ── P&L кэш ──────────────────────────────────────────────────────────────────
_pnl_cache: dict = {}
_pnl_cache_ts: float = 0.0


def _f(v) -> float:
    try:
        return float(str(v).replace(",", ".") or 0)
    except (ValueError, TypeError):
        return 0.0


def _wb_reports_save_db(rows: list) -> None:
    try:
        import db
        import json
        db.execute("CREATE TABLE IF NOT EXISTS kv_cache (k TEXT PRIMARY KEY, v TEXT)")
        db.execute("DELETE FROM kv_cache WHERE k = 'wb_reports'")
        db.execute("INSERT INTO kv_cache (k, v) VALUES ('wb_reports', ?)",
                   (json.dumps(rows, ensure_ascii=False),))
    except Exception as exc:
        _log.warning("wb_reports save to db failed: %s", exc)


def _wb_reports_load_db() -> list:
    try:
        import db
        import json
        rows = db.fetchall("SELECT v FROM kv_cache WHERE k = 'wb_reports'")
        return json.loads(rows[0][0]) if rows else []
    except Exception:
        return []


async def _get_wb_reports_cached(months: int = 6, refresh: bool = False) -> list[dict]:
    """Внутренняя функция — возвращает нормализованные отчёты WB (с кешем)."""
    global _wb_cache, _wb_cache_ts

    if not refresh and _wb_cache and _time.monotonic() - _wb_cache_ts < _WB_TTL:
        return _wb_cache.get("reports", [])

    # После рестарта/сна память пуста, а WB Finance API (1 req/мин) мог только
    # что отработать в прогреве → запрос падал 502. Отдаём последний снапшот
    # из БД сразу, свежие данные подтянутся фоном.
    if not refresh and not _wb_cache:
        db_rows = await asyncio.to_thread(_wb_reports_load_db)
        if db_rows:
            _wb_cache = {"reports": db_rows}
            _wb_cache_ts = 0.0   # заведомо протухший — фон обновит
            _spawn(_get_wb_reports_cached(months, True), light=True)
            return db_rows

    async with _wb_lock:
        if not refresh and _wb_cache and _time.monotonic() - _wb_cache_ts < _WB_TTL:
            return _wb_cache.get("reports", [])

        dt_to   = datetime.utcnow() + timedelta(hours=3)
        dt_from = dt_to - timedelta(days=30 * months)
        min_date = datetime(2025, 1, 1)
        if dt_from < min_date:
            dt_from = min_date

        date_from = dt_from.strftime("%Y-%m-%d")
        date_to   = dt_to.strftime("%Y-%m-%d")

        try:
            raw = await wb_finance_client.get_sales_reports(date_from, date_to)
        except Exception as exc:
            _log.error("WB Finance API error: %s", exc)
            # есть устаревший кеш — отдаём его вместо ошибки (429 пройдёт сам)
            if _wb_cache.get("reports"):
                _log.warning("WB Finance: отдаём устаревший кеш (%d отчётов)",
                             len(_wb_cache["reports"]))
                return _wb_cache["reports"]
            db_rows = await asyncio.to_thread(_wb_reports_load_db)
            if db_rows:
                _log.warning("WB Finance: отдаём снапшот из БД (%d отчётов)", len(db_rows))
                _wb_cache = {"reports": db_rows}
                _wb_cache_ts = 0.0
                return db_rows
            raise HTTPException(status_code=502, detail=f"WB Finance API: {exc}")

        rows = []
        for r in sorted(raw, key=lambda x: x.get("dateFrom", ""), reverse=True):
            rows.append({
                "reportId":          r.get("reportId"),
                "dateFrom":          r.get("dateFrom", "")[:10],
                "dateTo":            r.get("dateTo", "")[:10],
                "createDate":        r.get("createDate", "")[:10],
                "currency":          r.get("currency", "RUB"),
                "reportType":        r.get("reportType", 1),
                "retailAmount":      _f(r.get("retailAmountSum")),
                "forPay":            _f(r.get("forPaySum")),
                "avgSalePercent":    r.get("avgSalePercent", 0),
                "deliveryService":   _f(r.get("deliveryServiceSum")),
                "paidStorage":       _f(r.get("paidStorageSum")),
                "paidAcceptance":    _f(r.get("paidAcceptanceSum")),
                "deduction":         _f(r.get("deductionSum")),
                "penalty":           _f(r.get("penaltySum")),
                "additionalPayment": _f(r.get("additionalPaymentSum")),
                "cashbackAmount":    _f(r.get("cashbackAmountSum")),
                "cashbackDiscount":  _f(r.get("cashbackDiscountSum")),
                "cashbackCommission":_f(r.get("cashbackCommissionChangeSum")),
                "bankPayment":       _f(r.get("bankPaymentSum")),
            })

        result = {"reports": rows, "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}
        _wb_cache = result
        _wb_cache_ts = _time.monotonic()
        if rows:
            await asyncio.to_thread(_wb_reports_save_db, rows)
        return rows


@router.get("/wb/reports")
async def get_wb_reports(
    months: int = Query(default=6, ge=1, le=24),
    refresh: bool = Query(default=False),
):
    reports = await _get_wb_reports_cached(months, refresh)
    return {"reports": reports, "fetched_at": _wb_cache.get("fetched_at", "")}


@router.post("/wb/reports/invalidate", include_in_schema=False)
async def invalidate_wb_finance():
    global _wb_cache, _wb_cache_ts, _pnl_cache, _pnl_cache_ts
    _wb_cache = {}
    _wb_cache_ts = 0.0
    _pnl_cache = {}
    _pnl_cache_ts = 0.0
    return {"status": "ok"}


_detail_cache: dict = {}
_detail_cache_ts: float = 0.0
_DETAIL_TTL = 86400  # 24ч — детальный отчёт меняется редко
_detail_lock = asyncio.Lock()


_detail_fetching: bool = False  # идёт ли фоновая загрузка
_detail_last_error: str = ""


def _normalize_stat_rows(stat_rows: list[dict]) -> list[dict]:
    """statistics-api reportDetailByPeriod (v5) → формат finance-api detailed.

    Кабинетная «Финансовая аналитика» строится из этого же отчёта,
    поэтому цифры сходятся с ЛК.
    """
    out = []
    for r in stat_rows:
        qty = r.get("quantity") or 0
        # «до СПП»: retail_amount в v5 — фактическая сумма ПОСЛЕ СПП;
        # цена продавца до СПП — retail_price_withdisc_rub (за единицу)
        pre_spp = (r.get("retail_price_withdisc_rub") or 0) * (qty or 1)
        out.append({
            "rrDate":          (r.get("rr_dt") or r.get("sale_dt") or "")[:10],
            "saleDate":        (r.get("sale_dt") or r.get("rr_dt") or "")[:10],
            "docTypeName":     r.get("doc_type_name") or "",
            "operName":        r.get("supplier_oper_name") or "",
            "retailAmount":    r.get("retail_amount") or 0,
            "retailPreSpp":    pre_spp,
            "forPay":          r.get("ppvz_for_pay") or 0,
            "deliveryService": r.get("delivery_rub") or 0,
            "paidStorage":     r.get("storage_fee") or 0,
            "paidAcceptance":  r.get("acceptance") or 0,
            "penalty":         r.get("penalty") or 0,
            "deduction":       r.get("deduction") or 0,
            "cashbackAmount":  0,
            "acquiringFee":    r.get("acquiring_fee") or 0,
            "vendorCode":      (r.get("sa_name") or "").strip(),
            "nmId":            r.get("nm_id"),
            "quantity":        qty,
        })
    return out


async def _fetch_detail_bg(date_from: str, date_to: str) -> None:
    """Фоновая задача: детальные строки для P&L.

    Основной источник — statistics-api reportDetailByPeriod (отдельный
    rate-limit, весь период за 1-2 запроса). Фолбэк — finance-api
    detailed (общий лимит 1 req/мин, загрузка занимает минуты).
    """
    global _detail_cache, _detail_cache_ts, _detail_fetching, _detail_last_error
    global _pnl_cache, _pnl_cache_ts
    if _detail_fetching:
        return
    _detail_fetching = True
    try:
        cache_key = f"{date_from}_{date_to}"
        async with _detail_lock:
            rows: list[dict] = []
            try:
                import wb_client
                stat_rows = await wb_client.get_report_detail(
                    datetime.strptime(date_from, "%Y-%m-%d"),
                    datetime.strptime(date_to, "%Y-%m-%d"),
                )
                rows = _normalize_stat_rows(stat_rows)
                del stat_rows  # сырой ответ statistics-api больше не нужен
                if rows:
                    _log.info("Detail via statistics-api: %d rows", len(rows))
            except Exception as e:
                _log.warning("statistics-api detail failed (%s) — пробуем finance-api", e)

            if not rows:
                rows = await wb_finance_client.get_detailed_report(date_from, date_to)
                _log.info("Detail via finance-api: %d rows", len(rows))

            _detail_cache = {"key": cache_key, "rows": rows}
            _detail_cache_ts = _time.monotonic()
            _detail_last_error = ""
        # детали готовы → сбрасываем P&L-кэш, чтобы следующий запрос пересобрал
        # отчёт по точным датам (иначе weekly-версия жила бы до конца TTL)
        _pnl_cache = {}
        _pnl_cache_ts = 0.0
    except Exception as e:
        _detail_last_error = str(e)[:300]
        _log.error("Detail report fetch failed: %s", e)
    finally:
        _detail_fetching = False


def _get_detail_if_ready(date_from: str, date_to: str) -> list[dict]:
    """Возвращает детальный отчёт из кеша (без ожидания)."""
    cache_key = f"{date_from}_{date_to}"
    if _detail_cache.get("key") == cache_key \
            and _time.monotonic() - _detail_cache_ts < _DETAIL_TTL:
        return _detail_cache.get("rows", [])
    return []


@router.get("/wb/debug", include_in_schema=False)
async def wb_finance_debug():
    """Диагностика WB P&L: состояние кешей + живая проба reportDetailByPeriod."""
    out = {
        "detail_cache_key": _detail_cache.get("key"),
        "detail_rows": len(_detail_cache.get("rows", [])),
        "detail_fetching": _detail_fetching,
        "detail_error": _detail_last_error,
        "pnl_cache_source": _pnl_cache.get("source") if _pnl_cache else None,
        "pnl_cache_age_sec": round(_time.monotonic() - _pnl_cache_ts) if _pnl_cache else None,
        "wb_summary_cached": len(_wb_cache.get("reports", [])) if _wb_cache else 0,
    }
    rows = _detail_cache.get("rows", [])
    if rows:
        # разбивка июня по типам операций: что входит в выручку, а что пропускается
        june = [r for r in rows if (r.get("saleDate") or r.get("rrDate") or "").startswith("2026-06")]
        out["june_rows"] = len(june)
        by_oper: dict = {}
        for r in june:
            key = f"{r.get('docTypeName') or '—'} / {r.get('operName') or '—'}"
            b = by_oper.setdefault(key, {"rows": 0, "qty": 0, "pre_spp": 0.0, "post_spp": 0.0, "forPay": 0.0})
            b["rows"] += 1
            b["qty"] += int(r.get("quantity") or 0)
            b["pre_spp"] += abs(float(r.get("retailPreSpp") or 0))
            b["post_spp"] += abs(float(r.get("retailAmount") or 0))
            b["forPay"] += float(r.get("forPay") or 0)
        out["june_by_oper"] = {k: {kk: round(vv) if isinstance(vv, float) else vv
                                   for kk, vv in v.items()}
                               for k, v in sorted(by_oper.items())}
    # живая проба v5 (7 дней) — какие поля реально приходят
    try:
        import wb_client
        stat = await wb_client.get_report_detail(
            datetime.utcnow() - timedelta(days=7), datetime.utcnow())
        out["v5_live_rows"] = len(stat)
        if stat:
            r0 = stat[0]
            out["v5_live_keys"] = sorted(r0.keys())
            out["v5_live_sample"] = {k: r0.get(k) for k in
                ["rr_dt", "doc_type_name", "retail_amount", "retail_price_withdisc_rub",
                 "ppvz_for_pay", "delivery_rub", "storage_fee", "acceptance", "penalty",
                 "deduction", "sa_name", "nm_id", "quantity", "supplier_oper_name"]}
    except Exception as e:
        out["v5_live_error"] = str(e)[:300]
    return out


@router.get("/wb/pnl")
async def get_wb_pnl(
    months: int = Query(default=6, ge=1, le=24),
    refresh: bool = Query(default=False),
):
    """P&L по месяцам. Сводные данные — сразу, COGS — после фоновой загрузки детального отчёта."""
    global _pnl_cache, _pnl_cache_ts, _detail_cache_ts

    if not refresh and _pnl_cache and _time.monotonic() - _pnl_cache_ts < _WB_TTL:
        # ГОНКА: P&L мог закешироваться как weekly, пока детали ещё качались.
        # Если детальный кеш уже готов — не отдаём устаревший weekly, пересобираем.
        if _pnl_cache.get("source") == "detail" or not _detail_cache.get("rows"):
            return _pnl_cache
        _log.info("P&L cache is weekly but detail is ready — rebuilding")

    # Диапазон дат
    dt_to   = datetime.utcnow() + timedelta(hours=3)
    dt_from = dt_to - timedelta(days=30 * months)
    min_date = datetime(2025, 1, 1)
    if dt_from < min_date:
        dt_from = min_date
    date_from = dt_from.strftime("%Y-%m-%d")
    date_to   = dt_to.strftime("%Y-%m-%d")

    # Сводные отчёты — быстро (кеш 1ч)
    summary_reports = await _get_wb_reports_cached(months, refresh)

    # Детальный отчёт — берём из кеша если есть, иначе запускаем фоновую загрузку.
    # ВАЖНО: refresh НЕ сбрасывает детальный кеш — его загрузка идёт с лимитом
    # 1 запрос/мин и занимает минуты; сброс возвращал бы weekly-цифры и
    # перезапускал загрузку по кругу. Детали протухают сами по TTL (24ч).
    # dateTo сдвигаем на неделю вперёд: отчёт реализации отдаёт только
    # СФОРМИРОВАННЫЕ недели, и хвост месяца появляется в отчёте следующей
    # недели — так он подтянется сразу после формирования.
    detail_to = (dt_to + timedelta(days=7)).strftime("%Y-%m-%d")
    cache_key = f"{date_from}_{detail_to}"
    detail_ready = (_detail_cache.get("key") == cache_key
                    and _time.monotonic() - _detail_cache_ts < _DETAIL_TTL)
    if not detail_ready:
        _spawn(_fetch_detail_bg(date_from, detail_to))
    detail_rows = _get_detail_if_ready(date_from, detail_to)

    # ── Агрегируем сводные отчёты по месяцам (пропорционально дням) ────────────
    SUM_KEYS = ["retailAmount","forPay","deliveryService","paidStorage",
                "paidAcceptance","penalty","deduction","cashbackAmount",
                "additionalPayment","cashbackCommission","bankPayment"]

    month_totals: dict[str, dict] = {}

    def ensure(mk):
        if mk not in month_totals:
            month_totals[mk] = {k: 0.0 for k in SUM_KEYS}
        return month_totals[mk]

    for r in summary_reports:
        try:
            df = datetime.strptime(r["dateFrom"], "%Y-%m-%d").date()
            dt = datetime.strptime(r["dateTo"],   "%Y-%m-%d").date()
        except (ValueError, KeyError):
            continue
        segs: list[tuple[str, int]] = []
        cur = df
        while cur <= dt:
            mk = f"{cur.year}-{cur.month:02d}"
            if segs and segs[-1][0] == mk:
                segs[-1] = (mk, segs[-1][1] + 1)
            else:
                segs.append((mk, 1))
            cur += timedelta(days=1)
        total_days = sum(d for _, d in segs)
        for mk, days in segs:
            ratio = days / total_days
            m = ensure(mk)
            for k in SUM_KEYS:
                m[k] += r.get(k, 0.0) * ratio

    # ── Детальный отчёт: точные месяцы по датам операций (как в кабинете WB) ──
    # Если детальный отчёт загружен — пересобираем ВСЕ строки из него: недельный
    # отчёт размазывается по месяцам пропорционально дням и даёт расхождения
    # с финансовой аналитикой кабинета.
    costs = cost_store.get_costs()
    cogs_by_month: dict[str, float] = {}
    source = "weekly"
    last_detail_sale = ""
    tail_days = 0
    # SKU-разрез для юнит-экономики: собирается тем же проходом, что и P&L,
    # поэтому суммы по артикулам сходятся со строками вкладки Финансы.
    sku_data: dict[str, dict] = {}   # sku → mk → {revenue, forPay, qty, cogs, delivery, storage, acceptance, penalty}

    _SKU_KEYS = ("revenue", "forPay", "qty", "cogs", "delivery", "storage",
                 "acceptance", "penalty", "acquiring")

    def s_ensure(sku, mk):
        m = sku_data.setdefault(sku, {})
        if mk not in m:
            m[mk] = {k: 0.0 for k in _SKU_KEYS}
        return m[mk]

    def _f(v) -> float:
        try: return float(str(v).replace(",", ".") or 0)
        except: return 0.0

    if detail_rows:
        detail_totals: dict[str, dict] = {}

        def d_ensure(mk):
            if mk not in detail_totals:
                detail_totals[mk] = {k: 0.0 for k in SUM_KEYS}
            return detail_totals[mk]

        import catalog as _cat

        def _canon_sku(row) -> str:
            """Каноничный артикул: через nmId по каталогу, иначе sa_name."""
            nm = row.get("nmId")
            if nm:
                r = _cat.resolve_wb(nm)
                if r and not str(r).isdigit():
                    return r
            return (row.get("vendorCode") or "").strip().upper() or "—"

        # v5 отдаёт sa_name в нижнем регистре («al-01») — матчим без учёта регистра
        costs_upper = {k.upper(): v for k, v in costs.items()}

        def _unit_cost(row) -> float:
            """Закупочная цена: по артикулу (без учёта регистра), иначе по nmId."""
            sku = (row.get("vendorCode") or "").strip().upper()
            uc = costs_upper.get(sku, 0.0)
            if uc <= 0 and row.get("nmId"):
                uc = costs.get(_cat.resolve_wb(row.get("nmId")), 0.0)
            return uc

        min_mk = date_from[:7]  # хвосты недель прошлого периода отсекаем
        last_detail_sale = ""   # последняя дата продажи в отчёте реализации

        for row in detail_rows:
            rr = (row.get("rrDate") or row.get("saleDt") or "")[:10]
            if not rr:
                continue
            mk = rr[:7]                       # месяц операции (затраты)
            doc_type = row.get("docTypeName", "")
            oper = (row.get("operName") or "").lower()
            # Продажи бывают с разными операциями (в т.ч. ВБ Клуб) и иногда с
            # пустым doc_type — определяем по сумме продажи, а не только по типу
            is_return = doc_type == "Возврат" or "возврат" in oper
            if "сторно возврат" in oper:
                is_return = False   # сторно возврата = отмена возврата → плюс
            elif "сторно продаж" in oper:
                is_return = True    # сторно продажи = минус
            has_sale_amount = (_f(row.get("retailPreSpp")) != 0
                               or (_f(row.get("retailAmount")) != 0
                                   and int(row.get("quantity") or 0) > 0))
            sign = -1 if is_return else 1
            if has_sale_amount:
                # продажи кабинет группирует по ДАТЕ ПРОДАЖИ (sale_dt), а не по
                # дате операции — иначе недели на стыке месяцев уезжают не туда
                sale_date = (row.get("saleDate") or rr)[:10]
                if not is_return and sale_date > last_detail_sale:
                    last_detail_sale = sale_date
                mk_sale = sale_date[:7]
                if mk_sale >= min_mk:
                    ms = d_ensure(mk_sale)
                    base = abs(_f(row.get("retailPreSpp"))) or abs(_f(row.get("retailAmount")))
                    ms["retailAmount"] += sign * base
                    ms["forPay"]       += sign * abs(_f(row.get("forPay")))
                    # COGS: количество проданного × закупочная (возврат — минус)
                    qty = int(row.get("quantity") or 0)
                    uc = _unit_cost(row)
                    if uc > 0 and qty > 0:
                        cogs_by_month[mk_sale] = cogs_by_month.get(mk_sale, 0.0) + sign * uc * qty
                    # SKU-разрез (юнитка)
                    sd = s_ensure(_canon_sku(row), mk_sale)
                    sd["revenue"]   += sign * base
                    sd["forPay"]    += sign * abs(_f(row.get("forPay")))
                    sd["qty"]       += sign * qty
                    sd["acquiring"] += sign * abs(_f(row.get("acquiringFee")))
                    if uc > 0 and qty > 0:
                        sd["cogs"] += sign * uc * qty
            # операционные затраты — по дате операции, на любых строках
            if mk < min_mk:
                continue
            m = d_ensure(mk)
            m["deliveryService"] += _f(row.get("deliveryService"))
            m["paidStorage"]     += _f(row.get("paidStorage"))
            m["paidAcceptance"]  += _f(row.get("paidAcceptance"))
            m["penalty"]         += _f(row.get("penalty"))
            m["deduction"]       += _f(row.get("deduction"))
            m["cashbackAmount"]  += _f(row.get("cashbackAmount"))
            # SKU-разрез операционных затрат (строки несут артикул)
            if any(_f(row.get(k)) for k in ("deliveryService", "paidStorage", "paidAcceptance", "penalty")):
                sd = s_ensure(_canon_sku(row), mk)
                sd["delivery"]   += _f(row.get("deliveryService"))
                sd["storage"]    += _f(row.get("paidStorage"))
                sd["acceptance"] += _f(row.get("paidAcceptance"))
                sd["penalty"]    += _f(row.get("penalty"))

        # ── Хвост месяца: дни после последней сформированной недели отчёта
        # реализации дополняем ОПЕРАТИВНЫМИ продажами statistics-api (кабинетная
        # финаналитика делает так же — поэтому у неё месяц полный сразу).
        tail_days = 0
        if detail_totals and last_detail_sale:
            try:
                import cache as _wb_cache_mod
                import analytics as _an
                op_sales, _, _ = await _wb_cache_mod.get_raw_data(dt_from, dt_to)
                tail_dates = set()
                for s in op_sales:
                    d = (s.get("date") or "")[:10]
                    if not d or d <= last_detail_sale:
                        continue
                    mk = d[:7]
                    if mk < min_mk:
                        continue
                    sale_id = str(s.get("saleID") or "")
                    neg = -1 if sale_id.startswith("R") else 1   # R* = возврат
                    pre = _f(s.get("priceWithDisc")) or _f(s.get("finishedPrice"))
                    fp = _f(s.get("forPay")) or pre * 0.7
                    m = d_ensure(mk)
                    m["retailAmount"] += neg * pre
                    m["forPay"]       += neg * fp
                    raw = s.get("supplierArticle") or s.get("nmId") or ""
                    uc = _unit_cost({"vendorCode": str(raw), "nmId": s.get("nmId")})
                    if uc > 0:
                        cogs_by_month[mk] = cogs_by_month.get(mk, 0.0) + neg * uc
                    sd = s_ensure(_canon_sku({"vendorCode": str(raw), "nmId": s.get("nmId")}), mk)
                    sd["revenue"] += neg * pre
                    sd["forPay"]  += neg * fp
                    sd["qty"]     += neg
                    if uc > 0:
                        sd["cogs"] += neg * uc
                    tail_dates.add(d)
                tail_days = len(tail_dates)
                if tail_days:
                    _log.info("WB P&L: хвост из оперативных продаж — %d дней после %s",
                              tail_days, last_detail_sale)
            except Exception as e:
                _log.warning("WB tail sales failed: %s", e)

        if detail_totals:
            # Удержания (Джем, списания за отзывы, «другие») в детальном отчёте
            # ОТСУТСТВУЮТ — WB отдаёт их только в недельных сводных отчётах.
            # Переносим их из weekly-агрегации (месяцы разбиты пропорционально).
            for mk, dm in detail_totals.items():
                wm = month_totals.get(mk) or {}
                dm["deduction"]      = wm.get("deduction", 0.0)
                dm["cashbackAmount"] = wm.get("cashbackAmount", 0.0)
                dm["cashbackCommission"] = wm.get("cashbackCommission", 0.0)
            month_totals = detail_totals
            source = "detail"

    # ── Реклама (продвижение): списания из advert API по месяцам ──────────────
    advert_by_month = await _get_wb_advert_cached(dt_from, dt_to)
    # точная раскладка по nmId (для юнитки): сперва из БД, обновление — в фоне
    await _adv_nm_ensure_loaded()
    if not _adv_nm_building and (not _adv_nm_cache
                                 or _time.monotonic() - _adv_nm_ts >= _ADV_NM_TTL):
        _spawn(_build_adv_nm_bg(months), light=True)

    # ── Единая структура P&L ──────────────────────────────────────────────────
    # Комиссия WB (вкл. эквайринг) = выручка − «к перечислению за товар» (forPay).
    # Реклама с оплатой «с баланса продаж» уже сидит в deduction («удержания»),
    # поэтому выделяем её отдельной строкой и вычитаем из прочих удержаний —
    # без двойного счёта. Если реклама платилась со счёта (не из баланса),
    # прочие удержания просто останутся 0, а расход всё равно виден в P&L.
    advert_bonus_by_month: dict[str, float] = {}
    advert_balance_by_month: dict[str, float] = {}
    for mk, m in month_totals.items():
        m["commission"] = m.get("retailAmount", 0.0) - m.get("forPay", 0.0)
        adv_info = advert_by_month.get(mk) or {}
        adv_total   = adv_info.get("total", 0.0)
        adv_bonus   = adv_info.get("bonus", 0.0)
        adv_balance = adv_info.get("balance", 0.0)
        # промо-бонусы WB компенсируют часть рекламы — в затраты идёт чистое
        # списание, бонусная часть показывается справочной строкой
        m["advert"] = adv_total - adv_bonus
        advert_bonus_by_month[mk] = adv_bonus
        advert_balance_by_month[mk] = adv_balance
        # В «Удержаниях» WB сидит реклама, оплаченная С БАЛАНСА продаж —
        # вычитаем ровно её (не всю), чтобы не задвоить. Остаток — Джем,
        # списания за отзывы и прочие удержания.
        deductions_full = (m.get("deduction", 0.0) + m.get("cashbackAmount", 0.0)
                           + m.get("cashbackCommission", 0.0))
        m["deductions"] = max(deductions_full - adv_balance, 0.0)

    pnl_rows = _build_pnl_rows(
        month_totals, cogs_by_month,
        [
            ("retailAmount",   "📦 Выручка (до СПП)",          "header"),
            ("commission",     "  − Комиссия WB и эквайринг",   "cost"),
            ("deliveryService","  − Логистика",                 "cost"),
            ("paidStorage",    "  − Хранение",                  "cost"),
            ("paidAcceptance", "  − Приёмка",                   "cost"),
            ("penalty",        "  − Штрафы",                    "cost"),
            ("deductions",     "  − Прочие удержания (Джем, отзывы и др.)", "cost"),
            ("advert",         "  − Продвижение (за вычетом бонусов)", "cost"),
        ],
    )

    # Справочные строки под «Продвижением» (в затратах не задваиваются):
    #  - сколько списано с баланса продаж (= строка «Продвижение» в удержаниях кабинета)
    #  - сколько компенсировано промо-бонусами WB (уже вычтено из затрат)
    sorted_mks = sorted(month_totals.keys(), reverse=True)
    adv_idx = next((i for i, r in enumerate(pnl_rows) if r["key"] == "advert"), None)
    if adv_idx is not None:
        extra_rows = []
        if any(v > 0 for v in advert_balance_by_month.values()):
            extra_rows.append({"key": "advert_balance",
                               "label": "      ↳ из них с баланса продаж (в кабинете — «Продвижение» в удержаниях)",
                               "style": "note", "formula": "info",
                               "values": {mk: round(advert_balance_by_month.get(mk, 0.0)) for mk in sorted_mks}})
        if any(v > 0 for v in advert_bonus_by_month.values()):
            extra_rows.append({"key": "advert_bonus",
                               "label": "      ↳ компенсировано промо-бонусами WB",
                               "style": "note", "formula": "info",
                               "values": {mk: round(advert_bonus_by_month.get(mk, 0.0)) for mk in sorted_mks}})
        for j, er in enumerate(extra_rows):
            pnl_rows.insert(adv_idx + 1 + j, er)

    cogs_has_data = len(costs) > 0 and any(v > 0 for v in cogs_by_month.values())
    result = {
        "months": _months_out(month_totals.keys()),
        "rows":   pnl_rows,
        "cogs_loaded": len(costs) > 0,
        "cogs_has_data": cogs_has_data,
        "source": source,
        "detail_upto": last_detail_sale,
        "tail_days": tail_days,
        "detail_fetching": _detail_fetching,
        "detail_error": _detail_last_error,
        "fetched_at": _wb_cache.get("fetched_at", ""),
    }

    # SKU-разрез для юнит-экономики — из этого же прохода (суммы сходятся с P&L)
    global _wb_unit_data
    if source == "detail":
        _wb_unit_data = {
            "sku": sku_data,
            "totals": month_totals,
            "advert": {mk: (advert_by_month.get(mk) or {}).get("total", 0.0)
                            - (advert_by_month.get(mk) or {}).get("bonus", 0.0)
                       for mk in month_totals},
            "deductions": {mk: month_totals[mk].get("deductions", 0.0) for mk in month_totals},
            "detail_upto": last_detail_sale,
            "fetched_at": result["fetched_at"],
        }
    _pnl_cache = result
    _pnl_cache_ts = _time.monotonic()
    if result.get("source") == "detail":
        await asyncio.to_thread(_snapshot_pnl, "WB", result)
    return result


# кеш рекламных списаний WB (advert API, 6ч)
_wb_advert_cache: dict = {}
_wb_advert_ts: float = 0.0

# SKU-разрез для юнит-экономики (наполняется при сборке P&L из деталей)
_wb_unit_data: dict = {}

# Точная реклама по nmId из fullstats: {mk: {nmId: sum}} (кеш 12ч, сборка в фоне)
_adv_nm_cache: dict = {}
_adv_nm_ts: float = 0.0
_adv_nm_building: bool = False
_adv_nm_error: str = ""
_adv_nm_ids: int = -1
_adv_nm_db_checked: bool = False
_ADV_NM_TTL = 12 * 3600


def _adv_nm_save_month_db(mk: str, per_nm: dict) -> None:
    """Инкрементальное сохранение месяца — деплой/рестарт не теряет прогресс."""
    import db
    db.execute("CREATE TABLE IF NOT EXISTS adv_nm_spend "
               "(mk TEXT, nm_id BIGINT, spend REAL, PRIMARY KEY (mk, nm_id))")
    db.execute("DELETE FROM adv_nm_spend WHERE mk = ?", (mk,))
    rows = [(mk, int(nm), float(sp)) for nm, sp in per_nm.items()]
    if rows:
        db.executemany("INSERT INTO adv_nm_spend (mk, nm_id, spend) VALUES (?,?,?)", rows)


def _adv_nm_mark_built() -> None:
    import db
    built_at = datetime.utcnow().isoformat()
    db.execute("CREATE TABLE IF NOT EXISTS cogs_meta (key TEXT PRIMARY KEY, value TEXT)")
    if db.IS_PG:
        db.execute("INSERT INTO cogs_meta (key, value) VALUES ('adv_nm_built_at', ?) "
                   "ON CONFLICT (key) DO UPDATE SET value = excluded.value", (built_at,))
    else:
        db.execute("INSERT OR REPLACE INTO cogs_meta VALUES ('adv_nm_built_at', ?)", (built_at,))


def _adv_nm_load_db() -> None:
    """Поднимает раскладку из БД. Если она старше TTL — данные показываем,
    но помечаем устаревшими (сборка обновит в фоне)."""
    global _adv_nm_cache, _adv_nm_ts
    import db
    try:
        rows = db.fetchall("SELECT mk, nm_id, spend FROM adv_nm_spend")
    except Exception:
        return  # таблицы ещё нет
    cache: dict = {}
    for mk, nm, sp in rows:
        cache.setdefault(mk, {})[int(nm)] = float(sp)
    if not cache:
        return
    fresh = False
    try:
        r = db.fetchone("SELECT value FROM cogs_meta WHERE key = 'adv_nm_built_at'")
        if r and r[0]:
            built = datetime.fromisoformat(r[0])
            fresh = (datetime.utcnow() - built).total_seconds() < _ADV_NM_TTL
    except Exception:
        pass
    _adv_nm_cache = cache
    # устаревший кеш служит данными, но триггерит фоновое обновление
    _adv_nm_ts = _time.monotonic() if fresh else _time.monotonic() - _ADV_NM_TTL - 1
    _log.info("adv-by-nm поднят из БД: %d месяцев (fresh=%s)", len(cache), fresh)


async def _adv_nm_ensure_loaded() -> None:
    global _adv_nm_db_checked
    if _adv_nm_db_checked:
        return
    _adv_nm_db_checked = True
    if not _adv_nm_cache:
        await asyncio.to_thread(_adv_nm_load_db)


async def _build_adv_nm_bg(months: int) -> None:
    """Фоновая сборка рекламы по nmId: /adv/v3/fullstats помесячно.

    Лимит fullstats ~1 req/мин → полгода собирается несколько минут;
    до готовности юнитка распределяет рекламу по доле выручки.
    """
    global _adv_nm_cache, _adv_nm_ts, _adv_nm_building, _adv_nm_error, _adv_nm_ids
    if _adv_nm_building:
        return
    _adv_nm_building = True
    try:
        from config import USE_ADVERT_MOCK
        if USE_ADVERT_MOCK:
            _adv_nm_error = "advert mock mode"
            return
        import advert_client
        ids = await advert_client.get_all_campaign_ids_ext()
        _adv_nm_ids = len(ids)
        if not ids:
            _adv_nm_error = "campaign list is empty (проверьте доступ токена к продвижению)"
            return
        result: dict = {}
        today = (datetime.utcnow() + timedelta(hours=3)).date()
        for (y, m) in _last_months(months):
            mk = f"{y}-{m:02d}"
            # уже собран этим или прошлым запуском (свежие месяцы — приоритет)
            if mk in _adv_nm_cache and _time.monotonic() - _adv_nm_ts < _ADV_NM_TTL:
                continue
            begin = f"{mk}-01"
            last_day = (datetime(y + (m == 12), m % 12 + 1, 1) - timedelta(days=1)).date()
            end = min(last_day, today).strftime("%Y-%m-%d")
            per_nm = await advert_client.get_fullstats_nm(ids, begin, end)
            if per_nm:
                # инкрементально: месяц сразу доступен юнитке и сохранён в БД
                result[mk] = per_nm
                _adv_nm_cache = {**_adv_nm_cache, mk: per_nm}
                try:
                    await asyncio.to_thread(_adv_nm_save_month_db, mk, per_nm)
                except Exception as e:
                    _log.warning("adv-by-nm save month %s failed: %s", mk, e)
                _log.info("adv-by-nm: месяц %s готов (%d nm)", mk, len(per_nm))
            await asyncio.sleep(62)   # 1 req/мин между месяцами
        _adv_nm_ts = _time.monotonic()
        _adv_nm_error = "" if _adv_nm_cache else "fullstats вернул пусто по всем месяцам"
        try:
            await asyncio.to_thread(_adv_nm_mark_built)
        except Exception:
            pass
        _log.info("adv-by-nm собран: %s", {k: len(v) for k, v in _adv_nm_cache.items()})
    except Exception as e:
        _adv_nm_error = str(e)[:300]
        _log.warning("adv-by-nm build failed: %s", e)
    finally:
        _adv_nm_building = False


@router.get("/wb/adv_debug", include_in_schema=False)
async def wb_adv_debug():
    """Состояние точной раскладки рекламы: какие месяцы собраны, суммы по SKU."""
    import catalog as _cat
    out = {
        "building": _adv_nm_building,
        "error": _adv_nm_error,
        "campaign_ids": _adv_nm_ids,
        "age_sec": round(_time.monotonic() - _adv_nm_ts) if _adv_nm_cache else None,
        "months": {},
    }
    for mk, per_nm in _adv_nm_cache.items():
        by_sku: dict = {}
        for nm, s in per_nm.items():
            sku = _cat.resolve_wb(nm)
            by_sku[sku] = round(by_sku.get(sku, 0.0) + s)
        out["months"][mk] = {
            "fullstats_gross_total": round(sum(per_nm.values())),
            "upd_net_total": round((_wb_advert_cache.get(mk) or {}).get("total", 0)
                                   - (_wb_advert_cache.get(mk) or {}).get("bonus", 0)),
            "by_sku_gross": dict(sorted(by_sku.items(), key=lambda kv: -kv[1])),
        }
    return out


@router.get("/wb/unit")
async def get_wb_unit(
    months: int = Query(default=6, ge=1, le=12),
    refresh: bool = Query(default=False),
):
    """Юнит-экономика WB по SKU × месяц.

    Собирается из того же прохода, что и P&L — суммы столбцов сходятся
    со вкладкой Финансы. Продвижение и прочие удержания распределяются
    по SKU пропорционально доле выручки месяца (остаток округления —
    крупнейшему SKU, чтобы итог бился рубль в рубль).
    """
    # гарантируем свежую сборку P&L (она наполняет _wb_unit_data)
    pnl = await get_wb_pnl(months=months, refresh=refresh)
    if pnl.get("source") != "detail" or not _wb_unit_data:
        return {"months": [], "skus": [],
                "message": "⏳ Точные данные WB ещё собираются — юнитка появится через минуту"}

    sku_data   = _wb_unit_data["sku"]
    totals     = _wb_unit_data["totals"]
    advert_m   = _wb_unit_data["advert"]
    deduct_m   = _wb_unit_data["deductions"]
    month_keys = sorted(totals.keys())

    names = cost_store.get_names()
    import catalog as _cat

    # выручка месяца по SKU (только положительная — база распределения-фолбэка)
    rev_by_month: dict[str, float] = {
        mk: sum(max(m.get(mk, {}).get("revenue", 0.0), 0.0) for m in sku_data.values())
        for mk in month_keys
    }

    # Точная реклама по SKU из fullstats (nmId → артикул): сперва из БД
    await _adv_nm_ensure_loaded()
    if not _adv_nm_building and (not _adv_nm_cache
                                 or _time.monotonic() - _adv_nm_ts >= _ADV_NM_TTL):
        _spawn(_build_adv_nm_bg(months), light=True)
    adv_sku_by_month: dict[str, dict[str, float]] = {}
    for mk, per_nm in _adv_nm_cache.items():
        agg: dict[str, float] = {}
        for nm, s in per_nm.items():
            sku = _cat.resolve_wb(nm)
            agg[sku] = agg.get(sku, 0.0) + s
        adv_sku_by_month[mk] = agg
    adv_spend_total = {mk: sum(v.values()) for mk, v in adv_sku_by_month.items()}

    # «Точные или никакие» на уровне месяца: показываем готовые месяцы сразу,
    # несобранные (реклама есть, раскладки нет) скрываем, пока не подъедут.
    months_needing_ads = [mk for mk in month_keys if advert_m.get(mk, 0.0) > 0]
    imprecise = [mk for mk in months_needing_ads if adv_spend_total.get(mk, 0.0) <= 0]
    if imprecise:
        ready = [mk for mk in month_keys if mk not in imprecise]
        if not ready:
            if _adv_nm_error and not _adv_nm_building:
                msg = f"⚠ Раскладка рекламы по артикулам не собралась: {_adv_nm_error}"
            else:
                msg = ("⏳ Собираем точную раскладку рекламы по артикулам из статистики "
                       "кампаний WB (по минуте на месяц, свежие — первыми). "
                       "Юнитка появится автоматически.")
            return {"months": [], "skus": [], "message": msg,
                    "advert_building": _adv_nm_building}
        month_keys = ready

    skus_out = []
    # аллокация с корректировкой остатка: крупнейший SKU месяца добирает разницу
    alloc_sum = {mk: {"advert": 0.0, "deductions": 0.0} for mk in month_keys}
    top_sku_by_month = {
        mk: max(sku_data.items(),
                key=lambda kv: kv[1].get(mk, {}).get("revenue", 0.0), default=(None,))[0]
        for mk in month_keys
    }

    unit_costs = cost_store.get_costs()
    nmids = cost_store.get_nmids()

    for sku, mdata in sku_data.items():
        cat_info = _cat.lookup(sku)
        name = names.get(sku) or cat_info.get("name", "")
        row = {"sku": sku, "name": name,
               "brand": cat_info.get("brand", ""),
               "nmId": nmids.get(sku),
               "unitCost": unit_costs.get(sku, 0),
               "months": {}}
        for mk in month_keys:
            d = mdata.get(mk)
            if not d:
                continue
            rev = d["revenue"]
            share = (max(rev, 0.0) / rev_by_month[mk]) if rev_by_month[mk] > 0 else 0.0
            # реклама: точная доля из fullstats по nmId, иначе — по выручке
            if adv_spend_total.get(mk, 0.0) > 0:
                adv_share = adv_sku_by_month[mk].get(sku, 0.0) / adv_spend_total[mk]
            else:
                adv_share = share
            adv = advert_m.get(mk, 0.0) * adv_share
            ded = deduct_m.get(mk, 0.0) * share
            alloc_sum[mk]["advert"] += adv
            alloc_sum[mk]["deductions"] += ded
            commission_full = rev - d["forPay"]      # комиссия WB + эквайринг
            acq = d.get("acquiring", 0.0)
            costs_sum = (commission_full + d["delivery"] + d["storage"] + d["acceptance"]
                         + d["penalty"] + adv + ded)
            gross = rev - costs_sum - d["cogs"]
            row["months"][mk] = {
                "qty":        round(d["qty"]),
                "revenue":    round(rev),
                "commission": round(commission_full - acq),
                "acquiring":  round(acq),
                "delivery":   round(d["delivery"]),
                "storage":    round(d["storage"]),
                "acceptance": round(d["acceptance"]),
                "penalty":    round(d["penalty"]),
                "advert":     round(adv),
                "deductions": round(ded),
                "cogs":       round(d["cogs"]),
                "payout":     round(rev - costs_sum),
                "gross":      round(gross),
                "margin":     round(gross / rev * 100) if rev > 0 else 0,
            }
        if row["months"]:
            skus_out.append(row)

    # добор остатка распределения крупнейшему SKU месяца (итог = P&L)
    for mk in month_keys:
        top = top_sku_by_month.get(mk)
        if not top:
            continue
        row = next((r for r in skus_out if r["sku"] == top), None)
        if not row or mk not in row["months"]:
            continue
        cell = row["months"][mk]
        d_adv = round(advert_m.get(mk, 0.0) - alloc_sum[mk]["advert"])
        d_ded = round(deduct_m.get(mk, 0.0) - alloc_sum[mk]["deductions"])
        if d_adv or d_ded:
            cell["advert"] += d_adv
            cell["deductions"] += d_ded
            cell["payout"] -= d_adv + d_ded
            cell["gross"] -= d_adv + d_ded
            cell["margin"] = round(cell["gross"] / cell["revenue"] * 100) if cell["revenue"] else 0

    skus_out.sort(key=lambda r: -sum(m["revenue"] for m in r["months"].values()))

    # итоги месяца — прямо из P&L-агрегации (сходимость с Финансами)
    totals_out = {}
    for mk in month_keys:
        t = totals[mk]
        rev = t.get("retailAmount", 0.0)
        commission_full = t.get("commission", rev - t.get("forPay", 0.0))
        acq = sum(m.get(mk, {}).get("acquiring", 0.0) for m in sku_data.values())
        costs_sum = (commission_full + t.get("deliveryService", 0.0) + t.get("paidStorage", 0.0)
                     + t.get("paidAcceptance", 0.0) + t.get("penalty", 0.0)
                     + t.get("advert", 0.0) + t.get("deductions", 0.0))
        cg = _wb_unit_totals_cogs(mk)
        gross = rev - costs_sum - cg
        totals_out[mk] = {
            "revenue": round(rev),
            "commission": round(commission_full - acq),
            "acquiring": round(acq),
            "delivery": round(t.get("deliveryService", 0.0)),
            "storage": round(t.get("paidStorage", 0.0)),
            "acceptance": round(t.get("paidAcceptance", 0.0)),
            "penalty": round(t.get("penalty", 0.0)),
            "advert": round(t.get("advert", 0.0)),
            "deductions": round(t.get("deductions", 0.0)),
            "cogs": round(cg),
            "payout": round(rev - costs_sum),
            "gross": round(gross),
            "margin": round(gross / rev * 100) if rev else 0,
            "qty": round(sum(m.get(mk, {}).get("qty", 0) for m in sku_data.values())),
        }

    return {
        "months": [{"key": mk, "label": RU_MON[int(mk[5:7])] + " " + mk[:4]} for mk in month_keys],
        "skus": skus_out,
        "totals": totals_out,
        "advert_precise_months": sorted(mk for mk, t in adv_spend_total.items() if t > 0),
        "months_pending": imprecise,
        "advert_building": _adv_nm_building,
        "detail_upto": _wb_unit_data.get("detail_upto", ""),
        "fetched_at": _wb_unit_data.get("fetched_at", ""),
    }


def _wb_unit_totals_cogs(mk: str) -> float:
    sku_data = _wb_unit_data.get("sku", {})
    return sum(m.get(mk, {}).get("cogs", 0.0) for m in sku_data.values())


async def _get_wb_advert_cached(dt_from: datetime, dt_to: datetime) -> dict[str, float]:
    global _wb_advert_cache, _wb_advert_ts
    if _wb_advert_cache and _time.monotonic() - _wb_advert_ts < 6 * 3600:
        return _wb_advert_cache
    try:
        from config import USE_ADVERT_MOCK
        if USE_ADVERT_MOCK:
            return {}
        import advert_client
        spend = await advert_client.get_spend_by_month(dt_from, dt_to)
        _wb_advert_cache = spend
        _wb_advert_ts = _time.monotonic()
        return spend
    except Exception as e:
        _log.warning("WB advert spend failed: %s", e)
        return _wb_advert_cache or {}


RU_MON = ["", "Янв", "Фев", "Мар", "Апр", "Май", "Июн",
          "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]


def _months_out(month_keys) -> list[dict]:
    out = []
    for mk in sorted(month_keys, reverse=True):
        y, m = mk.split("-")
        out.append({"key": mk, "label": RU_MON[int(m)] + " " + y})
    return out


def _month_range(months: int) -> tuple[str, str]:
    dt_to = datetime.utcnow() + timedelta(hours=3)
    dt_from = (dt_to - timedelta(days=30 * months)).replace(day=1)
    return dt_from.strftime("%Y-%m-%d"), dt_to.strftime("%Y-%m-%d")


# ══ YM P&L ═════════════════════════════════════════════════════════════════════

_ym_pnl_cache: dict = {}
_ym_pnl_ts: float = 0.0
_YM_PNL_TTL = 3600
_ym_srv_cache: dict = {}      # mk → {группа: ₽} из «Отчёта по стоимости услуг»
_ym_srv_building: bool = False
_ym_srv_error: str = ""
_ym_unit_sku: dict = {}       # mk → {sku: {группа: ₽}} — SKU-разрез услуг
_ym_unit_rev: dict = {}       # mk → {sku: {revenue, qty, cogs}} — из заказов

# лист XLSX «Отчёта по стоимости услуг» → строка P&L (по ключевым словам)
_YM_SHEET_GROUPS = [
    ("commission", ("размещение товаров",)),
    ("acquiring",  ("приём платежа", "прием платежа", "перевод платежа")),
    ("delivery",   ("доставка",)),                    # покупателю + средняя миля
    ("advert",     ("буст",)),
    ("loyalty",    ("лояльност",)),
    ("storage",    ("хранение",)),
    ("processing", ("транзит", "обработка заказов", "возврат", "утилизац")),
]


def _ym_parse_services_xlsx(data: bytes,
                            sku_sums: dict | None = None) -> dict[str, float]:
    """Суммирует «Стоимость услуги, ₽» по листам отчёта → {группа: ₽}.

    Берётся последняя колонка «Стоимость …, ₽» листа — это итог по акту
    после скидок и наценок (в кабинетном XLSX ровно она).
    Если передан sku_sums — параллельно копит SKU-разрез {sku: {группа: ₽}}
    по колонке «Ваш SKU» (для юнит-экономики)."""
    import io
    import openpyxl
    import catalog as _cat
    # read_only (потоковое чтение, ~8× меньше памяти — важно на Render free).
    # У кабинетного XLSX битые metadata размеров листов — reset_dimensions()
    # заставляет openpyxl читать фактические строки.
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: dict[str, float] = {}
    for sh in wb.sheetnames:
        low = sh.lower()
        if low.startswith("сводка"):
            continue
        wb[sh].reset_dimensions()
        grp = next((g for g, keys in _YM_SHEET_GROUPS
                    if any(k in low for k in keys)), "otherServices")
        cost_idx = None
        sku_idx = None
        total = 0.0
        for r in wb[sh].iter_rows(values_only=True):
            if cost_idx is None:
                if r and any(c and str(c).strip().startswith("Стоимость") for c in r):
                    idxs = [j for j, c in enumerate(r)
                            if c and str(c).strip().startswith("Стоимость") and "₽" in str(c)]
                    if idxs:
                        cost_idx = idxs[-1]
                    skus = [j for j, c in enumerate(r) if c and str(c).strip() == "Ваш SKU"]
                    sku_idx = skus[0] if skus else None
                continue
            v = r[cost_idx] if len(r) > cost_idx else None
            if isinstance(v, (int, float)):
                total += float(v)
                if sku_sums is not None and sku_idx is not None and len(r) > sku_idx and r[sku_idx]:
                    sku = _cat.resolve_ym(str(r[sku_idx]).strip())
                    cell = sku_sums.setdefault(sku, {})
                    cell[grp] = cell.get(grp, 0.0) + float(v)
        if total:
            out[grp] = out.get(grp, 0.0) + total
    wb.close()
    return out


def _ym_srv_save_db(mk: str, groups: dict, sku_sums: dict | None = None) -> None:
    import db
    db.execute("CREATE TABLE IF NOT EXISTS ym_srv_groups "
               "(mk TEXT, grp TEXT, amount REAL, PRIMARY KEY (mk, grp))")
    db.execute("DELETE FROM ym_srv_groups WHERE mk = ?", (mk,))
    rows = [(mk, g, float(v)) for g, v in groups.items()]
    rows.append((mk, "__built__", 1.0))
    db.executemany("INSERT INTO ym_srv_groups (mk, grp, amount) VALUES (?,?,?)", rows)
    if sku_sums is not None:
        db.execute("CREATE TABLE IF NOT EXISTS ym_unit_sku "
                   "(mk TEXT, sku TEXT, grp TEXT, amount REAL, PRIMARY KEY (mk, sku, grp))")
        db.execute("DELETE FROM ym_unit_sku WHERE mk = ?", (mk,))
        srows = [(mk, s, g, float(v)) for s, cols in sku_sums.items() for g, v in cols.items()]
        srows.append((mk, "__built__", "__built__", 1.0))
        db.executemany("INSERT INTO ym_unit_sku (mk, sku, grp, amount) VALUES (?,?,?,?)", srows)


def _ym_srv_load_db() -> tuple[dict, dict]:
    import db
    try:
        rows = db.fetchall("SELECT mk, grp, amount FROM ym_srv_groups")
    except Exception:
        return {}, {}
    out: dict = {}
    for mk, g, v in rows:
        if g == "__built__":
            out.setdefault(mk, {})
            continue
        out.setdefault(mk, {})[g] = float(v or 0)
    sku: dict = {}
    try:
        for mk, s, g, v in db.fetchall("SELECT mk, sku, grp, amount FROM ym_unit_sku"):
            if s == "__built__":
                sku.setdefault(mk, {})
                continue
            sku.setdefault(mk, {}).setdefault(s, {})[g] = float(v or 0)
    except Exception:
        sku = {}
    # месяцы без SKU-разреза пересобираются (отчёт перетянется заново)
    for mk in [mk for mk in out if mk not in sku]:
        out.pop(mk, None)
    return out, sku


async def _build_ym_srv_bg(months: int) -> None:
    """Фоново тянет «Отчёт по стоимости услуг» помесячно (свежие — первыми)."""
    global _ym_srv_building, _ym_srv_error, _ym_pnl_ts
    if _ym_srv_building:
        return
    _ym_srv_building = True
    _ym_srv_error = ""
    try:
        import ym_client
        if not _ym_srv_cache:
            g, s = await asyncio.to_thread(_ym_srv_load_db)
            _ym_srv_cache.update(g)
            _ym_unit_sku.update(s)
        today = (datetime.utcnow() + timedelta(hours=3)).date()
        cur_mk = today.strftime("%Y-%m")
        prev_mk = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
        for (y, m) in _last_months(months):
            mk = f"{y}-{m:02d}"
            if mk in _ym_srv_cache and mk not in (cur_mk, prev_mk):
                continue
            try:
                data = await ym_client.get_services_report_month(y, m)
            except Exception as e:
                _ym_srv_error = str(e)[:300]
                _log.warning("YM services report %s: %s", mk, e)
                continue
            if not data:
                _ym_srv_cache.setdefault(mk, {})
                _ym_unit_sku.setdefault(mk, {})
                try:
                    await asyncio.to_thread(_ym_srv_save_db, mk, {}, {})
                except Exception:
                    pass
                continue
            sku_sums: dict = {}
            groups = await asyncio.to_thread(_ym_parse_services_xlsx, data, sku_sums)
            _ym_srv_cache[mk] = groups
            _ym_unit_sku[mk] = sku_sums
            try:
                await asyncio.to_thread(_ym_srv_save_db, mk, groups, sku_sums)
            except Exception as e:
                _log.warning("YM srv save %s: %s", mk, e)
            _ym_pnl_ts = 0.0   # готовый месяц сразу виден при следующем запросе
            _log.info("YM services %s: %s", mk, {k: round(v) for k, v in groups.items()})
            await asyncio.sleep(20)   # пауза между генерациями — лимит YM (420)
    finally:
        _ym_srv_building = False


_ym_orders_state: dict = {}     # {"mt": {mk:{key:₽}}, "cogs": {mk:₽}, "unit_rev": …, "built_at": iso}
_ym_orders_building: bool = False


def _ym_orders_save_db(state: dict) -> None:
    import db
    import json
    db.execute("CREATE TABLE IF NOT EXISTS kv_cache (k TEXT PRIMARY KEY, v TEXT)")
    db.execute("DELETE FROM kv_cache WHERE k = 'ym_orders_state'")
    db.execute("INSERT INTO kv_cache (k, v) VALUES ('ym_orders_state', ?)",
               (json.dumps(state, ensure_ascii=False),))


def _ym_orders_load_db() -> dict:
    import db
    import json
    try:
        rows = db.fetchall("SELECT v FROM kv_cache WHERE k = 'ym_orders_state'")
        return json.loads(rows[0][0]) if rows else {}
    except Exception:
        return {}


async def _build_ym_orders_bg(months: int) -> None:
    """Фоново тянет выкупленные заказы YM и агрегирует по месяцам/SKU."""
    global _ym_orders_building, _ym_pnl_ts
    if _ym_orders_building:
        return
    _ym_orders_building = True
    try:
        import ym_client
        import catalog as _cat
        date_from, date_to = _month_range(months)
        # захватываем заказы, созданные до начала периода, но доставленные внутри него
        pad_from = (datetime.strptime(date_from, "%Y-%m-%d") - timedelta(days=45)).strftime("%Y-%m-%d")
        orders = await ym_client.get_orders_stats(
            pad_from, date_to,
            statuses=["DELIVERED", "PARTIALLY_DELIVERED", "PARTIALLY_RETURNED"],
        )
        costs = cost_store.get_costs()
        KEYS = ["retailAmount", "subsidies", "commission", "acquiring", "delivery",
                "processing", "advert", "loyalty", "storage", "otherServices"]
        mt: dict[str, dict] = {}
        cogs_by_month: dict[str, float] = {}
        unit_rev: dict = {}   # mk → sku → {revenue, qty, cogs}
        min_mk = date_from[:7]
        for o in orders:
            # месяц реализации = дата последнего статуса (доставки), fallback — создание
            d = (o.get("statusUpdateDate") or o.get("creationDate") or "")[:7]
            if not d or d < min_mk:
                continue
            m = mt.setdefault(d, {k: 0.0 for k in KEYS})
            for it in o.get("items") or []:
                sku = _cat.resolve_ym(it.get("shopSku") or "")
                cell = unit_rev.setdefault(d, {}).setdefault(
                    sku, {"revenue": 0.0, "qty": 0.0, "cogs": 0.0})
                for p in it.get("prices") or []:
                    # Выручка — только реальные платежи покупателей (BUYER), как в
                    # «Балансе кабинета». Субсидии Маркета (акции/баллы) деньгами
                    # не приходят — Маркет компенсирует их скидками на услуги
                    # (уже учтены в актах отчёта услуг) — показываем справочно.
                    if p.get("type") == "BUYER":
                        m["retailAmount"] += float(p.get("total") or 0)
                        cell["revenue"] += float(p.get("total") or 0)
                    elif p.get("type") in ("CASHBACK", "MARKETPLACE"):
                        m["subsidies"] += float(p.get("total") or 0)
                qty = int(it.get("count") or 0)
                cell["qty"] += qty
                uc = costs.get(sku, 0.0)
                if uc > 0 and qty > 0:
                    cogs_by_month[d] = cogs_by_month.get(d, 0.0) + uc * qty
                    cell["cogs"] += uc * qty
        state = {"mt": mt, "cogs": cogs_by_month, "unit_rev": unit_rev,
                 "built_at": datetime.utcnow().isoformat()}
        _ym_orders_state.clear()
        _ym_orders_state.update(state)
        _ym_unit_rev.clear()
        _ym_unit_rev.update(unit_rev)
        try:
            await asyncio.to_thread(_ym_orders_save_db, state)
        except Exception as e:
            _log.warning("YM orders save: %s", e)
        _ym_pnl_ts = 0.0   # свежие заказы видны при следующем запросе
        _log.info("YM orders built: %d months", len(mt))
    except Exception as e:
        _log.error("YM orders build failed: %s", e)
    finally:
        _ym_orders_building = False


@router.get("/cogs_debug", include_in_schema=False)
async def cogs_debug(reapply: bool = Query(default=False)):
    """Диагностика себестоимости: сид-файл кабинета, состояние БД.

    ?reapply=1 — принудительно перечитать сид в БД (не трогая ручные правки
    сверх сида: set_costs делает upsert)."""
    from config import CABINET
    out: dict = {"cabinet": CABINET, "seed_path": str(cost_store._SEED_PATH),
                 "seed_exists": cost_store._SEED_PATH.exists()}
    try:
        costs, names, nmids = cost_store._parse_seed()
        out["seed_rows"] = {k: v for k, v in list(costs.items())[:20]}
    except Exception as e:
        out["seed_parse_error"] = str(e)[:200]
    if reapply:
        try:
            costs, names, nmids = cost_store._parse_seed()
            cost_store.set_costs(costs, names, nmids)
            out["reapplied"] = len(costs)
        except Exception as e:
            out["reapply_error"] = str(e)[:300]
    cur = cost_store.get_costs()
    out["db_count"] = len(cur)
    out["db_sample"] = {k: v for k, v in list(cur.items())[:20]}
    return out


@router.get("/ym/srv_debug", include_in_schema=False)
async def ym_srv_debug():
    """Состояние сборки «Отчёта по стоимости услуг» YM."""
    return {
        "building": _ym_srv_building,
        "error": _ym_srv_error,
        "months": {mk: {k: round(v) for k, v in g.items()}
                   for mk, g in sorted(_ym_srv_cache.items(), reverse=True)},
    }


@router.get("/ym/pnl")
async def get_ym_pnl(
    months: int = Query(default=6, ge=1, le=12),
    refresh: bool = Query(default=False),
):
    """P&L Яндекс Маркета: выручка из stats/orders (выкупы), затраты — из
    официального «Отчёта по стоимости услуг» (там ВСЁ: комиссия, буст,
    доставка, приём/перевод платежа, хранение, обработка).

    Отдаёт мгновенно из сохранённого состояния (как WB/Ozon):
    заказы и отчёты услуг пересобираются в фоне."""
    global _ym_pnl_cache, _ym_pnl_ts
    if not refresh and _ym_pnl_cache and _time.monotonic() - _ym_pnl_ts < _YM_PNL_TTL:
        return _ym_pnl_cache

    # затраты: из БД + фоновая сборка отчётов по месяцам
    if not _ym_srv_cache:
        g, s = await asyncio.to_thread(_ym_srv_load_db)
        _ym_srv_cache.update(g)
        _ym_unit_sku.update(s)
    if not _ym_srv_building:
        _spawn(_build_ym_srv_bg(months))

    # выручка/заказы: из БД + фоновая пересборка (stale-while-revalidate)
    if not _ym_orders_state:
        st = await asyncio.to_thread(_ym_orders_load_db)
        if st:
            _ym_orders_state.update(st)
            _ym_unit_rev.clear()
            _ym_unit_rev.update(st.get("unit_rev") or {})
    built_at = _ym_orders_state.get("built_at") or ""
    stale = True
    if built_at:
        try:
            age = (datetime.utcnow() - datetime.fromisoformat(built_at)).total_seconds()
            stale = age > 1800
        except ValueError:
            pass
    if (stale or refresh) and not _ym_orders_building:
        _spawn(_build_ym_orders_bg(months))
    if not _ym_orders_state.get("mt"):
        return {"months": [], "rows": [],
                "message": "⏳ Загружаем заказы YM в фоне — страница обновится сама"}

    mt = _ym_orders_state["mt"]
    cogs_by_month = _ym_orders_state.get("cogs") or {}

    # «точные или никакие»: месяц показываем, когда собрался отчёт услуг
    shown = {mk: dict(v) for mk, v in mt.items() if mk in _ym_srv_cache}
    if not shown:
        msg = (f"⚠ Отчёт услуг YM не собрался: {_ym_srv_error}"
               if _ym_srv_error and not _ym_srv_building else
               "⏳ Тянем «Отчёт по стоимости услуг» YM по месяцам (генерация ~минута на месяц) — страница обновится сама")
        return {"months": [], "rows": [], "message": msg}
    for mk in shown:
        for g, v in (_ym_srv_cache.get(mk) or {}).items():
            shown[mk][g] = v

    pnl_rows = _build_pnl_rows(
        shown, {mk: v for mk, v in cogs_by_month.items() if mk in shown},
        [
            ("retailAmount",  "📦 Выручка (платежи покупателей)", "header"),
            ("commission",    "  − Комиссия размещения",          "cost"),
            ("acquiring",     "  − Приём и перевод платежа",      "cost"),
            ("delivery",      "  − Доставка (вкл. среднюю милю)", "cost"),
            ("advert",        "  − Буст продаж",                  "cost"),
            ("loyalty",       "  − Лояльность и отзывы",          "cost"),
            ("storage",       "  − Платное хранение",             "cost"),
            ("processing",    "  − Обработка и поставки",         "cost"),
            ("otherServices", "  − Прочие услуги",                "cost"),
        ],
    )
    # субсидии Маркета — справочно (компенсируются скидками на услуги, не деньгами)
    if any(v.get("subsidies") for v in shown.values()):
        pnl_rows.insert(1, {
            "key": "subsidies",
            "label": "      ↳ Скидки за счёт Маркета (акции/баллы, компенсированы скидками на услуги)",
            "style": "note", "formula": "info",
            "values": {mk: round(shown[mk].get("subsidies", 0.0))
                       for mk in sorted(shown, reverse=True)}})

    pending = sorted(mk for mk in mt if mk not in shown)
    costs = cost_store.get_costs()
    cogs_has_data = len(costs) > 0 and any(v > 0 for v in cogs_by_month.values())
    result = {
        "months": _months_out(shown.keys()),
        "rows": pnl_rows,
        "cogs_loaded": len(costs) > 0,
        "cogs_has_data": cogs_has_data,
        "building": _ym_srv_building,
        "months_pending": pending,
        "note": "Затраты — из официального «Отчёта по стоимости услуг» YM"
                + (f"; месяцы {', '.join(pending)} ещё собираются" if pending else ""),
        "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC"),
    }
    _ym_pnl_cache = result
    _ym_pnl_ts = _time.monotonic()
    await asyncio.to_thread(_snapshot_pnl, "YM", result)
    return result


# группа услуг YM → колонка юнитки (фронтовые ключи как у WB/Ozon)
_YM_GRP_TO_COL = {
    "commission":    "commission",
    "acquiring":     "acquiring",
    "delivery":      "delivery",
    "storage":       "storage",
    "advert":        "advert",
    "loyalty":       "deductions",
    "processing":    "deductions",
    "otherServices": "deductions",
}


@router.get("/ym/unit")
async def get_ym_unit(
    months: int = Query(default=6, ge=1, le=12),
    refresh: bool = Query(default=False),
):
    """Юнит-экономика YM по SKU × месяц.

    Затраты — SKU-разрез «Отчёта по стоимости услуг» (доставка, комиссия,
    буст и т.д. привязаны к товару построчно). Услуги без SKU (буст показов,
    транзитные поставки) распределяются по доле выручки; остаток округления —
    крупнейшему SKU. Итоги сходятся со вкладкой Финансы."""
    pnl = await get_ym_pnl(months=months, refresh=refresh)
    month_keys = sorted(mk for mk in _ym_unit_sku
                        if mk in _ym_unit_rev and _ym_unit_sku[mk])
    if not month_keys:
        return {"months": [], "skus": [],
                "message": pnl.get("message")
                or "⏳ SKU-разрез услуг YM собирается — юнитка появится автоматически"}

    COST_COLS = ("commission", "acquiring", "delivery", "storage", "advert", "deductions")
    names = cost_store.get_names()
    unit_costs = cost_store.get_costs()
    import catalog as _cat

    # итоги месяца по колонкам — из тех же групп, что и Финансы
    totals_cols: dict[str, dict[str, float]] = {}
    for mk in month_keys:
        t = {c: 0.0 for c in COST_COLS}
        for g, v in (_ym_srv_cache.get(mk) or {}).items():
            t[_YM_GRP_TO_COL.get(g, "deductions")] += v
        t["revenue"] = sum(c["revenue"] for c in _ym_unit_rev[mk].values())
        totals_cols[mk] = t

    rows_by_sku: dict[str, dict] = {}
    for mk in month_keys:
        t = totals_cols[mk]
        rev_map = _ym_unit_rev[mk]
        # прямые затраты по SKU из отчёта услуг
        direct: dict[str, dict[str, float]] = {}
        for sku, grps in _ym_unit_sku[mk].items():
            cell = direct.setdefault(sku, {c: 0.0 for c in COST_COLS})
            for g, v in grps.items():
                cell[_YM_GRP_TO_COL.get(g, "deductions")] += v
        skus = sorted(set(rev_map) | set(direct))
        rev_direct = {s: max(rev_map.get(s, {}).get("revenue", 0.0), 0.0) for s in skus}
        rev_base = sum(rev_direct.values())
        top_sku = max(rev_direct, key=rev_direct.get, default=None)

        alloc = {s: dict(direct.get(s) or {c: 0.0 for c in COST_COLS}) for s in skus}
        for col in COST_COLS:
            residual = t[col] - sum(c[col] for c in alloc.values())
            if abs(residual) < 0.5 or rev_base <= 0:
                continue
            for s in skus:
                alloc[s][col] += residual * rev_direct[s] / rev_base

        rounded_sum: dict[str, int] = {}
        for s in skus:
            rm = rev_map.get(s, {})
            cell = {c: round(alloc[s][c]) for c in COST_COLS}
            cell["revenue"] = round(rm.get("revenue", 0.0))
            cell["qty"] = round(rm.get("qty", 0.0))
            cell["cogs"] = round(rm.get("cogs", 0.0))
            for c in COST_COLS:
                rounded_sum[c] = rounded_sum.get(c, 0) + cell[c]
            row = rows_by_sku.setdefault(s, {
                "sku": s, "name": names.get(s) or _cat.lookup(s).get("name", ""),
                "brand": _cat.lookup(s).get("brand", ""),
                "nmId": None, "unitCost": unit_costs.get(s, 0), "months": {}})
            row["months"][mk] = cell
        if top_sku and mk in rows_by_sku.get(top_sku, {}).get("months", {}):
            cell = rows_by_sku[top_sku]["months"][mk]
            for c in COST_COLS:
                cell[c] += round(t[c]) - rounded_sum.get(c, 0)

        for s in skus:
            cell = rows_by_sku[s]["months"][mk]
            costs_sum = sum(cell[c] for c in COST_COLS)
            cell["acceptance"] = 0
            cell["penalty"] = 0
            cell["payout"] = cell["revenue"] - costs_sum
            cell["gross"] = cell["payout"] - cell["cogs"]
            cell["margin"] = round(cell["gross"] / cell["revenue"] * 100) if cell["revenue"] else 0

    skus_out = [r for r in rows_by_sku.values()
                if any(any(round(v or 0) for v in c.values()) for c in r["months"].values())]
    skus_out.sort(key=lambda r: -sum(m.get("revenue", 0) for m in r["months"].values()))

    totals_out = {}
    for mk in month_keys:
        t = totals_cols[mk]
        costs_sum = sum(t[c] for c in COST_COLS)
        cg = sum(c["cogs"] for c in _ym_unit_rev[mk].values())
        payout = t["revenue"] - costs_sum
        gross = payout - cg
        totals_out[mk] = {
            "revenue": round(t["revenue"]),
            **{c: round(t[c]) for c in COST_COLS},
            "acceptance": 0, "penalty": 0,
            "cogs": round(cg),
            "payout": round(payout),
            "gross": round(gross),
            "margin": round(gross / t["revenue"] * 100) if t["revenue"] else 0,
            "qty": round(sum(c["qty"] for c in _ym_unit_rev[mk].values())),
        }

    pending = sorted(mk for mk in _ym_srv_cache if mk not in month_keys and _ym_srv_cache[mk])
    return {
        "months": [{"key": mk, "label": RU_MON[int(mk[5:7])] + " " + mk[:4]} for mk in month_keys],
        "skus": skus_out,
        "totals": totals_out,
        "months_pending": pending,
        "building": _ym_srv_building,
        "fetched_at": (_ym_pnl_cache or {}).get("fetched_at", ""),
    }


def _snapshot_pnl(mp: str, result: dict) -> None:
    """Архив P&L в БД: {площадка, месяц, строка} → ₽.

    История копится навсегда, независимо от того, что API площадок отдадут
    потом (WB, например, свою детализацию мы нигде больше не храним)."""
    try:
        rows = result.get("rows") or []
        if not rows:
            return
        import db
        db.execute("CREATE TABLE IF NOT EXISTS fin_history "
                   "(mp TEXT, mk TEXT, row_key TEXT, label TEXT, amount REAL, "
                   "PRIMARY KEY (mp, mk, row_key))")
        out = []
        for r in rows:
            for mk, v in (r.get("values") or {}).items():
                out.append((mp, mk, r.get("key") or "", r.get("label") or "", float(v or 0)))
        if out:
            db.executemany(
                "INSERT INTO fin_history (mp, mk, row_key, label, amount) VALUES (?,?,?,?,?) "
                "ON CONFLICT (mp, mk, row_key) DO UPDATE "
                "SET amount = excluded.amount, label = excluded.label", out)
    except Exception as e:
        _log.warning("fin_history snapshot %s: %s", mp, e)


def _build_pnl_rows(month_totals: dict, cogs_by_month: dict, cost_lines: list) -> list[dict]:
    """Единый скелет P&L: выручка → затраты → к перечислению → COGS → валовая → маржа.

    cost_lines: [(key, label, style)], первая строка — выручка (header),
    остальные — затраты (вычитаются из выручки для строки bankPayment).
    """
    sorted_months = sorted(month_totals.keys(), reverse=True)
    rows = []

    def add(key, label, style, vals):
        rows.append({"key": key, "label": label, "style": style,
                     "formula": "direct", "values": vals})

    revenue_key = cost_lines[0][0]
    add(revenue_key, cost_lines[0][1], cost_lines[0][2],
        {mk: round(month_totals[mk].get(revenue_key, 0.0)) for mk in sorted_months})

    for key, label, style in cost_lines[1:]:
        add(key, label, style,
            {mk: -round(month_totals[mk].get(key, 0.0)) for mk in sorted_months})

    payout = {}
    for mk in sorted_months:
        m = month_totals[mk]
        payout[mk] = round(m.get(revenue_key, 0.0)
                           - sum(m.get(k, 0.0) for k, _, _ in cost_lines[1:]))
    add("bankPayment", "💳 К перечислению", "subtotal", payout)

    add("cogs", "  − Себестоимость", "cost",
        {mk: -round(cogs_by_month.get(mk, 0.0)) for mk in sorted_months})

    gross = {mk: payout[mk] - round(cogs_by_month.get(mk, 0.0)) for mk in sorted_months}
    add("gross", "✅ Валовая прибыль", "total", gross)

    pct = {}
    for mk in sorted_months:
        rev = month_totals[mk].get(revenue_key, 0.0)
        pct[mk] = round(gross[mk] / rev * 100) if rev else 0
    rows.append({"key": "gross_pct", "label": "   Маржа %", "style": "pct",
                 "formula": "gross_pct", "values": pct})
    return rows


# ══ Ozon P&L ══════════════════════════════════════════════════════════════════

_oz_pnl_cache: dict = {}
_oz_pnl_ts: float = 0.0
_OZ_PNL_TTL = 6 * 3600
_oz_pnl_building: bool = False
_oz_pnl_error: str = ""
_oz_pnl_progress: str = ""
# состояние юнит-экономики Ozon: {groups: {mk:{грп:₽}}, sku: {mk:{sku:{col:₽}}}, cogs: {mk:₽}}
_oz_unit_state: dict = {}


def _oz_save_month_db(mk: str, groups: dict, cogs: float,
                      sku_sums: dict | None = None) -> None:
    """Помесячное сохранение статей Ozon — рестарты не теряют прогресс."""
    import db
    db.execute("CREATE TABLE IF NOT EXISTS oz_pnl_groups "
               "(mk TEXT, grp TEXT, amount REAL, PRIMARY KEY (mk, grp))")
    db.execute("DELETE FROM oz_pnl_groups WHERE mk = ?", (mk,))
    rows = [(mk, g, float(v)) for g, v in groups.items()]
    rows.append((mk, "__cogs__", float(cogs)))
    db.executemany("INSERT INTO oz_pnl_groups (mk, grp, amount) VALUES (?,?,?)", rows)
    if sku_sums is not None:
        db.execute("CREATE TABLE IF NOT EXISTS oz_unit_sku "
                   "(mk TEXT, sku TEXT, col TEXT, amount REAL, PRIMARY KEY (mk, sku, col))")
        db.execute("DELETE FROM oz_unit_sku WHERE mk = ?", (mk,))
        srows = [(mk, sku, col, float(v))
                 for sku, cols in sku_sums.items() for col, v in cols.items()]
        srows.append((mk, "__built__", "__built__", 1.0))   # маркер: SKU-разрез собран
        db.executemany("INSERT INTO oz_unit_sku (mk, sku, col, amount) VALUES (?,?,?,?)", srows)


def _oz_load_months_db() -> tuple[dict, dict, dict]:
    """{mk: {группа: сумма}}, {mk: cogs}, {mk: {sku: {col: сумма}}} из БД."""
    import db
    try:
        rows = db.fetchall("SELECT mk, grp, amount FROM oz_pnl_groups")
    except Exception:
        return {}, {}, {}
    groups: dict = {}
    cogs: dict = {}
    for mk, g, v in rows:
        if g == "__cogs__":
            cogs[mk] = float(v or 0)
        else:
            groups.setdefault(mk, {})[g] = float(v or 0)
    sku: dict = {}
    try:
        for mk, s, col, v in db.fetchall("SELECT mk, sku, col, amount FROM oz_unit_sku"):
            if s == "__built__":
                sku.setdefault(mk, {})
                continue
            sku.setdefault(mk, {}).setdefault(s, {})[col] = float(v or 0)
    except Exception:
        sku = {}
    # инвалидация старых схем: баллы отдельной статьёй (двойной счёт) или
    # месяц без SKU-разреза — такие месяцы пересоберутся заново
    stale = [mk for mk, g in groups.items()
             if "Баллы за скидки" in g or mk not in sku]
    for mk in stale:
        groups.pop(mk, None)
        cogs.pop(mk, None)
        sku.pop(mk, None)
    return groups, cogs, sku


def _last_months(n: int) -> list[tuple[int, int]]:
    """[(year, month)] — последние n месяцев включая текущий, новые первыми."""
    d = datetime.utcnow() + timedelta(hours=3)
    out = []
    y, m = d.year, d.month
    for _ in range(n):
        out.append((y, m))
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    return out


def _money(v) -> float:
    if isinstance(v, dict):
        v = v.get("amount")
    try:
        return float(v or 0)
    except (ValueError, TypeError):
        return 0.0


def _accrual_amount(a: dict) -> float:
    return _money(a.get("total_amount"))


def _accrual_type_id(a: dict):
    return a.get("type_id") or a.get("accrual_id") or a.get("type")


# Группировка элементарных типов начислений в кабинетные статьи (по описанию)
_OZ_GROUP_RULES = [
    ("Продвижение и реклама", ("продвижен", "реклам", "brand", "трафарет", "полка",
                               "advertising", "promotion", "буст", "отзыв",
                               "оплата за клик", "за клик")),
    ("Услуги доставки",       ("доставк", "магистрал", "миля", "drop-off", "курьер",
                               "логистик", "перевозк", "shipment")),
    ("Услуги FBO",            ("сборка", "fulfillment", "размещени", "кросс-док",
                               "приёмк", "приемк", "обработка отправ", "хранени")),
    ("Обработка возвратов и отмен", ("возврат", "отмен", "невостреб", "утилизац")),
    ("Эквайринг",             ("эквайринг", "acquiring", "рассрочк", "installment")),
    ("Компенсации",           ("компенсац", "претензи", "claim")),
    ("Другие услуги и штрафы", ("штраф", "ошибок", "defect", "инвентаризац", "пожертвован")),
]


def _oz_group_for(desc: str) -> str:
    low = (desc or "").lower()
    for group, keys in _OZ_GROUP_RULES:
        if any(k in low for k in keys):
            return group
    return "Прочие услуги"


# колонки юнит-экономики Ozon и соответствие кабинетным статьям
_OZ_GROUP_TO_COL = {
    "Продажи и возвраты":           "revenue",
    "Программы партнёров":          "revenue",
    "Вознаграждение Ozon":          "commission",
    "Услуги доставки":              "delivery",
    "Услуги FBO":                   "storage",
    "Продвижение и реклама":        "advert",
    "Эквайринг":                    "acquiring",
    "Обработка возвратов и отмен":  "returns",
}


def _oz_col_for(group: str) -> str:
    return _OZ_GROUP_TO_COL.get(group, "other")


def _oz_sku_of(obj: dict) -> str:
    """Артикул из объекта начисления: пробуем offer_id, потом ozon sku id."""
    import catalog as _cat
    for k in ("offer_id", "offer", "item_code", "article"):
        v = obj.get(k)
        if v:
            return str(v).strip().upper()
    item = obj.get("item") or {}
    if isinstance(item, dict):
        for k in ("offer_id", "sku"):
            v = item.get(k)
            if v:
                return _cat.resolve_ozon(v) if str(v).isdigit() else str(v).strip().upper()
    for k in ("sku", "item_sku", "product_id"):
        v = obj.get(k)
        if v:
            return _cat.resolve_ozon(v)
    return ""


def _oz_decompose_accrual(a: dict, types: dict, sums: dict,
                          sku_sums: dict | None = None) -> None:
    """Раскладывает начисление на кабинетные статьи. Остаток строки уходит в
    «Продажи и возвраты»/«Прочие начисления», чтобы Итого всегда сходился.
    Если передан sku_sums — параллельно копит SKU-разрез для юнит-экономики
    (те же суммы, поэтому юнитка сходится с Финансами по построению)."""
    total = _accrual_amount(a)
    accounted = 0.0

    def add(group: str, amount: float, sku: str = ""):
        nonlocal accounted
        if not amount:
            return
        sums[group] = sums.get(group, 0.0) + amount
        accounted += amount
        if sku_sums is not None and sku:
            cell = sku_sums.setdefault(sku, {})
            col = _oz_col_for(group)
            cell[col] = cell.get(col, 0.0) + amount

    def s_qty(sku: str, q: float):
        if sku_sums is not None and sku and q:
            cell = sku_sums.setdefault(sku, {})
            cell["qty"] = cell.get("qty", 0.0) + q

    def fee_group(tid) -> str:
        t = types.get(tid) or {}
        return _oz_group_for(t.get("description") or t.get("name") or "")

    row_sku = _oz_sku_of(a)

    # товарные и нетоварные услуги
    item_fees = (a.get("item_fees") or {}).get("fees") or []
    for by_sku in item_fees:
        f_sku = _oz_sku_of(by_sku) or row_sku
        for f in by_sku.get("fees") or []:
            add(fee_group(f.get("type_id")), _money(f.get("accrued")), f_sku)
    nif = a.get("non_item_fee")
    for f in (nif if isinstance(nif, list) else [nif] if nif else []):
        add(fee_group(f.get("type_id")), _money(f.get("accrued")), row_sku)
    cf = a.get("container_fees")
    for f in (cf if isinstance(cf, list) else [cf] if cf else []):
        if isinstance(f, dict):
            add(fee_group(f.get("type_id")), _money(f.get("accrued")), row_sku)

    # отправление: выручка, вознаграждение, баллы, услуги доставки
    posting = a.get("posting") or {}
    prod_skus: list[tuple[str, float]] = []   # (sku, |sale|) — для раздачи остатка
    for prod in posting.get("products") or []:
        c = prod.get("commission") or {}
        p_sku = _oz_sku_of(prod) or row_sku
        sale = _money(c.get("sale_amount"))
        add("Продажи и возвраты", sale, p_sku)
        add("Вознаграждение Ozon", -abs(_money(c.get("commission")) or _money(c.get("sale_commission"))), p_sku)
        prod_skus.append((p_sku, abs(sale)))
        qty = _money(prod.get("quantity")) or (1 if sale else 0)
        s_qty(p_sku, qty if sale >= 0 else -qty)
        # Баллы за скидки (соинвест Ozon) уже входят в сумму строки продаж —
        # НЕ отдельная статья (иначе двойной счёт). Копим справочно, мимо accounted:
        # остаток строки уйдёт в «Продажи и возвраты», как в кабинете.
        for bonus_key in ("bonus", "coinvestment", "stars"):
            v = _money(c.get(bonus_key))
            if v:
                sums["__points__"] = sums.get("__points__", 0.0) + v
        d = prod.get("delivery") or {}
        for svc in d.get("services") or []:
            add(fee_group(svc.get("type_id")), _money(svc.get("accrued")), p_sku)
        if d.get("total_accrued") is not None and not d.get("services"):
            add("Услуги доставки", _money(d.get("total_accrued")), p_sku)

    # остаток строки — чтобы Итого совпало с кабинетом копейка в копейку
    rest = total - accounted
    if abs(rest) > 0.005:
        cat = a.get("accrued_category") or ""
        group = "Продажи и возвраты" if cat == "POSTING" and rest > 0 else "Прочие начисления"
        # остаток POSTING (обычно баллы) — товарам отправления по доле продаж
        base = sum(w for _, w in prod_skus)
        if prod_skus and base > 0:
            left = rest
            for i, (p_sku, w) in enumerate(prod_skus):
                part = rest * w / base if i < len(prod_skus) - 1 else left
                add(group, part, p_sku)
                left -= part
        else:
            add(group, rest, prod_skus[0][0] if prod_skus else row_sku)


async def _fetch_ozon_accruals_month(y: int, m: int, today,
                                     sku_sums: dict | None = None) -> dict[str, float]:
    """Суммы начислений месяца по кабинетным статьям (+SKU-разрез в sku_sums)."""
    import ozon_client
    types = {t.get("id"): t for t in await ozon_client.get_accrual_types()}
    sums: dict[str, float] = {}
    last_day = (datetime(y + (m == 12), m % 12 + 1, 1) - timedelta(days=1)).date()
    d = datetime(y, m, 1).date()
    while d <= min(last_day, today):
        try:
            for a in await ozon_client.get_accruals_day(d.isoformat()):
                _oz_decompose_accrual(a, types, sums, sku_sums)
        except Exception as e:
            _log.warning("Ozon accruals %s: %s", d, e)
        d += timedelta(days=1)
    return {k: v for k, v in sums.items() if abs(v) >= 0.5}


@router.get("/ozon/accrual_debug", include_in_schema=False)
async def ozon_accrual_debug(date: str = Query(default="")):
    """Живая проба accruals API: справочник типов + один день начислений."""
    import ozon_client
    out: dict = {}
    try:
        types = await ozon_client.get_accrual_types()
        out["types_count"] = len(types)
        out["types_sample"] = types[:25]
    except Exception as e:
        out["types_error"] = str(e)[:300]
    probe = date or (datetime.utcnow() + timedelta(hours=3) - timedelta(days=3)).strftime("%Y-%m-%d")
    try:
        rows = await ozon_client.get_accruals_day(probe)
        out["day"] = probe
        out["rows"] = len(rows)
        if rows:
            out["row_keys"] = sorted(rows[0].keys())
            # по одному примеру каждой категории — видно структуру posting/non_item
            seen_cat: dict = {}
            for a in rows:
                cat = a.get("accrued_category") or "—"
                if cat not in seen_cat:
                    seen_cat[cat] = a
            out["samples_by_category"] = seen_cat
            # проверка разложения: сумма статей должна равняться сумме строк
            types = {t.get("id"): t for t in await ozon_client.get_accrual_types()}
            sums: dict = {}
            for a in rows:
                _oz_decompose_accrual(a, types, sums)
            out["day_decomposed"] = {k: round(v) for k, v in sorted(sums.items(), key=lambda kv: -abs(kv[1]))}
            out["day_total_rows"] = round(sum(_accrual_amount(a) for a in rows))
            out["day_total_decomposed"] = round(sum(sums.values()))
    except Exception as e:
        out["day_error"] = str(e)[:300]
    return out


async def _build_ozon_pnl(months: int) -> None:
    """Фоновая сборка P&L Ozon.

    Статьи — из «Детализации начислений» (accruals API): динамические группы
    ровно как в кабинете («Вознаграждение Ozon», «Услуги доставки», «Услуги
    FBO», «Продвижение и реклама», ...). Фолбэк для месяцев без начислений —
    отчёт о реализации + cash-flow. Себестоимость: qty из отчёта о реализации,
    для текущего месяца — из доставленных отправлений.
    """
    global _oz_pnl_cache, _oz_pnl_ts, _oz_pnl_building, _oz_pnl_error, _oz_pnl_progress
    if _oz_pnl_building:
        return
    _oz_pnl_building = True
    _oz_pnl_error = ""
    try:
        import ozon_client
        import catalog as _cat

        costs = cost_store.get_costs()
        today = (datetime.utcnow() + timedelta(hours=3)).date()
        # поднимаем уже собранные месяцы из БД (свежие два — пересобираем,
        # они ещё пополняются начислениями)
        db_groups, db_cogs, db_sku = await asyncio.to_thread(_oz_load_months_db)
        cur_mk = today.strftime("%Y-%m")
        prev_mk = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
        month_groups: dict[str, dict[str, float]] = {}   # mk → {группа: сумма}
        cogs_by_month: dict[str, float] = {}
        sku_by_month: dict[str, dict] = {}               # mk → {sku: {col: сумма}}

        # stale-while-revalidate: сразу публикуем ВСЁ, что есть в БД (включая
        # свежие месяцы прошлой сборки) — страница открывается мгновенно,
        # а текущий/прошлый месяц тихо пересобираются ниже
        want = {f"{y}-{m:02d}" for (y, m) in _last_months(months)}
        for mk in sorted(db_groups.keys() & want, reverse=True):
            month_groups[mk] = db_groups[mk]
            cogs_by_month[mk] = db_cogs.get(mk, 0.0)
            sku_by_month[mk] = db_sku.get(mk, {})
        if month_groups:
            _oz_pnl_partial(month_groups, cogs_by_month, costs)
            _oz_unit_state.update(groups=dict(month_groups), sku=dict(sku_by_month),
                                  cogs=dict(cogs_by_month))

        for (y, m) in _last_months(months):
            mk = f"{y}-{m:02d}"
            if mk in db_groups and mk not in (cur_mk, prev_mk):
                continue   # уже опубликован из БД, пересборка не нужна
            _oz_pnl_progress = mk
            month_start = f"{mk}-01"
            month_end = (datetime(y + (m == 12), m % 12 + 1, 1) - timedelta(days=1)).strftime("%Y-%m-%d")

            # ── COGS: отчёт о реализации (официальные количества по SKU) ──
            realization_ok = False
            try:
                res = await ozon_client.get_realization_report(y, m)
                rows = res.get("rows") or []
                if rows:
                    cogs = 0.0
                    for r in rows:
                        dc = r.get("delivery_commission") or {}
                        rc = r.get("return_commission") or {}
                        qty = int(dc.get("quantity") or 0) - int(rc.get("quantity") or 0)
                        offer = ((r.get("item") or {}).get("offer_id") or "").strip()
                        uc = costs.get(_cat.resolve_ozon(offer), 0.0)
                        if uc > 0 and qty > 0:
                            cogs += uc * qty
                    cogs_by_month[mk] = cogs
                    realization_ok = True
            except Exception as e:
                _log.warning("Ozon realization %s: %s", mk, e)

            # текущий месяц: COGS из доставленных отправлений
            if not realization_ok:
                try:
                    rows = await ozon_client.get_sales_detail(month_start, month_end)
                    cogs = 0.0
                    for r in rows:
                        if (r.get("status") or "").lower() != "delivered":
                            continue
                        dd = (r.get("delivered_date") or r.get("date") or "")[:7]
                        if dd != mk:
                            continue
                        uc = costs.get(_cat.resolve_ozon(r.get("offer_id") or ""), 0.0)
                        if uc > 0:
                            cogs += uc * int(r.get("qty") or 0)
                    cogs_by_month[mk] = cogs
                except Exception as e:
                    _log.warning("Ozon delivered COGS %s: %s", mk, e)

            # ── Статьи: детализация начислений (как в кабинете) + SKU-разрез ──
            sku_sums: dict = {}
            accr = await _fetch_ozon_accruals_month(y, m, today, sku_sums)
            if accr:
                month_groups[mk] = accr
                sku_by_month[mk] = sku_sums
                try:
                    await asyncio.to_thread(_oz_save_month_db, mk, accr,
                                            cogs_by_month.get(mk, 0.0), sku_sums)
                except Exception as e:
                    _log.warning("Ozon save month %s failed: %s", mk, e)
                # промежуточный кэш: собранные месяцы видны сразу
                _oz_pnl_partial(month_groups, cogs_by_month, costs)
                _oz_unit_state.update(groups=dict(month_groups), sku=dict(sku_by_month),
                                      cogs=dict(cogs_by_month))
                continue

            # фолбэк: реализация + cash-flow старыми строками
            fb: dict[str, float] = {}
            try:
                if realization_ok:
                    rev = comm = 0.0
                    for r in (res.get("rows") or []):
                        dc = r.get("delivery_commission") or {}
                        rc = r.get("return_commission") or {}
                        rev  += float(dc.get("amount") or 0) - float(rc.get("amount") or 0)
                        comm += float(dc.get("commission") or 0) - float(rc.get("commission") or 0)
                    fb["Выручка (реализация)"] = rev
                    fb["Вознаграждение Ozon"] = -comm
                flows = await ozon_client.get_cash_flow(month_start, month_end)
                for f in flows or []:
                    fb["Услуги доставки"] = fb.get("Услуги доставки", 0.0) - float(f.get("item_delivery_and_return_amount") or 0)
                    fb["Услуги Ozon"] = fb.get("Услуги Ozon", 0.0) - float(f.get("services_amount") or 0)
                    if not realization_ok:
                        fb["Выручка (реализация)"] = fb.get("Выручка (реализация)", 0.0) + \
                            float(f.get("orders_amount") or 0) + float(f.get("returns_amount") or 0)
                        fb["Вознаграждение Ozon"] = fb.get("Вознаграждение Ozon", 0.0) - float(f.get("commission_amount") or 0)
            except Exception as e:
                _log.warning("Ozon fallback %s: %s", mk, e)
            if fb:
                month_groups[mk] = fb

        # финальная сборка строк
        _oz_pnl_partial(month_groups, cogs_by_month, costs)
        _oz_unit_state.update(groups=dict(month_groups), sku=dict(sku_by_month),
                              cogs=dict(cogs_by_month))
        _oz_pnl_ts = _time.monotonic()
        _oz_pnl_progress = ""
        _log.info("Ozon P&L built: %d months", len(month_groups))
    except Exception as e:
        _oz_pnl_error = str(e)[:300]
        _log.error("Ozon P&L build failed: %s", e)
    finally:
        _oz_pnl_building = False


def _oz_pnl_partial(month_groups: dict, cogs_by_month: dict, costs: dict) -> None:
    """Собирает строки P&L из уже готовых месяцев и кладёт в кэш."""
    global _oz_pnl_cache
    if not month_groups:
        return
    month_keys = sorted(month_groups.keys(), reverse=True)
    all_groups: dict[str, float] = {}
    for mk in month_keys:
        for g, v in month_groups[mk].items():
            if g.startswith("__"):
                continue   # служебные ключи (__points__) — не статьи
            all_groups[g] = all_groups.get(g, 0.0) + v

    # Блок «Продажи и возвраты» кабинета УЖЕ включает баллы за скидки
    # (соинвест Ozon) — баллы показываем только справочной подстрокой
    REVENUE_SET = {"Продажи и возвраты", "Программы партнёров"}
    rev_components = [g for g in all_groups if g in REVENUE_SET]
    other_pos = sorted((g for g, v in all_groups.items() if v > 0 and g not in REVENUE_SET),
                       key=lambda g: -all_groups[g])
    cost_groups = sorted((g for g, v in all_groups.items() if v <= 0 and g not in REVENUE_SET),
                         key=lambda g: all_groups[g])

    revenue = {mk: round(sum(month_groups[mk].get(g, 0.0) for g in rev_components))
               for mk in month_keys}
    points = {mk: round(month_groups[mk].get("__points__", 0.0)) for mk in month_keys}
    pnl_rows = [{"key": "retailAmount", "label": "📦 Выручка (продажи и возвраты, вкл. баллы)",
                 "style": "header", "formula": "direct", "values": revenue}]
    # составляющие выручки — справочно
    if any(points.values()):
        sales_wo = {mk: round(month_groups[mk].get("Продажи и возвраты", 0.0)) - points[mk]
                    for mk in month_keys}
        pnl_rows.append({"key": "revc_sales", "label": "      ↳ Продажи (оплачено покупателями)",
                         "style": "note", "formula": "info", "values": sales_wo})
        pnl_rows.append({"key": "revc_points", "label": "      ↳ Баллы за скидки (соинвест Ozon)",
                         "style": "note", "formula": "info", "values": points})
        rev_notes = [g for g in rev_components if g != "Продажи и возвраты"]
    else:
        rev_notes = list(rev_components)
    for i, g in enumerate(sorted(rev_notes, key=lambda g: -all_groups[g])):
        pnl_rows.append({"key": f"revc_{i}", "label": f"      ↳ {g}", "style": "note",
                         "formula": "info",
                         "values": {mk: round(month_groups[mk].get(g, 0.0)) for mk in month_keys}})
    for i, g in enumerate(other_pos):
        pnl_rows.append({"key": f"pos_{i}", "label": f"  + {g}", "style": "normal",
                         "formula": "direct",
                         "values": {mk: round(month_groups[mk].get(g, 0.0)) for mk in month_keys}})
    for i, g in enumerate(cost_groups):
        pnl_rows.append({"key": f"cost_{i}", "label": f"  − {g}", "style": "cost",
                         "formula": "direct",
                         "values": {mk: round(month_groups[mk].get(g, 0.0)) for mk in month_keys}})
    payout = {mk: round(sum(v for g, v in month_groups[mk].items() if not g.startswith("__")))
              for mk in month_keys}
    pnl_rows.append({"key": "bankPayment", "label": "💳 Итого начислено (к перечислению)",
                     "style": "subtotal", "formula": "direct", "values": payout})
    pnl_rows.append({"key": "cogs", "label": "  − Себестоимость", "style": "cost",
                     "formula": "direct",
                     "values": {mk: -round(cogs_by_month.get(mk, 0.0)) for mk in month_keys}})
    gross = {mk: payout[mk] - round(cogs_by_month.get(mk, 0.0)) for mk in month_keys}
    pnl_rows.append({"key": "gross", "label": "✅ Валовая прибыль", "style": "total",
                     "formula": "direct", "values": gross})
    pct = {}
    for mk in month_keys:
        pct[mk] = round(gross[mk] / revenue[mk] * 100) if revenue.get(mk) else 0
    pnl_rows.append({"key": "gross_pct", "label": "   Маржа %", "style": "pct",
                     "formula": "gross_pct", "values": pct})

    cogs_has_data = len(costs) > 0 and any(v > 0 for v in cogs_by_month.values())
    _oz_pnl_cache = {
        "months": _months_out(month_groups.keys()),
        "rows": pnl_rows,
        "cogs_loaded": len(costs) > 0,
        "cogs_has_data": cogs_has_data,
        "building": _oz_pnl_building,
        "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC"),
    }
    _snapshot_pnl("OZON", _oz_pnl_cache)


@router.get("/ozon/pnl")
async def get_ozon_pnl(
    months: int = Query(default=6, ge=1, le=12),
    refresh: bool = Query(default=False),
):
    """P&L Ozon по месяцам. Сборка в фоне (несколько запросов к API)."""
    global _oz_pnl_ts
    fresh = _oz_pnl_cache and _time.monotonic() - _oz_pnl_ts < _OZ_PNL_TTL
    if not refresh and fresh:
        return _oz_pnl_cache
    if refresh:
        _oz_pnl_ts = 0.0
    _spawn(_build_ozon_pnl(months))
    if _oz_pnl_cache:
        return _oz_pnl_cache  # отдаём собранное, остальное доедет в фоне
    if _oz_pnl_error and not _oz_pnl_building:
        return {"months": [], "rows": [],
                "message": f"⚠ Ozon: сборка упала — {_oz_pnl_error}"}
    prog = f" (сейчас: {_oz_pnl_progress})" if _oz_pnl_progress else ""
    return {"months": [], "rows": [],
            "message": f"⏳ Отчёт Ozon собирается в фоне{prog} — страница обновится сама"}


@router.get("/ozon/unit")
async def get_ozon_unit(
    months: int = Query(default=6, ge=1, le=12),
    refresh: bool = Query(default=False),
):
    """Юнит-экономика Ozon по SKU × месяц.

    SKU-разрез копится тем же проходом по детализации начислений, что и P&L,
    поэтому итоги столбцов сходятся со вкладкой Финансы. Начисления без
    привязки к товару (промо «оплата за клик», размещение FBO и т.п.)
    распределяются по SKU пропорционально доле выручки месяца; остаток
    округления — крупнейшему SKU."""
    pnl = await get_ozon_pnl(months=months, refresh=refresh)
    st = _oz_unit_state
    sku_m: dict = st.get("sku") or {}
    month_keys = sorted(mk for mk, s in sku_m.items() if s)
    if not month_keys:
        return {"months": [], "skus": [],
                "message": pnl.get("message")
                or "⏳ SKU-разрез Ozon собирается по мере пересборки месяцев — юнитка появится автоматически"}
    groups_m: dict = st.get("groups") or {}
    cogs_m: dict = st.get("cogs") or {}
    pending = sorted(mk for mk in groups_m if mk not in month_keys)

    COST_COLS = ("commission", "delivery", "storage", "advert", "acquiring", "returns", "other")

    # итоги месяца по колонкам — прямо из статей P&L (сходимость с Финансами)
    totals_cols: dict[str, dict[str, float]] = {}
    for mk in month_keys:
        t = {c: 0.0 for c in ("revenue",) + COST_COLS}
        for grp, v in (groups_m.get(mk) or {}).items():
            if grp.startswith("__"):
                continue
            col = _oz_col_for(grp)
            t[col] += v if col == "revenue" else -v
        totals_cols[mk] = t

    # прямые суммы по SKU (знак → положительные затраты)
    direct: dict[str, dict[str, dict[str, float]]] = {}   # mk → sku → col
    for mk in month_keys:
        d = direct[mk] = {}
        for sku, cols in sku_m[mk].items():
            cell = d.setdefault(sku, {})
            for col, v in cols.items():
                if col == "qty":
                    cell["qty"] = cell.get("qty", 0.0) + v
                elif col == "revenue":
                    cell["revenue"] = cell.get("revenue", 0.0) + v
                else:
                    cell[col] = cell.get(col, 0.0) - v

    names = cost_store.get_names()
    unit_costs = cost_store.get_costs()
    import catalog as _cat
    oz_ids = {art: oid for oid, art in _cat.OZON_ID_TO_ART.items()}

    all_skus = sorted({s for mk in month_keys for s in direct[mk]})
    rows_by_sku: dict[str, dict] = {}
    for mk in month_keys:
        t = totals_cols[mk]
        d = direct[mk]
        rev_direct = {s: max(c.get("revenue", 0.0), 0.0) for s, c in d.items()}
        rev_base = sum(rev_direct.values())
        top_sku = max(rev_direct, key=rev_direct.get, default=None)

        # распределение недоатрибуцированного остатка каждой колонки по доле выручки
        alloc: dict[str, dict[str, float]] = {s: dict(c) for s, c in d.items()}
        for col in ("revenue",) + COST_COLS:
            residual = t[col] - sum(c.get(col, 0.0) for c in d.values())
            if abs(residual) < 0.5 or rev_base <= 0:
                continue
            for s in alloc:
                alloc[s][col] = alloc[s].get(col, 0.0) + residual * rev_direct[s] / rev_base

        # себестоимость: цена × шт, остаток к официальному COGS — топ-SKU
        for s, c in alloc.items():
            qty = c.get("qty", 0.0)
            c["cogs"] = unit_costs.get(s, 0.0) * max(qty, 0.0)
        cg_target = cogs_m.get(mk, 0.0)
        if cg_target > 0 and top_sku:
            diff = cg_target - sum(c["cogs"] for c in alloc.values())
            alloc[top_sku]["cogs"] = max(alloc[top_sku]["cogs"] + diff, 0.0)

        # округление + добор разницы топ-SKU (итог = P&L рубль в рубль)
        rounded_sum: dict[str, int] = {}
        for s, c in alloc.items():
            cell = {k: round(c.get(k, 0.0)) for k in ("qty", "revenue", "cogs") + COST_COLS}
            for k in ("revenue",) + COST_COLS:
                rounded_sum[k] = rounded_sum.get(k, 0) + cell[k]
            row = rows_by_sku.setdefault(s, {
                "sku": s, "name": names.get(s) or _cat.lookup(s).get("name", ""),
                "brand": _cat.lookup(s).get("brand", ""),
                "nmId": oz_ids.get(s), "unitCost": unit_costs.get(s, 0), "months": {}})
            row["months"][mk] = cell
        if top_sku and top_sku in rows_by_sku and mk in rows_by_sku[top_sku]["months"]:
            cell = rows_by_sku[top_sku]["months"][mk]
            for k in ("revenue",) + COST_COLS:
                cell[k] += round(t[k]) - rounded_sum.get(k, 0)

        # производные значения ячеек
        for s in list(d.keys()):
            cell = rows_by_sku[s]["months"].get(mk)
            if not cell:
                continue
            costs_sum = sum(cell[k] for k in COST_COLS)
            cell["deductions"] = cell.pop("returns") + cell.pop("other")
            cell["acceptance"] = 0
            cell["penalty"] = 0
            cell["payout"] = cell["revenue"] - costs_sum
            cell["gross"] = cell["payout"] - cell["cogs"]
            cell["margin"] = round(cell["gross"] / cell["revenue"] * 100) if cell["revenue"] else 0

    skus_out = [r for r in rows_by_sku.values()
                if any(any(round(v or 0) for v in c.values()) for c in r["months"].values())]
    skus_out.sort(key=lambda r: -sum(m.get("revenue", 0) for m in r["months"].values()))

    totals_out = {}
    for mk in month_keys:
        t = totals_cols[mk]
        costs_sum = sum(t[c] for c in COST_COLS)
        cg = cogs_m.get(mk, 0.0)
        if cg <= 0:
            cg = sum(r["months"].get(mk, {}).get("cogs", 0) for r in skus_out)
        payout = t["revenue"] - costs_sum       # == «Итого начислено» в Финансах
        gross = payout - cg
        totals_out[mk] = {
            "revenue": round(t["revenue"]),
            "commission": round(t["commission"]),
            "acquiring": round(t["acquiring"]),
            "delivery": round(t["delivery"]),
            "storage": round(t["storage"]),
            "acceptance": 0, "penalty": 0,
            "advert": round(t["advert"]),
            "deductions": round(t["returns"] + t["other"]),
            "cogs": round(cg),
            "payout": round(payout),
            "gross": round(gross),
            "margin": round(gross / t["revenue"] * 100) if t["revenue"] else 0,
            "qty": round(sum(r["months"].get(mk, {}).get("qty", 0) for r in skus_out)),
        }

    return {
        "months": [{"key": mk, "label": RU_MON[int(mk[5:7])] + " " + mk[:4]} for mk in month_keys],
        "skus": skus_out,
        "totals": totals_out,
        "months_pending": pending,
        "building": _oz_pnl_building,
        "fetched_at": (_oz_pnl_cache or {}).get("fetched_at", ""),
    }


# ══ График выплат (вкладка Тотал) ══════════════════════════════════════════════

_payouts_cache: dict = {}
_payouts_ts: float = 0.0


def _row_val(cache: dict, key: str, mk: str) -> float:
    for r in (cache or {}).get("rows") or []:
        if r.get("key") == key:
            return float((r.get("values") or {}).get(mk) or 0)
    return 0.0


@router.get("/payouts")
async def get_payouts(refresh: bool = Query(default=False)):
    """Балансы кабинетов и предстоящие поступления по площадкам.

    WB — живой баланс продавца из finance-api (к выводу = невыплаченное).
    Ozon и YM публичного баланса в API не имеют — показываем оценку:
    «к перечислению», начисленное в текущем месяце (и хвост прошлого),
    с пояснением графика выплат площадки."""
    global _payouts_cache, _payouts_ts
    if not refresh and _payouts_cache and _time.monotonic() - _payouts_ts < 1800:
        return _payouts_cache

    from config import CABINET_MARKETPLACES
    today = (datetime.utcnow() + timedelta(hours=3)).date()
    cur_mk = today.strftime("%Y-%m")
    prev_mk = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    items = []

    # ── WB: реальный баланс ──
    if "WB" in CABINET_MARKETPLACES:
        wb: dict = {"mp": "WB", "color": "#c026d3"}
        try:
            import wb_finance_client
            bal = await wb_finance_client.get_balance()
            # имена полей варьируются: current/balance, for_withdraw/forWithdraw
            cur = next((float(bal[k]) for k in ("current", "balance", "value") if bal.get(k) is not None), None)
            wd = next((float(bal[k]) for k in ("for_withdraw", "forWithdraw", "forwithdraw") if bal.get(k) is not None), None)
            if cur is not None or wd is not None:
                wb["balance"] = round(cur if cur is not None else wd)
                if wd is not None:
                    wb["for_withdraw"] = round(wd)
        except Exception as e:
            _log.warning("WB balance: %s", e)
        # WB не платит сам: деньги копятся на балансе, вывод — вручную кнопкой
        # в кабинете, зачисление ~неделя. «Ожидается» = весь баланс.
        wb["upcoming"] = wb.get("balance") if wb.get("balance") is not None \
            else round(_row_val(_pnl_cache, "bankPayment", cur_mk))
        wb["note"] = ("Баланс — из API WB. Деньги копятся на балансе, вывод — вручную "
                      "в кабинете («Вывести»), зачисление ~неделя после заявки.")
        items.append(wb)

    # Обе площадки платят еженедельно с отсрочкой ~4 недели (сверено с
    # кабинетами). Невыплачено ≈ начисления после «отсечки» = сегодня − лаг.
    def _prev_unpaid_share(lag_days: int) -> float:
        cutoff = today - timedelta(days=lag_days)
        prev_last = today.replace(day=1) - timedelta(days=1)
        if (cutoff.year, cutoff.month) < (prev_last.year, prev_last.month):
            return 1.0    # отсечка раньше прошлого месяца — он весь не выплачен
        if (cutoff.year, cutoff.month) > (prev_last.year, prev_last.month):
            return 0.0
        return max(0, prev_last.day - cutoff.day) / prev_last.day

    # ── Ozon: недельные периоды, выплата ~через 24 дня после конца недели ──
    if "OZON" in CABINET_MARKETPLACES:
        upcoming = (_row_val(_oz_pnl_cache, "bankPayment", cur_mk)
                    + _row_val(_oz_pnl_cache, "bankPayment", prev_mk) * _prev_unpaid_share(24))
        items.append({
            "mp": "Ozon", "color": "#3b82f6",
            "upcoming": round(upcoming),
            "note": "Оценка: начисления последних ~4 недель (стандартный график — еженедельно с отсрочкой ~4 недели).",
        })

    # ── YM: недельные периоды, выплата через 4 недели после конца недели ──
    if "YM" in CABINET_MARKETPLACES:
        upcoming = (_row_val(_ym_pnl_cache, "bankPayment", cur_mk)
                    + _row_val(_ym_pnl_cache, "bankPayment", prev_mk) * _prev_unpaid_share(35))
        items.append({
            "mp": "ЯМ", "color": "#b45309",
            "upcoming": round(upcoming),
            "note": "Оценка ≈ балансу кабинета: начисления последних ~5 недель (еженедельно с отсрочкой 4 недели).",
        })

    result = {
        "items": items,
        "total_upcoming": round(sum(i.get("upcoming") or 0 for i in items)),
        "total_balance": round(sum(i.get("balance") or 0 for i in items)),
        "fetched_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC"),
    }
    _payouts_cache = result
    _payouts_ts = _time.monotonic()
    return result


# ══ Ручные статьи затрат (вкладка Тотал) ═══════════════════════════════════════

def _manual_costs_init():
    import db
    # автоинкремент по-разному: SERIAL в Postgres, AUTOINCREMENT в SQLite
    id_col = ("id SERIAL PRIMARY KEY" if db.IS_PG
              else "id INTEGER PRIMARY KEY AUTOINCREMENT")
    db.execute(f"CREATE TABLE IF NOT EXISTS manual_costs "
               f"({id_col}, mk TEXT, label TEXT, amount REAL)")


@router.get("/manual_costs")
async def get_manual_costs():
    """Ручные статьи затрат: {items: [{id, mk, label, amount}]}."""
    import db
    def _load():
        _manual_costs_init()
        return db.fetchall("SELECT id, mk, label, amount FROM manual_costs ORDER BY mk, id")
    rows = await asyncio.to_thread(_load)
    return {"items": [{"id": r[0], "mk": r[1], "label": r[2], "amount": float(r[3] or 0)}
                      for r in rows]}


@router.post("/manual_costs")
async def add_manual_cost(payload: dict):
    """Добавить статью на один или несколько месяцев:
    {mk: '2026-06', ...} или {mks: ['2026-05', '2026-06'], label, amount}."""
    mks = payload.get("mks") or ([payload.get("mk")] if payload.get("mk") else [])
    mks = [str(m).strip() for m in mks if m and len(str(m).strip()) == 7]
    label = str(payload.get("label") or "").strip()
    try:
        amount = float(payload.get("amount") or 0)
    except (ValueError, TypeError):
        amount = 0.0
    if not mks or not label or not amount:
        raise HTTPException(status_code=400, detail="Нужны месяцы (ГГГГ-ММ), название и сумма")
    import db
    def _save():
        _manual_costs_init()
        db.executemany("INSERT INTO manual_costs (mk, label, amount) VALUES (?,?,?)",
                       [(mk, label, amount) for mk in mks])
        return db.fetchall("SELECT MAX(id) FROM manual_costs")[0][0]
    new_id = await asyncio.to_thread(_save)
    return {"id": new_id, "added": len(mks)}


@router.delete("/manual_costs/{cost_id}")
async def delete_manual_cost(cost_id: int):
    import db
    await asyncio.to_thread(db.execute, "DELETE FROM manual_costs WHERE id = ?", (cost_id,))
    return {"ok": True}


@router.get("/ozon/reports")
async def get_ozon_reports():
    return {"reports": [], "message": "OZON финансовые отчёты будут добавлены позже"}


@router.get("/ym/reports")
async def get_ym_reports():
    return {"reports": [], "message": "YM финансовые отчёты будут добавлены позже"}
