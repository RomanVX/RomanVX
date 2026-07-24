"""Ущерб от пожаров на складах WB: сколько денег сгорело по себестоимости.

Пожары июля 2026 (Шушары, Электросталь, Котовск, Краснодар, Невинномысск).
Считаем по остаткам из отчёта платного хранения WB на дату пожара, умноженным
на себестоимость. Розница — справочно, для суммы претензии.

ВАЖНО по Электростали: часть остатков WB формально вернул в кабинет продавца,
но физически сгорело всё. Поэтому возвращённое НЕ вычитаем — считаем полный
объём, что и отражено в поле note.

Данные лежат в БД и правятся: строки можно догрузить файлом остатков по
складу, себестоимость подтягивается из cost_store.
"""
import asyncio
import json
import logging
import os
from datetime import datetime, timedelta

import db

_log = logging.getLogger("damage")

_SEED = os.path.join(os.path.dirname(__file__), "data", "damage_seed.json")

# Склады, по которым ведём учёт. Порядок = порядок вывода на экране.
WAREHOUSES = ["Шушары", "Электросталь", "Котовск", "Краснодар", "Невинномысск"]

# дата пожара → срез остатков берём за предыдущий день
FIRE_DATES = {"Краснодар": "2026-07-22", "Невинномысск": "2026-07-22",
              "Шушары": "2026-07-24",
              "Электросталь": "2026-07-18", "Котовск": "2026-07-18"}

_NOTES = {
    "Электросталь": "Остатки на 17.07 + принятая поставка 580 шт. "
                    "Возвращённое WB в кабинет НЕ вычитаем — физически "
                    "сгорело всё.",
    "Котовск": "Остатки из отчёта платного хранения, срез 18.07.2026.",
    "Шушары": "Склад СПб, Московское ш. 153 к. 2, пос. Шушары. "
              "Пожар 24.07.2026.",
    "Краснодар": "Пожар 22.07.2026.",
    "Невинномысск": "Пожар 22.07.2026.",
}

# в отчёте WB склад может называться иначе — синонимы для поиска строк
_WH_ALIASES = {
    "Шушары": ("шушары", "санкт-петербург", "спб"),
    "Невинномысск": ("невинномысск",),
    "Краснодар": ("краснодар",),
}


def _init() -> None:
    db.execute("""CREATE TABLE IF NOT EXISTS damage (
        id TEXT PRIMARY KEY, warehouse TEXT, sku TEXT, nm TEXT, name TEXT,
        qty INTEGER, cost REAL, retail REAL, source TEXT, note TEXT,
        updated TEXT)""")
    try:      # таблица могла быть создана до появления колонки источника
        db.execute("ALTER TABLE damage ADD COLUMN source TEXT")
    except Exception:
        pass


def _msk() -> str:
    return (datetime.utcnow() + timedelta(hours=3)).strftime("%Y-%m-%d %H:%M")


def _seed_if_empty() -> None:
    """Заливаем расчёты, уже сделанные по отчётам WB.

    Ключ включает источник: один артикул может гореть и в складских
    остатках, и в поставке, принятой перед пожаром (AL-01: 30 + 440 шт).
    Без этого вторая строка затиралась и сумма занижалась.
    """
    _init()
    try:
        rows = json.load(open(_SEED, encoding="utf-8"))
    except Exception as e:
        _log.warning("seed: %s", e)
        return
    seeded = {r["wh"] for r in rows}
    marks = ",".join("?" * len(seeded))
    have = db.fetchone(
        f"SELECT COUNT(*) FROM damage WHERE warehouse IN ({marks})",
        tuple(seeded))
    have_n = int((have or [0])[0] or 0)
    if have_n == len(rows):
        return
    if have_n:      # была залита урезанная версия — переписываем начисто
        _log.warning("damage: пересев (было %d строк, ожидается %d)",
                     have_n, len(rows))
        for wh in seeded:
            db.execute("DELETE FROM damage WHERE warehouse = ?", (wh,))
    now = _msk()
    for i, r in enumerate(rows):
        src = r.get("source") or ""
        db.execute(
            "INSERT INTO damage (id, warehouse, sku, nm, name, qty, cost, "
            "retail, source, note, updated) VALUES (?,?,?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT (id) DO NOTHING",
            (f"{r['wh']}|{r['sku']}|{i}", r["wh"], r["sku"], r.get("nm", ""),
             r.get("name", ""), int(r["qty"]), float(r.get("cost") or 0),
             float(r.get("retail") or 0), src, _NOTES.get(r["wh"], ""), now))
    _log.info("damage: залито %d позиций", len(rows))


def rows(warehouse: str = "") -> list[dict]:
    _seed_if_empty()
    where, params = [], []
    if warehouse:
        where.append("warehouse = ?"); params.append(warehouse)
    q = ("SELECT warehouse, sku, nm, name, qty, cost, retail, source, note, "
         "updated FROM damage"
         + (" WHERE " + " AND ".join(where) if where else "")
         + " ORDER BY warehouse, qty DESC")
    keys = ["warehouse", "sku", "nm", "name", "qty", "cost", "retail",
            "source", "note", "updated"]
    out = []
    for r in db.fetchall(q, tuple(params)):
        d = dict(zip(keys, r))
        d["cost_total"] = round((d["qty"] or 0) * (d["cost"] or 0))
        d["retail_total"] = round((d["qty"] or 0) * (d["retail"] or 0))
        out.append(d)
    return out


def summary() -> dict:
    """Свод по складам + итог. Склады без данных показываем пустыми строками —
    видно, что расчёт по ним ещё не сделан, а не что ущерба нет."""
    all_rows = rows()
    by_wh: dict = {}
    for r in all_rows:
        g = by_wh.setdefault(r["warehouse"], {
            "warehouse": r["warehouse"], "skus": 0, "qty": 0,
            "cost_total": 0, "retail_total": 0, "note": r.get("note") or "",
            "updated": r.get("updated")})
        g["skus"] += 1
        g["qty"] += r["qty"] or 0
        g["cost_total"] += r["cost_total"]
        g["retail_total"] += r["retail_total"]
    order = {w: i for i, w in enumerate(WAREHOUSES)}
    for w in WAREHOUSES:
        by_wh.setdefault(w, {"warehouse": w, "skus": 0, "qty": 0,
                             "cost_total": 0, "retail_total": 0,
                             "note": "расчёт ещё не загружен", "updated": None})
    warehouses = sorted(by_wh.values(),
                        key=lambda x: order.get(x["warehouse"], 99))
    no_cost = [r["sku"] for r in all_rows if not r.get("cost")]
    # контроль: сходится ли с исходными расчётами по файлам WB
    control = {}
    try:
        for r in json.load(open(_SEED, encoding="utf-8")):
            c = control.setdefault(r["wh"], {"qty": 0, "cost_total": 0})
            c["qty"] += int(r["qty"])
            c["cost_total"] += int(r["qty"]) * float(r.get("cost") or 0)
        for w in by_wh.values():
            exp = control.get(w["warehouse"])
            if exp:
                w["control_qty"] = exp["qty"]
                w["control_cost"] = round(exp["cost_total"])
                w["matches"] = (w["qty"] == exp["qty"]
                                and abs(w["cost_total"] - exp["cost_total"]) < 1)
    except Exception:
        pass
    return {
        "warehouses": warehouses,
        "total": {
            "skus": len({r["sku"] for r in all_rows}),
            "qty": sum(r["qty"] or 0 for r in all_rows),
            "cost_total": sum(r["cost_total"] for r in all_rows),
            "retail_total": sum(r["retail_total"] for r in all_rows),
        },
        "rows": all_rows,
        "missing_cost": sorted(set(no_cost)),
        "pending": [w["warehouse"] for w in warehouses if not w["qty"]],
    }


async def fetch_from_storage(warehouse: str, snap_date: str = "") -> dict:
    """Остатки склада на дату из отчёта платного хранения WB → в расчёт.

    Тот же метод, которым считали Электросталь и Котовск: строки отчёта
    date × warehouse × nmId с barcodesCount (штук на хранении). Берём срез
    за день до пожара, себестоимость и розницу — из своих справочников.
    """
    import wb_client
    _init()
    if not snap_date:
        fire = FIRE_DATES.get(warehouse)
        if not fire:
            return {"error": f"не знаю дату пожара склада {warehouse} — "
                             "передай дату среза явно"}
        d = datetime.strptime(fire, "%Y-%m-%d") - timedelta(days=1)
        snap_date = d.strftime("%Y-%m-%d")
    rows_raw = await wb_client.get_paid_storage(snap_date, snap_date)
    if not rows_raw:
        return {"error": f"отчёт хранения за {snap_date} пуст или не собрался "
                         "(лимит WB — 1 запрос в минуту, попробуй ещё раз)"}
    # какие склады вообще есть в отчёте — пригодится, если имя не совпало
    seen_wh = sorted({str(r.get("warehouse") or "") for r in rows_raw} - {""})
    want = warehouse.lower()
    match = [w for w in seen_wh if want in w.lower() or w.lower() in want]
    if not match:
        for al in _WH_ALIASES.get(warehouse, ()):
            match = [w for w in seen_wh if al in w.lower()]
            if match:
                break
    if not match:
        return {"error": f"склад «{warehouse}» не найден в отчёте за {snap_date}",
                "warehouses_in_report": seen_wh}
    if len(match) > 1:
        # у WB бывает несколько площадок с похожим именем — суммировать
        # их молча нельзя, в претензии нужен конкретный адрес
        return {"error": f"нашёл несколько складов: {', '.join(match)} — "
                         "уточни, какой именно (наш — Московское ш. 153 к. 2)",
                "warehouses_in_report": seen_wh}
    # агрегируем: nmId → штук (строк может быть несколько на баркоды)
    agg: dict = {}
    for r in rows_raw:
        if str(r.get("warehouse") or "") not in match:
            continue
        nm = str(r.get("nmId") or "")
        if not nm:
            continue
        g = agg.setdefault(nm, {"qty": 0,
                                "vendor": str(r.get("vendorCode") or "")})
        g["qty"] += int(r.get("barcodesCount") or 0)
    if not agg:
        return {"error": f"по складам {match} нет строк с товаром"}

    try:
        import cost_store
        costs = cost_store.get_costs() or {}
    except Exception:
        costs = {}
    retail_map: dict = {}
    try:      # розница — живые цены WB (до СПП)
        live = await asyncio.wait_for(wb_client.get_current_prices(), timeout=30)
        retail_map = {k.upper(): (v or {}).get("discounted") or 0
                      for k, v in (live or {}).items()}
    except Exception:
        pass
    try:
        import catalog as _cat
        nm_to_art = {str(k): v for k, v in
                     getattr(_cat, "WB_ID_TO_ART", {}).items()}
    except Exception:
        nm_to_art = {}

    db.execute("DELETE FROM damage WHERE warehouse = ?", (warehouse,))
    now = _msk()
    note = (f"Остатки из отчёта платного хранения, срез {snap_date} "
            f"(пожар {FIRE_DATES.get(warehouse, '—')}). "
            f"Склады отчёта: {', '.join(match)}.")
    added, qty_total = 0, 0
    for nm, g in agg.items():
        art = nm_to_art.get(nm) or g["vendor"] or nm
        try:
            import catalog as _cat2
            art = _cat2.canon(art)
        except Exception:
            pass
        cost = float(costs.get(art) or costs.get(str(art).upper()) or 0)
        retail = float(retail_map.get(str(art).upper()) or 0)
        db.execute(
            "INSERT INTO damage (id, warehouse, sku, nm, name, qty, cost, "
            "retail, source, note, updated) VALUES (?,?,?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT (id) DO UPDATE SET qty=excluded.qty, "
            "cost=excluded.cost, retail=excluded.retail, updated=excluded.updated",
            (f"{warehouse}|{art}|{nm}", warehouse, str(art), nm, "",
             g["qty"], cost, retail, f"хранение {snap_date}", note, now))
        added += 1
        qty_total += g["qty"]
    return {"warehouse": warehouse, "snap_date": snap_date,
            "skus": added, "qty": qty_total,
            "wb_warehouses": match}


async def upload(warehouse: str, raw: bytes) -> dict:
    """Файл остатков по складу (выгрузка «Платное хранение» WB или свой xlsx):
    берём артикул и количество, себестоимость подставляем свою."""
    import io
    import openpyxl
    _init()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        return {"error": f"не смог открыть файл: {str(e)[:200]}"}
    try:
        import cost_store
        costs = cost_store.get_costs() or {}
    except Exception:
        costs = {}
    try:
        import catalog as _cat
        nm_by_art = {v.upper(): k for k, v in
                     getattr(_cat, "WB_ID_TO_ART", {}).items()}
    except Exception:
        nm_by_art = {}

    added, skipped = 0, 0
    now = _msk()
    for ws in wb:
        hdr_i, hdr = -1, []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15,
                                             values_only=True)):
            cells = [str(c or "").strip().lower() for c in row]
            if any("артикул" in c for c in cells) and \
                    any(("кол" in c or "остат" in c or "шт" in c) for c in cells):
                hdr_i, hdr = i, cells
                break
        if hdr_i < 0:
            continue

        def col(*words, exclude=()):
            for j, h in enumerate(hdr):
                if any(w in h for w in words) and not any(x in h for x in exclude):
                    return j
            return -1

        c_art = col("артикул", exclude=("wb", "nmid"))
        c_qty = col("кол", "остат", "шт", exclude=("штрих",))
        c_name = col("назван", "предмет")
        c_cost = col("себес")
        c_ret = col("розниц", "цена")
        if c_art < 0 or c_qty < 0:
            continue
        for row in ws.iter_rows(min_row=hdr_i + 2, values_only=True):
            art = str(row[c_art] or "").strip()
            if not art or art.upper().startswith("ИТОГО"):
                continue
            try:
                qty = int(float(str(row[c_qty]).strip()))
            except (TypeError, ValueError):
                skipped += 1
                continue
            if qty <= 0:
                continue
            cost = 0.0
            if c_cost >= 0 and row[c_cost]:
                try:
                    cost = float(row[c_cost])
                except (TypeError, ValueError):
                    cost = 0.0
            if not cost:
                cost = float(costs.get(art) or costs.get(art.upper()) or 0)
            retail = 0.0
            if c_ret >= 0 and row[c_ret]:
                try:
                    retail = float(row[c_ret])
                except (TypeError, ValueError):
                    retail = 0.0
            name = str(row[c_name] or "")[:120] if c_name >= 0 else ""
            db.execute(
                "INSERT INTO damage (id, warehouse, sku, nm, name, qty, cost, "
                "retail, source, note, updated) VALUES (?,?,?,?,?,?,?,?,?,?,?) "
                "ON CONFLICT (id) DO UPDATE SET qty=excluded.qty, "
                "cost=excluded.cost, retail=excluded.retail, "
                "name=excluded.name, updated=excluded.updated",
                (f"{warehouse}|{art}|upload", warehouse, art,
                 nm_by_art.get(art.upper(), ""), name, qty, cost, retail,
                 "загружено файлом", _NOTES.get(warehouse, ""), now))
            added += 1
    if not added:
        return {"error": "не нашёл строк с артикулом и количеством — "
                         "проверь, что это выгрузка остатков"}
    return {"warehouse": warehouse, "added": added, "skipped": skipped}


def clear(warehouse: str) -> dict:
    _init()
    db.execute("DELETE FROM damage WHERE warehouse = ?", (warehouse,))
    return {"cleared": warehouse}


def burned_names() -> tuple:
    """Имена сгоревших складов — для фильтрации живых остатков.

    ТОЛЬКО точные имена, без городских синонимов: «спб» скрыл бы и живую
    Уткину Заводь. «СПБ Шушары» ловится по слову «шушары»."""
    return tuple(sorted(w.lower() for w in FIRE_DATES))


def is_burned(warehouse_name: str) -> bool:
    w = str(warehouse_name or "").lower()
    if not w:
        return False
    return any(b in w for b in burned_names())
